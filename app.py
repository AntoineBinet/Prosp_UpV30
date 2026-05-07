from __future__ import annotations

import json
import sqlite3
from io import BytesIO
from copy import copy
import datetime
import csv
import re
import sys
import unicodedata
import shutil
import zipfile
import threading
import time
import difflib
import math
from pathlib import Path
from typing import Any, Dict, List

from flask import Flask, jsonify, request, send_from_directory, send_file, redirect, session, g, Response, render_template, stream_with_context

# ReportLab pour génération PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
from markupsafe import escape as escape_html
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import secrets
import hmac
import base64
from services.dashboard_goals import build_goals_payload as _build_goals_payload, get_goals_config as _get_goals_config

import os
import uuid
import subprocess
import traceback
import hashlib
import urllib.error
import urllib.request
import logging
from logging.handlers import RotatingFileHandler

# Configuration centrale (Phase A1 modularisation — voir config.py).
from config import (
    APP_DIR,
    APP_VERSION,
    AVATARS_DIR,
    DATA_DIR,
    DB_PATH,
    INITIAL_JSON,
    OLLAMA_MODEL,
    OLLAMA_TIMEOUT,
    OLLAMA_URL,
    OUTLOOK_AVAILABLE,
    PHOTOS_DIR,
    SNAPSHOT_DIR,
    TAVILY_API_KEY,
    TAVILY_URL,
    TEMPLATE_PATH,
)
# Helpers DB / common / validation / auth (Phase A1 + A2 — voir utils/).
from utils.common import _now_iso, _parse_linkedin_name, _row_to_dict, _today_iso
from utils.db import _auth_conn, _conn, _conn_for_user, _user_db_path
from utils.validation import (
    _check_table_exists,
    _safe_execute_insert,
    _safe_execute_update,
    _safe_row_to_dict,
    _validate_optional_positive_int,
    _validate_positive_int,
)
from utils.snapshots import (
    _is_safe_snapshot_name,
    _snapshot_path,
    create_snapshot,
    list_snapshots,
    restore_snapshot,
)
from utils.files import (
    _MAGIC_BYTES,
    _UPLOAD_RULES,
    _attachment_dir,
    _extract_pdf_text,
    _generate_thumbnail,
    _sniff_mime,
    _thumb_dir,
    _validate_upload,
)
from utils.ai_helpers import (
    _AI_CONFIG_FILE,
    _build_web_enriched_prompt,
    _call_ai,
    _call_ai_provider,
    _call_ai_web,
    _call_ollama_direct,
    _call_tavily_search,
    _clear_ai_config_cache,
    _compute_semantic_similarity,
    _cosine_similarity,
    _embedding_mem_cache,
    _get_embedding_for_text,
    _get_text_embedding_simple,
    _load_ai_config,
    _save_ai_config,
    _stream_ai_sse,
    _stream_ai_web_sse,
    _stream_ollama_sse,
    _stream_tavily_ollama_sse,
)
from utils.candidates import (
    _build_candidate_descriptions,
    _generate_candidate_description_ai,
    _resolve_dc_pdf_path,
)
from utils.push import (
    _apply_call_note,
    _apply_candidates,
    _apply_salutation,
    _generate_eml_file,
    _personalize_html_body,
    _read_msg_body,
    _remove_signature,
    _resolve_dc_path,
    _save_to_outlook_drafts,
)
from utils.auth import (
    ROLE_LEVELS,
    _ALLOWED_ORIGINS,
    _b64url_decode,
    _b64url_encode,
    _candidate_owned,
    _check_login_rate_limit,
    _company_owned,
    _generate_access_token,
    _generate_refresh_token,
    _get_current_user,
    _get_user_prefix,
    _JWT_ACCESS_EXPIRY,
    _JWT_REFRESH_EXPIRY,
    _jwt_decode,
    _jwt_encode,
    _LOGIN_MAX_ATTEMPTS,
    _LOGIN_WINDOW_SECONDS,
    _login_attempts,
    _login_lock,
    _prospect_owned,
    _record_login_attempt,
    _require_same_origin,
    _uid,
    _verify_access_token,
    _verify_refresh_token,
    login_required,
    role_required,
    validate_payload,
)

# ═══════════════════════════════════════════════════════════════════
# v24.1: Structured logging with file rotation (24/7 production)
# ═══════════════════════════════════════════════════════════════════
_log_dir = APP_DIR / "logs"
_log_dir.mkdir(exist_ok=True)
_log_handler = RotatingFileHandler(
    str(_log_dir / "prospup.log"), maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"
)
_log_handler.setFormatter(logging.Formatter(
    "[%(asctime)s] %(levelname)s %(name)s: %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
))
_log_handler.setLevel(logging.INFO)
logging.getLogger().addHandler(_log_handler)
logging.getLogger().setLevel(logging.INFO)
logger = logging.getLogger("prospup")

logger.info("Outlook disponible (win32com): %s", OUTLOOK_AVAILABLE)

# IA (Ollama + Tavily + embeddings + streaming SSE) — voir utils/ai_helpers.py
# _parse_linkedin_name — voir utils/common.py

# Cache temporaire en mémoire pour les analyses RDV (clé: "{uid}_{prospect_id}")
# Utilisé à la place de session[] car la session n'est pas persistée dans les réponses SSE streaming
_rdv_analysis_cache: Dict[str, str] = {}

app = Flask(__name__, static_folder=str(APP_DIR / 'static'), static_url_path='/static', template_folder=str(APP_DIR / 'templates'))

# ═══════════════════════════════════════════════════════════════════
# Session & Auth configuration
# ═══════════════════════════════════════════════════════════════════
_secret_file = APP_DIR / ".secret_key"
if _secret_file.exists():
    app.secret_key = _secret_file.read_text().strip()
else:
    app.secret_key = secrets.token_hex(32)
    _secret_file.write_text(app.secret_key)

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True   # v23.4: requires HTTPS (Cloudflare Tunnel)
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(hours=8)  # v23.4: reduced from 30d for security
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # v32.1 : 500 MB pour les uploads audio
app.json.ensure_ascii = False  # v27.12: caractères Unicode non échappés dans les réponses JSON

# v22: Compute content hashes for static assets (auto cache busters)
_static_hashes: Dict[str, str] = {}


def _compute_static_hashes():
    static_dir = APP_DIR / 'static'
    for f in static_dir.rglob('*'):
        if f.is_file() and f.suffix in ('.css', '.js', '.png', '.ico', '.json'):
            h = hashlib.md5(f.read_bytes()).hexdigest()[:8]
            rel = str(f.relative_to(static_dir)).replace('\\', '/')
            _static_hashes[rel] = h


_compute_static_hashes()

# Helper function for Jinja2 templates to get static file hash
def _get_static_hash(static_path: str) -> str:
    """Get the hash for a static file path (e.g., 'css/style.css' -> 'a1b2c3d4')."""
    return _static_hashes.get(static_path, '')

# Register the helper in Jinja2
app.jinja_env.globals['static_hash'] = _get_static_hash

# Regex to match ?v=XXXX in /static/ paths
_CACHE_BUSTER_RE = re.compile(r'(/static/[^"\'?]+)\?v=\d+')

# Validation uploads / MIME / miniatures / extraction PDF — voir utils/files.py

@app.after_request
def _after_request(response):
    # ── Security headers (v23.4) ──
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    # HSTS: only when served behind Cloudflare Tunnel (HTTPS)
    if request.is_secure or request.headers.get('X-Forwarded-Proto') == 'https':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # CSP: restrictive but allows inline styles/scripts (needed for current architecture)
    # v32.26: jsdelivr ajouté à style-src (Leaflet CSS), tuiles OSM autorisées en img-src
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
        "img-src 'self' data: blob: https://*.tile.openstreetmap.org https://cdn.jsdelivr.net; "
        "connect-src 'self' https://api.tavily.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "frame-ancestors 'self'"
    )

    # ── CORS for mobile JWT auth (v24.0) ──
    if request.headers.get("Authorization", "").startswith("Bearer ") or request.method == "OPTIONS":
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"

    # Cache headers for API GET responses (30s private cache)
    if request.path.startswith('/api/') and request.method == 'GET':
        response.headers.setdefault('Cache-Control', 'private, max-age=30')
    # Auto cache busters: replace ?v=XXXX with ?v=<hash> in HTML responses
    if response.content_type and 'text/html' in response.content_type and _static_hashes:
        try:
            data = response.get_data(as_text=True)

            def _replace_hash(m):
                path = m.group(1)  # e.g. /static/css/style.css
                rel = path.lstrip('/static/')
                # Try to find the hash for this path
                h = _static_hashes.get(rel)
                if not h:
                    # Try without leading slash
                    rel2 = path.replace('/static/', '', 1)
                    h = _static_hashes.get(rel2)
                return f'{path}?v={h}' if h else m.group(0)

            data = _CACHE_BUSTER_RE.sub(_replace_hash, data)
            response.set_data(data)
        except Exception:
            pass  # Don't break pages if hash replacement fails
    return response


@app.before_request
def _require_auth():
    """Protect all routes except login, static, and favicon.
    Supports both session cookies (web) and JWT Bearer tokens (mobile v24.0).
    CSRF is enforced for cookie auth; JWT Bearer is inherently CSRF-safe."""
    # ── CORS preflight ──
    if request.method == "OPTIONS":
        return

    allowed = ('/login', '/v30/login', '/static/', '/favicon.ico', '/api/auth/', '/api/app-version', '/api/system/check-deployment', '/api/system/logs',
               '/api/deploy/health', '/api/deploy/pull-from-404', '/api/deploy/rollback',
               '/api/deploy/validation-status', '/api/deploy/confirm-validation',
               '/prospects/mode-prosp', '/api/mode-prosp/')
    if any(request.path.startswith(p) for p in allowed):
        return

    # ── JWT auth (mobile v24.0) ──
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:]
        result = _verify_access_token(token)
        if result == "expired":
            return jsonify(ok=False, error="token_expired"), 401
        if result is None:
            return jsonify(ok=False, error="invalid_token"), 401
        # Populate session-like context so _uid() and g.user work transparently
        g.jwt_user = result
        g.user = {"id": result["user_id"], "role": result["user_role"],
                   "display_name": result["user_name"], "username": result["user_name"]}
        session["user_id"] = result["user_id"]
        session["user_role"] = result["user_role"]
        session["user_name"] = result["user_name"]
        return  # JWT auth OK — skip session & CSRF checks

    # ── CSRF protection on mutations (cookie auth only, v23.4) ──
    if request.method in ('POST', 'PUT', 'DELETE') and request.path.startswith('/api/'):
        chk = _require_same_origin()
        if chk:
            return chk

    # ── Session auth (web, unchanged) ──
    if not session.get('user_id'):
        if request.path.startswith('/api/'):
            return jsonify(ok=False, error="Non authentifié"), 401
        return redirect('/login')
    g.user = _get_current_user()
    if not g.user:
        session.clear()
        return redirect('/login')


@app.route("/api/<path:path>", methods=["OPTIONS"])
def api_cors_preflight(path):
    """Handle CORS preflight requests for mobile app (v24.0)."""
    resp = app.make_default_options_response()
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return resp

@app.context_processor
def _inject_user():
    """Make current user available to all templates."""
    return {'current_user': getattr(g, 'user', None)}

@app.get("/favicon.ico")
def favicon():
    # Serve app icon (tab favicon)
    return send_from_directory(str(APP_DIR / "static"), "favicon.ico", mimetype="image/vnd.microsoft.icon")


# /api/deploy/pull-from-404 et /api/deploy/rollback — déplacés dans routes/deploy.py


@app.errorhandler(404)
def page_not_found(e):
    """Custom 404 page (v23.4)."""
    if request.path.startswith('/api/'):
        return jsonify(ok=False, error="Endpoint introuvable"), 404
    return send_from_directory(APP_DIR, "404.html"), 404


@app.errorhandler(400)
def bad_request(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Requête invalide', 'detail': str(e)}), 400
    return send_from_directory(APP_DIR, "400.html"), 400


@app.errorhandler(500)
def server_error(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Erreur serveur interne'}), 500
    return send_from_directory(APP_DIR, "500.html"), 500


# ═══════════════════════════════════════════════════════════════════
# Auth routes
# ═══════════════════════════════════════════════════════════════════

@app.get("/login")
def page_login():
    if session.get('user_id'):
        # v30 est l'interface par defaut depuis 30.1. L'opt-out client
        # (localStorage.prospup_ui_mode === 'v29') est gere par base.html
        # legacy qui redirige vers l'equivalent legacy si necessaire.
        return redirect('/v30/dashboard')
    return send_from_directory(APP_DIR, "login.html")

# Auth helpers (rate limit + JWT) — voir utils/auth.py
# /api/auth/* — déplacé dans routes/auth.py (Blueprint enregistré en bas de ce fichier)

# ═══════════════════════════════════════════════════════════════════
# User management (admin only)
# ═══════════════════════════════════════════════════════════════════

@app.get("/users")
@login_required
@role_required('admin')
def page_users():
    return redirect("/v30/users", code=302)

@app.get("/api/users")
@login_required
@role_required('admin')
def api_users_list():
    user = getattr(g, 'user', None) or _get_current_user()
    with _auth_conn() as conn:
        rows = conn.execute("SELECT id, username, display_name, role, is_active, createdAt, lastLoginAt FROM users ORDER BY id;").fetchall()
    is_admin = user and user.get('role') == 'admin'
    current_user_id = int(user["id"]) if user and user.get("id") is not None else None
    users_list = []
    for r in rows:
        u = dict(r)
        u['lastLoginAt'] = u.get('lastLoginAt') or ''
        users_list.append(u)
    return jsonify(ok=True, users=users_list, is_admin=is_admin, current_user_id=current_user_id)

@app.get("/api/users/for-push")
@login_required
def api_users_for_push():
    """Liste minimale des utilisateurs actifs pour la sélection de consultants dans le push modal.
    Accessible à tous les utilisateurs authentifiés (admin et editor)."""
    user = getattr(g, 'user', None) or _get_current_user()
    current_user_id = int(user["id"]) if user and user.get("id") is not None else None
    with _auth_conn() as conn:
        rows = conn.execute(
            "SELECT id, username, display_name FROM users WHERE is_active=1 ORDER BY display_name, username;"
        ).fetchall()
    users_list = [dict(r) for r in rows]
    return jsonify(ok=True, users=users_list, current_user_id=current_user_id)

@app.post("/api/users/save")
@login_required
@role_required('admin')
def api_users_save():
    payload, err = validate_payload({'username': str})
    if err:
        return err
    uid = payload.get("id")
    username = (payload.get("username") or "").strip().lower()
    display_name = (payload.get("display_name") or "").strip()
    role = payload.get("role", "editor")
    password = payload.get("password", "")
    is_active = 1 if payload.get("is_active", True) else 0

    if not username:
        return jsonify(ok=False, error="Username requis"), 400
    if role not in ROLE_LEVELS:
        return jsonify(ok=False, error="Rôle invalide"), 400

    new_user_id = None
    with _auth_conn() as conn:
        if uid:
            existing = conn.execute("SELECT * FROM users WHERE id=?;", (uid,)).fetchone()
            if not existing:
                return jsonify(ok=False, error="Utilisateur introuvable"), 404
            conn.execute("UPDATE users SET username=?, display_name=?, role=?, is_active=? WHERE id=?;",
                         (username, display_name, role, is_active, uid))
            if password and password.strip():
                conn.execute("UPDATE users SET password_hash=? WHERE id=?;",
                             (generate_password_hash(password), uid))
            return jsonify(ok=True, action="updated")
        else:
            if not password or len(password) < 8:
                return jsonify(ok=False, error="Mot de passe requis (min 8 caractères, avec au moins 1 chiffre et 1 lettre)"), 400
            if not any(c.isdigit() for c in password):
                return jsonify(ok=False, error="Le mot de passe doit contenir au moins un chiffre"), 400
            if not any(c.isalpha() for c in password):
                return jsonify(ok=False, error="Le mot de passe doit contenir au moins une lettre"), 400
            dup = conn.execute("SELECT id FROM users WHERE LOWER(username)=?;", (username,)).fetchone()
            if dup:
                return jsonify(ok=False, error="Username déjà pris"), 409
            now = datetime.datetime.now().isoformat(timespec="seconds")
            cur = conn.execute(
                "INSERT INTO users (username, display_name, password_hash, role, is_active, createdAt, must_change_password) VALUES (?, ?, ?, ?, ?, ?, 1);",
                (username, display_name, generate_password_hash(password), role, is_active, now)
            )
            new_user_id = cur.lastrowid

    if new_user_id:
        try:
            _init_user_db(new_user_id)
        except Exception as e:
            print(f"[WARN] Erreur creation DB utilisateur {new_user_id}: {e}")
        return jsonify(ok=True, action="created", user_id=new_user_id)

    return jsonify(ok=True, action="created")

@app.post("/api/users/delete")
@login_required
@role_required('admin')
def api_users_delete():
    """Supprime un utilisateur et toutes ses données, en gérant correctement les données collaboratives."""
    payload = request.get_json(force=True, silent=True) or {}
    uid = payload.get("id")
    if not uid:
        return jsonify(ok=False, error="ID requis"), 400
    if uid == session.get('user_id'):
        return jsonify(ok=False, error="Impossible de supprimer votre propre compte"), 400
    
    try:
        uid = int(uid)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="ID invalide"), 400
    
    with _auth_conn() as conn:
        # Vérifier que l'utilisateur existe
        user = conn.execute("SELECT id, username, display_name FROM users WHERE id=?;", (uid,)).fetchone()
        if not user:
            return jsonify(ok=False, error="Utilisateur introuvable"), 404
        
        username = user["username"] or user["display_name"] or f"user_{uid}"
        
        # 1. Nettoyer shared_companies
        # - Supprimer les partages où from_user_id = uid (partages envoyés)
        #   Les données restent dans la DB du collaborateur (to_user_id)
        sent_shares = conn.execute(
            "SELECT id, company_id, to_user_id FROM shared_companies WHERE from_user_id=?;",
            (uid,)
        ).fetchall()
        conn.execute("DELETE FROM shared_companies WHERE from_user_id=?;", (uid,))
        
        # - Supprimer les partages où to_user_id = uid (partages reçus)
        #   Supprimer aussi les données copiées dans la DB de l'utilisateur supprimé
        received_shares = conn.execute(
            "SELECT id, company_id, from_user_id FROM shared_companies WHERE to_user_id=?;",
            (uid,)
        ).fetchall()
        conn.execute("DELETE FROM shared_companies WHERE to_user_id=?;", (uid,))
        
        # Supprimer les entreprises et prospects partagés de la DB de l'utilisateur supprimé
        # (ces données ont été copiées via _sync_shared_company_to_collaborator)
        if received_shares:
            user_db_path = _user_db_path(uid)
            if user_db_path.exists():
                try:
                    user_conn = sqlite3.connect(user_db_path)
                    user_conn.row_factory = sqlite3.Row
                    user_conn.execute("PRAGMA foreign_keys = OFF;")
                    try:
                        # Supprimer les prospects des entreprises partagées
                        company_ids = [s["company_id"] for s in received_shares]
                        if company_ids:
                            placeholders = ','.join(['?'] * len(company_ids))
                            user_conn.execute(
                                f"DELETE FROM prospects WHERE company_id IN ({placeholders}) AND owner_id=?;",
                                (*company_ids, uid)
                            )
                            # Supprimer les entreprises partagées
                            user_conn.execute(
                                f"DELETE FROM companies WHERE id IN ({placeholders}) AND owner_id=?;",
                                (*company_ids, uid)
                            )
                        user_conn.commit()
                    finally:
                        user_conn.execute("PRAGMA foreign_keys = ON;")
                        user_conn.close()
                except Exception as e:
                    logger.warning(f"Erreur nettoyage données partagées user {uid}: {e}")
        
        # 2. Nettoyer audit_log
        conn.execute("DELETE FROM audit_log WHERE user_id=?;", (uid,))
        
        # 3. Nettoyer refresh_tokens (CASCADE devrait le faire, mais on le fait explicitement)
        conn.execute("DELETE FROM refresh_tokens WHERE user_id=?;", (uid,))
        
        # 4. Supprimer l'utilisateur (CASCADE supprimera aussi refresh_tokens)
        conn.execute("DELETE FROM users WHERE id=?;", (uid,))
        
        logger.info(f"Utilisateur {uid} ({username}) supprimé : {len(sent_shares)} partages envoyés, {len(received_shares)} partages reçus nettoyés")
    
    # 5. Supprimer le dossier utilisateur avec retry
    user_dir = DATA_DIR / f"user_{uid}"
    if user_dir.exists():
        max_retries = 3
        retry_delay = 1.0  # secondes
        for attempt in range(max_retries):
            try:
                shutil.rmtree(user_dir)
                logger.info(f"DB utilisateur supprimée : {user_dir}")
                break
            except (OSError, PermissionError) as e:
                if attempt < max_retries - 1:
                    logger.warning(f"Tentative {attempt + 1}/{max_retries} échouée pour {user_dir}, retry dans {retry_delay}s...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    logger.warning(f"Impossible de supprimer {user_dir} après {max_retries} tentatives: {e}")
                    # Le dossier sera nettoyé au prochain redémarrage par _migrate_all_user_dbs()

    return jsonify(ok=True, message=f"Utilisateur {username} supprimé avec succès")

# Admin: View another user's data (read-only)
@app.get("/api/users/<int:target_user_id>/data")
@login_required
@role_required('admin')
def api_users_view_data(target_user_id):
    """Admin can view another user's prospect/candidate data in read-only mode."""
    with _auth_conn() as conn:
        user = conn.execute("SELECT id, username, display_name, role FROM users WHERE id=?;", (target_user_id,)).fetchone()
        if not user:
            return jsonify(ok=False, error="Utilisateur introuvable"), 404

    with _conn_for_user(target_user_id) as uconn:
        prospects = uconn.execute("SELECT COUNT(*) AS n FROM prospects;").fetchone()["n"]
        candidates = uconn.execute("SELECT COUNT(*) AS n FROM candidates;").fetchone()["n"]
    has_own_db = (DATA_DIR / f"user_{target_user_id}" / "prospects.db").exists()
    return jsonify(ok=True, user=dict(user), stats={"prospects": prospects, "candidates": candidates}, has_own_db=has_own_db)


@app.post("/api/admin/reassign-ownership")
@login_required
@role_required('admin')
def api_admin_reassign_ownership():
    """Admin endpoint: reassign prospects/companies ownership from one user to another."""
    chk = _require_same_origin()
    if chk:
        return chk

    payload = request.get_json(force=True, silent=True) or {}
    from_user_id = payload.get("from_user_id")
    to_user_id = payload.get("to_user_id")
    try:
        from_user_id = int(from_user_id)
        to_user_id = int(to_user_id)
    except Exception:
        return jsonify(ok=False, error="from_user_id et to_user_id requis"), 400

    if from_user_id == to_user_id:
        return jsonify(ok=False, error="Les utilisateurs source et destination doivent être différents"), 400

    src_has_own_db = (DATA_DIR / f"user_{from_user_id}" / "prospects.db").exists()
    dst_has_own_db = (DATA_DIR / f"user_{to_user_id}" / "prospects.db").exists()
    if src_has_own_db or dst_has_own_db:
        return jsonify(ok=False, error="Réattribution non supportée entre utilisateurs ayant des bases de données séparées. Contactez l'administrateur."), 400

    with _auth_conn() as conn:
        src = conn.execute(
            "SELECT id, username, display_name FROM users WHERE id=?;",
            (from_user_id,),
        ).fetchone()
        dst = conn.execute(
            "SELECT id, username, display_name FROM users WHERE id=?;",
            (to_user_id,),
        ).fetchone()
        if not src:
            return jsonify(ok=False, error="Utilisateur source introuvable"), 404
        if not dst:
            return jsonify(ok=False, error="Utilisateur destination introuvable"), 404

    with _conn() as conn:
        cur = conn.cursor()
        cur.execute("BEGIN;")
        try:
            prospects_n = int(cur.execute(
                "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=?;",
                (from_user_id,),
            ).fetchone()["n"])
            companies_n = int(cur.execute(
                "SELECT COUNT(*) AS n FROM companies WHERE owner_id=?;",
                (from_user_id,),
            ).fetchone()["n"])

            cur.execute(
                "UPDATE prospects SET owner_id=? WHERE owner_id=?;",
                (to_user_id, from_user_id),
            )
            cur.execute(
                "UPDATE companies SET owner_id=? WHERE owner_id=?;",
                (to_user_id, from_user_id),
            )
            cur.execute("COMMIT;")
        except Exception:
            cur.execute("ROLLBACK;")
            raise

    return jsonify(
        ok=True,
        moved={"prospects": prospects_n, "companies": companies_n},
        from_user={"id": int(src["id"]), "username": src["username"], "display_name": src["display_name"]},
        to_user={"id": int(dst["id"]), "username": dst["username"], "display_name": dst["display_name"]},
    )


def init_db() -> None:
    SNAPSHOT_DIR.mkdir(exist_ok=True)
    # Créer le dossier pour les dossiers de compétences
    (APP_DIR / "dossiers_competence").mkdir(exist_ok=True)

    with _conn() as conn:
        conn.executescript(
            '''
            CREATE TABLE IF NOT EXISTS companies (
                id        INTEGER PRIMARY KEY,
                groupe    TEXT NOT NULL,
                site      TEXT NOT NULL,
                phone     TEXT,
                notes     TEXT,
                tags      TEXT
            );

            CREATE TABLE IF NOT EXISTS prospects (
                id            INTEGER PRIMARY KEY,
                name          TEXT NOT NULL,
                company_id    INTEGER NOT NULL,
                fonction      TEXT,
                telephone     TEXT,
                email         TEXT,
                linkedin      TEXT,
                pertinence    TEXT,
                statut        TEXT,
                lastContact   TEXT,
                nextFollowUp  TEXT,
                priority      INTEGER,
                notes         TEXT,
                callNotes     TEXT,
                pushEmailSentAt TEXT,
                tags          TEXT,
                template_id   INTEGER,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE RESTRICT
            );

            CREATE TABLE IF NOT EXISTS candidates (
                id        INTEGER PRIMARY KEY,
                name      TEXT NOT NULL,
                role      TEXT,
                location  TEXT,
                seniority TEXT,
                tech      TEXT,
                linkedin  TEXT,
                source    TEXT,
                status    TEXT,
                notes     TEXT,
                createdAt TEXT,
                updatedAt TEXT
            );

            CREATE TABLE IF NOT EXISTS push_logs (
                id            INTEGER PRIMARY KEY,
                prospect_id   INTEGER NOT NULL,
                sentAt        TEXT NOT NULL,
                channel       TEXT,
                to_email      TEXT,
                subject       TEXT,
                body          TEXT,
                template_id   INTEGER,
                template_name TEXT,
                createdAt     TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS call_logs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL,
                owner_id    INTEGER NOT NULL,
                date        TEXT NOT NULL,
                called_at   TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_call_logs_owner_date ON call_logs(owner_id, date);
            CREATE INDEX IF NOT EXISTS idx_call_logs_prospect ON call_logs(prospect_id);

            CREATE TABLE IF NOT EXISTS templates (
                id         INTEGER PRIMARY KEY,
                name       TEXT NOT NULL,
                subject    TEXT,
                body       TEXT,
                is_default INTEGER DEFAULT 0,
                createdAt  TEXT,
                updatedAt  TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_push_logs_prospect_id ON push_logs(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_push_logs_sentAt ON push_logs(sentAt);
            CREATE INDEX IF NOT EXISTS idx_templates_default ON templates(is_default);

            CREATE TABLE IF NOT EXISTS push_variants (
                id            INTEGER PRIMARY KEY,
                push_log_id   INTEGER NOT NULL,
                variant_id    TEXT NOT NULL,
                subject       TEXT,
                body          TEXT,
                sent_at       TEXT,
                opened_at     TEXT,
                clicked_at    TEXT,
                replied_at    TEXT,
                createdAt     TEXT NOT NULL,
                FOREIGN KEY(push_log_id) REFERENCES push_logs(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_push_variants_push_log_id ON push_variants(push_log_id);
            CREATE INDEX IF NOT EXISTS idx_push_variants_variant_id ON push_variants(variant_id);

CREATE TABLE IF NOT EXISTS saved_views (
    id        INTEGER PRIMARY KEY,
    page      TEXT NOT NULL,
    name      TEXT NOT NULL,
    state     TEXT NOT NULL,
    createdAt TEXT,
    updatedAt TEXT
);
CREATE INDEX IF NOT EXISTS idx_saved_views_page ON saved_views(page);

CREATE TABLE IF NOT EXISTS opportunities (
    id             INTEGER PRIMARY KEY,
    company_id      INTEGER NOT NULL,
    title          TEXT NOT NULL,
    stage          TEXT NOT NULL,
    candidate_name TEXT,
    candidate_link TEXT,
    amount         REAL,
    notes          TEXT,
    createdAt      TEXT,
    updatedAt      TEXT,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_opportunities_company ON opportunities(company_id);

CREATE TABLE IF NOT EXISTS company_events (
    id        INTEGER PRIMARY KEY,
    company_id INTEGER NOT NULL,
    date      TEXT NOT NULL,
    type      TEXT,
    title     TEXT,
    content   TEXT,
    meta      TEXT,
    createdAt TEXT,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_company_events_company ON company_events(company_id);
CREATE INDEX IF NOT EXISTS idx_company_events_date ON company_events(date);

CREATE TABLE IF NOT EXISTS prospect_events (
    id         INTEGER PRIMARY KEY,
    prospect_id INTEGER NOT NULL,
    date       TEXT NOT NULL,
    type       TEXT,
    title      TEXT,
    content    TEXT,
    meta       TEXT,
    createdAt  TEXT,
    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_prospect_events_prospect ON prospect_events(prospect_id);
CREATE INDEX IF NOT EXISTS idx_prospect_events_date ON prospect_events(date);
CREATE UNIQUE INDEX IF NOT EXISTS idx_prospect_events_unique ON prospect_events(prospect_id, type, date);

CREATE TABLE IF NOT EXISTS prospect_attachments (
    id            INTEGER PRIMARY KEY,
    prospect_id   INTEGER NOT NULL,
    owner_id      INTEGER NOT NULL,
    filename      TEXT NOT NULL,
    original_name TEXT NOT NULL,
    size          INTEGER,
    mime_type     TEXT,
    description   TEXT,
    meeting_id    INTEGER,
    tags          TEXT,
    thumbnail     TEXT,
    extracted_text TEXT,
    title         TEXT,
    createdAt     TEXT NOT NULL,
    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_prospect_attachments_prospect ON prospect_attachments(prospect_id);
CREATE INDEX IF NOT EXISTS idx_prospect_attachments_owner ON prospect_attachments(owner_id);

CREATE TABLE IF NOT EXISTS prospect_summaries (
    prospect_id INTEGER PRIMARY KEY,
    owner_id    INTEGER NOT NULL,
    summary     TEXT,
    generatedAt TEXT,
    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS candidate_events (
    id           INTEGER PRIMARY KEY,
    candidate_id INTEGER NOT NULL,
    date         TEXT NOT NULL,
    type         TEXT,
    title        TEXT,
    content      TEXT,
    meta         TEXT,
    createdAt    TEXT,
    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_candidate_events_candidate ON candidate_events(candidate_id);
CREATE INDEX IF NOT EXISTS idx_candidate_events_date ON candidate_events(date);
CREATE UNIQUE INDEX IF NOT EXISTS idx_candidate_events_unique ON candidate_events(candidate_id, type, date);

CREATE TABLE IF NOT EXISTS push_categories (
    id             INTEGER PRIMARY KEY,
    name           TEXT NOT NULL,
    keywords       TEXT,
    auto_detected  INTEGER DEFAULT 0,
    owner_id       INTEGER,
    candidate1_id  INTEGER,
    candidate2_id  INTEGER,
    no_candidates  INTEGER DEFAULT 0,
    createdAt      TEXT,
    updatedAt      TEXT,
    UNIQUE(name, owner_id)
);
CREATE INDEX IF NOT EXISTS idx_push_categories_name ON push_categories(name);

CREATE TABLE IF NOT EXISTS rdv_checklists (
    id          INTEGER PRIMARY KEY,
    prospect_id INTEGER NOT NULL UNIQUE,
    data        TEXT,
    updatedAt   TEXT,
    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_rdv_checklists_prospect ON rdv_checklists(prospect_id);

-- Candidate EC1 checklist (v15.1)
CREATE TABLE IF NOT EXISTS candidate_ec1_checklists (
    id           INTEGER PRIMARY KEY,
    candidate_id INTEGER NOT NULL UNIQUE,
    interviewAt  TEXT,
    data         TEXT,
    updatedAt    TEXT,
    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_candidate_ec1_candidate ON candidate_ec1_checklists(candidate_id);
CREATE INDEX IF NOT EXISTS idx_candidate_ec1_interviewAt ON candidate_ec1_checklists(interviewAt);

-- Candidate tabs (EC1 + note libre, v25) — onglets fiche candidat
CREATE TABLE IF NOT EXISTS candidate_tabs (
    id           INTEGER PRIMARY KEY,
    candidate_id INTEGER NOT NULL,
    sort_order   INTEGER NOT NULL DEFAULT 0,
    type         TEXT NOT NULL,
    title        TEXT NOT NULL,
    payload      TEXT,
    updated_at   TEXT,
    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_candidate_tabs_candidate ON candidate_tabs(candidate_id);
CREATE INDEX IF NOT EXISTS idx_candidate_tabs_sort ON candidate_tabs(candidate_id, sort_order);

-- Candidate experiences (v26.6: enrichissement IA structuré)
CREATE TABLE IF NOT EXISTS candidate_experiences (
    id           INTEGER PRIMARY KEY,
    candidate_id INTEGER NOT NULL,
    company_name TEXT NOT NULL,
    role         TEXT,
    start_date   TEXT,
    end_date     TEXT,
    description  TEXT,
    technologies TEXT,
    owner_id     INTEGER,
    createdAt    TEXT,
    updatedAt    TEXT,
    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_candidate_experiences_candidate ON candidate_experiences(candidate_id);
CREATE INDEX IF NOT EXISTS idx_candidate_experiences_owner ON candidate_experiences(owner_id);

-- Candidate educations (v26.6: enrichissement IA structuré)
CREATE TABLE IF NOT EXISTS candidate_educations (
    id            INTEGER PRIMARY KEY,
    candidate_id  INTEGER NOT NULL,
    degree        TEXT,
    school        TEXT NOT NULL,
    year          TEXT,
    specialization TEXT,
    owner_id      INTEGER,
    createdAt     TEXT,
    updatedAt     TEXT,
    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_candidate_educations_candidate ON candidate_educations(candidate_id);
CREATE INDEX IF NOT EXISTS idx_candidate_educations_owner ON candidate_educations(owner_id);

-- Candidate certifications (v26.6: enrichissement IA structuré)
CREATE TABLE IF NOT EXISTS candidate_certifications (
    id            INTEGER PRIMARY KEY,
    candidate_id  INTEGER NOT NULL,
    name          TEXT NOT NULL,
    issuer        TEXT,
    obtained_date TEXT,
    expiry_date   TEXT,
    owner_id      INTEGER,
    createdAt     TEXT,
    updatedAt     TEXT,
    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_candidate_certifications_candidate ON candidate_certifications(candidate_id);
CREATE INDEX IF NOT EXISTS idx_candidate_certifications_owner ON candidate_certifications(owner_id);

CREATE TABLE IF NOT EXISTS tasks (
    id          INTEGER PRIMARY KEY,
    title       TEXT NOT NULL,
    comment     TEXT,
    due_date    TEXT,
    status      TEXT NOT NULL DEFAULT 'pending',
    linked_ids  TEXT,
    createdAt   TEXT,
    updatedAt   TEXT
);
CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);

-- Embeddings cache (Phase 1: matching sémantique)
CREATE TABLE IF NOT EXISTS embeddings_cache (
    id          INTEGER PRIMARY KEY,
    entity_type TEXT NOT NULL,  -- 'prospect', 'candidate', 'tag', 'metier'
    entity_id   INTEGER,        -- ID de l'entité (prospect_id, candidate_id, ou NULL pour tag/metier)
    text_key    TEXT NOT NULL,  -- Texte ou tag pour lequel on a l'embedding
    embedding   TEXT NOT NULL,  -- JSON array de floats
    created_at  TEXT DEFAULT (datetime('now')),
    UNIQUE(entity_type, entity_id, text_key)
);
CREATE INDEX IF NOT EXISTS idx_embeddings_lookup ON embeddings_cache(entity_type, text_key);
CREATE INDEX IF NOT EXISTS idx_tasks_due_date ON tasks(due_date);

CREATE TABLE IF NOT EXISTS task_rules (
    id            INTEGER PRIMARY KEY,
    name          TEXT NOT NULL,
    trigger_type  TEXT NOT NULL,
    conditions    TEXT NOT NULL,
    template_title TEXT NOT NULL,
    template_comment TEXT,
    priority      INTEGER DEFAULT 2,
    enabled       INTEGER DEFAULT 1,
    owner_id      INTEGER,
    createdAt     TEXT,
    updatedAt     TEXT
);
CREATE INDEX IF NOT EXISTS idx_task_rules_trigger ON task_rules(trigger_type, enabled);
CREATE INDEX IF NOT EXISTS idx_task_rules_owner ON task_rules(owner_id);

-- v23.5: search performance indexes (columns that exist in CREATE TABLE)
CREATE INDEX IF NOT EXISTS idx_prospects_name ON prospects(name);
CREATE INDEX IF NOT EXISTS idx_prospects_email ON prospects(email);
CREATE INDEX IF NOT EXISTS idx_companies_groupe ON companies(groupe);
CREATE INDEX IF NOT EXISTS idx_push_logs_sentAt ON push_logs(sentAt);

-- v23.5: audit trail table
CREATE TABLE IF NOT EXISTS audit_log (
    id        INTEGER PRIMARY KEY,
    user_id   INTEGER NOT NULL,
    action    TEXT NOT NULL,
    entity    TEXT NOT NULL,
    entity_id INTEGER,
    old_value TEXT,
    new_value TEXT,
    ip        TEXT,
    createdAt TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_audit_log_user ON audit_log(user_id);
CREATE INDEX IF NOT EXISTS idx_audit_log_entity ON audit_log(entity, entity_id);
CREATE INDEX IF NOT EXISTS idx_audit_log_date ON audit_log(createdAt);

-- v27.2: Assistant virtuel — historique des conversations
CREATE TABLE IF NOT EXISTS assistant_history (
    id        INTEGER PRIMARY KEY,
    user_id   INTEGER NOT NULL,
    session_id TEXT,
    role      TEXT NOT NULL,
    content   TEXT NOT NULL,
    metadata  TEXT,
    createdAt TEXT NOT NULL DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_assistant_history_user ON assistant_history(user_id);
CREATE INDEX IF NOT EXISTS idx_assistant_history_session ON assistant_history(session_id);
CREATE INDEX IF NOT EXISTS idx_assistant_history_date ON assistant_history(createdAt);

-- v27.10: journal d'activité multi-utilisateurs
CREATE TABLE IF NOT EXISTS activity_logs (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id      INTEGER NOT NULL,
    username     TEXT NOT NULL,
    action       TEXT NOT NULL,
    entity_type  TEXT,
    entity_id    INTEGER,
    entity_label TEXT,
    details      TEXT,
    ip_address   TEXT,
    created_at   DATETIME DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_activity_logs_user   ON activity_logs(user_id);
CREATE INDEX IF NOT EXISTS idx_activity_logs_action ON activity_logs(action);
CREATE INDEX IF NOT EXISTS idx_activity_logs_date   ON activity_logs(created_at);
'''
        )

        # --- Migrations légères ---
        def _add_col(table: str, col: str, ddl: str):
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl};")

        # Companies (v6) – champs enrichis pour la fiche entreprise
        ccols = [r["name"] for r in conn.execute("PRAGMA table_info(companies);").fetchall()]
        if "website" not in ccols:
            _add_col("companies", "website", "TEXT")
        if "linkedin" not in ccols:
            _add_col("companies", "linkedin", "TEXT")
        if "industry" not in ccols:
            _add_col("companies", "industry", "TEXT")
        if "size" not in ccols:
            _add_col("companies", "size", "TEXT")
        if "address" not in ccols:
            _add_col("companies", "address", "TEXT")
        if "city" not in ccols:
            _add_col("companies", "city", "TEXT")
        if "country" not in ccols:
            _add_col("companies", "country", "TEXT")
        if "stack" not in ccols:
            _add_col("companies", "stack", "TEXT")
        if "pain_points" not in ccols:
            _add_col("companies", "pain_points", "TEXT")
        if "budget" not in ccols:
            _add_col("companies", "budget", "TEXT")
        if "urgency" not in ccols:
            _add_col("companies", "urgency", "TEXT")
        if "owner_id" not in ccols:
            _add_col("companies", "owner_id", "INTEGER")

        # Prospects (v4→v6)
        cols = [r["name"] for r in conn.execute("PRAGMA table_info(prospects);").fetchall()]

        if "nextFollowUp" not in cols:
            _add_col("prospects", "nextFollowUp", "TEXT")
        if "priority" not in cols:
            _add_col("prospects", "priority", "INTEGER")
        if "pushEmailSentAt" not in cols:
            _add_col("prospects", "pushEmailSentAt", "TEXT")
        if "tags" not in cols:
            _add_col("prospects", "tags", "TEXT")
        if "template_id" not in cols:
            _add_col("prospects", "template_id", "INTEGER")
        if "nextAction" not in cols:
            _add_col("prospects", "nextAction", "TEXT")
        if "pushLinkedInSentAt" not in cols:
            _add_col("prospects", "pushLinkedInSentAt", "TEXT")
        if "photo_url" not in cols:
            _add_col("prospects", "photo_url", "TEXT")
        if "push_category_id" not in cols:
            _add_col("prospects", "push_category_id", "INTEGER")
        if "fixedMetier" not in cols:
            _add_col("prospects", "fixedMetier", "TEXT")
        if "rdvDate" not in cols:
            _add_col("prospects", "rdvDate", "TEXT")
        # Migration: renommer is_contact en is_archived
        if "is_contact" in cols and "is_archived" not in cols:
            conn.execute("ALTER TABLE prospects ADD COLUMN is_archived INTEGER")
            conn.execute("UPDATE prospects SET is_archived = is_contact")
        elif "is_archived" not in cols:
            _add_col("prospects", "is_archived", "INTEGER")
        # Cas où les deux colonnes coexistent (is_archived ajouté vide avant la copie)
        if "is_contact" in cols and "is_archived" in cols:
            conn.execute("UPDATE prospects SET is_archived = is_contact WHERE is_contact = 1 AND (is_archived IS NULL OR is_archived = 0)")
        if "owner_id" not in cols:
            _add_col("prospects", "owner_id", "INTEGER")

        # Custom metiers (user-added tags/specialties)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS custom_metiers (
                id        INTEGER PRIMARY KEY,
                type      TEXT NOT NULL,
                category  TEXT NOT NULL,
                specialty TEXT,
                tech_group TEXT,
                value     TEXT NOT NULL,
                createdAt TEXT
            );
        ''')

        lcols = [r["name"] for r in conn.execute("PRAGMA table_info(push_logs);").fetchall()]
        if "template_id" not in lcols:
            _add_col("push_logs", "template_id", "INTEGER")
        if "template_name" not in lcols:
            _add_col("push_logs", "template_name", "TEXT")
        # v25.3: Traçabilité candidats et consultants dans push_logs
        if "candidate_id1" not in lcols:
            _add_col("push_logs", "candidate_id1", "INTEGER")
        if "candidate_id2" not in lcols:
            _add_col("push_logs", "candidate_id2", "INTEGER")
        if "consultant1_id" not in lcols:
            _add_col("push_logs", "consultant1_id", "INTEGER")
        if "consultant2_id" not in lcols:
            _add_col("push_logs", "consultant2_id", "INTEGER")
        # v26.6: Optimisation mailing (timing et A/B testing)
        if "sent_at_hour" not in lcols:
            _add_col("push_logs", "sent_at_hour", "INTEGER")
        if "sent_at_day_of_week" not in lcols:
            _add_col("push_logs", "sent_at_day_of_week", "INTEGER")
        if "variant_id" not in lcols:
            _add_col("push_logs", "variant_id", "TEXT")
        if "opened_at" not in lcols:
            _add_col("push_logs", "opened_at", "TEXT")
        if "clicked_at" not in lcols:
            _add_col("push_logs", "clicked_at", "TEXT")
        if "replied_at" not in lcols:
            _add_col("push_logs", "replied_at", "TEXT")
        if "tracking_pixel_id" not in lcols:
            _add_col("push_logs", "tracking_pixel_id", "TEXT")

        cand_cols = [r["name"] for r in conn.execute("PRAGMA table_info(candidates);").fetchall()]
        # Links & matching (v5.1+)
        if "onenote_url" not in cand_cols:
            _add_col("candidates", "onenote_url", "TEXT")
        if "vsa_url" not in cand_cols:
            _add_col("candidates", "vsa_url", "TEXT")
        # JSON list of skills (e.g. ["c++","rtos"]) stored as TEXT
        if "skills" not in cand_cols:
            _add_col("candidates", "skills", "TEXT")
        # JSON list of company ids (e.g. [12, 42]) stored as TEXT
        if "company_ids" not in cand_cols:
            _add_col("candidates", "company_ids", "TEXT")
        if "is_archived" not in cand_cols:
            _add_col("candidates", "is_archived", "INTEGER")
        # v11: years of experience (numeric) replaces seniority text
        if "years_experience" not in cand_cols:
            _add_col("candidates", "years_experience", "INTEGER")
        if "sector" not in cand_cols:
            _add_col("candidates", "sector", "TEXT")
        if "phone" not in cand_cols:
            _add_col("candidates", "phone", "TEXT")
        if "email" not in cand_cols:
            _add_col("candidates", "email", "TEXT")
        if "dossier_competence_pdf" not in cand_cols:
            _add_col("candidates", "dossier_competence_pdf", "TEXT")
        if "owner_id" not in cand_cols:
            _add_col("candidates", "owner_id", "INTEGER")
        # v27.x PARTIE 3: champs pour extraction DC + génération push .msg
        if "prenom" not in cand_cols:
            _add_col("candidates", "prenom", "TEXT")
        if "titre" not in cand_cols:
            _add_col("candidates", "titre", "TEXT")
        if "annees_experience" not in cand_cols:
            _add_col("candidates", "annees_experience", "INTEGER")
        if "domaine_principal" not in cand_cols:
            _add_col("candidates", "domaine_principal", "TEXT")
        # v27.4: description push IA (cache Ollama)
        if "description_push" not in cand_cols:
            _add_col("candidates", "description_push", "TEXT")
        # v28.1: champs fiche entretien candidat
        for _col, _ddl in [
            ("disponibilite", "TEXT"),
            ("mobilite", "TEXT"),
            ("permis_conduire", "INTEGER"),
            ("vehicule", "INTEGER"),
            ("permis_travail", "TEXT"),
            ("fonctions_recherchees", "TEXT"),
            ("motif_recherche", "TEXT"),
            ("avancement_recherches", "TEXT"),
            ("remuneration_actuelle", "TEXT"),
            ("pretentions_salariales", "TEXT"),
            ("propal_a", "TEXT"),
            ("eval_technique", "TEXT"),
            ("eval_personnalite", "TEXT"),
            ("eval_communication", "TEXT"),
            ("langues", "TEXT"),
            ("references_candidat", "TEXT"),
            ("avis_perso", "TEXT"),
        ]:
            if _col not in cand_cols:
                _add_col("candidates", _col, _ddl)

        # v30.0: champs entretien inline
        for _col, _ddl in [
            ("entretien_date", "TEXT"),
            ("entretien_lieu", "TEXT"),
            ("entretien_notes", "TEXT"),
        ]:
            if _col not in cand_cols:
                _add_col("candidates", _col, _ddl)

        # v29.0: DC Generator — dossier généré par le générateur interne
        if "dossier_path" not in cand_cols:
            _add_col("candidates", "dossier_path", "TEXT")
        if "dossier_generated_at" not in cand_cols:
            _add_col("candidates", "dossier_generated_at", "DATETIME")

        # v23.5: Soft delete — add deleted_at column to main tables
        for tbl in ("prospects", "companies", "candidates"):
            tbl_cols = [r["name"] for r in conn.execute(f"PRAGMA table_info({tbl});").fetchall()]
            if "deleted_at" not in tbl_cols:
                _add_col(tbl, "deleted_at", "TEXT")

        # Migration: prospect_events.meta — colonne absente dans certaines DBs créées avant l'ajout
        # du champ meta. Sans cette colonne, les INSERT rdv_taken échouent silencieusement
        # (try/except) et les stats/gamification ne comptent pas les changements de date RDV.
        ev_cols = [r["name"] for r in conn.execute("PRAGMA table_info(prospect_events);").fetchall()]
        if ev_cols and "meta" not in ev_cols:
            _add_col("prospect_events", "meta", "TEXT")

        # v23.3+: indexes on owner_id (created after migration adds the column)
        conn.executescript('''
            CREATE INDEX IF NOT EXISTS idx_prospects_owner ON prospects(owner_id);
            CREATE INDEX IF NOT EXISTS idx_prospects_owner_statut ON prospects(owner_id, statut);
            CREATE INDEX IF NOT EXISTS idx_companies_owner ON companies(owner_id);
            CREATE INDEX IF NOT EXISTS idx_candidates_owner ON candidates(owner_id);
        ''')

        # v27.9: indexes on high-frequency filter/join columns (added after deleted_at migration)
        conn.executescript('''
            CREATE INDEX IF NOT EXISTS idx_prospects_company_id ON prospects(company_id);
            CREATE INDEX IF NOT EXISTS idx_prospects_next_followup ON prospects(nextFollowUp);
            CREATE INDEX IF NOT EXISTS idx_prospects_last_contact ON prospects(lastContact);
            CREATE INDEX IF NOT EXISTS idx_prospects_owner_deleted ON prospects(owner_id, deleted_at);
            CREATE INDEX IF NOT EXISTS idx_companies_owner_deleted ON companies(owner_id, deleted_at);
            CREATE INDEX IF NOT EXISTS idx_candidates_status ON candidates(status);
            CREATE INDEX IF NOT EXISTS idx_candidates_owner_deleted ON candidates(owner_id, deleted_at);
            CREATE INDEX IF NOT EXISTS idx_push_logs_prospect_sentAt ON push_logs(prospect_id, sentAt);
        ''')
        
        # v25.9: Add owner_id to push_categories for per-user categories
        pc_cols = [r["name"] for r in conn.execute("PRAGMA table_info(push_categories);").fetchall()]
        if "owner_id" not in pc_cols:
            _add_col("push_categories", "owner_id", "INTEGER")
            # Migrate existing categories: assign to first admin user, or NULL if no users
            with _auth_conn() as auth_conn:
                admin_user = auth_conn.execute("SELECT id FROM users WHERE role='admin' LIMIT 1;").fetchone()
                if admin_user:
                    conn.execute("UPDATE push_categories SET owner_id=? WHERE owner_id IS NULL;", (admin_user["id"],))
            conn.executescript('''
                CREATE INDEX IF NOT EXISTS idx_push_categories_owner ON push_categories(owner_id);
            ''')
        # v27.3: Add default candidate slots to push categories
        if "candidate1_id" not in pc_cols:
            _add_col("push_categories", "candidate1_id", "INTEGER")
        if "candidate2_id" not in pc_cols:
            _add_col("push_categories", "candidate2_id", "INTEGER")

        # App settings (v11) — key/value config store
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );
        ''')

        # Users table (v15) — multi-user auth
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id           INTEGER PRIMARY KEY,
                username     TEXT NOT NULL UNIQUE,
                display_name TEXT,
                password_hash TEXT NOT NULL,
                role         TEXT NOT NULL DEFAULT 'reader',
                is_active    INTEGER DEFAULT 1,
                createdAt    TEXT,
                lastLoginAt  TEXT
            );
        ''')

        # Refresh tokens for mobile JWT auth (v24.0)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS refresh_tokens (
                id         INTEGER PRIMARY KEY,
                user_id    INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                expires_at TEXT NOT NULL,
                revoked    INTEGER DEFAULT 0,
                device     TEXT,
                createdAt  TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
        ''')

        # Mode Prosp sessions persistées (v28.1) — survivent aux redémarrages serveur
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS mode_prosp_sessions (
                token      TEXT PRIMARY KEY,
                user_id    INTEGER NOT NULL,
                ids        TEXT NOT NULL,
                created_at REAL NOT NULL
            );
        ''')

        # Paires de prospects marquées "pas un doublon" (v27.3)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS duplicate_ignores (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                owner_id       INTEGER NOT NULL,
                prospect_id_a  INTEGER NOT NULL,
                prospect_id_b  INTEGER NOT NULL,
                created_at     TEXT DEFAULT (datetime(\'now\')),
                UNIQUE(owner_id, prospect_id_a, prospect_id_b)
            );
            CREATE INDEX IF NOT EXISTS idx_dup_ignores_owner ON duplicate_ignores(owner_id);
        ''')

        # Shared companies for collaboration (v25.5+)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS shared_companies (
                id          INTEGER PRIMARY KEY,
                company_id  INTEGER NOT NULL,
                from_user_id INTEGER NOT NULL,
                to_user_id  INTEGER NOT NULL,
                shared_at   TEXT NOT NULL,
                FOREIGN KEY(from_user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(to_user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_shared_companies_from ON shared_companies(from_user_id);
            CREATE INDEX IF NOT EXISTS idx_shared_companies_to ON shared_companies(to_user_id);
            CREATE INDEX IF NOT EXISTS idx_shared_companies_company ON shared_companies(company_id);
        ''')

        # Manual KPI entries table (v16.5)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS manual_kpi (
                id          INTEGER PRIMARY KEY,
                user_id     INTEGER,
                type        TEXT NOT NULL,
                date        TEXT NOT NULL,
                count       INTEGER DEFAULT 1,
                description TEXT,
                createdAt   TEXT
            );
        ''')
        
        # Meetings table (v25.10) — historique des réunions avec grille de qualification
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS meetings (
                id            INTEGER PRIMARY KEY,
                prospect_id  INTEGER NOT NULL,
                owner_id     INTEGER NOT NULL,
                date         TEXT NOT NULL,
                title        TEXT NOT NULL,
                checklist_data TEXT,
                notes        TEXT,
                createdAt    TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_meetings_prospect ON meetings(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meetings_owner ON meetings(owner_id);
            CREATE INDEX IF NOT EXISTS idx_meetings_date ON meetings(date);
            
            CREATE TABLE IF NOT EXISTS meeting_action_items (
                id            INTEGER PRIMARY KEY,
                meeting_id    INTEGER NOT NULL,
                prospect_id   INTEGER NOT NULL,
                task          TEXT NOT NULL,
                assignee      TEXT,
                due_date      TEXT,
                priority      TEXT,
                status        TEXT NOT NULL DEFAULT 'pending',
                owner_id      INTEGER NOT NULL,
                createdAt     TEXT NOT NULL,
                FOREIGN KEY(meeting_id) REFERENCES meetings(id) ON DELETE CASCADE,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_meeting ON meeting_action_items(meeting_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_prospect ON meeting_action_items(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_owner ON meeting_action_items(owner_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_status ON meeting_action_items(status);
            
            CREATE TABLE IF NOT EXISTS meeting_opportunities (
                id              INTEGER PRIMARY KEY,
                meeting_id      INTEGER NOT NULL,
                prospect_id     INTEGER NOT NULL,
                type            TEXT NOT NULL,
                estimated_value REAL,
                probability     INTEGER,
                description     TEXT,
                owner_id        INTEGER NOT NULL,
                createdAt       TEXT NOT NULL,
                FOREIGN KEY(meeting_id) REFERENCES meetings(id) ON DELETE CASCADE,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_meeting_opportunities_meeting ON meeting_opportunities(meeting_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_opportunities_prospect ON meeting_opportunities(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_opportunities_owner ON meeting_opportunities(owner_id);
        ''')

        # LinkedIn InMails (dashboard objectifs sourcing)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS linkedin_inmails (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                url        TEXT NOT NULL,
                note       TEXT,
                sent_at    TEXT NOT NULL,
                owner_id   INTEGER NOT NULL,
                created_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_linkedin_inmails_owner_date ON linkedin_inmails(owner_id, sent_at);

            -- v31: persistent DC generation history (replaces in-session list)
            CREATE TABLE IF NOT EXISTS dc_generations (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                candidate_id INTEGER,
                filename     TEXT,
                file_path    TEXT NOT NULL,
                used_ollama  INTEGER DEFAULT 0,
                generated_at TEXT NOT NULL,
                owner_id     INTEGER NOT NULL,
                deleted_at   TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_dc_gen_owner_date ON dc_generations(owner_id, generated_at);
            CREATE INDEX IF NOT EXISTS idx_dc_gen_candidate ON dc_generations(candidate_id, owner_id);

            -- v31: standalone calendar events (créés depuis l'UI v30)
            CREATE TABLE IF NOT EXISTS calendar_events (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                title        TEXT NOT NULL,
                event_date   TEXT NOT NULL,
                event_time   TEXT,
                duration_min INTEGER,
                location     TEXT,
                notes        TEXT,
                status       TEXT DEFAULT 'planifie',
                event_type   TEXT DEFAULT 'rdv',
                prospect_id  INTEGER,
                candidate_id INTEGER,
                company_id   INTEGER,
                owner_id     INTEGER NOT NULL,
                created_at   TEXT,
                updated_at   TEXT,
                deleted_at   TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_cal_evt_owner_date ON calendar_events(owner_id, event_date);

            -- v32.1 : Transcription de réunions (pipeline Whisper + pyannote + Claude)
            CREATE TABLE IF NOT EXISTS transcriptions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                title           TEXT NOT NULL,
                audio_filename  TEXT NOT NULL,
                audio_path      TEXT NOT NULL,
                audio_size      INTEGER,
                duration_sec    REAL,
                language        TEXT,
                status          TEXT NOT NULL DEFAULT 'pending',
                progress        INTEGER NOT NULL DEFAULT 0,
                stage           TEXT,
                error_message   TEXT,
                transcript_text TEXT,
                segments_json   TEXT,
                speakers_json   TEXT,
                analysis_json   TEXT,
                whisper_model   TEXT,
                analysis_model  TEXT,
                owner_id        INTEGER NOT NULL,
                created_at      TEXT NOT NULL,
                updated_at      TEXT,
                completed_at    TEXT,
                deleted_at      TEXT,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_transcriptions_owner ON transcriptions(owner_id, created_at);
            CREATE INDEX IF NOT EXISTS idx_transcriptions_status ON transcriptions(status);

            -- v32.x : Traitement Besoin (fiches besoin client + suivi candidats)
            CREATE TABLE IF NOT EXISTS besoins (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                client          TEXT,
                localisation    TEXT,
                contact         TEXT,
                date_appel      TEXT,
                intitule        TEXT,
                date_besoin     TEXT,
                duree_mission   TEXT,
                descriptif      TEXT,
                competences     TEXT,
                connaissances   TEXT,
                experience      TEXT,
                profil_type     TEXT,
                commentaires    TEXT,
                statut          TEXT NOT NULL DEFAULT 'ouvert',
                priority        INTEGER,
                candidats_json  TEXT,
                prospect_id     INTEGER,
                company_id      INTEGER,
                owner_id        INTEGER NOT NULL,
                created_at      TEXT NOT NULL,
                updated_at      TEXT,
                deleted_at      TEXT,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_besoins_owner ON besoins(owner_id, created_at);
            CREATE INDEX IF NOT EXISTS idx_besoins_status ON besoins(statut);
            CREATE INDEX IF NOT EXISTS idx_besoins_prospect ON besoins(prospect_id);
        ''')

        # Seed default admin if no users exist
        user_count = conn.execute("SELECT COUNT(*) AS n FROM users;").fetchone()["n"]
        if user_count == 0:
            now = datetime.datetime.now().isoformat(timespec="seconds")
            conn.execute(
                "INSERT INTO users (username, display_name, password_hash, role, is_active, createdAt) VALUES (?, ?, ?, ?, 1, ?);",
                ("admin", "Antoine (Admin)", generate_password_hash("admin"), "admin", now)
            )
            print("Compte admin cree — login: admin / mdp: admin (a changer !)")

        # Migration: reader -> editor (rôles simplifiés à admin + editor uniquement)
        try:
            conn.execute("UPDATE users SET role='editor' WHERE role='reader';")
        except Exception:
            pass

        # Migration: attribuer prospects/candidats/companies/saved_views/tasks sans owner au premier utilisateur (admin)
        try:
            first_user = conn.execute("SELECT id FROM users WHERE is_active=1 ORDER BY id LIMIT 1;").fetchone()
            if first_user:
                uid = first_user["id"]
                conn.execute("UPDATE prospects SET owner_id=? WHERE owner_id IS NULL;", (uid,))
                conn.execute("UPDATE candidates SET owner_id=? WHERE owner_id IS NULL;", (uid,))
                conn.execute("UPDATE companies SET owner_id=? WHERE owner_id IS NULL;", (uid,))
                try:
                    vcols = [r["name"] for r in conn.execute("PRAGMA table_info(saved_views);").fetchall()]
                    if "owner_id" in vcols:
                        conn.execute("UPDATE saved_views SET owner_id=? WHERE owner_id IS NULL;", (uid,))
                except Exception:
                    pass
                try:
                    tcols = [r["name"] for r in conn.execute("PRAGMA table_info(tasks);").fetchall()]
                    if "owner_id" in tcols:
                        conn.execute("UPDATE tasks SET owner_id=? WHERE owner_id IS NULL;", (uid,))
                except Exception:
                    pass
        except Exception:
            pass

        # saved_views et tasks : colonne owner_id (isolation par user)
        try:
            vcols = [r["name"] for r in conn.execute("PRAGMA table_info(saved_views);").fetchall()]
            if "owner_id" not in vcols:
                _add_col("saved_views", "owner_id", "INTEGER")
                first = conn.execute("SELECT id FROM users WHERE is_active=1 ORDER BY id LIMIT 1;").fetchone()
                if first:
                    conn.execute("UPDATE saved_views SET owner_id=? WHERE owner_id IS NULL;", (first["id"],))
        except Exception:
            pass
        try:
            tcols = [r["name"] for r in conn.execute("PRAGMA table_info(tasks);").fetchall()]
            if "owner_id" not in tcols:
                _add_col("tasks", "owner_id", "INTEGER")
                first = conn.execute("SELECT id FROM users WHERE is_active=1 ORDER BY id LIMIT 1;").fetchone()
                if first:
                    conn.execute("UPDATE tasks SET owner_id=? WHERE owner_id IS NULL;", (first["id"],))
        except Exception:
            pass

        tpl_cols = [r["name"] for r in conn.execute("PRAGMA table_info(templates);").fetchall()]
        if "linkedin_body" not in tpl_cols:
            _add_col("templates", "linkedin_body", "TEXT")

        # Users: must_change_password flag (v25.1)
        try:
            ucols_check = [r["name"] for r in conn.execute("PRAGMA table_info(users);").fetchall()]
            if "must_change_password" not in ucols_check:
                _add_col("users", "must_change_password", "INTEGER DEFAULT 0")
        except Exception:
            pass

        # Users: onboarding (popup bienvenue / visite guidée pour nouveaux utilisateurs uniquement)
        try:
            ucols = [r["name"] for r in conn.execute("PRAGMA table_info(users);").fetchall()]
            if "onboarding_seen" not in ucols:
                _add_col("users", "onboarding_seen", "INTEGER DEFAULT 0")
                conn.execute("UPDATE users SET onboarding_seen=1;")
            else:
                # Backfill une seule fois : utilisateurs existants = déjà vus (évite popup aux collègues actuels)
                try:
                    row = conn.execute("SELECT value FROM app_settings WHERE key=?;", ("onboarding_backfill_done",)).fetchone()
                    if not row or row["value"] != "1":
                        conn.execute("UPDATE users SET onboarding_seen=1;")
                        conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES (?, ?);", ("onboarding_backfill_done", "1"))
                except Exception:
                    pass
        except Exception:
            pass

        # Meetings (v31.8) — colonnes pour CR détaillés (raw transcript, summary IA, next_action, tags snapshot, documents)
        try:
            mcols = [r["name"] for r in conn.execute("PRAGMA table_info(meetings);").fetchall()]
            if "summary" not in mcols:
                _add_col("meetings", "summary", "TEXT")
            if "raw_transcript" not in mcols:
                _add_col("meetings", "raw_transcript", "TEXT")
            if "next_action" not in mcols:
                _add_col("meetings", "next_action", "TEXT")
            if "tags" not in mcols:
                _add_col("meetings", "tags", "TEXT")
            if "documents" not in mcols:
                _add_col("meetings", "documents", "TEXT")
        except Exception:
            pass

        # prospect_attachments (v31.9) — pièces jointes par prospect, isolées par owner
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS prospect_attachments (
                    id            INTEGER PRIMARY KEY,
                    prospect_id   INTEGER NOT NULL,
                    owner_id      INTEGER NOT NULL,
                    filename      TEXT NOT NULL,
                    original_name TEXT NOT NULL,
                    size          INTEGER,
                    mime_type     TEXT,
                    description   TEXT,
                    meeting_id    INTEGER,
                    createdAt     TEXT NOT NULL,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_prospect_attachments_prospect ON prospect_attachments(prospect_id);")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_prospect_attachments_owner ON prospect_attachments(owner_id);")
        except Exception:
            pass

        # prospect_attachments (v32.0) — colonnes tags / thumbnail / extracted_text
        # prospect_attachments (v32.2) — colonne title (titre éditable affiché en timeline)
        try:
            acols = [r["name"] for r in conn.execute("PRAGMA table_info(prospect_attachments);").fetchall()]
            if "tags" not in acols:
                _add_col("prospect_attachments", "tags", "TEXT")
            if "thumbnail" not in acols:
                _add_col("prospect_attachments", "thumbnail", "TEXT")
            if "extracted_text" not in acols:
                _add_col("prospect_attachments", "extracted_text", "TEXT")
            if "title" not in acols:
                _add_col("prospect_attachments", "title", "TEXT")
        except Exception:
            pass

        # prospect_summaries (v32.0) — cache résumés IA des fiches
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS prospect_summaries (
                    prospect_id INTEGER PRIMARY KEY,
                    owner_id    INTEGER NOT NULL,
                    summary     TEXT,
                    generatedAt TEXT,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );
            """)
        except Exception:
            pass

        # push_categories (v32.3) — colonnes candidats par défaut + flag no_candidates
        try:
            pc_cols = [r["name"] for r in conn.execute("PRAGMA table_info(push_categories);").fetchall()]
            if "candidate1_id" not in pc_cols:
                _add_col("push_categories", "candidate1_id", "INTEGER")
            if "candidate2_id" not in pc_cols:
                _add_col("push_categories", "candidate2_id", "INTEGER")
            if "no_candidates" not in pc_cols:
                _add_col("push_categories", "no_candidates", "INTEGER DEFAULT 0")
        except Exception:
            pass

        # Seed template par défaut si besoin
        n = conn.execute("SELECT COUNT(*) AS n FROM templates;").fetchone()["n"]
        if n == 0:
            now = datetime.datetime.now().isoformat(timespec="seconds")
            default_subject = "Prospection - {{entreprise}}"

            default_body = """Bonjour {{civilite}} {{nom}},

Je vous contacte en tant qu'ingénieur d'affaires au sein d'Up Technologies.

Up Technologies est une société de conseil en ingénierie spécialisée en électronique (hardware et software), informatique et systèmes.

Je souhaiterais échanger avec vous pour étudier des pistes de collaboration.
Pourriez-vous me communiquer vos prochains créneaux disponibles afin de planifier un échange s’il vous plaît ?

Cordialement,"""
            conn.execute(
                '''
                INSERT INTO templates (name, subject, body, linkedin_body, is_default, createdAt, updatedAt)
                        VALUES (?, ?, ?, ?, 1, ?, ?);
                ''',
                (
                    "Template par défaut",
                    default_subject,
                    default_body,
                    # Par défaut, on réutilise le body email pour LinkedIn (modifiable ensuite)
                    default_body,
                    now,
                    now,
                ),
            )

        # Migration: corriger les templates dont la phrase d'intro est incomplète.
        # Remplace toute variante de "au sein … Technologies." (avec ou sans HTML, avec ou sans "d'Up")
        # par la version correcte avec mise en forme orange.
        try:
            import re as _re_tmpl
            _PHRASE_PATTERN = _re_tmpl.compile(
                r"au sein\s*(?:<[^>]*>\s*)*(?:d['']Up\s+)?(?:\s*)Technologies\.",
                _re_tmpl.DOTALL | _re_tmpl.IGNORECASE
            )
            _PHRASE_REPLACE = "au sein <span style=\"color:#E07020;font-weight:bold;\">d'Up Technologies.</span>"
            _tmpl_rows = conn.execute("SELECT id, body, linkedin_body FROM templates;").fetchall()
            _tmpl_updated = False
            for _tr in _tmpl_rows:
                _bid, _body, _lbody = _tr["id"], _tr["body"] or "", _tr["linkedin_body"] or ""
                _new_body  = _PHRASE_PATTERN.sub(_PHRASE_REPLACE, _body)
                _new_lbody = _PHRASE_PATTERN.sub(_PHRASE_REPLACE, _lbody)
                if _new_body != _body or _new_lbody != _lbody:
                    conn.execute(
                        "UPDATE templates SET body=?, linkedin_body=? WHERE id=?;",
                        (_new_body, _new_lbody, _bid)
                    )
                    _tmpl_updated = True
            if _tmpl_updated:
                conn.commit()
        except Exception:
            pass

        # v25: candidate_tabs (onglets fiche candidat) — migration depuis candidate_ec1_checklists
        _migrate_candidate_tabs(conn)



def _migrate_candidate_tabs(conn: sqlite3.Connection) -> None:
    """Crée candidate_tabs si absent et migre les données depuis candidate_ec1_checklists (v25)."""
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS candidate_tabs (
                id           INTEGER PRIMARY KEY,
                candidate_id INTEGER NOT NULL,
                sort_order   INTEGER NOT NULL DEFAULT 0,
                type         TEXT NOT NULL,
                title        TEXT NOT NULL,
                payload      TEXT,
                updated_at   TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidate_tabs_candidate ON candidate_tabs(candidate_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidate_tabs_sort ON candidate_tabs(candidate_id, sort_order);")
        conn.commit()
    except Exception:
        pass
    # Migrer chaque EC1 existant vers le premier onglet
    try:
        rows = conn.execute(
            "SELECT candidate_id, interviewAt, data, updatedAt FROM candidate_ec1_checklists;"
        ).fetchall()
        now = datetime.datetime.now().isoformat(timespec="seconds")
        for row in rows:
            cid = row["candidate_id"]
            existing = conn.execute(
                "SELECT id FROM candidate_tabs WHERE candidate_id=? AND sort_order=0;",
                (cid,),
            ).fetchone()
            if existing:
                continue
            try:
                data = json.loads(row["data"]) if row["data"] else {}
            except Exception:
                data = {}
            payload = json.dumps(
                {"interviewAt": row["interviewAt"] or None, "data": data},
                ensure_ascii=False,
            )
            conn.execute(
                """INSERT INTO candidate_tabs (candidate_id, sort_order, type, title, payload, updated_at)
                   VALUES (?, 0, 'ec1', 'EC1', ?, ?);""",
                (cid, payload, row["updatedAt"] or now),
            )
        conn.commit()
    except Exception:
        pass


def _migrate_user_db_schema(db_path: Path) -> None:
    """Ajoute deleted_at aux tables companies, prospects, candidates si absent (v23.5).
    Ajoute aussi dossier_competence_pdf à la table candidates si absent.
    v25: candidate_tabs + migration depuis candidate_ec1_checklists.
    v27: migration is_contact → is_archived pour les DBs per-user."""
    if not db_path.exists():
        return
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        for tbl in ("companies", "prospects", "candidates"):
            try:
                cols = [r["name"] for r in conn.execute(f"PRAGMA table_info({tbl});").fetchall()]
            except Exception:
                continue
            # PRAGMA renvoie [] si la table n'existe pas (DB user vide) — skip
            # pour éviter un ALTER TABLE qui échoue.
            if not cols:
                continue
            if "deleted_at" not in cols:
                conn.execute(f"ALTER TABLE {tbl} ADD COLUMN deleted_at TEXT;")
                conn.commit()
        # Migration: is_contact → is_archived pour prospects (per-user DBs)
        try:
            pros_cols = [r["name"] for r in conn.execute("PRAGMA table_info(prospects);").fetchall()]
            if "is_contact" in pros_cols and "is_archived" not in pros_cols:
                conn.execute("ALTER TABLE prospects ADD COLUMN is_archived INTEGER;")
                conn.execute("UPDATE prospects SET is_archived = is_contact;")
                conn.commit()
                print(f"[OK] Migration is_contact -> is_archived (add+copy) sur {db_path}")
            elif "is_archived" not in pros_cols:
                conn.execute("ALTER TABLE prospects ADD COLUMN is_archived INTEGER;")
                conn.commit()
            # Cas où les deux colonnes coexistent : copier les valeurs manquantes
            if "is_contact" in pros_cols and "is_archived" in pros_cols:
                conn.execute("UPDATE prospects SET is_archived = is_contact WHERE is_contact = 1 AND (is_archived IS NULL OR is_archived = 0);")
                conn.commit()
        except Exception as e:
            print(f"[WARN] Migration is_archived prospects ({db_path}): {e}")
        # Migration: ajouter colonnes candidates (schéma complet aligné sur la main DB)
        try:
            cand_cols = [r["name"] for r in conn.execute("PRAGMA table_info(candidates);").fetchall()]
            _cand_migrations = [
                ("dossier_competence_pdf", "TEXT"),
                ("description_push", "TEXT"),
                ("prenom", "TEXT"),
                ("titre", "TEXT"),
                ("annees_experience", "INTEGER"),
                ("domaine_principal", "TEXT"),
                ("disponibilite", "TEXT"),
                ("mobilite", "TEXT"),
                ("permis_conduire", "INTEGER"),
                ("vehicule", "INTEGER"),
                ("permis_travail", "TEXT"),
                ("fonctions_recherchees", "TEXT"),
                ("motif_recherche", "TEXT"),
                ("avancement_recherches", "TEXT"),
                ("remuneration_actuelle", "TEXT"),
                ("pretentions_salariales", "TEXT"),
                ("propal_a", "TEXT"),
                ("eval_technique", "TEXT"),
                ("eval_personnalite", "TEXT"),
                ("eval_communication", "TEXT"),
                ("langues", "TEXT"),
                ("references_candidat", "TEXT"),
                ("avis_perso", "TEXT"),
                ("entretien_date", "TEXT"),
                ("entretien_lieu", "TEXT"),
                ("entretien_notes", "TEXT"),
                ("dossier_path", "TEXT"),
                ("dossier_generated_at", "DATETIME"),
            ]
            for col, typ in _cand_migrations:
                if col not in cand_cols:
                    conn.execute(f"ALTER TABLE candidates ADD COLUMN {col} {typ};")
            conn.commit()
        except Exception as e:
            print(f"[WARN] Migration candidates columns ({db_path}): {e}")
        _migrate_candidate_tabs(conn)
        # Migration: créer custom_metiers si absent (v27.22 — tag management)
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS custom_metiers (
                id        INTEGER PRIMARY KEY,
                type      TEXT NOT NULL,
                category  TEXT NOT NULL,
                specialty TEXT,
                tech_group TEXT,
                value     TEXT NOT NULL,
                createdAt TEXT
            )''')
            conn.commit()
        except Exception:
            pass
        # Migration: créer call_logs si absent (stats appels per-user)
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS call_logs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL,
                owner_id    INTEGER NOT NULL,
                date        TEXT NOT NULL,
                called_at   TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            )''')
            conn.execute("CREATE INDEX IF NOT EXISTS idx_call_logs_owner_date ON call_logs(owner_id, date);")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_call_logs_prospect ON call_logs(prospect_id);")
            conn.commit()
        except Exception:
            pass
        # Migration v25.3+: push_logs (traçabilité + analytics + tracking pixel) per-user
        try:
            pl_cols = {r["name"] for r in conn.execute("PRAGMA table_info(push_logs);").fetchall()}
            for col, typ in (
                ("candidate_id1", "INTEGER"),
                ("candidate_id2", "INTEGER"),
                ("consultant1_id", "INTEGER"),
                ("consultant2_id", "INTEGER"),
                ("sent_at_hour", "INTEGER"),
                ("sent_at_day_of_week", "INTEGER"),
                ("variant_id", "TEXT"),
                ("opened_at", "TEXT"),
                ("clicked_at", "TEXT"),
                ("replied_at", "TEXT"),
                ("tracking_pixel_id", "TEXT"),
                ("campaign_id", "INTEGER"),
            ):
                if col not in pl_cols:
                    conn.execute(f"ALTER TABLE push_logs ADD COLUMN {col} {typ};")
            conn.commit()
        except Exception as e:
            print(f"[WARN] Migration push_logs columns ({db_path}): {e}")

        # Migration: prospect_events.meta (colonnes absentes dans les vieilles per-user DBs)
        try:
            ev_cols = [r["name"] for r in conn.execute("PRAGMA table_info(prospect_events);").fetchall()]
            if ev_cols and "meta" not in ev_cols:
                conn.execute("ALTER TABLE prospect_events ADD COLUMN meta TEXT;")
                conn.commit()
        except Exception as e:
            print(f"[WARN] Migration prospect_events.meta ({db_path}): {e}")

        # Migration v27.3: push_categories default candidate slots
        # Migration v32.2: no_candidates flag (push categorie "sans consultant")
        try:
            pc_cols = {r["name"] for r in conn.execute("PRAGMA table_info(push_categories);").fetchall()}
            for col, typ in (
                ("candidate1_id", "INTEGER"),
                ("candidate2_id", "INTEGER"),
                ("no_candidates", "INTEGER DEFAULT 0"),
            ):
                if col not in pc_cols:
                    conn.execute(f"ALTER TABLE push_categories ADD COLUMN {col} {typ};")
            conn.commit()
        except Exception as e:
            print(f"[WARN] Migration push_categories candidates ({db_path}): {e}")

        # Migration: tables diverses souvent absentes des vieilles per-user DBs
        try:
            conn.executescript('''
                CREATE TABLE IF NOT EXISTS mode_prosp_sessions (
                    token      TEXT PRIMARY KEY,
                    user_id    INTEGER NOT NULL,
                    ids        TEXT NOT NULL,
                    created_at REAL NOT NULL
                );
                CREATE TABLE IF NOT EXISTS candidate_skills (
                    id           INTEGER PRIMARY KEY,
                    candidate_id INTEGER NOT NULL,
                    name         TEXT NOT NULL,
                    category     TEXT,
                    level        INTEGER DEFAULT 3,
                    UNIQUE(candidate_id, name),
                    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS candidate_availability (
                    id           INTEGER PRIMARY KEY,
                    candidate_id INTEGER NOT NULL,
                    week_iso     TEXT NOT NULL,
                    status       TEXT NOT NULL,
                    UNIQUE(candidate_id, week_iso),
                    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS duplicate_ignores (
                    id             INTEGER PRIMARY KEY,
                    owner_id       INTEGER NOT NULL,
                    prospect_id_a  INTEGER NOT NULL,
                    prospect_id_b  INTEGER NOT NULL,
                    created_at     TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_duplicate_ignores_owner ON duplicate_ignores(owner_id);
                CREATE TABLE IF NOT EXISTS embeddings_cache (
                    id          INTEGER PRIMARY KEY,
                    entity_type TEXT NOT NULL,
                    entity_id   INTEGER,
                    text_key    TEXT NOT NULL,
                    embedding   TEXT NOT NULL,
                    created_at  TEXT DEFAULT (datetime('now')),
                    UNIQUE(entity_type, entity_id, text_key)
                );
                CREATE INDEX IF NOT EXISTS idx_embeddings_lookup ON embeddings_cache(entity_type, text_key);
            ''')
            conn.commit()
        except Exception as e:
            print(f"[WARN] Migration tables diverses ({db_path}): {e}")
        # Migration v29+: créer toutes les tables d'événements et KPI manquantes dans les DBs per-user existantes
        # Ces tables sont essentielles pour les KPIs et la gamification (prospect_events, candidate_events, etc.)
        try:
            conn.executescript('''
                CREATE TABLE IF NOT EXISTS company_events (
                    id        INTEGER PRIMARY KEY,
                    company_id INTEGER NOT NULL,
                    date      TEXT NOT NULL,
                    type      TEXT,
                    title     TEXT,
                    content   TEXT,
                    meta      TEXT,
                    createdAt TEXT,
                    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_company_events_company ON company_events(company_id);
                CREATE INDEX IF NOT EXISTS idx_company_events_date ON company_events(date);

                CREATE TABLE IF NOT EXISTS prospect_events (
                    id         INTEGER PRIMARY KEY,
                    prospect_id INTEGER NOT NULL,
                    date       TEXT NOT NULL,
                    type       TEXT,
                    title      TEXT,
                    content    TEXT,
                    meta       TEXT,
                    createdAt  TEXT,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_prospect_events_prospect ON prospect_events(prospect_id);
                CREATE INDEX IF NOT EXISTS idx_prospect_events_date ON prospect_events(date);
                CREATE UNIQUE INDEX IF NOT EXISTS idx_prospect_events_unique ON prospect_events(prospect_id, type, date);

                CREATE TABLE IF NOT EXISTS candidate_events (
                    id           INTEGER PRIMARY KEY,
                    candidate_id INTEGER NOT NULL,
                    date         TEXT NOT NULL,
                    type         TEXT,
                    title        TEXT,
                    content      TEXT,
                    meta         TEXT,
                    createdAt    TEXT,
                    FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_candidate_events_candidate ON candidate_events(candidate_id);
                CREATE INDEX IF NOT EXISTS idx_candidate_events_date ON candidate_events(date);
                CREATE UNIQUE INDEX IF NOT EXISTS idx_candidate_events_unique ON candidate_events(candidate_id, type, date);

                CREATE TABLE IF NOT EXISTS push_logs (
                    id            INTEGER PRIMARY KEY,
                    prospect_id   INTEGER NOT NULL,
                    sentAt        TEXT NOT NULL,
                    channel       TEXT,
                    to_email      TEXT,
                    subject       TEXT,
                    body          TEXT,
                    template_id   INTEGER,
                    template_name TEXT,
                    createdAt     TEXT NOT NULL,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_push_logs_prospect_id ON push_logs(prospect_id);
                CREATE INDEX IF NOT EXISTS idx_push_logs_sentAt ON push_logs(sentAt);

                CREATE TABLE IF NOT EXISTS push_categories (
                    id            INTEGER PRIMARY KEY,
                    name          TEXT NOT NULL,
                    keywords      TEXT,
                    auto_detected INTEGER DEFAULT 0,
                    owner_id      INTEGER,
                    candidate1_id INTEGER,
                    candidate2_id INTEGER,
                    no_candidates INTEGER DEFAULT 0,
                    createdAt     TEXT,
                    updatedAt     TEXT,
                    UNIQUE(name, owner_id)
                );
                CREATE INDEX IF NOT EXISTS idx_push_categories_name ON push_categories(name);
                CREATE INDEX IF NOT EXISTS idx_push_categories_owner ON push_categories(owner_id);

                CREATE TABLE IF NOT EXISTS push_variants (
                    id            INTEGER PRIMARY KEY,
                    push_log_id   INTEGER NOT NULL,
                    variant_id    TEXT NOT NULL,
                    subject       TEXT,
                    body          TEXT,
                    sent_at       TEXT,
                    opened_at     TEXT,
                    clicked_at    TEXT,
                    replied_at    TEXT,
                    createdAt     TEXT NOT NULL,
                    FOREIGN KEY(push_log_id) REFERENCES push_logs(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_push_variants_push_log_id ON push_variants(push_log_id);
                CREATE INDEX IF NOT EXISTS idx_push_variants_variant_id ON push_variants(variant_id);

                CREATE TABLE IF NOT EXISTS rdv_checklists (
                    id          INTEGER PRIMARY KEY,
                    prospect_id INTEGER NOT NULL UNIQUE,
                    data        TEXT,
                    updatedAt   TEXT,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS manual_kpi (
                    id          INTEGER PRIMARY KEY,
                    user_id     INTEGER,
                    type        TEXT NOT NULL,
                    date        TEXT NOT NULL,
                    count       INTEGER DEFAULT 1,
                    description TEXT,
                    createdAt   TEXT
                );

                CREATE TABLE IF NOT EXISTS app_settings (
                    key   TEXT PRIMARY KEY,
                    value TEXT
                );

                CREATE TABLE IF NOT EXISTS tasks (
                    id          INTEGER PRIMARY KEY,
                    title       TEXT NOT NULL,
                    comment     TEXT,
                    due_date    TEXT,
                    status      TEXT NOT NULL DEFAULT 'pending',
                    linked_ids  TEXT,
                    createdAt   TEXT,
                    updatedAt   TEXT,
                    owner_id    INTEGER
                );

                CREATE TABLE IF NOT EXISTS task_rules (
                    id            INTEGER PRIMARY KEY,
                    name          TEXT NOT NULL,
                    trigger_type  TEXT NOT NULL,
                    conditions    TEXT NOT NULL,
                    template_title TEXT NOT NULL,
                    template_comment TEXT,
                    priority      INTEGER DEFAULT 2,
                    enabled       INTEGER DEFAULT 1,
                    owner_id      INTEGER,
                    createdAt     TEXT,
                    updatedAt     TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_task_rules_trigger ON task_rules(trigger_type, enabled);
                CREATE INDEX IF NOT EXISTS idx_task_rules_owner ON task_rules(owner_id);

                CREATE TABLE IF NOT EXISTS meetings (
                    id            INTEGER PRIMARY KEY,
                    prospect_id  INTEGER NOT NULL,
                    owner_id     INTEGER NOT NULL,
                    date         TEXT NOT NULL,
                    title        TEXT NOT NULL,
                    checklist_data TEXT,
                    notes        TEXT,
                    createdAt    TEXT NOT NULL,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS meeting_action_items (
                    id            INTEGER PRIMARY KEY,
                    meeting_id    INTEGER NOT NULL,
                    prospect_id   INTEGER NOT NULL,
                    task          TEXT NOT NULL,
                    assignee      TEXT,
                    due_date      TEXT,
                    priority      TEXT,
                    status        TEXT NOT NULL DEFAULT 'pending',
                    owner_id      INTEGER NOT NULL,
                    createdAt     TEXT NOT NULL,
                    FOREIGN KEY(meeting_id) REFERENCES meetings(id) ON DELETE CASCADE,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS meeting_opportunities (
                    id              INTEGER PRIMARY KEY,
                    meeting_id      INTEGER NOT NULL,
                    prospect_id     INTEGER NOT NULL,
                    type            TEXT NOT NULL,
                    estimated_value REAL,
                    probability     INTEGER,
                    description     TEXT,
                    owner_id        INTEGER NOT NULL,
                    createdAt       TEXT NOT NULL,
                    FOREIGN KEY(meeting_id) REFERENCES meetings(id) ON DELETE CASCADE,
                    FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS linkedin_inmails (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    url        TEXT NOT NULL,
                    note       TEXT,
                    sent_at    TEXT NOT NULL,
                    owner_id   INTEGER NOT NULL,
                    created_at REAL
                );
                CREATE INDEX IF NOT EXISTS idx_linkedin_inmails_owner_date ON linkedin_inmails(owner_id, sent_at);

                CREATE TABLE IF NOT EXISTS dc_generations (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    candidate_id INTEGER,
                    filename     TEXT,
                    file_path    TEXT NOT NULL,
                    used_ollama  INTEGER DEFAULT 0,
                    generated_at TEXT NOT NULL,
                    owner_id     INTEGER NOT NULL,
                    deleted_at   TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_dc_gen_owner_date ON dc_generations(owner_id, generated_at);
                CREATE INDEX IF NOT EXISTS idx_dc_gen_candidate ON dc_generations(candidate_id, owner_id);

                CREATE TABLE IF NOT EXISTS calendar_events (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    title        TEXT NOT NULL,
                    event_date   TEXT NOT NULL,
                    event_time   TEXT,
                    duration_min INTEGER,
                    location     TEXT,
                    notes        TEXT,
                    status       TEXT DEFAULT 'planifie',
                    event_type   TEXT DEFAULT 'rdv',
                    prospect_id  INTEGER,
                    candidate_id INTEGER,
                    company_id   INTEGER,
                    owner_id     INTEGER NOT NULL,
                    created_at   TEXT,
                    updated_at   TEXT,
                    deleted_at   TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_cal_evt_owner_date ON calendar_events(owner_id, event_date);
            ''')
        except Exception as e:
            print(f"[WARN] Migration tables KPI manquantes ({db_path}): {e}")
    finally:
        conn.close()


_CANDIDATE_STATUS_MIGRATION = {
    "a_sourcer": "nouveau",
    "a_contacter": "proposition",
    "a_contacter_relance": "proposition",
    "en_cours": "entretien",
    "ec1": "entretien",
    "ec2": "entretien",
    "ed": "a_faire",
    "interesse": "oksi",
    "mission": "freelance_mission",
    "embauche": "valide_contrat",
    "refuse": "nok",
    "archive": "plus_disponible",
}


def _migrate_candidate_statuses(db_path: Path) -> None:
    """Migre les anciens statuts candidats vers les nouveaux slugs."""
    try:
        conn = sqlite3.connect(db_path)
        for old, new in _CANDIDATE_STATUS_MIGRATION.items():
            conn.execute(
                "UPDATE candidates SET status=? WHERE status=? AND deleted_at IS NULL;",
                (new, old),
            )
        # Sync is_archived for archive statuses
        archive_statuses = ("nok_prequal", "nok", "plus_disponible", "refus_contrat")
        placeholders = ",".join("?" * len(archive_statuses))
        conn.execute(
            f"UPDATE candidates SET is_archived=1 WHERE status IN ({placeholders}) AND deleted_at IS NULL;",
            archive_statuses,
        )
        conn.execute(
            f"UPDATE candidates SET is_archived=0 WHERE status NOT IN ({placeholders}) AND deleted_at IS NULL;",
            archive_statuses,
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.warning("migrate_candidate_statuses %s: %s", db_path, e)


def _migrate_call_logs_to_user_db(user_id: int, user_db: Path) -> None:
    """Copie les call_logs de la DB globale vers la DB per-user si la table est vide."""
    try:
        if not DB_PATH.exists():
            return
        conn = sqlite3.connect(user_db)
        conn.row_factory = sqlite3.Row
        count = conn.execute("SELECT COUNT(*) AS n FROM call_logs WHERE owner_id=?;", (user_id,)).fetchone()["n"]
        if count > 0:
            conn.close()
            return  # Déjà des données, pas besoin de migrer
        conn.close()
        global_conn = sqlite3.connect(DB_PATH)
        global_conn.row_factory = sqlite3.Row
        try:
            rows = global_conn.execute(
                "SELECT prospect_id, owner_id, date, called_at FROM call_logs WHERE owner_id=?;",
                (user_id,)
            ).fetchall()
        except Exception:
            rows = []
        finally:
            global_conn.close()
        if not rows:
            return
        conn = sqlite3.connect(user_db)
        conn.execute("PRAGMA foreign_keys = OFF;")
        conn.executemany(
            "INSERT OR IGNORE INTO call_logs (prospect_id, owner_id, date, called_at) VALUES (?, ?, ?, ?);",
            [(r["prospect_id"], r["owner_id"], r["date"], r["called_at"]) for r in rows]
        )
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.commit()
        conn.close()
        print(f"[OK] {len(rows)} call_logs récupérés depuis DB globale vers {user_db}")
    except Exception as e:
        print(f"[WARN] Migration call_logs ({user_db}): {e}")


def _v30_schema_sql() -> str:
    """Schémas additifs v30 (push_campaigns, saved_views colonnes, skills, availability)."""
    return '''
    CREATE TABLE IF NOT EXISTS push_campaigns (
        id           INTEGER PRIMARY KEY,
        owner_id     INTEGER NOT NULL,
        name         TEXT NOT NULL,
        category_id  INTEGER,
        template_id  INTEGER,
        filters_json TEXT,
        scheduled_at TEXT,
        sent_at      TEXT,
        stats_json   TEXT,
        created_at   TEXT NOT NULL,
        updated_at   TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_push_campaigns_owner ON push_campaigns(owner_id);

    CREATE TABLE IF NOT EXISTS candidate_skills (
        id           INTEGER PRIMARY KEY,
        candidate_id INTEGER NOT NULL,
        name         TEXT NOT NULL,
        category     TEXT,
        level        INTEGER NOT NULL DEFAULT 3,
        UNIQUE(candidate_id, name)
    );
    CREATE INDEX IF NOT EXISTS idx_cand_skills_cid ON candidate_skills(candidate_id);

    CREATE TABLE IF NOT EXISTS candidate_availability (
        id           INTEGER PRIMARY KEY,
        candidate_id INTEGER NOT NULL,
        week_iso     TEXT NOT NULL,
        status       TEXT NOT NULL,
        UNIQUE(candidate_id, week_iso)
    );
    CREATE INDEX IF NOT EXISTS idx_cand_avail_cid ON candidate_availability(candidate_id);

    -- Traitement Besoin : fiches besoin client + suivi candidats
    CREATE TABLE IF NOT EXISTS besoins (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        client          TEXT,
        localisation    TEXT,
        contact         TEXT,
        date_appel      TEXT,
        intitule        TEXT,
        date_besoin     TEXT,
        duree_mission   TEXT,
        descriptif      TEXT,
        competences     TEXT,
        connaissances   TEXT,
        experience      TEXT,
        profil_type     TEXT,
        commentaires    TEXT,
        statut          TEXT NOT NULL DEFAULT 'ouvert',
        priority        INTEGER,
        candidats_json  TEXT,
        prospect_id     INTEGER,
        company_id      INTEGER,
        owner_id        INTEGER NOT NULL,
        created_at      TEXT NOT NULL,
        updated_at      TEXT,
        deleted_at      TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_besoins_owner ON besoins(owner_id, created_at);
    CREATE INDEX IF NOT EXISTS idx_besoins_status ON besoins(statut);
    CREATE INDEX IF NOT EXISTS idx_besoins_prospect ON besoins(prospect_id);
    '''


def _v30_apply_migrations(conn) -> list[str]:
    """Applique les migrations v30 sur une connexion DB. Retourne la liste des changements effectués."""
    done: list[str] = []
    # 1. Tables additives
    cur = conn.cursor()
    existing = {r[0] for r in cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()}
    for t in ("push_campaigns", "candidate_skills", "candidate_availability", "besoins"):
        if t not in existing:
            done.append(f"create:{t}")
    conn.executescript(_v30_schema_sql())
    conn.commit()

    # 2. saved_views : ajouter colonnes v30 (owner_id, filters_json, columns_json, is_shared)
    try:
        sv_cols = {r[1] for r in cur.execute("PRAGMA table_info(saved_views);").fetchall()}
        if sv_cols:
            if "owner_id" not in sv_cols:
                cur.execute("ALTER TABLE saved_views ADD COLUMN owner_id INTEGER;")
                done.append("alter:saved_views.owner_id")
            if "filters_json" not in sv_cols:
                cur.execute("ALTER TABLE saved_views ADD COLUMN filters_json TEXT;")
                done.append("alter:saved_views.filters_json")
            if "columns_json" not in sv_cols:
                cur.execute("ALTER TABLE saved_views ADD COLUMN columns_json TEXT;")
                done.append("alter:saved_views.columns_json")
            if "is_shared" not in sv_cols:
                cur.execute("ALTER TABLE saved_views ADD COLUMN is_shared INTEGER DEFAULT 0;")
                done.append("alter:saved_views.is_shared")
            # Backfill filters_json depuis state si ancienne colonne présente
            if "state" in sv_cols and "filters_json" in sv_cols:
                cur.execute(
                    "UPDATE saved_views SET filters_json = state "
                    "WHERE filters_json IS NULL AND state IS NOT NULL;"
                )
                if cur.rowcount:
                    done.append(f"backfill:saved_views.filters_json({cur.rowcount})")
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_saved_views_owner_page "
                "ON saved_views(owner_id, page);"
            )
    except Exception as e:
        print(f"[v30_migrate] WARN saved_views ({e})")

    # 3. push_logs : ajouter colonne campaign_id
    try:
        pl_cols = {r[1] for r in cur.execute("PRAGMA table_info(push_logs);").fetchall()}
        if pl_cols and "campaign_id" not in pl_cols:
            cur.execute("ALTER TABLE push_logs ADD COLUMN campaign_id INTEGER;")
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_push_logs_campaign "
                "ON push_logs(campaign_id);"
            )
            done.append("alter:push_logs.campaign_id")
    except Exception as e:
        print(f"[v30_migrate] WARN push_logs ({e})")

    # 4. linkedin_inmails : ajouter colonne name (enrichissement Tavily)
    try:
        li_cols = {r[1] for r in cur.execute("PRAGMA table_info(linkedin_inmails);").fetchall()}
        if li_cols and "name" not in li_cols:
            cur.execute("ALTER TABLE linkedin_inmails ADD COLUMN name TEXT;")
            done.append("alter:linkedin_inmails.name")
    except Exception as e:
        print(f"[v30_migrate] WARN linkedin_inmails ({e})")

    # 5. v32.26 — Carte géographique : lat/long sur companies + prospects
    try:
        co_cols = {r[1] for r in cur.execute("PRAGMA table_info(companies);").fetchall()}
        if co_cols:
            for col, ddl in (
                ("latitude", "REAL"),
                ("longitude", "REAL"),
                ("geocoded_at", "TEXT"),
            ):
                if col not in co_cols:
                    cur.execute(f"ALTER TABLE companies ADD COLUMN {col} {ddl};")
                    done.append(f"alter:companies.{col}")
    except Exception as e:
        print(f"[v30_migrate] WARN companies geo ({e})")
    try:
        pr_cols = {r[1] for r in cur.execute("PRAGMA table_info(prospects);").fetchall()}
        if pr_cols:
            for col, ddl in (
                ("address", "TEXT"),
                ("city", "TEXT"),
                ("country", "TEXT"),
                ("latitude", "REAL"),
                ("longitude", "REAL"),
                ("geocoded_at", "TEXT"),
            ):
                if col not in pr_cols:
                    cur.execute(f"ALTER TABLE prospects ADD COLUMN {col} {ddl};")
                    done.append(f"alter:prospects.{col}")
    except Exception as e:
        print(f"[v30_migrate] WARN prospects geo ({e})")

    conn.commit()
    return done


def _v30_needs_migration() -> bool:
    """Retourne True si au moins une des nouvelles tables v30 n'existe pas dans la DB principale."""
    try:
        with _conn() as conn:
            rows = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name IN ('push_campaigns','candidate_skills','candidate_availability','besoins');"
            ).fetchall()
            return len(rows) < 4
    except Exception:
        return True


def _migrate_v30_all() -> None:
    """Orchestre la migration v30 : backup si nécessaire, puis apply sur toutes les DB."""
    need = _v30_needs_migration()
    if need:
        try:
            from scripts.v30_backup import backup_all_databases
            backup_path = backup_all_databases(reason="v30_auto_migration")
            if backup_path:
                print(f"[v30_migrate] Backup pre-migration : {backup_path}")
        except Exception as e:
            print(f"[v30_migrate] WARN backup impossible ({e}) — on continue quand meme")

    # DB principale
    try:
        with _conn() as conn:
            changes = _v30_apply_migrations(conn)
            if changes:
                print(f"[v30_migrate] main: {', '.join(changes)}")
    except Exception as e:
        print(f"[v30_migrate] ERR main DB: {e}")

    # DBs per-user (filtre sur existence + user valide)
    if not DATA_DIR.exists():
        return
    try:
        valid_ids: set[int] = set()
        try:
            with _auth_conn() as ac:
                for row in ac.execute("SELECT id FROM users;").fetchall():
                    valid_ids.add(int(row["id"]))
        except Exception:
            pass
        for p in DATA_DIR.iterdir():
            if not p.is_dir() or not p.name.startswith("user_"):
                continue
            try:
                uid = int(p.name.replace("user_", "", 1))
            except ValueError:
                continue
            if valid_ids and uid not in valid_ids:
                continue
            user_db = p / "prospects.db"
            if not user_db.exists():
                continue
            try:
                c2 = sqlite3.connect(user_db)
                c2.row_factory = sqlite3.Row
                changes = _v30_apply_migrations(c2)
                c2.close()
                if changes:
                    print(f"[v30_migrate] user_{uid}: {', '.join(changes)}")
            except Exception as e:
                print(f"[v30_migrate] ERR user_{uid}: {e}")
    except Exception as e:
        print(f"[v30_migrate] WARN per-user loop ({e})")


def _migrate_all_user_dbs() -> None:
    """Migre toutes les DB per-user (deleted_at) et supprime les dossiers orphelins."""
    if not DATA_DIR.exists():
        return
    valid_ids = set()
    with _auth_conn() as conn:
        for row in conn.execute("SELECT id FROM users;").fetchall():
            valid_ids.add(int(row["id"]))
    for p in DATA_DIR.iterdir():
        if not p.is_dir() or not p.name.startswith("user_"):
            continue
        try:
            uid = int(p.name.replace("user_", "", 1))
        except ValueError:
            continue
        user_db = p / "prospects.db"
        if user_db.exists():
            if uid not in valid_ids:
                try:
                    shutil.rmtree(p)
                    print(f"[OK] Dossier orphelin supprime : {p}")
                except Exception as e:
                    print(f"[WARN] Impossible de supprimer {p}: {e}")
            else:
                _migrate_user_db_schema(user_db)
                _migrate_candidate_statuses(user_db)
                _migrate_call_logs_to_user_db(uid, user_db)


def _migrate_users_schema() -> None:
    """Ajoute les colonnes email, phone, avatar à la table users si absentes (v27.7)."""
    with _auth_conn() as conn:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(users);")}
        if "email" not in cols:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT;")
        if "phone" not in cols:
            conn.execute("ALTER TABLE users ADD COLUMN phone TEXT;")
        if "avatar" not in cols:
            conn.execute("ALTER TABLE users ADD COLUMN avatar TEXT;")


def _init_user_db(user_id: int) -> Path:
    """Crée et initialise la DB isolée d'un nouvel utilisateur dans data/user_<id>/prospects.db."""
    user_dir = DATA_DIR / f"user_{user_id}"
    user_dir.mkdir(parents=True, exist_ok=True)
    user_db = user_dir / "prospects.db"

    conn = sqlite3.connect(user_db)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA busy_timeout = 20000;")
    conn.execute("PRAGMA journal_mode = WAL;")
    try:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS companies (
                id        INTEGER PRIMARY KEY,
                groupe    TEXT NOT NULL,
                site      TEXT NOT NULL,
                phone     TEXT,
                notes     TEXT,
                tags      TEXT,
                website   TEXT,
                linkedin  TEXT,
                industry  TEXT,
                size      TEXT,
                address   TEXT,
                city      TEXT,
                country   TEXT,
                stack     TEXT,
                pain_points TEXT,
                budget    TEXT,
                urgency   TEXT,
                owner_id  INTEGER,
                deleted_at TEXT
            );

            CREATE TABLE IF NOT EXISTS prospects (
                id            INTEGER PRIMARY KEY,
                name          TEXT NOT NULL,
                company_id    INTEGER NOT NULL,
                fonction      TEXT,
                telephone     TEXT,
                email         TEXT,
                linkedin      TEXT,
                pertinence    TEXT,
                statut        TEXT,
                lastContact   TEXT,
                nextFollowUp  TEXT,
                priority      INTEGER,
                notes         TEXT,
                callNotes     TEXT,
                pushEmailSentAt TEXT,
                tags          TEXT,
                template_id   INTEGER,
                nextAction    TEXT,
                pushLinkedInSentAt TEXT,
                photo_url     TEXT,
                push_category_id INTEGER,
                fixedMetier   TEXT,
                rdvDate       TEXT,
                is_archived   INTEGER,
                owner_id      INTEGER,
                deleted_at    TEXT,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE RESTRICT DEFERRABLE INITIALLY DEFERRED
            );

            CREATE TABLE IF NOT EXISTS candidates (
                id        INTEGER PRIMARY KEY,
                name      TEXT NOT NULL,
                role      TEXT,
                location  TEXT,
                seniority TEXT,
                tech      TEXT,
                linkedin  TEXT,
                source    TEXT,
                status    TEXT,
                notes     TEXT,
                createdAt TEXT,
                updatedAt TEXT,
                onenote_url TEXT,
                vsa_url   TEXT,
                skills    TEXT,
                company_ids TEXT,
                is_archived INTEGER,
                years_experience INTEGER,
                sector    TEXT,
                phone     TEXT,
                email     TEXT,
                dossier_competence_pdf TEXT,
                owner_id  INTEGER,
                deleted_at TEXT
            );

            CREATE TABLE IF NOT EXISTS push_logs (
                id            INTEGER PRIMARY KEY,
                prospect_id   INTEGER NOT NULL,
                sentAt        TEXT NOT NULL,
                channel       TEXT,
                to_email      TEXT,
                subject       TEXT,
                body          TEXT,
                template_id   INTEGER,
                template_name TEXT,
                createdAt     TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS templates (
                id         INTEGER PRIMARY KEY,
                name       TEXT NOT NULL,
                subject    TEXT,
                body       TEXT,
                is_default INTEGER DEFAULT 0,
                createdAt  TEXT,
                updatedAt  TEXT,
                linkedin_body TEXT
            );

            CREATE TABLE IF NOT EXISTS saved_views (
                id        INTEGER PRIMARY KEY,
                page      TEXT NOT NULL,
                name      TEXT NOT NULL,
                state     TEXT NOT NULL,
                createdAt TEXT,
                updatedAt TEXT,
                owner_id  INTEGER
            );

            CREATE TABLE IF NOT EXISTS opportunities (
                id             INTEGER PRIMARY KEY,
                company_id      INTEGER NOT NULL,
                title          TEXT NOT NULL,
                stage          TEXT NOT NULL,
                candidate_name TEXT,
                candidate_link TEXT,
                amount         REAL,
                notes          TEXT,
                createdAt      TEXT,
                updatedAt      TEXT,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS company_events (
                id        INTEGER PRIMARY KEY,
                company_id INTEGER NOT NULL,
                date      TEXT NOT NULL,
                type      TEXT,
                title     TEXT,
                content   TEXT,
                meta      TEXT,
                createdAt TEXT,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS prospect_events (
                id         INTEGER PRIMARY KEY,
                prospect_id INTEGER NOT NULL,
                date       TEXT NOT NULL,
                type       TEXT,
                title      TEXT,
                content    TEXT,
                meta       TEXT,
                createdAt  TEXT,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS candidate_events (
                id           INTEGER PRIMARY KEY,
                candidate_id INTEGER NOT NULL,
                date         TEXT NOT NULL,
                type         TEXT,
                title        TEXT,
                content      TEXT,
                meta         TEXT,
                createdAt    TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS push_categories (
                id            INTEGER PRIMARY KEY,
                name          TEXT NOT NULL,
                keywords      TEXT,
                auto_detected INTEGER DEFAULT 0,
                owner_id      INTEGER,
                candidate1_id INTEGER,
                candidate2_id INTEGER,
                no_candidates INTEGER DEFAULT 0,
                createdAt     TEXT,
                updatedAt     TEXT,
                UNIQUE(name, owner_id)
            );

            CREATE TABLE IF NOT EXISTS push_variants (
                id            INTEGER PRIMARY KEY,
                push_log_id   INTEGER NOT NULL,
                variant_id    TEXT NOT NULL,
                subject       TEXT,
                body          TEXT,
                sent_at       TEXT,
                opened_at     TEXT,
                clicked_at    TEXT,
                replied_at    TEXT,
                createdAt     TEXT NOT NULL,
                FOREIGN KEY(push_log_id) REFERENCES push_logs(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_push_variants_push_log_id ON push_variants(push_log_id);
            CREATE INDEX IF NOT EXISTS idx_push_variants_variant_id ON push_variants(variant_id);

            CREATE TABLE IF NOT EXISTS rdv_checklists (
                id          INTEGER PRIMARY KEY,
                prospect_id INTEGER NOT NULL UNIQUE,
                data        TEXT,
                updatedAt   TEXT,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS candidate_ec1_checklists (
                id           INTEGER PRIMARY KEY,
                candidate_id INTEGER NOT NULL UNIQUE,
                interviewAt  TEXT,
                data         TEXT,
                updatedAt    TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS candidate_tabs (
                id           INTEGER PRIMARY KEY,
                candidate_id INTEGER NOT NULL,
                sort_order   INTEGER NOT NULL DEFAULT 0,
                type         TEXT NOT NULL,
                title        TEXT NOT NULL,
                payload      TEXT,
                updated_at   TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_candidate_tabs_candidate ON candidate_tabs(candidate_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_tabs_sort ON candidate_tabs(candidate_id, sort_order);

            -- Candidate experiences (v26.6: enrichissement IA structuré)
            CREATE TABLE IF NOT EXISTS candidate_experiences (
                id           INTEGER PRIMARY KEY,
                candidate_id INTEGER NOT NULL,
                company_name TEXT NOT NULL,
                role         TEXT,
                start_date   TEXT,
                end_date     TEXT,
                description  TEXT,
                technologies TEXT,
                owner_id     INTEGER,
                createdAt    TEXT,
                updatedAt    TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_candidate_experiences_candidate ON candidate_experiences(candidate_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_experiences_owner ON candidate_experiences(owner_id);

            -- Candidate educations (v26.6: enrichissement IA structuré)
            CREATE TABLE IF NOT EXISTS candidate_educations (
                id            INTEGER PRIMARY KEY,
                candidate_id  INTEGER NOT NULL,
                degree        TEXT,
                school        TEXT NOT NULL,
                year          TEXT,
                specialization TEXT,
                owner_id      INTEGER,
                createdAt     TEXT,
                updatedAt     TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_candidate_educations_candidate ON candidate_educations(candidate_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_educations_owner ON candidate_educations(owner_id);

            -- Candidate certifications (v26.6: enrichissement IA structuré)
            CREATE TABLE IF NOT EXISTS candidate_certifications (
                id            INTEGER PRIMARY KEY,
                candidate_id  INTEGER NOT NULL,
                name          TEXT NOT NULL,
                issuer        TEXT,
                obtained_date TEXT,
                expiry_date   TEXT,
                owner_id      INTEGER,
                createdAt     TEXT,
                updatedAt     TEXT,
                FOREIGN KEY(candidate_id) REFERENCES candidates(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_candidate_certifications_candidate ON candidate_certifications(candidate_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_certifications_owner ON candidate_certifications(owner_id);

            CREATE TABLE IF NOT EXISTS tasks (
                id          INTEGER PRIMARY KEY,
                title       TEXT NOT NULL,
                comment     TEXT,
                due_date    TEXT,
                status      TEXT NOT NULL DEFAULT 'pending',
                linked_ids  TEXT,
                createdAt   TEXT,
                updatedAt   TEXT,
                owner_id    INTEGER
            );

            CREATE TABLE IF NOT EXISTS task_rules (
                id            INTEGER PRIMARY KEY,
                name          TEXT NOT NULL,
                trigger_type  TEXT NOT NULL,
                conditions    TEXT NOT NULL,
                template_title TEXT NOT NULL,
                template_comment TEXT,
                priority      INTEGER DEFAULT 2,
                enabled       INTEGER DEFAULT 1,
                owner_id      INTEGER,
                createdAt     TEXT,
                updatedAt     TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_task_rules_trigger ON task_rules(trigger_type, enabled);
            CREATE INDEX IF NOT EXISTS idx_task_rules_owner ON task_rules(owner_id);

            CREATE TABLE IF NOT EXISTS custom_metiers (
                id        INTEGER PRIMARY KEY,
                type      TEXT NOT NULL,
                category  TEXT NOT NULL,
                specialty TEXT,
                tech_group TEXT,
                value     TEXT NOT NULL,
                createdAt TEXT
            );

            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS manual_kpi (
                id          INTEGER PRIMARY KEY,
                user_id     INTEGER,
                type        TEXT NOT NULL,
                date        TEXT NOT NULL,
                count       INTEGER DEFAULT 1,
                description TEXT,
                createdAt   TEXT
            );

            CREATE TABLE IF NOT EXISTS meetings (
                id            INTEGER PRIMARY KEY,
                prospect_id  INTEGER NOT NULL,
                owner_id     INTEGER NOT NULL,
                date         TEXT NOT NULL,
                title        TEXT NOT NULL,
                checklist_data TEXT,
                notes        TEXT,
                createdAt    TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            
            CREATE TABLE IF NOT EXISTS meeting_action_items (
                id            INTEGER PRIMARY KEY,
                meeting_id    INTEGER NOT NULL,
                prospect_id   INTEGER NOT NULL,
                task          TEXT NOT NULL,
                assignee      TEXT,
                due_date      TEXT,
                priority      TEXT,
                status        TEXT NOT NULL DEFAULT 'pending',
                owner_id      INTEGER NOT NULL,
                createdAt     TEXT NOT NULL,
                FOREIGN KEY(meeting_id) REFERENCES meetings(id) ON DELETE CASCADE,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );
            
            CREATE TABLE IF NOT EXISTS meeting_opportunities (
                id              INTEGER PRIMARY KEY,
                meeting_id      INTEGER NOT NULL,
                prospect_id     INTEGER NOT NULL,
                type            TEXT NOT NULL,
                estimated_value REAL,
                probability     INTEGER,
                description     TEXT,
                owner_id        INTEGER NOT NULL,
                createdAt       TEXT NOT NULL,
                FOREIGN KEY(meeting_id) REFERENCES meetings(id) ON DELETE CASCADE,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY(owner_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS call_logs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                prospect_id INTEGER NOT NULL,
                owner_id    INTEGER NOT NULL,
                date        TEXT NOT NULL,
                called_at   TEXT NOT NULL,
                FOREIGN KEY(prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_push_logs_prospect_id ON push_logs(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_push_logs_sentAt ON push_logs(sentAt);
            CREATE INDEX IF NOT EXISTS idx_templates_default ON templates(is_default);
            CREATE INDEX IF NOT EXISTS idx_saved_views_page ON saved_views(page);
            CREATE INDEX IF NOT EXISTS idx_opportunities_company ON opportunities(company_id);
            CREATE INDEX IF NOT EXISTS idx_company_events_company ON company_events(company_id);
            CREATE INDEX IF NOT EXISTS idx_company_events_date ON company_events(date);
            CREATE INDEX IF NOT EXISTS idx_prospect_events_prospect ON prospect_events(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_prospect_events_date ON prospect_events(date);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_prospect_events_unique ON prospect_events(prospect_id, type, date);
            CREATE INDEX IF NOT EXISTS idx_candidate_events_candidate ON candidate_events(candidate_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_events_date ON candidate_events(date);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_candidate_events_unique ON candidate_events(candidate_id, type, date);
            CREATE INDEX IF NOT EXISTS idx_push_categories_name ON push_categories(name);
            CREATE INDEX IF NOT EXISTS idx_push_categories_owner ON push_categories(owner_id);
            CREATE INDEX IF NOT EXISTS idx_rdv_checklists_prospect ON rdv_checklists(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meetings_prospect ON meetings(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meetings_owner ON meetings(owner_id);
            CREATE INDEX IF NOT EXISTS idx_meetings_date ON meetings(date);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_meeting ON meeting_action_items(meeting_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_prospect ON meeting_action_items(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_owner ON meeting_action_items(owner_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_action_items_status ON meeting_action_items(status);
            CREATE INDEX IF NOT EXISTS idx_meeting_opportunities_meeting ON meeting_opportunities(meeting_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_opportunities_prospect ON meeting_opportunities(prospect_id);
            CREATE INDEX IF NOT EXISTS idx_meeting_opportunities_owner ON meeting_opportunities(owner_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_ec1_candidate ON candidate_ec1_checklists(candidate_id);
            CREATE INDEX IF NOT EXISTS idx_candidate_ec1_interviewAt ON candidate_ec1_checklists(interviewAt);
            CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
            CREATE INDEX IF NOT EXISTS idx_tasks_due_date ON tasks(due_date);
            CREATE INDEX IF NOT EXISTS idx_call_logs_owner_date ON call_logs(owner_id, date);
            CREATE INDEX IF NOT EXISTS idx_call_logs_prospect ON call_logs(prospect_id);

            CREATE TABLE IF NOT EXISTS linkedin_inmails (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                url        TEXT NOT NULL,
                note       TEXT,
                sent_at    TEXT NOT NULL,
                owner_id   INTEGER NOT NULL,
                created_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_linkedin_inmails_owner_date ON linkedin_inmails(owner_id, sent_at);

            CREATE TABLE IF NOT EXISTS dc_generations (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                candidate_id INTEGER,
                filename     TEXT,
                file_path    TEXT NOT NULL,
                used_ollama  INTEGER DEFAULT 0,
                generated_at TEXT NOT NULL,
                owner_id     INTEGER NOT NULL,
                deleted_at   TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_dc_gen_owner_date ON dc_generations(owner_id, generated_at);
            CREATE INDEX IF NOT EXISTS idx_dc_gen_candidate ON dc_generations(candidate_id, owner_id);

            CREATE TABLE IF NOT EXISTS calendar_events (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                title        TEXT NOT NULL,
                event_date   TEXT NOT NULL,
                event_time   TEXT,
                duration_min INTEGER,
                location     TEXT,
                notes        TEXT,
                status       TEXT DEFAULT 'planifie',
                event_type   TEXT DEFAULT 'rdv',
                prospect_id  INTEGER,
                candidate_id INTEGER,
                company_id   INTEGER,
                owner_id     INTEGER NOT NULL,
                created_at   TEXT,
                updated_at   TEXT,
                deleted_at   TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_cal_evt_owner_date ON calendar_events(owner_id, event_date);
        ''')

        now = datetime.datetime.now().isoformat(timespec="seconds")
        n = conn.execute("SELECT COUNT(*) AS n FROM templates;").fetchone()["n"]
        if n == 0:
            default_subject = "Prospection - {{entreprise}}"
            default_body = """Bonjour {{civilite}} {{nom}},

Je vous contacte en tant qu'ingénieur d'affaires.

Cordialement,"""
            conn.execute(
                "INSERT INTO templates (name, subject, body, linkedin_body, is_default, createdAt, updatedAt) VALUES (?, ?, ?, ?, 1, ?, ?);",
                ("Template par défaut", default_subject, default_body, default_body, now, now),
            )
        conn.commit()
    finally:
        conn.close()

    snap_dir = user_dir / "snapshots"
    snap_dir.mkdir(exist_ok=True)

    try:
        _migrate_user_db_schema(user_db)
    except Exception as e:
        print(f"[WARN] _migrate_user_db_schema sur nouvelle DB {user_db}: {e}")

    print(f"[OK] DB utilisateur creee : {user_db}")
    return user_db


def db_is_empty() -> bool:
    with _conn() as conn:
        c1 = conn.execute("SELECT COUNT(*) AS n FROM companies;").fetchone()["n"]
        c2 = conn.execute("SELECT COUNT(*) AS n FROM prospects;").fetchone()["n"]
    return (c1 == 0) and (c2 == 0)


def read_all(owner_id: int | None = None) -> Dict[str, Any]:
    """Retourne companies et prospects filtrés par owner_id si fourni (tout privé par user)."""
    def _parse_tags(v) -> List[str]:
        if v is None:
            return []
        if isinstance(v, list):
            return [str(x).strip() for x in v if str(x).strip()]
        s = str(v).strip()
        if not s:
            return []
        try:
            j = json.loads(s)
            if isinstance(j, list):
                return [str(x).strip() for x in j if str(x).strip()]
        except Exception:
            pass
        # fallback "tag1, tag2"
        return [t.strip() for t in s.split(",") if t.strip()]

    with _conn() as conn:
        if owner_id is not None:
            companies = [dict(r) for r in conn.execute("SELECT * FROM companies WHERE owner_id=? AND deleted_at IS NULL ORDER BY id;", (owner_id,)).fetchall()]
        else:
            companies = [dict(r) for r in conn.execute("SELECT * FROM companies WHERE deleted_at IS NULL ORDER BY id;").fetchall()]
        if owner_id is not None:
            prospects_rows = conn.execute(
                "SELECT * FROM prospects WHERE owner_id=? AND deleted_at IS NULL ORDER BY id;", (owner_id,)
            ).fetchall()
        else:
            prospects_rows = conn.execute("SELECT * FROM prospects WHERE deleted_at IS NULL ORDER BY id;").fetchall()

    for c in companies:
        c["tags"] = _parse_tags(c.get("tags"))

    prospects: List[Dict[str, Any]] = []
    for r in prospects_rows:
        d = dict(r)
        try:
            d["callNotes"] = json.loads(d.get("callNotes") or "[]")
        except Exception:
            d["callNotes"] = []
        d["tags"] = _parse_tags(d.get("tags"))
        d["is_archived"] = int(d.get("is_archived") or 0)
        prospects.append(d)

    return {"companies": companies, "prospects": prospects}






def upsert_all(data: Dict[str, Any]) -> None:
    """SAFE save: upsert companies/prospects and delete only missing ids.
    Prospects sont isolés par owner_id (utilisateur connecté)."""
    uid = _uid()
    if uid is None:
        raise ValueError("Authentification requise pour enregistrer")

    companies = data.get("companies") or []
    prospects = data.get("prospects") or []

    if not isinstance(companies, list) or not isinstance(prospects, list):
        raise ValueError("Invalid payload: companies/prospects must be lists")

    # Forcer owner_id sur tous les prospects et companies du payload
    for p in prospects:
        p["owner_id"] = uid
    for c in companies:
        c["owner_id"] = uid

    def _dump_tags(v) -> str:
        if v is None:
            return "[]"
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return "[]"
            if s.startswith("["):
                return s
            parts = [t.strip() for t in s.split(",") if t.strip()]
            return json.dumps(parts, ensure_ascii=False)
        if isinstance(v, list):
            parts = [str(t).strip() for t in v if str(t).strip()]
            return json.dumps(parts, ensure_ascii=False)
        return "[]"

    def _safe_int(v, default=0) -> int:
        try:
            return int(v)
        except Exception:
            return default

    for c in companies:
        c["tags"] = _dump_tags(c.get("tags"))

    for p in prospects:
        if isinstance(p.get("callNotes"), list):
            p["callNotes"] = json.dumps(p["callNotes"], ensure_ascii=False)
        else:
            # preserve if string else default
            p["callNotes"] = p.get("callNotes") if isinstance(p.get("callNotes"), str) else "[]"
        p["tags"] = _dump_tags(p.get("tags"))
        tid = p.get("template_id")
        try:
            p["template_id"] = int(tid) if tid not in (None, "", "null") else None
        except Exception:
            p["template_id"] = None

    with _conn() as conn:
        # Ownership safety: if an incoming id already belongs to another owner,
        # remap it to a fresh global id so we never "steal" rows via ON CONFLICT.
        incoming_company_ids = sorted({
            _safe_int(c.get("id"))
            for c in companies
            if c.get("id") is not None and _safe_int(c.get("id")) > 0
        })
        incoming_prospect_ids = sorted({
            _safe_int(p.get("id"))
            for p in prospects
            if p.get("id") is not None and _safe_int(p.get("id")) > 0
        })

        existing_company_owner: Dict[int, int | None] = {}
        existing_prospect_owner: Dict[int, int | None] = {}
        if incoming_company_ids:
            q_marks = ",".join("?" for _ in incoming_company_ids)
            rows = conn.execute(
                f"SELECT id, owner_id FROM companies WHERE id IN ({q_marks});",
                incoming_company_ids,
            ).fetchall()
            existing_company_owner = {
                int(r["id"]): (int(r["owner_id"]) if r["owner_id"] is not None else None)
                for r in rows
            }
        if incoming_prospect_ids:
            q_marks = ",".join("?" for _ in incoming_prospect_ids)
            rows = conn.execute(
                f"SELECT id, owner_id FROM prospects WHERE id IN ({q_marks});",
                incoming_prospect_ids,
            ).fetchall()
            existing_prospect_owner = {
                int(r["id"]): (int(r["owner_id"]) if r["owner_id"] is not None else None)
                for r in rows
            }

        next_company_id = int(conn.execute(
            "SELECT COALESCE(MAX(id), 0) AS n FROM companies;"
        ).fetchone()["n"])
        next_prospect_id = int(conn.execute(
            "SELECT COALESCE(MAX(id), 0) AS n FROM prospects;"
        ).fetchone()["n"])

        remapped_company_ids: Dict[int, int] = {}
        used_company_ids: set[int] = set()
        for c in companies:
            cid = _safe_int(c.get("id"), 0)
            conflict_owner = cid > 0 and cid in existing_company_owner and existing_company_owner[cid] != uid
            duplicate_payload = cid > 0 and cid in used_company_ids
            missing_or_invalid = cid <= 0
            if conflict_owner or duplicate_payload or missing_or_invalid:
                next_company_id += 1
                while next_company_id in used_company_ids:
                    next_company_id += 1
                if cid > 0 and cid not in remapped_company_ids:
                    remapped_company_ids[cid] = next_company_id
                cid = next_company_id
            c["id"] = cid
            used_company_ids.add(cid)

        if remapped_company_ids:
            for p in prospects:
                old_company_id = _safe_int(p.get("company_id"), 0)
                if old_company_id in remapped_company_ids:
                    p["company_id"] = remapped_company_ids[old_company_id]

        used_prospect_ids: set[int] = set()
        for p in prospects:
            pid = _safe_int(p.get("id"), 0)
            conflict_owner = pid > 0 and pid in existing_prospect_owner and existing_prospect_owner[pid] != uid
            duplicate_payload = pid > 0 and pid in used_prospect_ids
            missing_or_invalid = pid <= 0
            if conflict_owner or duplicate_payload or missing_or_invalid:
                next_prospect_id += 1
                while next_prospect_id in used_prospect_ids:
                    next_prospect_id += 1
                pid = next_prospect_id
            p["id"] = pid
            used_prospect_ids.add(pid)

        company_ids = [int(c["id"]) for c in companies if c.get("id") is not None]
        prospect_ids = [int(p["id"]) for p in prospects if p.get("id") is not None]

        # Guard: comptage des données de l'utilisateur courant uniquement (hors soft-deleted)
        try:
            existing_companies_n = int(conn.execute("SELECT COUNT(*) AS n FROM companies WHERE owner_id=? AND deleted_at IS NULL;", (uid,)).fetchone()["n"])
            existing_prospects_n = int(conn.execute(
                "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND deleted_at IS NULL;", (uid,)
            ).fetchone()["n"])
        except Exception:
            existing_companies_n = 0
            existing_prospects_n = 0

        force = bool(data.get("force")) or bool(data.get("confirm_mass_delete"))
        if not force:
            incoming_companies_n = len(company_ids)
            incoming_prospects_n = len(prospect_ids)
            if existing_companies_n >= 25 and incoming_companies_n < max(5, int(existing_companies_n * 0.5)):
                raise ValueError("Payload incomplet (entreprises). Refus de supprimer en masse. Rechargez la page puis réessayez, ou envoyez confirm_mass_delete=true.")
            if existing_prospects_n >= 50 and incoming_prospects_n < max(10, int(existing_prospects_n * 0.5)):
                raise ValueError("Payload incomplet (prospects). Refus de supprimer en masse. Rechargez la page puis réessayez, ou envoyez confirm_mass_delete=true.")

        cur = conn.cursor()
        # Désactiver temporairement les FK pour permettre l'ordre DELETE prospects puis DELETE companies (RESTRICT sinon bloquant au commit)
        try:
            cur.execute("PRAGMA foreign_keys = OFF;")
        except Exception:
            pass
        cur.execute("BEGIN;")
        # Snapshot previous prospect statuses/contacts for RDV events and resume hints
        old_prospect_map = {}
        try:
            if prospect_ids:
                q_marks = ",".join("?" for _ in prospect_ids)
                rows0 = cur.execute(
                    f"SELECT id, statut, rdvDate, lastContact FROM prospects WHERE owner_id=? AND id IN ({q_marks});",
                    [uid] + prospect_ids,
                ).fetchall()
                old_prospect_map = {
                    int(r["id"]): {
                        "statut": r["statut"],
                        "rdvDate": r["rdvDate"],
                        "lastContact": r["lastContact"],
                    }
                    for r in rows0
                }
        except Exception:
            old_prospect_map = {}
        try:
            # Capturer les prospects qui seront supprimés pour le journal d'activité
            _deleted_prospects_for_log = []
            # Les suppressions par omission ne se font que si confirm_mass_delete=true (suppression explicite en masse).
            # Cela évite qu'un onglet en mode prospection efface les prospects ajoutés depuis un autre onglet.
            if force:
                try:
                    if prospect_ids:
                        _dm = ",".join("?" for _ in prospect_ids)
                        _del_rows = cur.execute(
                            f"SELECT id, name FROM prospects WHERE owner_id=? AND deleted_at IS NULL AND id NOT IN ({_dm});",
                            [uid] + prospect_ids
                        ).fetchall()
                    else:
                        _del_rows = cur.execute("SELECT id, name FROM prospects WHERE owner_id=? AND deleted_at IS NULL;", (uid,)).fetchall()
                    _deleted_prospects_for_log = [(int(r["id"]), r["name"]) for r in _del_rows]
                except Exception:
                    _deleted_prospects_for_log = []

                # 1) Supprimer d'abord les prospects qui référencent des entreprises qu'on va supprimer (évite FK RESTRICT au commit)
                # v27.10: AND deleted_at IS NULL — ne pas toucher aux enregistrements soft-deleted (fenêtre d'annulation)
                if company_ids:
                    q_marks = ",".join("?" for _ in company_ids)
                    cur.execute(
                        f"DELETE FROM prospects WHERE owner_id=? AND deleted_at IS NULL AND company_id IN (SELECT id FROM companies WHERE owner_id=? AND id NOT IN ({q_marks}));",
                        [uid, uid] + company_ids,
                    )
                else:
                    cur.execute("DELETE FROM prospects WHERE owner_id=? AND deleted_at IS NULL;", (uid,))

                # 2) Supprimer les prospects de l'utilisateur courant qui ne sont plus dans le payload
                if prospect_ids:
                    q_marks = ",".join("?" for _ in prospect_ids)
                    cur.execute(
                        f"DELETE FROM prospects WHERE owner_id=? AND deleted_at IS NULL AND id NOT IN ({q_marks});",
                        [uid] + prospect_ids,
                    )
                else:
                    cur.execute("DELETE FROM prospects WHERE owner_id=? AND deleted_at IS NULL;", (uid,))

                # 3) Supprimer les entreprises de l'utilisateur courant absentes du payload
                if company_ids:
                    q_marks = ",".join("?" for _ in company_ids)
                    cur.execute(
                        f"DELETE FROM companies WHERE owner_id=? AND deleted_at IS NULL AND id NOT IN ({q_marks});",
                        [uid] + company_ids,
                    )
                else:
                    cur.execute("DELETE FROM companies WHERE owner_id=? AND deleted_at IS NULL;", (uid,))

            # Upsert companies (owner_id forcé à l'utilisateur connecté)
            cur.executemany(
                '''
                INSERT INTO companies (id, groupe, site, phone, notes, tags, website, linkedin, industry, size, address, city, country, stack, pain_points, budget, urgency, owner_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    groupe=excluded.groupe,
                    site=excluded.site,
                    phone=excluded.phone,
                    notes=excluded.notes,
                    tags=excluded.tags,
                    website=excluded.website,
                    linkedin=excluded.linkedin,
                    industry=excluded.industry,
                    size=excluded.size,
                    address=excluded.address,
                    city=excluded.city,
                    country=excluded.country,
                    stack=excluded.stack,
                    pain_points=excluded.pain_points,
                    budget=excluded.budget,
                    urgency=excluded.urgency,
                    owner_id=excluded.owner_id,
                    deleted_at=NULL
                ;
                ''',
                [
                    (
                        int(c["id"]),
                        str(c.get("groupe", "")),
                        str(c.get("site", "")),
                        c.get("phone"),
                        c.get("notes"),
                        c.get("tags"),
                        c.get("website"),
                        c.get("linkedin"),
                        c.get("industry"),
                        c.get("size"),
                        c.get("address"),
                        c.get("city"),
                        c.get("country"),
                        c.get("stack"),
                        c.get("pain_points"),
                        c.get("budget"),
                        c.get("urgency"),
                        int(c.get("owner_id", uid)),
                    )
                    for c in companies
                ],
            )

            # Quand un statut change via /api/save sans lastContact explicite, forcer le datetime courant.
            # Cela permet une reprise mobile plus robuste basée sur le "dernier contact".
            now_iso = _now_iso()
            for p in prospects:
                try:
                    pid = int(p.get("id"))
                except Exception:
                    continue
                old_row = old_prospect_map.get(pid) or {}
                old_statut = str(old_row.get("statut") or "").strip()
                old_last = str(old_row.get("lastContact") or "").strip()
                new_statut = str(p.get("statut") or "").strip()
                incoming_last = str(p.get("lastContact") or "").strip()

                if not incoming_last:
                    p["lastContact"] = now_iso
                    continue
                if old_statut and new_statut and old_statut != new_statut and incoming_last == old_last:
                    p["lastContact"] = now_iso

            # Upsert prospects (owner_id forcé à l'utilisateur connecté)
            cur.executemany(
                '''
                INSERT INTO prospects
                (id, name, company_id, fonction, telephone, email, linkedin, pertinence, statut, lastContact, nextFollowUp, priority, notes, callNotes, pushEmailSentAt, tags, template_id, nextAction, pushLinkedInSentAt, photo_url, push_category_id, fixedMetier, rdvDate, is_archived, owner_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    name=excluded.name,
                    company_id=excluded.company_id,
                    fonction=excluded.fonction,
                    telephone=excluded.telephone,
                    email=excluded.email,
                    linkedin=excluded.linkedin,
                    pertinence=excluded.pertinence,
                    statut=excluded.statut,
                    lastContact=excluded.lastContact,
                    nextFollowUp=excluded.nextFollowUp,
                    priority=excluded.priority,
                    notes=excluded.notes,
                    callNotes=excluded.callNotes,
                    pushEmailSentAt=excluded.pushEmailSentAt,
                    tags=excluded.tags,
                    template_id=excluded.template_id,
                    nextAction=excluded.nextAction,
                    pushLinkedInSentAt=excluded.pushLinkedInSentAt,
                    photo_url=excluded.photo_url,
                    push_category_id=excluded.push_category_id,
                    fixedMetier=excluded.fixedMetier,
                    rdvDate=excluded.rdvDate,
                    is_archived=excluded.is_archived,
                    owner_id=excluded.owner_id,
                    deleted_at=NULL
                ;
                ''',
                [
                    (
                        int(p["id"]),
                        str(p.get("name", "")),
                        _safe_int(p.get("company_id"), 0),
                        p.get("fonction"),
                        p.get("telephone"),
                        p.get("email"),
                        p.get("linkedin"),
                        p.get("pertinence"),
                        p.get("statut"),
                        p.get("lastContact"),
                        p.get("nextFollowUp"),
                        p.get("priority"),
                        p.get("notes"),
                        p.get("callNotes"),
                        p.get("pushEmailSentAt"),
                        p.get("tags"),
                        p.get("template_id"),
                        p.get("nextAction"),
                        p.get("pushLinkedInSentAt"),
                        p.get("photo_url"),
                        p.get("push_category_id"),
                        p.get("fixedMetier"),
                        p.get("rdvDate"),
                        p.get("is_archived", 0),
                        int(p.get("owner_id", uid)),
                    )
                    for p in prospects
                ],
            )

            # Log statut changes for debugging persistence issues
            for p in prospects:
                try:
                    pid = int(p.get("id"))
                    old_row = old_prospect_map.get(pid) or {}
                    old_s = str(old_row.get("statut") or "").strip()
                    new_s = str(p.get("statut") or "").strip()
                    if old_s != new_s:
                        logger.info("[upsert_all] prospect %d statut: %r → %r", pid, old_s, new_s)
                        row = cur.execute("SELECT statut FROM prospects WHERE id=?", (pid,)).fetchone()
                        if row:
                            saved_s = str(row[0] or "").strip()
                            if saved_s != new_s:
                                logger.warning("[upsert_all] statut DB mismatch pour prospect %d : attendu %r, trouvé %r", pid, new_s, saved_s)
                except Exception as _log_err:
                    logger.debug("[upsert_all] erreur log statut: %s", _log_err)

            # Log "RDV pris" events for gamified goals (deduped by unique index)
            try:
                now_ev = datetime.datetime.now().isoformat(timespec="seconds")
                ev_date = now_ev[:10]
                for p in prospects:
                    pid = int(p.get("id"))
                    new_statut = (p.get("statut") or "").strip()
                    new_rdv = (p.get("rdvDate") or "").strip()
                    old_row = old_prospect_map.get(pid) or {}
                    old_statut = old_row.get("statut")
                    old_rdv = old_row.get("rdvDate")
                    if new_statut == "Rendez-vous" and new_rdv:
                        if old_statut != "Rendez-vous" or (str(old_rdv or "").strip() != new_rdv):
                            cur.execute(
                                "INSERT OR IGNORE INTO prospect_events (prospect_id, date, type, title, content, meta, createdAt) VALUES (?,?,?,?,?,?,?)",
                                (pid, ev_date, "rdv_taken", "RDV pris", None, json.dumps({"rdvDate": new_rdv}, ensure_ascii=False), now_ev),
                            )
                            # Teams webhook: RDV pris (v22.1)
                            try:
                                _p_name = (p.get("name") or "").strip()
                                _c_row = cur.execute("SELECT groupe FROM companies WHERE id=? AND owner_id=?;", (p.get("company_id"), uid)).fetchone()
                                _c_name = _c_row[0] if _c_row else ""
                                _prefix = _get_user_prefix(uid)
                                _card = _build_adaptive_card(
                                    "RDV pris",
                                    [("Prospect", _p_name), ("Entreprise", _c_name), ("Date RDV", new_rdv), ("Consultant", _prefix)],
                                    [{"title": "Voir prospect", "url": f"https://prospup.work/entreprises?highlight={pid}"}]
                                )
                                _send_teams_webhook(_card, "rdv_taken")
                            except Exception:
                                pass
            except Exception:
                pass

            cur.execute("COMMIT;")

            # Journal d'activité — suppressions et créations de prospects
            try:
                for (_dp_id, _dp_name) in _deleted_prospects_for_log:
                    log_activity('delete', 'prospect', _dp_id, _dp_name)
                for _p in prospects:
                    _pid = int(_p.get("id"))
                    if _pid not in old_prospect_map:
                        log_activity('create', 'prospect', _pid, _p.get("name"))
            except Exception:
                pass

            # Hooks pour création automatique de tâches
            # Détecter les nouveaux prospects et changements de statut
            for p in prospects:
                try:
                    pid = int(p.get("id"))
                    old_row = old_prospect_map.get(pid) or {}
                    is_new = pid not in old_prospect_map
                    statut_changed = old_row.get("statut") != p.get("statut")
                    
                    # Construire le contexte pour les règles
                    context = {
                        "prospect_id": pid,
                        "name": p.get("name", ""),
                        "email": p.get("email"),
                        "telephone": p.get("telephone"),
                        "linkedin": p.get("linkedin"),
                        "statut": p.get("statut"),
                        "pertinence": p.get("pertinence"),
                        "nextFollowUp": p.get("nextFollowUp"),
                        "company_id": p.get("company_id"),
                    }
                    
                    # Récupérer le nom de l'entreprise si disponible
                    if context.get("company_id"):
                        try:
                            c_row = conn.execute(
                                "SELECT groupe FROM companies WHERE id=? AND owner_id=?;",
                                (context["company_id"], uid)
                            ).fetchone()
                            if c_row:
                                context["company_groupe"] = c_row["groupe"] or ""
                        except Exception:
                            pass
                    
                    # Hook: prospect créé
                    if is_new:
                        _create_auto_task("prospect_created", context)
                    
                    # Hook: statut changé
                    if statut_changed and p.get("statut"):
                        context["old_statut"] = old_row.get("statut")
                        _create_auto_task("status_changed", context)
                except Exception as e:
                    logger.warning("Erreur hook tâche auto pour prospect %s: %s", p.get("id"), e)
        except Exception:
            cur.execute("ROLLBACK;")
            raise
        finally:
            try:
                cur.execute("PRAGMA foreign_keys = ON;")
            except Exception:
                pass

def replace_all(data: Dict[str, Any]) -> None:
    companies = data.get("companies") or []
    prospects = data.get("prospects") or []

    if not isinstance(companies, list) or not isinstance(prospects, list):
        raise ValueError("Invalid payload: companies/prospects must be lists")

    def _dump_tags(v) -> str:
        if v is None:
            return "[]"
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return "[]"
            # allow comma separated string
            if s.startswith("["):
                return s
            parts = [t.strip() for t in s.split(",") if t.strip()]
            return json.dumps(parts, ensure_ascii=False)
        if isinstance(v, list):
            parts = [str(t).strip() for t in v if str(t).strip()]
            return json.dumps(parts, ensure_ascii=False)
        return "[]"

    for c in companies:
        c["tags"] = _dump_tags(c.get("tags"))

    for p in prospects:
        if isinstance(p.get("callNotes"), list):
            p["callNotes"] = json.dumps(p["callNotes"], ensure_ascii=False)
        else:
            p["callNotes"] = "[]"
        p["tags"] = _dump_tags(p.get("tags"))
        # template_id should be int or None
        tid = p.get("template_id")
        try:
            p["template_id"] = int(tid) if tid not in (None, "", "null") else None
        except Exception:
            p["template_id"] = None

    try:
        with _auth_conn() as aconn:
            first_user = aconn.execute("SELECT id FROM users WHERE is_active=1 ORDER BY id LIMIT 1;").fetchone()
            seed_owner_id = int(first_user["id"]) if first_user else None
    except Exception:
        seed_owner_id = None

    with _conn() as conn:
        # IMPORTANT: /api/save remplace complètement companies/prospects.
        # On préserve donc l'historique des push pour éviter de le perdre à chaque sauvegarde.
        try:
            existing_push_logs = [
                dict(r)
                for r in conn.execute(
                    "SELECT id, prospect_id, sentAt, channel, to_email, subject, body, template_id, template_name, createdAt FROM push_logs;"
                ).fetchall()
            ]
        except sqlite3.OperationalError as e:
            logger.warning("push_logs read failed (table may not exist yet): %s", e)
            existing_push_logs = []

        cur = conn.cursor()
        cur.execute("BEGIN;")
        try:
            cur.execute("DELETE FROM prospects;")
            cur.execute("DELETE FROM companies;")

            cur.executemany(
                "INSERT INTO companies (id, groupe, site, phone, notes, tags, owner_id) VALUES (?, ?, ?, ?, ?, ?, ?);",
                [
                    (
                        int(c["id"]),
                        str(c.get("groupe", "")),
                        str(c.get("site", "")),
                        c.get("phone"),
                        c.get("notes"),
                        c.get("tags"),
                        seed_owner_id,
                    )
                    for c in companies
                ],
            )

            cur.executemany(
                '''
                INSERT INTO prospects
                (id, name, company_id, fonction, telephone, email, linkedin, pertinence, statut, lastContact, nextFollowUp, priority, notes, callNotes, pushEmailSentAt, tags, template_id, nextAction, pushLinkedInSentAt, photo_url, push_category_id, fixedMetier, rdvDate, is_archived, owner_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                ''',
                [
                    (
                        int(p["id"]),
                        str(p.get("name", "")),
                        _safe_int(p.get("company_id"), 0),
                        p.get("fonction"),
                        p.get("telephone"),
                        p.get("email"),
                        p.get("linkedin"),
                        p.get("pertinence"),
                        p.get("statut"),
                        p.get("lastContact"),
                        p.get("nextFollowUp"),
                        p.get("priority"),
                        p.get("notes"),
                        p.get("callNotes"),
                        p.get("pushEmailSentAt"),
                        p.get("tags"),
                        p.get("template_id"),
                        p.get("nextAction"),
                        p.get("pushLinkedInSentAt"),
                        p.get("photo_url"),
                        p.get("push_category_id"),
                        p.get("fixedMetier"),
                        p.get("rdvDate"),
                        p.get("is_archived", 0),
                        seed_owner_id,
                    )
                    for p in prospects
                ],
            )

            # Restaurer l'historique des push pour les prospects encore présents
            if existing_push_logs:
                kept_prospect_ids = {int(p.get("id")) for p in prospects if p.get("id") is not None}
                logs_to_restore = [
                    l for l in existing_push_logs
                    if (l.get("prospect_id") is not None and int(l["prospect_id"]) in kept_prospect_ids)
                ]

                if logs_to_restore:
                    cur.executemany(
                        '''
                        INSERT INTO push_logs (id, prospect_id, sentAt, channel, to_email, subject, body, template_id, template_name, createdAt)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                        ''',
                        [
                            (
                                int(l["id"]),
                                int(l["prospect_id"]),
                                str(l.get("sentAt") or ""),
                                l.get("channel"),
                                l.get("to_email"),
                                l.get("subject"),
                                l.get("body"),
                                l.get("template_id"),
                                l.get("template_name"),
                                str(l.get("createdAt") or l.get("sentAt") or ""),
                            )
                            for l in logs_to_restore
                        ],
                    )

            cur.execute("COMMIT;")
        except Exception:
            cur.execute("ROLLBACK;")
            raise





def load_initial_data_if_needed() -> None:
    if not db_is_empty():
        return
    if not INITIAL_JSON.exists():
        return
    replace_all(json.loads(INITIAL_JSON.read_text(encoding="utf-8")))
    # Neutraliser le fichier après le seed pour éviter un écrasement accidentel
    try:
        bak = INITIAL_JSON.with_suffix(".json.bak")
        INITIAL_JSON.rename(bak)
    except Exception:
        pass  # pas grave si le rename échoue (permissions, etc.)

def seed_from_initial() -> dict:
    """Force seed the DB from initial_data.json (used by /api/reset).
    Returns info about the seed source for the UI to display."""
    # Try initial_data.json first, then .bak as fallback
    source = None
    if INITIAL_JSON.exists():
        source = INITIAL_JSON
    else:
        bak = INITIAL_JSON.with_suffix(".json.bak")
        if bak.exists():
            source = bak

    if source is None:
        return {"seeded": False, "reason": "Aucun fichier initial_data.json trouvé"}

    payload = json.loads(source.read_text(encoding="utf-8"))
    replace_all(payload)

    mtime = datetime.datetime.fromtimestamp(source.stat().st_mtime)
    nb_c = len(payload.get("companies", []))
    nb_p = len(payload.get("prospects", []))
    return {
        "seeded": True,
        "source": source.name,
        "source_date": mtime.isoformat(timespec="seconds"),
        "companies": nb_c,
        "prospects": nb_p,
    }



def _audit_log(action: str, entity: str, entity_id: int | None = None,
               old_value: str | None = None, new_value: str | None = None):
    """v23.5: Write an entry to the audit trail."""
    uid = _uid()
    if not uid:
        return
    ip = request.remote_addr or "unknown"
    try:
        with _conn() as conn:
            conn.execute(
                "INSERT INTO audit_log (user_id, action, entity, entity_id, old_value, new_value, ip, createdAt) VALUES (?,?,?,?,?,?,?,?);",
                (uid, action, entity, entity_id, old_value, new_value, ip, _now_iso())
            )
    except Exception as e:
        logger.warning("audit_log failed: %s", e)  # Never break the main flow


def log_activity(action: str, entity_type: str = None, entity_id: int = None,
                 entity_label: str = None, details: dict = None):
    """v27.10: Enregistre une action dans activity_logs. Non-bloquant, toujours en DB principale."""
    try:
        uid = _uid()
        if not uid:
            return
        username = session.get('user_name', 'inconnu')
        ip = request.remote_addr or 'unknown'
        details_json = json.dumps(details, ensure_ascii=False) if details else None
        with _auth_conn() as conn:
            conn.execute(
                "INSERT INTO activity_logs (user_id, username, action, entity_type, entity_id, entity_label, details, ip_address) "
                "VALUES (?,?,?,?,?,?,?,?);",
                (uid, username, action, entity_type, entity_id, entity_label, details_json, ip)
            )
    except Exception as e:
        logger.warning("log_activity failed: %s", e)


def _create_auto_task(trigger_type: str, context: Dict[str, Any]) -> None:
    """Crée automatiquement des tâches basées sur les règles actives.
    
    Args:
        trigger_type: Type de déclencheur ('prospect_created', 'status_changed', 'meeting_done', 'daily_check')
        context: Contexte avec les données nécessaires (prospect_id, statut, etc.)
    """
    uid = _uid()
    if not uid:
        return
    
    try:
        with _conn() as conn:
            # Récupérer les règles actives pour ce trigger
            rules = conn.execute(
                "SELECT * FROM task_rules WHERE trigger_type=? AND enabled=1 AND (owner_id IS NULL OR owner_id=?);",
                (trigger_type, uid)
            ).fetchall()
            
            if not rules:
                return
            
            for rule in rules:
                rule_dict = dict(rule)
                conditions = json.loads(rule_dict.get("conditions") or "{}")
                
                # Évaluer les conditions
                if not _evaluate_task_conditions(conditions, context):
                    continue
                
                # Générer le titre et commentaire
                template_title = rule_dict.get("template_title") or ""
                template_comment = rule_dict.get("template_comment") or ""
                
                # Remplacer les variables du contexte dans les templates
                title = _render_task_template(template_title, context)
                comment = _render_task_template(template_comment, context)
                
                # Générer via IA si le template contient {{IA:...}}
                if "{{IA:" in title:
                    title = _generate_task_text_ia(title, context)
                if "{{IA:" in comment:
                    comment = _generate_task_text_ia(comment, context)
                
                # Déterminer la date d'échéance
                due_date = _calculate_task_due_date(context, conditions)
                
                # Construire linked_ids
                linked_ids = {}
                if context.get("prospect_id"):
                    linked_ids["prospect_id"] = context["prospect_id"]
                if context.get("company_id"):
                    linked_ids["company_id"] = context["company_id"]
                
                # Créer la tâche
                now = _now_iso()
                priority = rule_dict.get("priority") or 2
                conn.execute(
                    "INSERT INTO tasks (title, comment, due_date, status, linked_ids, priority, createdAt, updatedAt, owner_id) VALUES (?, ?, ?, 'pending', ?, ?, ?, ?, ?);",
                    (title, comment, due_date, json.dumps(linked_ids, ensure_ascii=False), priority, now, now, uid)
                )
                logger.info("Tâche auto-créée: %s (règle: %s)", title, rule_dict.get("name"))
    except Exception as e:
        logger.warning("Erreur création tâche auto: %s", e)


def _evaluate_task_conditions(conditions: Dict[str, Any], context: Dict[str, Any]) -> bool:
    """Évalue si les conditions sont remplies pour créer une tâche.
    
    Exemples de conditions:
    - {"has_email": False} : prospect sans email
    - {"statut": "Rendez-vous"} : statut spécifique
    - {"nextFollowUp_days": 2} : nextFollowUp dans N jours
    - {"has_linkedin": False} : pas de LinkedIn
    """
    if not conditions:
        return True
    
    # Condition: prospect sans email
    if conditions.get("has_email") is False:
        if context.get("email"):
            return False
    
    # Condition: statut spécifique
    if "statut" in conditions:
        if context.get("statut") != conditions["statut"]:
            return False
    
    # Condition: nextFollowUp dans N jours
    if "nextFollowUp_days" in conditions:
        next_follow = context.get("nextFollowUp")
        if not next_follow:
            return False
        try:
            follow_date = datetime.datetime.fromisoformat(next_follow.replace("Z", "+00:00")[:10])
            today = datetime.date.today()
            days_diff = (follow_date.date() - today).days
            if days_diff != conditions["nextFollowUp_days"]:
                return False
        except Exception:
            return False
    
    # Condition: pas de LinkedIn
    if conditions.get("has_linkedin") is False:
        if context.get("linkedin"):
            return False
    
    # Condition: pertinence spécifique
    if "pertinence" in conditions:
        if context.get("pertinence") != conditions["pertinence"]:
            return False
    
    return True


def _render_task_template(template: str, context: Dict[str, Any]) -> str:
    """Remplace les variables {{var}} dans le template par les valeurs du contexte."""
    if not template:
        return ""
    
    result = template
    # Remplacer {{variable}} par la valeur du contexte
    for key, value in context.items():
        placeholder = f"{{{{{key}}}}}"
        if placeholder in result:
            result = result.replace(placeholder, str(value or ""))
    
    return result


def _generate_task_text_ia(template: str, context: Dict[str, Any]) -> str:
    """Génère un texte via IA si le template contient {{IA:prompt}}.
    
    Exemple: "{{IA:Génère un titre de tâche pour : relancer {{name}} de {{company_groupe}} concernant {{context}}. Titre court (max 50 caractères).}}"
    """
    if "{{IA:" not in template:
        return template
    
    # Extraire le prompt IA
    import re
    match = re.search(r'\{\{IA:([^}]+)\}\}', template)
    if not match:
        return template
    
    prompt_template = match.group(1)
    
    # Remplacer les variables du contexte dans le prompt
    prompt = _render_task_template(prompt_template, context)
    
    try:
        # Appeler l'IA
        generated = _call_ai(prompt, timeout=30)
        # Remplacer {{IA:...}} par le texte généré
        result = template.replace(match.group(0), generated.strip())
        return result
    except Exception as e:
        logger.warning("Erreur génération IA pour tâche: %s", e)
        # Fallback: retourner le template sans {{IA:...}}
        return template.replace(match.group(0), "")


def _calculate_task_due_date(context: Dict[str, Any], conditions: Dict[str, Any]) -> str | None:
    """Calcule la date d'échéance de la tâche basée sur le contexte et les conditions."""
    # Si nextFollowUp est défini, utiliser cette date
    if context.get("nextFollowUp"):
        return context["nextFollowUp"]
    
    # Si une condition spécifie un délai
    if "due_days" in conditions:
        days = int(conditions.get("due_days", 0))
        due_date = datetime.date.today() + datetime.timedelta(days=days)
        return due_date.isoformat()
    
    # Par défaut: aujourd'hui
    return _today_iso()


def _optimize_task_schedule(tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Suggère l'ordre optimal des tâches basé sur :
    - Priorité
    - Date d'échéance
    - Type de tâche (regroupement)
    - Estimation de temps
    
    Retourne les tâches triées avec des suggestions d'ordre.
    """
    if not tasks:
        return []
    
    # Estimation de temps par type de tâche (en minutes)
    task_time_estimates = {
        "relance": 15,
        "appel": 20,
        "email": 10,
        "rdv": 30,
        "suivi": 15,
        "default": 20,
    }
    
    # Calculer un score pour chaque tâche
    scored_tasks = []
    for task in tasks:
        score = 0
        priority = task.get("priority") or 2
        due_date = task.get("due_date")
        title = (task.get("title") or "").lower()
        
        # Score basé sur la priorité (plus bas = plus urgent)
        score += (4 - priority) * 100
        
        # Score basé sur la date d'échéance
        if due_date:
            try:
                due = datetime.datetime.fromisoformat(due_date.replace("Z", "+00:00")[:10]).date()
                today = datetime.date.today()
                days_until = (due - today).days
                if days_until < 0:
                    score += 200  # En retard
                elif days_until == 0:
                    score += 150  # Aujourd'hui
                elif days_until == 1:
                    score += 100  # Demain
                elif days_until <= 3:
                    score += 50  # Cette semaine
            except Exception:
                pass
        
        # Estimation de temps
        time_estimate = task_time_estimates.get("default")
        for task_type, minutes in task_time_estimates.items():
            if task_type in title:
                time_estimate = minutes
                break
        task["estimated_minutes"] = time_estimate
        
        # Score négatif pour les tâches longues (prioriser les courtes)
        score -= time_estimate / 10
        
        scored_tasks.append((score, task))
    
    # Trier par score décroissant
    scored_tasks.sort(key=lambda x: x[0], reverse=True)
    
    # Regrouper les tâches similaires (même type, même entreprise)
    grouped = []
    current_group = []
    for score, task in scored_tasks:
        if not current_group:
            current_group.append(task)
        else:
            # Vérifier si on peut regrouper
            prev_task = current_group[0]
            prev_title = (prev_task.get("title") or "").lower()
            curr_title = (task.get("title") or "").lower()
            
            # Regrouper si même type de tâche
            can_group = False
            for task_type in task_time_estimates.keys():
                if task_type in prev_title and task_type in curr_title:
                    can_group = True
                    break
            
            # Regrouper si même entreprise (via linked_ids)
            try:
                prev_linked = json.loads(prev_task.get("linked_ids") or "{}")
                curr_linked = json.loads(task.get("linked_ids") or "{}")
                if prev_linked.get("company_id") and curr_linked.get("company_id"):
                    if prev_linked.get("company_id") == curr_linked.get("company_id"):
                        can_group = True
            except Exception:
                pass
            
            if can_group and len(current_group) < 3:  # Max 3 tâches par groupe
                current_group.append(task)
            else:
                grouped.append(current_group)
                current_group = [task]
    
    if current_group:
        grouped.append(current_group)
    
    # Aplatir et retourner
    result = []
    for group in grouped:
        result.extend(group)
    
    return result


def _snapshot_dir_for_user(user_id: int | None = None) -> Path:
    """Répertoire de snapshots pour un utilisateur (per-user ou global)."""
    if user_id:
        user_db = DATA_DIR / f"user_{user_id}" / "prospects.db"
        if user_db.exists():
            snap = user_db.parent / "snapshots"
            snap.mkdir(parents=True, exist_ok=True)
            return snap
    SNAPSHOT_DIR.mkdir(exist_ok=True)
    return SNAPSHOT_DIR


def _current_user_db_path() -> Path:
    """DB path de l'utilisateur courant (pour snapshots)."""
    try:
        uid = session.get("user_id")
        if uid:
            return _user_db_path(uid)
    except RuntimeError:
        pass
    return DB_PATH


def _auto_snapshot_if_needed() -> None:
    """Crée au plus 1 snapshot auto par jour. Garde les 14 derniers. Per-user aware."""
    try:
        uid = None
        try:
            uid = session.get("user_id")
        except RuntimeError:
            pass
        snap_dir = _snapshot_dir_for_user(uid)
        src_db = _current_user_db_path()

        today = _today_iso()
        existing = sorted(snap_dir.glob(f"auto_{today}_*.db"))
        if existing:
            return

        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"auto_{today}_{ts}.db"
        path = snap_dir / filename

        src = sqlite3.connect(src_db)
        try:
            dst = sqlite3.connect(path)
            try:
                src.backup(dst)
            finally:
                dst.close()
        finally:
            src.close()

        autos = sorted(snap_dir.glob("auto_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
        for p in autos[14:]:
            try:
                p.unlink()
            except Exception:
                pass
    except Exception:
        return


# Snapshots SQLite — voir utils/snapshots.py


# ─────────────────────────────────────────────────────────────────
# Redirects legacy → v30 (v31.7+) : v29 archivée dans archives/v29/.
# Les routes ci-dessous restent pour préserver bookmarks, partages
# externes et PWA shortcuts. Toute URL sans équivalent v30 → 404.
# ─────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────
# Routes de rendu HTML (legacy redirects + v30 pages) → routes/pages.py
# ─────────────────────────────────────────────────────────────────


# ── Mode Prosp: server-side token sessions ──
_MODE_PROSP_TTL = 3600 * 8  # 8 heures (survit aux redémarrages grâce à la DB)


def _mode_prosp_cleanup():
    """Supprime les sessions Mode Prosp expirées de la DB."""
    try:
        with _conn() as conn:
            conn.execute(
                "DELETE FROM mode_prosp_sessions WHERE ? - created_at > ?;",
                (time.time(), _MODE_PROSP_TTL)
            )
    except Exception as e:
        logger.warning("mode_prosp_cleanup: %s", e)


def _mode_prosp_auth(token: str):
    """Valide un token Mode Prosp depuis la DB et retourne le dict de session ou None."""
    if not token:
        return None
    try:
        with _conn() as conn:
            row = conn.execute(
                "SELECT user_id, ids, created_at FROM mode_prosp_sessions WHERE token=?;",
                (token,)
            ).fetchone()
    except Exception:
        return None
    if not row:
        return None
    row = dict(row)
    if time.time() - row['created_at'] > _MODE_PROSP_TTL:
        try:
            with _conn() as conn:
                conn.execute("DELETE FROM mode_prosp_sessions WHERE token=?;", (token,))
        except Exception:
            pass
        return None
    try:
        ids = json.loads(row['ids'])
    except Exception:
        ids = []
    return {
        'user_id': row['user_id'],
        'ids': ids,
        'created_at': row['created_at']
    }


@app.post("/api/mode-prosp/start")
def mode_prosp_start():
    """Create a mode-prosp session: store filtered prospect IDs in DB, return a token."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    ids = request.json.get('ids', [])
    if not isinstance(ids, list) or len(ids) == 0:
        return jsonify(ok=False, error="Aucun prospect"), 400
    # Validate IDs are integers
    ids = [int(i) for i in ids if str(i).isdigit()]
    if not ids:
        return jsonify(ok=False, error="IDs invalides"), 400
    token = secrets.token_urlsafe(16)
    try:
        with _conn() as conn:
            conn.execute(
                "INSERT INTO mode_prosp_sessions (token, user_id, ids, created_at) VALUES (?, ?, ?, ?);",
                (token, uid, json.dumps(ids), time.time())
            )
    except Exception as e:
        logger.error("mode_prosp_start: %s", e)
        return jsonify(ok=False, error="Erreur serveur"), 500
    _mode_prosp_cleanup()
    return jsonify(ok=True, token=token)


@app.get("/api/prospects/quick-filter")
def api_prospects_quick_filter():
    """Retourne des IDs de prospects selon un preset de filtres (usage: dashboard objectifs)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    preset = request.args.get('preset', '')
    try:
        with _conn() as conn:
            if preset == 'push_ready':
                rows = conn.execute(
                    "SELECT id FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') "
                    "AND (is_archived IS NULL OR is_archived=0) "
                    "AND (pushEmailSentAt IS NULL OR pushEmailSentAt='') "
                    "AND (pushLinkedInSentAt IS NULL OR pushLinkedInSentAt='') "
                    "AND email IS NOT NULL AND email!='' "
                    "AND (telephone IS NULL OR telephone='') "
                    "ORDER BY RANDOM() LIMIT 1",
                    (uid,)
                ).fetchall()
            elif preset == 'rdv_ready':
                rows = conn.execute(
                    "SELECT id FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') "
                    "AND (is_archived IS NULL OR is_archived=0) "
                    "AND statut IN ('Messagerie','Pas d''actions','À rappeler') "
                    "AND telephone IS NOT NULL AND telephone!=''",
                    (uid,)
                ).fetchall()
            else:
                return jsonify(ok=False, error="preset inconnu"), 400
        return jsonify(ok=True, ids=[r['id'] for r in rows])
    except Exception as e:
        logger.error("api_prospects_quick_filter: %s", e)
        return jsonify(ok=False, error="Erreur serveur"), 500


@app.get("/api/mode-prosp/data")
def mode_prosp_data():
    """Return only the selected prospects + their companies for a mode-prosp session."""
    token = request.args.get('t', '')
    sess = _mode_prosp_auth(token)
    if not sess:
        return jsonify(ok=False, error="Session expirée ou invalide"), 401
    uid = sess['user_id']
    ids = sess['ids']
    try:
        db_path = _user_db_path(uid)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA busy_timeout = 20000;")
        # Fetch prospects by IDs, respecting owner_id
        placeholders = ','.join('?' * len(ids))
        rows = conn.execute(
            f"SELECT * FROM prospects WHERE id IN ({placeholders}) AND owner_id=? AND deleted_at IS NULL",
            ids + [uid]
        ).fetchall()
        p_map = {r['id']: dict(r) for r in rows}
        # Maintain the original filter order
        prospects_list = [p_map[i] for i in ids if i in p_map]
        # Parse JSON fields
        for p in prospects_list:
            for jf in ('callNotes', 'tags'):
                raw = p.get(jf)
                if isinstance(raw, str) and raw:
                    try:
                        p[jf] = json.loads(raw)
                    except Exception:
                        p[jf] = []
                elif not raw:
                    p[jf] = []
        # Fetch related companies
        company_ids = list(set(p.get('company_id') for p in prospects_list if p.get('company_id')))
        companies_list = []
        if company_ids:
            cp = ','.join('?' * len(company_ids))
            c_rows = conn.execute(
                f"SELECT * FROM companies WHERE id IN ({cp}) AND owner_id=? AND deleted_at IS NULL",
                company_ids + [uid]
            ).fetchall()
            companies_list = [dict(r) for r in c_rows]
        conn.close()
        return jsonify(ok=True, prospects=prospects_list, companies=companies_list)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.post("/api/mode-prosp/save")
def mode_prosp_save():
    """Save a single prospect from mode-prosp. Updates only the editable fields."""
    token = request.args.get('t', '')
    sess = _mode_prosp_auth(token)
    if not sess:
        return jsonify(ok=False, error="Session expirée ou invalide"), 401
    uid = sess['user_id']
    prospect = request.json.get('prospect')
    if not prospect or not prospect.get('id'):
        return jsonify(ok=False, error="Prospect invalide"), 400
    pid = int(prospect['id'])
    # Verify ownership
    if pid not in sess['ids']:
        return jsonify(ok=False, error="Prospect non autorisé"), 403
    # Editable fields from Mode Prosp cards
    EDITABLE = ('statut', 'company_id', 'fonction', 'telephone', 'email', 'linkedin',
                'pertinence', 'priority', 'nextAction', 'nextFollowUp', 'rdvDate', 'lastContact', 'notes')
    try:
        db_path = _user_db_path(uid)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA busy_timeout = 20000;")
        conn.execute("PRAGMA journal_mode = WAL;")
        # Fetch current prospect state before update (for event diffing)
        old_row = conn.execute("SELECT statut, rdvDate FROM prospects WHERE id = ? AND owner_id = ?", (pid, uid)).fetchone()
        old_statut = str(old_row["statut"] or "").strip() if old_row else ""
        old_rdv = str(old_row["rdvDate"] or "").strip() if old_row else ""
        # Build SET clause for only editable fields present in payload
        sets = []
        vals = []
        for f in EDITABLE:
            if f in prospect:
                sets.append(f"{f} = ?")
                val = prospect[f]
                # Convert company_id and priority to int
                if f in ('company_id', 'priority', 'pertinence'):
                    val = int(val) if val is not None else None
                vals.append(val)
        if not sets:
            conn.close()
            return jsonify(ok=True)
        vals.extend([pid, uid])
        conn.execute(
            f"UPDATE prospects SET {', '.join(sets)} WHERE id = ? AND owner_id = ?",
            vals
        )
        # Log "RDV pris" event for gamified goals (same logic as upsert_all)
        try:
            new_statut = str(prospect.get("statut") or "").strip()
            new_rdv = str(prospect.get("rdvDate") or "").strip()
            if new_statut == "Rendez-vous" and new_rdv:
                if old_statut != "Rendez-vous" or old_rdv != new_rdv:
                    now_ev = datetime.datetime.now().isoformat(timespec="seconds")
                    ev_date = now_ev[:10]
                    conn.execute(
                        "INSERT OR IGNORE INTO prospect_events (prospect_id, date, type, title, content, meta, createdAt) VALUES (?,?,?,?,?,?,?)",
                        (pid, ev_date, "rdv_taken", "RDV pris", None, json.dumps({"rdvDate": new_rdv}, ensure_ascii=False), now_ev),
                    )
        except Exception:
            pass
        # Log status change event for timeline (notes & suivi)
        try:
            if "statut" in prospect:
                new_statut = str(prospect.get("statut") or "").strip()
                if new_statut and old_statut != new_statut:
                    ev_at = datetime.datetime.now().isoformat()
                    content_statut = f"{old_statut} → {new_statut}" if old_statut else new_statut
                    conn.execute(
                        "INSERT OR IGNORE INTO prospect_events (prospect_id, date, type, title, content, meta, createdAt) VALUES (?,?,?,?,?,?,?)",
                        (pid, ev_at, "status_change", "Changement de statut", content_statut, None, ev_at),
                    )
        except Exception:
            pass
        conn.commit()
        # Fetch updated prospect to return it
        row = conn.execute("SELECT * FROM prospects WHERE id = ? AND owner_id = ?", (pid, uid)).fetchone()
        conn.close()
        if row:
            updated = dict(row)
            for jf in ('callNotes', 'tags'):
                raw = updated.get(jf)
                if isinstance(raw, str) and raw:
                    try:
                        updated[jf] = json.loads(raw)
                    except Exception:
                        updated[jf] = []
                elif not raw:
                    updated[jf] = []
            return jsonify(ok=True, prospect=updated)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.get("/offline.html")
def page_offline():
    return send_from_directory(APP_DIR, "offline.html")


@app.post("/api/kpi/export/xlsx")
def api_kpi_export_xlsx():
    """Generate a simple KPI validation Excel from a checklist."""
    payload = request.get_json(force=True, silent=True) or {}
    date_iso = str(payload.get("date") or "").strip() or datetime.date.today().isoformat()
    title = str(payload.get("title") or "Validation KPI").strip() or "Validation KPI"
    items = payload.get("items")
    if not isinstance(items, list):
        items = []

    # sanitize items
    cleaned: List[Dict[str, Any]] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        label = str(it.get("label") or "").strip()
        if not label:
            continue
        checked = bool(it.get("checked"))
        note = str(it.get("note") or "").strip().replace("\r\n", "\n").replace("\r", "\n")
        cleaned.append({"label": label[:200], "checked": checked, "note": note[:500]})

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"

    ws.append(["Date", "KPI", "Coche", "Commentaire"])
    for row in cleaned:
        ws.append([date_iso, row["label"], "Oui" if row["checked"] else "Non", row["note"]])

    # basic sizing
    widths = [12, 45, 10, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Résumé")
    ws2.append(["Titre", title])
    ws2.append(["Date", date_iso])
    ws2.append([])
    total = len(cleaned)
    done = sum(1 for r in cleaned if r["checked"])
    ws2.append(["Total KPI", total])
    ws2.append(["Cochés", done])
    ws2.append(["Non cochés", total - done])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 40

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"kpi_{date_iso}.xlsx".replace(":", "-")
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )



# ====== Utils: keywords from prospect notes / fixedMetier (best-candidates) ======
NOTES_STOPWORDS = {
    "les", "des", "une", "que", "dans", "pour", "sur", "avec", "sans", "mais", "ou", "et",
    "son", "sa", "ses", "leur", "ce", "cet", "cette", "aux", "du", "de", "la", "le", "en", "au",
    "par", "est", "sont", "pas", "peu", "plutôt", "très", "plus", "tout", "tous", "toute",
    "comme", "donc", "ainsi", "chez", "entre", "après", "avant", "depuis",
}

def _keywords_from_notes(notes: str | None) -> List[str]:
    """Extract keywords from prospect notes (3+ chars, no stopwords). Weight ×1 in matching."""
    if not notes or not str(notes).strip():
        return []
    text = (notes or "").lower().strip()
    # Normalize accents for word boundaries
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    words = re.findall(r"[a-z0-9éèêëàâäùûüîïôöç]+", text)
    seen = set()
    out = []
    for w in words:
        if len(w) < 3 or w in NOTES_STOPWORDS or w in seen:
            continue
        seen.add(w)
        out.append(w)
    return out[:25]  # cap to avoid noise


def _keywords_from_fixed_metier(fixed_metier: str | None) -> List[str]:
    """Derive keywords from fixedMetier e.g. 'Project Manager > Achat / Industrialisation'."""
    if not fixed_metier or not str(fixed_metier).strip():
        return []
    raw = str(fixed_metier).strip()
    parts = re.split(r"\s*>\s*|\s*/\s*", raw)
    return [p.strip() for p in parts if len(p.strip()) >= 2]


# ====== Utils: JSON lists for candidates (skills, company_ids) ======
def _parse_json_str_list(v: Any) -> List[str]:
    """Accepts None | list | json string list | comma-separated string."""
    if v is None:
        return []
    if isinstance(v, list):
        out = [str(x).strip() for x in v if str(x).strip()]
        # dedupe (case-insensitive)
        seen = set()
        uniq: List[str] = []
        for s in out:
            k = s.lower()
            if k in seen:
                continue
            seen.add(k)
            uniq.append(s)
        return uniq
    s = str(v).strip()
    if not s:
        return []
    try:
        j = json.loads(s)
        if isinstance(j, list):
            return _parse_json_str_list(j)
    except Exception:
        pass
    # fallback "a, b, c"
    return _parse_json_str_list([x for x in (t.strip() for t in s.split(',')) if x])


def _parse_json_int_list(v: Any) -> List[int]:
    if v is None:
        return []
    if isinstance(v, list):
        out: List[int] = []
        for x in v:
            try:
                n = int(x)
                if n not in out:
                    out.append(n)
            except Exception:
                continue
        return out
    s = str(v).strip()
    if not s:
        return []
    try:
        j = json.loads(s)
        if isinstance(j, list):
            return _parse_json_int_list(j)
    except Exception:
        pass
    # fallback "1,2,3"
    parts = [p.strip() for p in s.split(',') if p.strip()]
    return _parse_json_int_list(parts)


def _dump_json_list(v: Any, *, as_int: bool = False) -> str | None:
    """Returns a compact JSON list string or None if empty."""
    if v is None:
        return None
    if as_int:
        arr = _parse_json_int_list(v)
        return json.dumps(arr, ensure_ascii=False) if arr else None
    arr = _parse_json_str_list(v)
    return json.dumps(arr, ensure_ascii=False) if arr else None


# ====== Candidates (Sourcing) API ======

# ── Candidate events (for KPI goals) ───────────────────────────────
# We define a simple rank to detect when a candidate crosses
# "contacted" (en_cours+) or "solid" (ec1+) states.
_CANDIDATE_STATUS_RANK = {
    "nouveau": 0,
    "proposition": 1,
    "entretien": 2,
    "a_faire": 3,
    "oksi": 4,
    "top_profil": 5,
    "reunion_tech": 6,
    "valide_contrat": 7,
    "freelance": 8,
    "freelance_mission": 9,
    "nok_prequal": -1,
    "nok": -2,
    "plus_disponible": -3,
    "refus_contrat": -4,
    "hors_aura": -5,
    # legacy — kept for migration compatibility
    "a_sourcer": 0,
    "a_contacter": 1,
    "en_cours": 2,
    "ec1": 2,
    "ec2": 2,
    "ed": 3,
    "interesse": 4,
    "mission": 9,
    "refuse": -2,
    "embauche": 7,
    "archive": -3,
}

_CANDIDATE_CONTACTED_RANK = 2  # entretien
_CANDIDATE_SOLID_RANK = 4      # oksi


def _candidate_status_rank(status: str | None) -> int:
    if not status:
        return -10
    s = str(status).strip().lower()
    return _CANDIDATE_STATUS_RANK.get(s, -10)


def _log_candidate_event(conn, candidate_id: int, date_iso: str, event_type: str, title: str, meta: dict | None = None):
    """Insert a candidate_event (deduped by unique index)."""
    now = datetime.datetime.now().isoformat(timespec="seconds")
    conn.execute(
        "INSERT OR IGNORE INTO candidate_events (candidate_id, date, type, title, content, meta, createdAt) VALUES (?,?,?,?,?,?,?)",
        (int(candidate_id), date_iso, event_type, title, None, json.dumps(meta or {}, ensure_ascii=False), now),
    )


def _maybe_log_candidate_events(conn, candidate_id: int, old_status: str | None, new_status: str | None, date_iso: str):
    old_r = _candidate_status_rank(old_status)
    new_r = _candidate_status_rank(new_status)
    if old_r < _CANDIDATE_CONTACTED_RANK <= new_r:
        _log_candidate_event(conn, candidate_id, date_iso, "candidate_contacted", "Candidat contacté", {"from": old_status, "to": new_status})
    if old_r < _CANDIDATE_SOLID_RANK <= new_r:
        _log_candidate_event(conn, candidate_id, date_iso, "candidate_solid", "Profil solide", {"from": old_status, "to": new_status})


# Routes /api/candidates/* (CRUD + sous-ressources) — voir routes/candidates.py


@app.post("/api/prospects/delete")
def api_prospects_delete():
    """v27.10: Soft delete a prospect (fenêtre d'annulation 10s via /api/soft-deleted/restore)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    pid = payload.get("id")
    if not pid:
        return jsonify(ok=False, error="id is required"), 400
    _name = None
    with _conn() as conn:
        _row = conn.execute("SELECT name FROM prospects WHERE id=? AND owner_id=?;", (int(pid), uid)).fetchone()
        _name = _row["name"] if _row else None
        conn.execute("UPDATE prospects SET deleted_at=? WHERE id=? AND owner_id=?;", (_now_iso(), int(pid), uid))
    _audit_log("soft_delete", "prospect", int(pid))
    log_activity('delete', 'prospect', int(pid), _name)
    return jsonify(ok=True)


# Routes /api/companies/{list,create,delete} — déplacées dans routes/companies.py


@app.post("/api/candidates/status")
def api_candidates_set_status():
    """Quick status update for a candidate (used by pipeline quick actions)."""
    payload = request.get_json(force=True, silent=True) or {}
    cid = payload.get("id") or payload.get("candidate_id")
    status = (payload.get("status") or "").strip()
    if not cid or not status:
        return jsonify(ok=False, error="id et status requis"), 400

    # Keep is_archived in sync with status when relevant
    st = status.lower()
    _ARCHIVE_STATUSES = {"nok_prequal", "nok", "plus_disponible", "refus_contrat", "hors_aura", "archive"}
    _ACTIVE_STATUSES = {
        "nouveau", "proposition", "entretien", "a_faire", "oksi", "top_profil",
        "reunion_tech", "valide_contrat", "freelance", "freelance_mission",
        # legacy
        "a_sourcer", "a_contacter", "en_cours", "ec1", "ec2", "ed",
        "interesse", "mission", "embauche", "refuse",
    }
    is_archived = None
    if st in _ARCHIVE_STATUSES:
        is_archived = 1
    elif st in _ACTIVE_STATUSES:
        is_archived = 0

    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    now = datetime.datetime.now().isoformat(timespec="seconds")
    event_date = now[:10]
    with _conn() as conn:
        # fetch previous status for goal events
        r0 = conn.execute("SELECT status FROM candidates WHERE id=? AND owner_id=?;", (int(cid), uid)).fetchone()
        if not r0:
            return jsonify(ok=False, error="Candidat non trouvé"), 404
        old_status = r0["status"]

        if is_archived is None:
            cur = conn.execute(
                "UPDATE candidates SET status=?, updatedAt=? WHERE id=? AND owner_id=?;",
                (status, now, int(cid), uid),
            )
        else:
            cur = conn.execute(
                "UPDATE candidates SET status=?, is_archived=?, updatedAt=? WHERE id=? AND owner_id=?;",
                (status, is_archived, now, int(cid), uid),
            )
        if cur.rowcount == 0:
            return jsonify(ok=False, error="not_found"), 404
        # goal events
        try:
            _maybe_log_candidate_events(conn, int(cid), old_status, status, event_date)
        except Exception:
            pass
    return jsonify(ok=True, updatedAt=now)


@app.post("/api/candidates/bulk-update")
def api_candidates_bulk_update():
    """Bulk update a whitelisted field for selected candidates (owner only).

    Currently supports field='status'. Keeps `is_archived` in sync with the new
    status (mirrors /api/candidates/status) and logs goal events per candidate.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    ids = payload.get("ids", [])
    field = payload.get("field", "")
    value = payload.get("value", "")
    ALLOWED_FIELDS = {"status": "status"}
    if field not in ALLOWED_FIELDS or not ids:
        return jsonify(ok=False, error="Requête invalide"), 400
    try:
        ids = [int(i) for i in ids if i is not None]
    except (TypeError, ValueError):
        return jsonify(ok=False, error="IDs invalides"), 400
    if not ids:
        return jsonify(ok=False, error="IDs invalides"), 400

    col = ALLOWED_FIELDS[field]
    value = (value or "").strip() if isinstance(value, str) else value
    now = datetime.datetime.now().isoformat(timespec="seconds")
    event_date = now[:10]

    # Sync is_archived when the bulk update targets the status column
    is_archived = None
    if col == "status":
        st = (value or "").lower()
        _ARCHIVE_STATUSES = {"nok_prequal", "nok", "plus_disponible", "refus_contrat", "hors_aura", "archive"}
        _ACTIVE_STATUSES = {
            "nouveau", "proposition", "entretien", "a_faire", "oksi", "top_profil",
            "reunion_tech", "valide_contrat", "freelance", "freelance_mission",
            # legacy
            "a_sourcer", "a_contacter", "en_cours", "ec1", "ec2", "ed",
            "interesse", "mission", "embauche", "refuse",
        }
        if st in _ARCHIVE_STATUSES:
            is_archived = 1
        elif st in _ACTIVE_STATUSES:
            is_archived = 0

    try:
        with _conn() as conn:
            placeholders = ",".join("?" * len(ids))
            old_rows = conn.execute(
                f"SELECT id, status FROM candidates WHERE id IN ({placeholders}) AND owner_id=? AND deleted_at IS NULL;",
                ids + [uid],
            ).fetchall()
            old_status_by_id = {r["id"]: r["status"] for r in old_rows}

            if col == "status" and is_archived is not None:
                cur = conn.execute(
                    f"UPDATE candidates SET status=?, is_archived=?, updatedAt=? "
                    f"WHERE id IN ({placeholders}) AND owner_id=? AND deleted_at IS NULL;",
                    [value, is_archived, now] + ids + [uid],
                )
            else:
                cur = conn.execute(
                    f"UPDATE candidates SET {col}=?, updatedAt=? "
                    f"WHERE id IN ({placeholders}) AND owner_id=? AND deleted_at IS NULL;",
                    [value, now] + ids + [uid],
                )
            updated = cur.rowcount

            if col == "status":
                for cid, old_status in old_status_by_id.items():
                    try:
                        _maybe_log_candidate_events(conn, int(cid), old_status, value, event_date)
                    except Exception:
                        pass
        return jsonify(ok=True, updated=updated, updatedAt=now)
    except Exception as exc:
        logger.exception("api_candidates_bulk_update failed")
        return jsonify(ok=False, error=str(exc)), 500


@app.post("/api/candidate-push")
def api_candidate_push_add():
    """Log a candidate → prospect push (simple history, reuses candidate_events)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    cid = payload.get("candidate_id")
    pid = payload.get("prospect_id")
    if not cid or not pid:
        return jsonify(ok=False, error="candidate_id et prospect_id requis"), 400
    try:
        cid_i = int(cid)
        pid_i = int(pid)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="IDs invalides"), 400
    if not _candidate_owned(cid_i) or not _prospect_owned(pid_i):
        return jsonify(ok=False, error="Accès refusé"), 403

    candidate_name = (payload.get("candidate_name") or "").strip()
    prospect_name = (payload.get("prospect_name") or "").strip()
    company_name = (payload.get("company_name") or "").strip()

    now = datetime.datetime.now().isoformat(timespec="seconds")
    event_date = now[:10]

    with _conn() as conn:
        meta = {
            "candidate_id": cid_i,
            "candidate_name": candidate_name,
            "prospect_id": pid_i,
            "prospect_name": prospect_name,
            "company_name": company_name,
        }
        # Store as generic candidate_event so it participates in KPIs if needed later.
        try:
            _log_candidate_event(
                conn,
                cid_i,
                event_date,
                "candidate_push",
                f"Proposé à {prospect_name or 'prospect'}",
                meta,
            )
        except Exception:
            # Even if logging fails, we don't block the UI.
            pass

    # Teams webhook: push candidat (v22.1)
    try:
        prefix = _get_user_prefix(uid)
        card = _build_adaptive_card(
            "Push candidat",
            [("Candidat", candidate_name), ("Prospect", prospect_name), ("Entreprise", company_name), ("Consultant", prefix), ("Date", event_date)],
            [{"title": "Voir dans Prosp'Up", "url": f"https://prospup.work/candidate?id={cid_i}"}]
        )
        _send_teams_webhook(card, "candidate_push")
    except Exception:
        pass

    return jsonify(ok=True, createdAt=now)


@app.get("/api/candidate-push")
def api_candidate_push_list():
    """Return full history of candidate → prospect pushes for a given candidate.

    Unions two data sources:
      1. push_logs (authoritative) where candidate_id1 or candidate_id2 match.
      2. candidate_events (type='candidate_push') for historical coverage.

    De-duplicates by (prospect_id, day) so entries shared across both sources
    appear once. The response also includes a per-company aggregate.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    cid = request.args.get("candidate_id", type=int)
    if not cid:
        return jsonify(ok=False, error="candidate_id requis"), 400
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403

    pushes: list[dict[str, Any]] = []
    seen: set[tuple] = set()

    def _day(ts: str | None) -> str:
        return (ts or "")[:10]

    with _conn() as conn:
        # 1) push_logs (primary source — every email/linkedin push with a candidate attached)
        try:
            rows = conn.execute(
                """
                SELECT pl.id            AS log_id,
                       pl.prospect_id   AS prospect_id,
                       pl.sentAt        AS sentAt,
                       pl.createdAt     AS createdAt,
                       pl.channel       AS channel,
                       pl.subject       AS subject,
                       p.name           AS prospect_name,
                       c.groupe         AS company_groupe,
                       c.site           AS company_site
                  FROM push_logs pl
                  LEFT JOIN prospects p ON p.id = pl.prospect_id AND p.owner_id = ?
                  LEFT JOIN companies c ON c.id = p.company_id
                 WHERE (pl.candidate_id1 = ? OR pl.candidate_id2 = ?)
                 ORDER BY COALESCE(pl.sentAt, pl.createdAt) DESC, pl.id DESC;
                """,
                (uid, int(cid), int(cid)),
            ).fetchall()
        except Exception:
            rows = []

        for r in rows:
            prospect_id = r["prospect_id"]
            sent_at = r["sentAt"] or r["createdAt"]
            company_parts = [r["company_groupe"] or "", r["company_site"] or ""]
            company_name = " · ".join([p for p in company_parts if p]).strip()
            key = (prospect_id, _day(sent_at))
            seen.add(key)
            pushes.append(
                {
                    "candidate_id": cid,
                    "prospect_id": prospect_id,
                    "prospect_name": r["prospect_name"],
                    "company_name": company_name or None,
                    "createdAt": sent_at,
                    "channel": r["channel"],
                    "subject": r["subject"],
                    "source": "push_log",
                }
            )

        # 2) candidate_events (fallback — legacy candidate-page triggered pushes)
        try:
            rows = conn.execute(
                """
                SELECT date, title, meta, createdAt
                  FROM candidate_events
                 WHERE candidate_id=? AND type='candidate_push'
                 ORDER BY COALESCE(createdAt, date) DESC, id DESC;
                """,
                (int(cid),),
            ).fetchall()
        except Exception:
            rows = []

        for r in rows:
            try:
                meta = json.loads(r["meta"] or "{}")
            except Exception:
                meta = {}
            prospect_id = meta.get("prospect_id")
            sent_at = r["createdAt"] or r["date"]
            key = (prospect_id, _day(sent_at))
            if key in seen:
                continue
            seen.add(key)
            pushes.append(
                {
                    "candidate_id": cid,
                    "prospect_id": prospect_id,
                    "prospect_name": meta.get("prospect_name"),
                    "company_name": meta.get("company_name"),
                    "createdAt": sent_at,
                    "channel": None,
                    "subject": None,
                    "title": r["title"],
                    "source": "event",
                }
            )

    pushes.sort(key=lambda p: (p.get("createdAt") or ""), reverse=True)

    # Aggregate per company for the summary badge.
    by_company: dict[str, int] = {}
    for p in pushes:
        label = (p.get("company_name") or "").strip() or "—"
        by_company[label] = by_company.get(label, 0) + 1
    companies = sorted(
        [{"company_name": k, "count": v} for k, v in by_company.items()],
        key=lambda x: (-x["count"], x["company_name"].lower()),
    )

    return jsonify(
        ok=True,
        pushes=pushes,
        total=len(pushes),
        companies=companies,
    )


# ====== Templates API ======
# Routes /api/templates + /api/push-categories + /api/push/* + /api/candidate-push + candidate save/generate-description — voir routes/push.py


# ═══════════════════════════════════════════════════════════════════
# v27.x PARTIE 2: Personnalisation .msg (win32com) ou .eml (fallback)
# OUTLOOK_AVAILABLE détecté au démarrage de l'app
# ═══════════════════════════════════════════════════════════════════

# _build_candidate_descriptions — voir utils/candidates.py


# Helpers push (Outlook + EML + parsing template .msg) — voir utils/push.py


# ────────────────────────────────────────────────────────────────────
# Best-match candidates – matching direct prospect.tags ↔ candidate.skills
# ────────────────────────────────────────────────────────────────────

@app.get("/api/prospect/<int:prospect_id>/best-candidates")
def api_prospect_best_candidates(prospect_id: int):
    """Find candidates whose skills best overlap with the prospect's tags.
    v11: weighted scoring — tags(×3), sector(×2), years_exp(×1.5), geo(×1)
    v12: fixedMetier keywords, notes keywords (×1), pertinence cap, push_category_id optional.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    push_category_id = request.args.get("push_category_id", type=int)

    with _conn() as conn:
        p_row = conn.execute(
            "SELECT name, tags, company_id, notes, fixedMetier, pertinence FROM prospects WHERE id=? AND owner_id=?;",
            (prospect_id, uid),
        ).fetchone()
        if not p_row:
            return jsonify(ok=False, error="prospect not found"), 404

        prospect_tags = _parse_json_str_list(p_row["tags"])
        company_id = p_row["company_id"]
        prospect_notes = (p_row["notes"] or "").strip()
        fixed_metier = (p_row["fixedMetier"] or "").strip()
        prospect_pertinence = (p_row["pertinence"] or "").strip()

        # Piste 2: keywords from fixedMetier (merge with tags; use as fallback when no tags)
        fixed_metier_keywords = _keywords_from_fixed_metier(fixed_metier)
        prospect_tags_effective = prospect_tags or [t for t in fixed_metier_keywords if t]
        if prospect_tags and fixed_metier_keywords:
            seen = {t.lower() for t in prospect_tags}
            for kw in fixed_metier_keywords:
                if kw.lower() not in seen:
                    prospect_tags_effective.append(kw)
                    seen.add(kw.lower())

        # Piste 1 option A: keywords from notes (weight ×1 in loop)
        notes_keywords = _keywords_from_notes(prospect_notes)
        notes_keywords_lower = [k.lower() for k in notes_keywords]
        notes_keywords_set = set(notes_keywords_lower)

        # Get company info for sector + location matching
        company_tags = []
        company_city = ""
        company_industry = ""
        company_groupe = ""
        if company_id:
            c_row = conn.execute(
                "SELECT groupe, tags, city, site, industry FROM companies WHERE id=? AND owner_id=?;",
                (company_id, uid),
            ).fetchone()
            if c_row:
                # sqlite3.Row n'a pas de méthode .get(), utiliser l'accès direct
                company_groupe = (c_row["groupe"] or "").strip() if c_row["groupe"] else ""
                company_tags = _parse_json_str_list(c_row["tags"])
                company_city = ((c_row["city"] or "") if c_row["city"] else (c_row["site"] or "")).lower().strip()
                company_industry = (c_row["industry"] or "").lower().strip() if c_row["industry"] else ""

        # Piste 5: optional push category keywords
        category_keywords = []
        if push_category_id:
            cat_row = conn.execute("SELECT keywords FROM push_categories WHERE id=? AND owner_id=?;", (push_category_id, uid)).fetchone()
            if cat_row:
                cat_dict = _row_to_dict(cat_row)
                if cat_dict and cat_dict.get("keywords"):
                    category_keywords = _parse_json_str_list(cat_dict["keywords"])

        all_sources = (
            [t.lower() for t in prospect_tags_effective]
            + [t.lower() for t in company_tags]
            + [t.lower() for t in category_keywords]
            + notes_keywords_lower
        )
        if not all_sources:
            return jsonify(ok=True, candidates=[], prospect_tags=prospect_tags)

        all_search_tags = list(dict.fromkeys(all_sources))  # preserve order, dedupe

        candidates = conn.execute("SELECT * FROM candidates WHERE owner_id=?;", (uid,)).fetchall()

    # Sector keywords (extracted from company tags + industry)
    SECTOR_KEYWORDS = {"automobile", "auto", "aéronautique", "aero", "ferroviaire", "défense", "defense",
                       "spatial", "médical", "medical", "énergie", "energie", "nucléaire", "nucleaire",
                       "iot", "telecom", "robotique", "naval", "industriel", "consumer", "domotique"}
    company_sectors = set()
    for t in company_tags:
        if t.lower() in SECTOR_KEYWORDS:
            company_sectors.add(t.lower())
    if company_industry:
        for s in SECTOR_KEYWORDS:
            if s in company_industry:
                company_sectors.add(s)

    scored = []
    for c in candidates:
        c_dict = dict(c)
        if c_dict.get("is_archived"):
            continue
        skills = _parse_json_str_list(c_dict.get("skills"))
        role = (c_dict.get("role") or "").lower()
        tech = (c_dict.get("tech") or "").lower()
        c_location = (c_dict.get("location") or "").lower().strip()
        c_sector = (c_dict.get("sector") or "").lower()
        c_notes = (c_dict.get("notes") or "").lower()
        c_years = c_dict.get("years_experience")
        skills_lower = [s.lower() for s in skills]
        haystack = " ".join(skills_lower) + " " + role + " " + tech + " " + c_notes

        # 1. Tags matching amélioré avec similarité sémantique (Phase 1)
        matched_tags = []
        tag_score = 0
        semantic_matches = []  # Tags matchés via similarité sémantique
        
        for tag_l in all_search_tags:
            exact_match = False
            # Match exact d'abord
            if tag_l in skills_lower:
                tag_score += 1 if tag_l in notes_keywords_set else 3
                matched_tags.append(tag_l)
                exact_match = True
            elif tag_l in haystack:
                tag_score += 1
                matched_tags.append(tag_l)
                exact_match = True
            
            # Si pas de match exact, essayer similarité sémantique (Phase 1)
            if not exact_match:
                best_similarity = 0.0
                best_skill = None
                for skill in skills_lower:
                    similarity = _compute_semantic_similarity(tag_l, skill, "tag")
                    if similarity > 0.7 and similarity > best_similarity:  # Seuil de 70%
                        best_similarity = similarity
                        best_skill = skill
                
                if best_skill:
                    # Score réduit pour match sémantique (×2 au lieu de ×3)
                    semantic_weight = 1 if tag_l in notes_keywords_set else 2
                    tag_score += semantic_weight
                    matched_tags.append(tag_l)
                    semantic_matches.append(f"{tag_l}≈{best_skill}")

        # 2. Sector matching (weight ×2)
        sector_score = 0
        if company_sectors:
            c_sectors_text = c_sector + " " + c_notes + " " + role
            for sec in company_sectors:
                if sec in c_sectors_text:
                    sector_score += 2

        # 3. Years experience (weight ×1.5)
        exp_score = 0
        if c_years is not None and c_years > 0:
            exp_score = min(c_years / 2, 5) * 1.5  # max ~7.5pts for 10+ years

        # 4. Geographic proximity (weight ×1)
        geo_score = 0
        if company_city and c_location:
            if company_city in c_location or c_location in company_city:
                geo_score = 3
            # Same region heuristic (e.g. both mention "lyon", "rhône", "69")
            elif any(w in c_location for w in company_city.split() if len(w) > 3):
                geo_score = 1

        total_score = tag_score + sector_score + exp_score + geo_score

        if total_score > 0:
            total_prospect = len(all_search_tags) if all_search_tags else 1
            pct = round(len(matched_tags) / total_prospect * 100) if total_prospect else 0
            # Piste 3: cap pct when few tags (avoid misleading 100%)
            if total_prospect < 4:
                pct = min(pct, 85)
            # Piste 3: global relevance score (score-based percentage)
            score_max_ref = 35.0  # ~ tag 15 + sector 6 + exp 7.5 + geo 3
            relevance_pct = min(100, round(total_score / score_max_ref * 100))
            # Piste 4: cap by prospect pertinence
            pertinence_cap = None
            if prospect_pertinence:
                pl = prospect_pertinence.lower()
                if "faible" in pl or "low" in pl:
                    pertinence_cap = 50
                elif "modérée" in pl or "moderee" in pl or "moderate" in pl:
                    pertinence_cap = 70
            if pertinence_cap is not None:
                pct = min(pct, pertinence_cap)
                relevance_pct = min(relevance_pct, pertinence_cap)
            scored.append({
                "id": c_dict["id"],
                "name": c_dict.get("name", ""),
                "role": c_dict.get("role", ""),
                "skills": skills,
                "tech": c_dict.get("tech", ""),
                "status": c_dict.get("status", ""),
                "linkedin": c_dict.get("linkedin", ""),
                "phone": c_dict.get("phone", ""),
                "years_experience": c_years,
                "location": c_dict.get("location", ""),
                "score": round(total_score, 1),
                "tag_score": tag_score,
                "sector_score": sector_score,
                "exp_score": round(exp_score, 1),
                "geo_score": geo_score,
                "pct": pct,
                "relevance_pct": relevance_pct,
                "matched_tags": list(set(matched_tags)),
                "semantic_matches": semantic_matches,  # Phase 1: matches sémantiques
            })

    scored.sort(key=lambda x: x["score"], reverse=True)
    top = scored[:8]
    
    # Phase 1: Génération d'explications IA pour chaque match
    use_ai_explanations = request.args.get("ai_explanations") == "1"
    if use_ai_explanations and top:
        try:
            p_dict = _row_to_dict(p_row)
            prospect_name = (p_dict.get("name") or "").strip() if p_dict else ""
            prospect_fonction = (p_dict.get("fonction") or "").strip() if p_dict else ""
            prospect_ctx = f"Prospect: {prospect_name}, entreprise {company_groupe}, fonction: {prospect_fonction}, tags: {prospect_tags_effective}"
            
            for candidate in top:
                matched_tags_str = ", ".join(candidate.get("matched_tags", [])[:10])
                semantic_str = ", ".join(candidate.get("semantic_matches", []))
                candidate_ctx = f"Candidat: {candidate.get('name')}, rôle: {candidate.get('role')}, compétences: {', '.join(candidate.get('skills', [])[:10])}, expérience: {candidate.get('years_experience', 'N/A')} ans"
                
                explanation_prompt = f"""Tu es un assistant de matching prospect/candidat. Explique en 2-3 phrases pourquoi ce candidat correspond bien à ce prospect.

{prospect_ctx}

{candidate_ctx}

Matches exacts: {matched_tags_str}
Matches sémantiques: {semantic_str if semantic_str else 'Aucun'}

Réponds UNIQUEMENT par une explication courte (2-3 phrases), sans formules de politesse, en expliquant les points forts du match."""
                
                try:
                    explanation = _call_ai(explanation_prompt, timeout=10)
                    candidate["ai_explanation"] = explanation.strip()
                except Exception:
                    candidate["ai_explanation"] = None
        except Exception as e:
            logger.warning("Erreur génération explications IA: %s", str(e))
    
    # Réordonnancement intelligent avec Ollama (existant, amélioré)
    use_ollama = request.args.get("use_ollama") == "1"
    if use_ollama and top:
        try:
            p_dict = _row_to_dict(p_row) if not isinstance(p_row, dict) else p_row
            prospect_name = (p_dict.get("name") or "").strip() if p_dict else ""
            prospect_ctx = f"Prospect: {prospect_name}, entreprise {company_groupe}, tags: {prospect_tags}"
            cand_lines = "\n".join(f"- {c.get('name') or '?'}: {', '.join((c.get('matched_tags') or [])[:8])}" for c in top)
            prompt = f"Contexte: {prospect_ctx}\n\nCandidats (nom + compétences matchées):\n{cand_lines}\n\nRéponds UNIQUEMENT par les noms des candidats, un par ligne, du meilleur au moins bon match. Pas d'autre texte."
            text = _call_ai(prompt, timeout=15)
            if text:
                order_names = [n.strip() for n in text.split("\n") if n.strip()]
                by_name = {c.get("name"): c for c in top}
                reordered = [by_name[n] for n in order_names if n in by_name]
                for c in top:
                    if c not in reordered:
                        reordered.append(c)
                top = reordered
        except Exception:
            pass
    
    return jsonify(ok=True, candidates=top, prospect_tags=prospect_tags)


# ====== Global search API ======
@app.get("/api/search")
def api_search():
    q = (request.args.get("q") or "").strip()
    try:
        limit = int(request.args.get("limit") or "50")
        limit = max(1, min(200, limit))
    except Exception:
        limit = 50
    # v23.5: pagination offset support
    try:
        offset = max(0, int(request.args.get("offset") or "0"))
    except Exception:
        offset = 0

    if not q:
        return jsonify({"prospects": [], "companies": [], "pushLogs": [], "candidates": [], "counts": {"prospects":0,"companies":0,"pushLogs":0,"candidates":0}, "limit": limit})

    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    like = f"%{q}%"
    with _conn() as conn:
        prospects = [
            dict(r)
            for r in conn.execute(
                '''
                SELECT p.*, c.groupe AS company_groupe, c.site AS company_site
                FROM prospects p
                LEFT JOIN companies c ON c.id = p.company_id
                WHERE p.owner_id=? AND p.deleted_at IS NULL AND (
                    p.name LIKE ? OR p.email LIKE ? OR p.telephone LIKE ?
                    OR p.linkedin LIKE ? OR p.fonction LIKE ? OR p.notes LIKE ?
                    OR p.callNotes LIKE ? OR p.tags LIKE ?
                    OR c.groupe LIKE ? OR c.site LIKE ?
                )
                ORDER BY p.id DESC
                LIMIT ? OFFSET ?;
                ''',
                (uid, like, like, like, like, like, like, like, like, like, like, limit, offset),
            ).fetchall()
        ]
        companies = [
            dict(r)
            for r in conn.execute(
                '''
                SELECT * FROM companies
                WHERE owner_id=? AND deleted_at IS NULL AND (groupe LIKE ? OR site LIKE ? OR phone LIKE ? OR notes LIKE ? OR tags LIKE ? OR website LIKE ? OR industry LIKE ? OR stack LIKE ? OR pain_points LIKE ?)
                ORDER BY id DESC
                LIMIT ? OFFSET ?;
                ''',
                (uid, like, like, like, like, like, like, like, like, like, limit, offset),
            ).fetchall()
        ]
        push_logs = [
            dict(r)
            for r in conn.execute(
                '''
                SELECT l.*, p.name AS prospect_name, p.email AS prospect_email, c.groupe AS company_groupe, c.site AS company_site
                FROM push_logs l
                JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=?
                LEFT JOIN companies c ON c.id = p.company_id
                WHERE l.to_email LIKE ? OR l.subject LIKE ? OR l.body LIKE ? OR p.name LIKE ? OR p.email LIKE ? OR c.groupe LIKE ? OR c.site LIKE ?
                ORDER BY l.id DESC
                LIMIT ? OFFSET ?;
                ''',
                (uid, like, like, like, like, like, like, like, limit, offset),
            ).fetchall()
        ]
        candidates = [
            dict(r)
            for r in conn.execute(
                '''
                SELECT * FROM candidates
                WHERE owner_id=? AND deleted_at IS NULL AND (name LIKE ? OR role LIKE ? OR location LIKE ? OR tech LIKE ? OR linkedin LIKE ? OR notes LIKE ?)
                ORDER BY COALESCE(updatedAt, createdAt) DESC, id DESC
                LIMIT ? OFFSET ?;
                ''',
                (uid, like, like, like, like, like, like, limit, offset),
            ).fetchall()
        ]

    out = {
        "prospects": prospects,
        "companies": companies,
        # camelCase for front v5+ (page-search.js)
        "pushLogs": push_logs,
        "candidates": candidates,
        "counts": {
            "prospects": len(prospects),
            "companies": len(companies),
            "pushLogs": len(push_logs),
            "candidates": len(candidates),
        },
        "limit": limit,
        "offset": offset,
        # legacy key for backward compatibility
        "push_logs": push_logs,
    }
    return jsonify(out)


# ====== Timeline API ====== ======
# Routes /api/prospect/timeline + /api/prospect/log-call + /api/prospect/log-stage + /api/prospect/events/add/update/delete — voir routes/prospects.py

@app.get("/api/dashboard/pipeline-stages")
def api_dashboard_pipeline_stages():
    """Retourne la distribution des prospects par étape de pipeline (frise chronologique)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        # Prospects de l'utilisateur non supprimés/archivés
        all_prospects = conn.execute(
            """SELECT id, name, statut, rdvDate, lastContact, nextFollowUp, company_id
               FROM prospects
               WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') AND (is_archived=0 OR is_archived IS NULL)
               ORDER BY id;""",
            (uid,),
        ).fetchall()

        # Prospects ayant au moins 1 réunion enregistrée
        meeting_pids = set(
            r[0]
            for r in conn.execute(
                "SELECT DISTINCT prospect_id FROM meetings WHERE owner_id=?;", (uid,)
            ).fetchall()
        )

        # Prospects avec event 'reunion_tech'
        rt_pids = set(
            r[0]
            for r in conn.execute(
                "SELECT DISTINCT prospect_id FROM prospect_events WHERE type='reunion_tech' AND prospect_id IN (SELECT id FROM prospects WHERE owner_id=?);",
                (uid,),
            ).fetchall()
        )

        # Prospects avec event 'contrat_signe'
        contrat_pids = set(
            r[0]
            for r in conn.execute(
                "SELECT DISTINCT prospect_id FROM prospect_events WHERE type='contrat_signe' AND prospect_id IN (SELECT id FROM prospects WHERE owner_id=?);",
                (uid,),
            ).fetchall()
        )

        # Map company_id → name
        company_names = {
            r[0]: r[1]
            for r in conn.execute(
                "SELECT id, groupe FROM companies WHERE owner_id=?;", (uid,)
            ).fetchall()
        }

        # Classement des prospects par stage
        stage_counts = {"appel": 0, "rdv": 0, "besoin": 0, "reunion_tech": 0, "contrat": 0}
        stage_prospects = {"appel": [], "rdv": [], "besoin": [], "reunion_tech": [], "contrat": []}

        RDV_STATUTS = {"Rendez-vous", "Prospecté"}

        for p in all_prospects:
            pid = p["id"]
            # Dériver l'étape la plus avancée
            if pid in contrat_pids:
                stage = "contrat"
            elif pid in rt_pids:
                stage = "reunion_tech"
            elif pid in meeting_pids:
                stage = "besoin"
            elif p["statut"] in RDV_STATUTS or (p["rdvDate"] and str(p["rdvDate"]).strip()):
                stage = "rdv"
            else:
                stage = "appel"

            stage_counts[stage] += 1
            stage_prospects[stage].append({
                "id": pid,
                "name": p["name"],
                "company": company_names.get(p["company_id"], ""),
                "statut": p["statut"],
                "lastContact": p["lastContact"] or "",
                "nextFollowUp": p["nextFollowUp"] or "",
                "stage": stage,
            })

        # Top prospects à pousser: stages besoin + reunion_tech, triés par lastContact (les plus anciens)
        priority = sorted(
            stage_prospects["besoin"] + stage_prospects["reunion_tech"],
            key=lambda x: (x["lastContact"] or "0"),
        )[:8]

    return jsonify(
        ok=True,
        stages=stage_counts,
        total=len(all_prospects),
        priority_prospects=priority,
    )


@app.get("/api/candidate/timeline")
def api_candidate_timeline():
    """Timeline des événements d'un candidat (candidate_events)."""
    cid = request.args.get("id")
    if not cid:
        return jsonify({"ok": False, "error": "id is required"}), 400
    uid = _uid()
    if not uid:
        return jsonify({"ok": False, "error": "Non authentifié"}), 401
    with _conn() as conn:
        row = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (int(cid), uid)).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "candidat not found"}), 404
        extra = [
            dict(r)
            for r in conn.execute(
                "SELECT date, type, title, content, meta, createdAt FROM candidate_events WHERE candidate_id=? ORDER BY date DESC, id DESC LIMIT 80;",
                (int(cid),),
            ).fetchall()
        ]
    events = []
    for e in extra:
        meta = None
        try:
            meta = json.loads(e.get("meta") or "null")
        except Exception:
            meta = None
        events.append({
            "type": e.get("type") or "event",
            "date": e.get("date") or e.get("createdAt") or "",
            "title": e.get("title") or "",
            "content": e.get("content") or "",
            "meta": meta,
        })
    def _key(ev):
        return str(ev.get("date") or "")
    events = sorted(events, key=_key, reverse=True)[:120]
    return jsonify({"ok": True, "events": events})


@app.post("/api/candidate/events/add")
def api_candidate_events_add():
    """Ajoute un événement manuel à la timeline d'un candidat."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    cid = payload.get("candidate_id")
    if not cid:
        return jsonify(ok=False, error="candidate_id requis"), 400
    try:
        cid_i = int(cid)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="candidate_id invalide"), 400
    if not _candidate_owned(cid_i):
        return jsonify(ok=False, error="Accès refusé"), 403
    title = (payload.get("title") or "").strip() or "Événement"
    content = (payload.get("content") or "").strip()
    etype = (payload.get("type") or "event").strip()
    date = (payload.get("date") or datetime.datetime.now().isoformat(timespec="seconds")).strip()
    if len(date) > 19:
        date = date[:19]
    now = datetime.datetime.now().isoformat(timespec="seconds")
    meta = payload.get("meta")
    meta_json = json.dumps(meta, ensure_ascii=False) if meta is not None else None
    with _conn() as conn:
        conn.execute(
            "INSERT INTO candidate_events (candidate_id, date, type, title, content, meta, createdAt) VALUES (?, ?, ?, ?, ?, ?, ?);",
            (cid_i, date, etype, title, content, meta_json, now),
        )
    return jsonify(ok=True)


# ====== Stats API ======
# Routes /api/stats* + /api/dashboard/pipeline-stages + /api/candidate/timeline + /api/candidate/events + /api/prospect/photo* — voir routes/dashboard.py
# Routes /api/duplicates/* + /api/prospects/check-duplicates + /api/prospects/create + /api/duplicates/merge-preview + /api/duplicates/merge — voir routes/duplicates.py

# ====== Company merge (duplicates) ======
@app.post("/api/companies/merge")
def api_companies_merge():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True) or {}
    keep_id = int(payload.get("keep_id") or 0)
    merge_id = int(payload.get("merge_id") or 0)
    if not keep_id or not merge_id or keep_id == merge_id:
        return jsonify({"ok": False, "error": "keep_id and merge_id are required"}), 400

    if not _company_owned(keep_id):
        return jsonify({"error": "Accès non autorisé à cette entreprise"}), 403
    if not _company_owned(merge_id):
        return jsonify({"error": "Accès non autorisé à cette entreprise"}), 403

    with _conn() as conn:
        keep = conn.execute("SELECT * FROM companies WHERE id=? AND owner_id=?;", (keep_id, uid)).fetchone()
        merg = conn.execute("SELECT * FROM companies WHERE id=? AND owner_id=?;", (merge_id, uid)).fetchone()
        if not keep or not merg:
            return jsonify({"ok": False, "error": "company not found"}), 404

        keep_d = dict(keep)
        merg_d = dict(merg)

        def _merge_text(a, b):
            a = (a or "").strip()
            b = (b or "").strip()
            if not a: return b
            if not b: return a
            if b in a: return a
            return a + "\n" + b

        def _to_tags(v):
            if v is None:
                return []
            if isinstance(v, list):
                return [str(x).strip() for x in v if str(x).strip()]
            s = str(v).strip()
            if not s:
                return []
            # try json array
            try:
                j = json.loads(s)
                if isinstance(j, list):
                    return [str(x).strip() for x in j if str(x).strip()]
            except Exception:
                pass
            return [x.strip() for x in s.split(",") if x.strip()]

        tags = sorted(set(_to_tags(keep_d.get("tags")) + _to_tags(merg_d.get("tags"))))
        notes = _merge_text(keep_d.get("notes"), merg_d.get("notes"))

        conn.execute("UPDATE companies SET notes=?, tags=? WHERE id=? AND owner_id=?;", (notes, json.dumps(tags), keep_id, uid))
        conn.execute("UPDATE prospects SET company_id=? WHERE company_id=? AND owner_id=?;", (keep_id, merge_id, uid))
        conn.execute("DELETE FROM companies WHERE id=? AND owner_id=?;", (merge_id, uid))

    _audit_log("merge_delete", "company", merge_id, new_value=str(keep_id))
    return jsonify({"ok": True})



# ====== Focus queue API ======
# Routes /api/focus_queue + /api/snapshots/* + /api/admin/* + /api/reset + /api/views/* + /api/rapport/export-pdf + /api/tasks/* + /api/companies/merge — voir routes/admin.py

# ====== Company / Opportunities API (v6) ======
# Routes /api/company/* + /api/audit-log + /api/activity + /api/soft-deleted/* + /api/opportunities/* + /api/ia-enrichment-log + /api/quickadd/* + /api/metiers/* + /api/system/* + /api/app-version + /api/health + /api/data + /api/save + /api/export/* + /api/rapport-hebdo — voir routes/misc.py

# ────────────────────────────────────────────────────────────────────
# Custom Métiers – ajout de compétences / spécialités / catégories
# ────────────────────────────────────────────────────────────────────

@app.get("/api/custom_metiers")
def api_custom_metiers_list():
    try:
        with _conn() as conn:
            conn.execute('''CREATE TABLE IF NOT EXISTS custom_metiers (
                id INTEGER PRIMARY KEY, type TEXT NOT NULL, category TEXT NOT NULL,
                specialty TEXT, tech_group TEXT, value TEXT NOT NULL, createdAt TEXT)''')
            rows = conn.execute("SELECT * FROM custom_metiers ORDER BY category, specialty, tech_group, value").fetchall()
            items = [dict(r) for r in rows]
        return jsonify(ok=True, items=items)
    except Exception as exc:
        logger.exception("Erreur api_custom_metiers_list")
        return jsonify(ok=False, error=str(exc)), 500


@app.post("/api/custom_metiers")
def api_custom_metiers_add():
    d = request.get_json(force=True)
    tp = d.get("type", "tech")  # tech | specialty | category | sector
    cat = (d.get("category") or "").strip()
    spec = (d.get("specialty") or "").strip() or None
    tg = (d.get("tech_group") or "").strip() or None
    val = d.get("value", "").strip()
    if not val:
        return jsonify(ok=False, error="value required"), 400
    now = datetime.datetime.now().isoformat(timespec="seconds")
    try:
        with _conn() as conn:
            conn.execute('''CREATE TABLE IF NOT EXISTS custom_metiers (
                id INTEGER PRIMARY KEY, type TEXT NOT NULL, category TEXT NOT NULL,
                specialty TEXT, tech_group TEXT, value TEXT NOT NULL, createdAt TEXT)''')
            # Check duplicate
            existing = conn.execute(
                "SELECT id FROM custom_metiers WHERE type=? AND category=? AND value=?",
                (tp, cat, val)
            ).fetchone()
            if existing:
                return jsonify(ok=False, error="duplicate"), 409
            conn.execute(
                "INSERT INTO custom_metiers (type, category, specialty, tech_group, value, createdAt) VALUES (?,?,?,?,?,?)",
                (tp, cat, spec, tg, val, now)
            )
        return jsonify(ok=True)
    except Exception as exc:
        logger.exception("Erreur api_custom_metiers_add")
        return jsonify(ok=False, error=str(exc)), 500


@app.delete("/api/custom_metiers/<int:item_id>")
def api_custom_metiers_delete(item_id):
    try:
        with _conn() as conn:
            conn.execute('''CREATE TABLE IF NOT EXISTS custom_metiers (
                id INTEGER PRIMARY KEY, type TEXT NOT NULL, category TEXT NOT NULL,
                specialty TEXT, tech_group TEXT, value TEXT NOT NULL, createdAt TEXT)''')
            conn.execute("DELETE FROM custom_metiers WHERE id=?", (item_id,))
        return jsonify(ok=True)
    except Exception as exc:
        logger.exception("Erreur api_custom_metiers_delete")
        return jsonify(ok=False, error=str(exc)), 500


# ═══════════════════════════════════════════════════════════════════
# v27.21 : Gestion des tags non référencés — classification IA batch
# ═══════════════════════════════════════════════════════════════════

@app.get("/api/prospects/tags-count")
def api_prospects_tags_count():
    """Retourne tous les tags utilisés dans les prospects avec leur nombre d'occurrences.
    Triés par count décroissant.
    Retourne: { "ok": true, "tags": [{"tag": "Python", "count": 12}, ...] }
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    counts = {}
    try:
        with _conn() as conn:
            rows = conn.execute(
                "SELECT tags FROM prospects WHERE owner_id=?",
                (uid,)
            ).fetchall()
        for row in rows:
            for tag in _parse_tags(row["tags"]):
                key = tag.strip()
                if key:
                    counts[key] = counts.get(key, 0) + 1
    except Exception as exc:
        logger.exception("Erreur api_prospects_tags_count")
        return jsonify(ok=False, error=str(exc)), 500
    sorted_tags = [{"tag": t, "count": c}
                   for t, c in sorted(counts.items(), key=lambda x: -x[1])]
    return jsonify(ok=True, tags=sorted_tags)


@app.post("/api/metiers/classify-tags-batch")
def api_metiers_classify_tags_batch():
    """Classifie une liste de tags non référencés via Ollama en un seul prompt batch.

    Body: { "tags": ["tag1", "tag2", ...] }
    Retourne: { "ok": true, "results": [{"tag":"...","category":"...","specialty":"...","techCategory":"...","confidence":0.9}, ...] }
    En cas d'erreur Ollama: { "ok": false, "error": "ollama_unavailable" }
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    payload = request.get_json(force=True, silent=True) or {}
    tags = payload.get("tags", [])
    if not tags or not isinstance(tags, list):
        return jsonify(ok=False, error="Liste de tags requise"), 400

    # Arbre métiers de référence pour le contexte Ollama
    metiers_ref = """Ingénierie Logicielle:
  Spécialités: Logiciel applicatif, Test/Validation/Qualification logicielle, Logiciels embarqués/IoT, Data Science/ML/Deep Learning, DevOps/Infrastructure/Cloud, Gestion de projet logiciel/Scrum Master, Développement Web/Fullstack

Ingénierie Électronique:
  Spécialités: Électronique analogique, Électronique numérique, Électronique de puissance, Génie électrique/Électrotechnique, Industrialisation, FPGA/ASIC/SoC

Ingénierie Système:
  Spécialités: Mécatronique/Robotique, Model Based Design (MBD), Safety/Sûreté de fonctionnement, Contrôle commande/Automatique, Simulation multiphysique/Modélisation, Mécanique, Ingénierie système, Test/Validation/Essais système

Life Science:
  Spécialités: Qualification d'équipements (Pharma & DM), Validation de systèmes automatisés (VSA), Validation de systèmes d'informations (VSI), Validation de produits (Dispositifs Médicaux)"""

    tech_groups = "Langages, Frameworks, Librairies, Outils, Bases de données, Systèmes, IDE, Protocoles, Microcontrôleurs, Capteurs, Outils CAO, Serveurs, Méthodologies, Matériel, Certifications"

    all_results = []
    # Lots de 5 : le frontend gère la boucle, chaque appel reste court (~15-30s)
    batch_size = 5
    try:
        for i in range(0, len(tags), batch_size):
            batch = tags[i:i + batch_size]
            tags_json = json.dumps(batch, ensure_ascii=False)
            prompt = f"""Tu es un expert en classification de compétences techniques pour l'ingénierie B2B (ESN/cabinet de conseil).

Arbre des métiers de référence:
{metiers_ref}

Groupes technologiques possibles: {tech_groups}

Voici une liste de tags à classifier. Pour chaque tag, détermine:
- La catégorie métier la plus appropriée (exactement l'un des 4 noms ci-dessus)
- La spécialité la plus appropriée dans cette catégorie
- Le groupe technologique le plus approprié
- Ta confiance de 0.0 à 1.0

Tags à classifier: {tags_json}

Réponds UNIQUEMENT avec un tableau JSON valide (sans markdown, sans texte avant ou après):
[
  {{"tag": "NomDuTag", "category": "Catégorie exacte", "specialty": "Spécialité exacte", "techCategory": "Groupe tech", "confidence": 0.9}},
  ...
]

Si un tag ne correspond à aucune catégorie connue, mets category null."""

            # Forcer Ollama : ce prompt ne nécessite pas de recherche web,
            # et l'utilisateur n'a peut-être plus de crédits Tavily
            response_text = _call_ai_provider("ollama", prompt, _load_ai_config(), 60)

            # Extraire le JSON du texte de réponse
            json_block = re.search(r'```(?:json)?\s*(\[.*?\])\s*```', response_text, re.DOTALL)
            if json_block:
                response_text = json_block.group(1)
            else:
                arr_match = re.search(r'\[[\s\S]*\]', response_text, re.DOTALL)
                if arr_match:
                    response_text = arr_match.group(0)

            try:
                batch_results = json.loads(response_text)
                if isinstance(batch_results, list):
                    all_results.extend(batch_results)
                else:
                    # Réponse invalide pour ce batch : retourner les tags non classés
                    for t in batch:
                        all_results.append({"tag": t, "category": None, "reason": "Réponse Ollama non parsable"})
            except json.JSONDecodeError:
                for t in batch:
                    all_results.append({"tag": t, "category": None, "reason": "JSON invalide"})

    except urllib.error.URLError:
        return jsonify(ok=False, error="ollama_unavailable")
    except Exception as e:
        logger.warning("Erreur classify-tags-batch: %s", e)
        return jsonify(ok=False, error=str(e))

    return jsonify(ok=True, results=all_results)


@app.post("/api/metiers/batch-confirm-tags")
def api_metiers_batch_confirm_tags():
    """Enregistre en lot les tags confirmés dans custom_metiers.

    Body: [{"tag":"Kubernetes","category":"Ingénierie Logicielle","specialty":"DevOps/Infrastructure/Cloud","tech_group":"Outils"}, ...]
    Retourne: { "ok": true, "saved": N, "skipped": M }
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    payload = request.get_json(force=True, silent=True) or []
    if not isinstance(payload, list):
        return jsonify(ok=False, error="Tableau JSON attendu"), 400

    saved = 0
    skipped = 0
    now = datetime.datetime.now().isoformat(timespec="seconds")

    try:
        with _conn() as conn:
            # Auto-créer la table si absente (per-user DBs créées avant v27.22)
            conn.execute('''CREATE TABLE IF NOT EXISTS custom_metiers (
                id INTEGER PRIMARY KEY, type TEXT NOT NULL, category TEXT NOT NULL,
                specialty TEXT, tech_group TEXT, value TEXT NOT NULL, createdAt TEXT)''')
            for item in payload:
                tag_val = str(item.get("tag", "")).strip()
                category = str(item.get("category", "")).strip()
                specialty = str(item.get("specialty", "")).strip()
                tech_group = str(item.get("tech_group", "")).strip() or None
                if not tag_val or not category:
                    skipped += 1
                    continue
                existing = conn.execute(
                    "SELECT id FROM custom_metiers WHERE type='tech' AND LOWER(category)=LOWER(?) AND LOWER(COALESCE(specialty,''))=LOWER(?) AND LOWER(value)=LOWER(?)",
                    (category, specialty, tag_val)
                ).fetchone()
                if existing:
                    skipped += 1
                else:
                    conn.execute(
                        "INSERT INTO custom_metiers (type, category, specialty, tech_group, value, createdAt) VALUES (?,?,?,?,?,?)",
                        ("tech", category, specialty, tech_group, tag_val, now)
                    )
                    saved += 1
            conn.commit()
    except Exception as exc:
        logger.exception("Erreur batch-confirm-tags")
        return jsonify(ok=False, error=str(exc)), 500

    return jsonify(ok=True, saved=saved, skipped=skipped)


# ────────────────────────────────────────────────────────────────────
# Calendar – vue calendrier des actions
# ────────────────────────────────────────────────────────────────────

@app.get("/calendrier")
def page_calendar():
    return redirect("/v30/calendrier", code=302)


@app.get("/collab")
@login_required
def page_collab():
    return redirect("/v30/collab", code=302)


# Routes /api/calendar_events* + _parse_ics_to_events — voir routes/calendar.py


# ────────────────────────────────────────────────────────────────────
# Dashboard – activité quotidienne / hebdo
# ────────────────────────────────────────────────────────────────────

@app.get("/dashboard")
def page_dashboard():
    return redirect("/v30/dashboard", code=302)


# Gamified goals helpers are extracted in services/dashboard_goals.py.


@app.get("/api/dashboard")
def api_dashboard():
    """Return KPIs for today + this week + trends. Accepts ?week=YYYY-WNN for historical navigation."""
    real_today = _today_iso()
    real_d_today = datetime.date.fromisoformat(real_today)

    week_param = request.args.get('week', '').strip()
    is_past_week = False
    d_today = real_d_today
    today = real_today
    if week_param and '-W' in week_param:
        try:
            yr_s, wn_s = week_param.split('-W')
            yr_p, wn_p = int(yr_s), int(wn_s)
            jan4_p = datetime.date(yr_p, 1, 4)
            w1_monday = jan4_p - datetime.timedelta(days=jan4_p.isoweekday() - 1)
            req_monday_d = w1_monday + datetime.timedelta(weeks=wn_p - 1)
            req_sunday_d = req_monday_d + datetime.timedelta(days=6)
            if req_monday_d <= real_d_today:
                d_today = min(req_sunday_d, real_d_today)
                today = d_today.isoformat()
                is_past_week = req_sunday_d < real_d_today
        except Exception:
            pass

    # Monday of the target week
    monday = (d_today - datetime.timedelta(days=d_today.weekday())).isoformat()
    # Previous week for trend comparison
    prev_monday = (d_today - datetime.timedelta(days=d_today.weekday() + 7)).isoformat()
    prev_sunday = (d_today - datetime.timedelta(days=d_today.weekday() + 1)).isoformat()

    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    with _conn() as conn:
        # BUG 27 : on exclut aussi les archivés pour cohérence KPIs dashboard
        prospects = conn.execute(
            "SELECT * FROM prospects WHERE owner_id=? "
            "AND (deleted_at IS NULL OR deleted_at='') "
            "AND (is_archived IS NULL OR is_archived=0);",
            (uid,),
        ).fetchall()
        push_logs = conn.execute(
            "SELECT l.* FROM push_logs l JOIN prospects p ON p.id=l.prospect_id AND p.owner_id=? AND (p.deleted_at IS NULL OR p.deleted_at='');",
            (uid,),
        ).fetchall()
        goals_cfg = _get_goals_config(conn)

        # Appels (call_logs) par jour de la semaine courante
        try:
            call_logs_rows = conn.execute(
                "SELECT date, COUNT(*) AS n FROM call_logs WHERE owner_id=? AND date >= ? AND date <= ? GROUP BY date;",
                (uid, monday, today),
            ).fetchall()
            calls_by_date = {r["date"]: r["n"] for r in call_logs_rows}
            calls_today = calls_by_date.get(today, 0)
            calls_week = sum(calls_by_date.values())
        except Exception:
            calls_by_date = {}
            calls_today = 0
            calls_week = 0

        # Event-based KPIs (for goals)
        # Fallback UNION: pour les prospects qui n'ont jamais eu d'event rdv_taken (DB ancienne sans prospect_events),
        # on comptabilise aussi les prospects statut='Rendez-vous' dont lastContact est dans la période
        # Condition "NOT EXISTS (ANY event)" pour éviter le surcômptage des RDV anciens déjà comptabilisés.
        try:
            rdv_taken_today = conn.execute(
                """SELECT COUNT(DISTINCT pid) FROM (
                    SELECT e.prospect_id AS pid
                    FROM prospect_events e
                    JOIN prospects p ON p.id=e.prospect_id AND p.owner_id=?
                    WHERE e.type='rdv_taken' AND e.date=?
                      AND (p.deleted_at IS NULL OR p.deleted_at='')
                    UNION
                    SELECT p.id AS pid
                    FROM prospects p
                    WHERE p.owner_id=? AND p.statut='Rendez-vous'
                      AND (p.deleted_at IS NULL OR p.deleted_at='')
                      AND p.rdvDate IS NOT NULL AND p.rdvDate != ''
                      AND substr(p.lastContact,1,10)=?
                      AND NOT EXISTS (
                          SELECT 1 FROM prospect_events e2
                          WHERE e2.prospect_id=p.id AND e2.type='rdv_taken'
                      )
                )""",
                (uid, today, uid, today),
            ).fetchone()[0]
            rdv_taken_week = conn.execute(
                """SELECT COUNT(DISTINCT pid) FROM (
                    SELECT e.prospect_id AS pid
                    FROM prospect_events e
                    JOIN prospects p ON p.id=e.prospect_id AND p.owner_id=?
                    WHERE e.type='rdv_taken' AND e.date BETWEEN ? AND ?
                      AND (p.deleted_at IS NULL OR p.deleted_at='')
                    UNION
                    SELECT p.id AS pid
                    FROM prospects p
                    WHERE p.owner_id=? AND p.statut='Rendez-vous'
                      AND (p.deleted_at IS NULL OR p.deleted_at='')
                      AND p.rdvDate IS NOT NULL AND p.rdvDate != ''
                      AND substr(p.lastContact,1,10) BETWEEN ? AND ?
                      AND NOT EXISTS (
                          SELECT 1 FROM prospect_events e2
                          WHERE e2.prospect_id=p.id AND e2.type='rdv_taken'
                      )
                )""",
                (uid, monday, today, uid, monday, today),
            ).fetchone()[0]
        except Exception:
            rdv_taken_today = 0
            rdv_taken_week = 0

        try:
            cand_contacted_today = conn.execute(
                "SELECT COUNT(*) FROM candidate_events e JOIN candidates c ON c.id=e.candidate_id AND c.owner_id=? WHERE e.type='candidate_contacted' AND e.date=?",
                (uid, today),
            ).fetchone()[0]
            cand_contacted_week = conn.execute(
                "SELECT COUNT(*) FROM candidate_events e JOIN candidates c ON c.id=e.candidate_id AND c.owner_id=? WHERE e.type='candidate_contacted' AND e.date BETWEEN ? AND ?",
                (uid, monday, today),
            ).fetchone()[0]
            cand_solid_week = conn.execute(
                "SELECT COUNT(*) FROM candidate_events e JOIN candidates c ON c.id=e.candidate_id AND c.owner_id=? WHERE e.type='candidate_solid' AND e.date BETWEEN ? AND ?",
                (uid, monday, today),
            ).fetchone()[0]
        except Exception:
            cand_contacted_today = 0
            cand_contacted_week = 0
            cand_solid_week = 0

        try:
            inmails_today = conn.execute(
                "SELECT COUNT(*) FROM linkedin_inmails WHERE owner_id=? AND sent_at=?",
                (uid, today),
            ).fetchone()[0]
            inmails_week = conn.execute(
                "SELECT COUNT(*) FROM linkedin_inmails WHERE owner_id=? AND sent_at BETWEEN ? AND ?",
                (uid, monday, today),
            ).fetchone()[0]
        except Exception:
            inmails_today = 0
            inmails_week = 0

        # RDV events par jour de la semaine (pour les barres d'activité)
        try:
            rdv_rows = conn.execute(
                """SELECT e.date, COUNT(DISTINCT e.prospect_id) AS n
                   FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id AND p.owner_id=?
                   WHERE e.type='rdv_taken' AND e.date BETWEEN ? AND ?
                     AND (p.deleted_at IS NULL OR p.deleted_at='')
                   GROUP BY e.date""",
                (uid, monday, today),
            ).fetchall()
            rdv_by_date = {r["date"]: r["n"] for r in rdv_rows}
        except Exception:
            rdv_by_date = {}

        # Prospects passés en RDV (aujourd'hui, ou toute la semaine pour une semaine passée)
        try:
            feed_start = monday if is_past_week else today
            today_rdv_rows = conn.execute(
                """SELECT DISTINCT p.id, p.name, COALESCE(c.groupe, '') AS company_name,
                          p.rdvDate, e.createdAt
                   FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id AND p.owner_id=?
                   LEFT JOIN companies c ON c.id=p.company_id
                   WHERE e.type='rdv_taken' AND e.date BETWEEN ? AND ?
                     AND (p.deleted_at IS NULL OR p.deleted_at='')
                   ORDER BY e.createdAt DESC LIMIT 20""",
                (uid, feed_start, today),
            ).fetchall()
            today_rdv_prospects = [dict(r) for r in today_rdv_rows]
        except Exception:
            today_rdv_prospects = []

        # Manual KPI ajustements pour la semaine courante
        try:
            mkpi_rows = conn.execute(
                "SELECT type, SUM(count) AS total FROM manual_kpi WHERE user_id=? AND date BETWEEN ? AND ? GROUP BY type",
                (uid, monday, today),
            ).fetchall()
            manual_kpi_week = {r["type"]: r["total"] for r in mkpi_rows}
            mkpi_today_rows = conn.execute(
                "SELECT type, SUM(count) AS total FROM manual_kpi WHERE user_id=? AND date=? GROUP BY type",
                (uid, today),
            ).fetchall()
            manual_kpi_today = {r["type"]: r["total"] for r in mkpi_today_rows}
            mkpi_calls_rows = conn.execute(
                "SELECT date, SUM(count) AS total FROM manual_kpi WHERE user_id=? AND date BETWEEN ? AND ? AND type='contact' GROUP BY date",
                (uid, monday, today),
            ).fetchall()
            manual_calls_by_date = {r["date"]: int(r["total"]) for r in mkpi_calls_rows}
        except Exception:
            manual_kpi_week = {}
            manual_kpi_today = {}
            manual_calls_by_date = {}

        # Notes stockées dans prospect_events (types note / note_libre / call_note)
        try:
            note_event_rows = conn.execute(
                """SELECT e.date, e.content, e.prospect_id, p.name AS prospect_name
                   FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id
                   WHERE p.owner_id=? AND e.type IN ('note','note_libre','call_note')
                     AND (p.deleted_at IS NULL OR p.deleted_at='')
                     AND (p.is_archived IS NULL OR p.is_archived=0);""",
                (uid,),
            ).fetchall()
            note_events = [dict(r) for r in note_event_rows]
        except Exception:
            note_events = []

    # Merge manual KPI "contact" adjustments into calls counts (for graph + totals)
    for _d, _cnt in manual_calls_by_date.items():
        calls_by_date[_d] = calls_by_date.get(_d, 0) + _cnt
    calls_today = max(0, calls_by_date.get(today, 0))
    calls_week = max(0, sum(calls_by_date.values()))

    prospects_list = [dict(r) for r in prospects]
    push_list = [dict(r) for r in push_logs]

    # Parse all call notes (callNotes JSON column + prospect_events de type note)
    all_notes = []
    for p in prospects_list:
        try:
            notes = json.loads(p.get("callNotes") or "[]")
            for n in (notes if isinstance(notes, list) else []):
                n["_prospect_id"] = p["id"]
                n["_prospect_name"] = p["name"]
                all_notes.append(n)
        except Exception:
            pass
    for ne in note_events:
        all_notes.append({
            "date": ne.get("date") or "",
            "content": ne.get("content") or "",
            "_prospect_id": ne.get("prospect_id"),
            "_prospect_name": ne.get("prospect_name") or "",
        })

    def count_relances(date_str):
        return sum(1 for p in prospects_list if (p.get("lastContact") or "") == date_str)

    def count_relances_range(start, end):
        return sum(1 for p in prospects_list if start <= (p.get("lastContact") or "") <= end)

    def count_notes(date_str):
        return sum(1 for n in all_notes if (n.get("date") or "")[:10] == date_str)

    def count_notes_range(start, end):
        return sum(1 for n in all_notes if start <= (n.get("date") or "")[:10] <= end)

    def count_push(date_str):
        return sum(1 for pl in push_list if (pl.get("sentAt") or "")[:10] == date_str)

    def count_push_range(start, end):
        return sum(1 for pl in push_list if start <= (pl.get("sentAt") or "")[:10] <= end)

    def count_push_channel(start, end, channel):
        return sum(1 for pl in push_list
                   if start <= (pl.get("sentAt") or "")[:10] <= end
                   and (pl.get("channel") or "") == channel)

    # Overdue / due today
    overdue = [p for p in prospects_list if (p.get("nextFollowUp") or "").strip() and p["nextFollowUp"].strip() < today]
    due_today = [p for p in prospects_list if (p.get("nextFollowUp") or "").strip() == today]
    due_week = [p for p in prospects_list if monday <= (p.get("nextFollowUp") or "").strip() <= today]

    # RDV count
    rdv_total = sum(1 for p in prospects_list if p.get("statut") == "Rendez-vous")

    # Notes/push for activity feed — full week range for past weeks, today only otherwise
    feed_start = monday if is_past_week else today
    today_notes = sorted(
        [n for n in all_notes if feed_start <= (n.get("date") or "")[:10] <= today],
        key=lambda x: x.get("date", ""), reverse=True
    )
    today_push = sorted(
        [pl for pl in push_list if feed_start <= (pl.get("sentAt") or "")[:10] <= today],
        key=lambda x: x.get("createdAt", ""), reverse=True
    )

    # Statut distribution
    statuts = {}
    for p in prospects_list:
        s = p.get("statut") or "Inconnu"
        statuts[s] = statuts.get(s, 0) + 1

    # Week daily breakdown for sparkline
    week_days = []
    for i in range(7):
        d = (datetime.date.fromisoformat(monday) + datetime.timedelta(days=i)).isoformat()
        if d > today:
            break
        week_days.append({
            "date": d,
            "relances": count_relances(d),
            "notes": count_notes(d),
            "push": count_push(d),
            "calls": calls_by_date.get(d, 0),
            "rdv": rdv_by_date.get(d, 0),
        })

    # Goals / gamification payload (daily + weekly)
    # Intègre les ajustements manual_kpi (peuvent être négatifs pour corriger les sur-comptages)
    goals_daily_counts = {
        "rdv": max(0, rdv_taken_today + int(manual_kpi_today.get("rdv", 0))),
        "push": max(0, count_push(today) + int(manual_kpi_today.get("push_email", 0)) + int(manual_kpi_today.get("push_linkedin", 0))),
        "sourcing_contacted": max(0, cand_contacted_today + inmails_today + int(manual_kpi_today.get("sourcing", 0))),
    }
    goals_weekly_counts = {
        "rdv": max(0, rdv_taken_week + int(manual_kpi_week.get("rdv", 0))),
        "push": max(0, count_push_range(monday, today) + int(manual_kpi_week.get("push_email", 0)) + int(manual_kpi_week.get("push_linkedin", 0))),
        "sourcing_contacted": max(0, cand_contacted_week + inmails_week + int(manual_kpi_week.get("sourcing", 0))),
        "sourcing_solid": max(0, cand_solid_week),
    }
    goals_payload = _build_goals_payload(
        goals_cfg=goals_cfg,
        daily_counts=goals_daily_counts,
        weekly_counts=goals_weekly_counts,
    )

    return jsonify(ok=True, data={
        "is_past_week": is_past_week,
        "today": {
            "date": today,
            "relances": count_relances(today),
            "notes": count_notes(today),
            "calls": calls_today,
            "push_total": count_push(today),
            "push_email": count_push_channel(today, today, "email"),
            "push_linkedin": count_push_channel(today, today, "linkedin"),
        },
        "goals": goals_payload,
        "week": {
            "start": monday,
            "end": today,
            "week_num": datetime.date.fromisoformat(monday).isocalendar()[1],
            "relances": count_relances_range(monday, today),
            "notes": count_notes_range(monday, today),
            "calls": calls_week,
            "push_total": count_push_range(monday, today),
            "push_email": count_push_channel(monday, today, "email"),
            "push_linkedin": count_push_channel(monday, today, "linkedin"),
            "rdv_total": rdv_taken_week,
            "days": week_days,
        },
        "prev_week": {
            "relances": count_relances_range(prev_monday, prev_sunday),
            "notes": count_notes_range(prev_monday, prev_sunday),
            "push_total": count_push_range(prev_monday, prev_sunday),
        },
        "pipeline": {
            "total": len(prospects_list),
            "rdv": rdv_total,
            "overdue": len(overdue),
            "due_today": len(due_today),
            "statuts": statuts,
        },
        "feed": {
            "notes": [{
                "prospect_id": n.get("_prospect_id"),
                "prospect_name": n.get("_prospect_name", ""),
                "content": n.get("content", ""),
                "date": n.get("date", ""),
            } for n in today_notes[:10]],
            "push": [{
                "prospect_id": pl.get("prospect_id"),
                "channel": pl.get("channel", ""),
                "subject": pl.get("subject", ""),
                "to_email": pl.get("to_email", ""),
                "createdAt": pl.get("createdAt", ""),
            } for pl in today_push[:10]],
            "rdv": [{
                "prospect_id": r.get("id"),
                "prospect_name": r.get("name", ""),
                "company_name": r.get("company_name", ""),
                "rdvDate": r.get("rdvDate", ""),
                "createdAt": r.get("createdAt", ""),
            } for r in today_rdv_prospects],
        },
        "overdue_list": [{
            "id": p["id"],
            "name": p["name"],
            "nextFollowUp": p.get("nextFollowUp", ""),
            "statut": p.get("statut", ""),
            "company_id": p.get("company_id"),
        } for p in sorted(overdue, key=lambda x: x.get("nextFollowUp", ""))[:10]],
        "today_appointments": [{
            "prospect_id": p["id"],
            "prospect_name": p.get("name", ""),
            "company_name": p.get("company_groupe") or p.get("company_site") or "",
            "rdvDate": p.get("rdvDate", ""),
        } for p in sorted(
            [p for p in prospects_list if (p.get("rdvDate") or "").strip()[:10] == today],
            key=lambda x: x.get("rdvDate", "")
        )],
        "upcoming_rdv": [{
            "id": p["id"],
            "name": p["name"],
            "rdvDate": p.get("rdvDate", ""),
            "statut": p.get("statut", ""),
        } for p in sorted(
            [p for p in prospects_list if (p.get("rdvDate") or "").strip()[:10] > today],
            key=lambda x: x.get("rdvDate", "")
        )[:5]],
    })


@app.get("/api/dashboard/stats")
def api_dashboard_stats():
    """Données Performance Pulse par semaine pour le dashboard v30.
    Accepte ?week=YYYY-Www. Retourne daily_rdv, daily_calls, insight + totaux."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    real_today = _today_iso()
    real_d = datetime.date.fromisoformat(real_today)

    week_param = request.args.get("week", "").strip()
    d_today = real_d
    if week_param and "-W" in week_param:
        try:
            yr_s, wn_s = week_param.split("-W")
            yr_p, wn_p = int(yr_s), int(wn_s)
            jan4_p = datetime.date(yr_p, 1, 4)
            w1_mon = jan4_p - datetime.timedelta(days=jan4_p.isoweekday() - 1)
            req_mon = w1_mon + datetime.timedelta(weeks=wn_p - 1)
            req_sun = req_mon + datetime.timedelta(days=6)
            if req_mon <= real_d:
                d_today = min(req_sun, real_d)
        except Exception:
            pass

    monday = d_today - datetime.timedelta(days=d_today.weekday())
    prev_monday = monday - datetime.timedelta(weeks=1)
    prev_sunday = monday - datetime.timedelta(days=1)
    today_iso = d_today.isoformat()
    monday_iso = monday.isoformat()
    prev_monday_iso = prev_monday.isoformat()
    prev_sunday_iso = prev_sunday.isoformat()

    week_num = monday.isocalendar()[1]
    week_label = f"{monday.year}-W{week_num:02d}"

    daily_rdv = [0] * 7
    daily_calls = [0] * 7

    with _conn() as conn:
        try:
            rdv_rows = conn.execute(
                """SELECT e.date, COUNT(DISTINCT e.prospect_id) AS n
                   FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id AND p.owner_id=?
                   WHERE e.type='rdv_taken' AND e.date BETWEEN ? AND ?
                     AND (p.deleted_at IS NULL OR p.deleted_at='')
                   GROUP BY e.date""",
                (uid, monday_iso, today_iso),
            ).fetchall()
            for r in rdv_rows:
                try:
                    d = datetime.date.fromisoformat(r["date"])
                    idx = (d - monday).days
                    if 0 <= idx < 7:
                        daily_rdv[idx] = r["n"]
                except Exception:
                    pass
        except Exception:
            pass

        try:
            call_rows = conn.execute(
                "SELECT date, COUNT(*) AS n FROM call_logs WHERE owner_id=? AND date BETWEEN ? AND ? GROUP BY date;",
                (uid, monday_iso, today_iso),
            ).fetchall()
            for r in call_rows:
                try:
                    d = datetime.date.fromisoformat(r["date"])
                    idx = (d - monday).days
                    if 0 <= idx < 7:
                        daily_calls[idx] = r["n"]
                except Exception:
                    pass
        except Exception:
            pass

        try:
            rdv_prev = conn.execute(
                """SELECT COUNT(DISTINCT e.prospect_id)
                   FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id AND p.owner_id=?
                   WHERE e.type='rdv_taken' AND e.date BETWEEN ? AND ?
                     AND (p.deleted_at IS NULL OR p.deleted_at='')""",
                (uid, prev_monday_iso, prev_sunday_iso),
            ).fetchone()[0]
        except Exception:
            rdv_prev = 0

    rdv_total = sum(daily_rdv)
    calls_total = sum(daily_calls)
    rdv_delta = rdv_total - rdv_prev
    sign = "+" if rdv_delta >= 0 else ""
    insight = f"{rdv_total} RDV cette semaine ({sign}{rdv_delta} vs semaine passée)"

    return jsonify(ok=True, data={
        "week": week_label,
        "daily_rdv": daily_rdv,
        "daily_calls": daily_calls,
        "rdv_total": rdv_total,
        "calls_total": calls_total,
        "rdv_prev_week": rdv_prev,
        "insight": insight,
    })


# ────────────────────────────────────────────────────────────────────
# RDV Checklist – grille de qualification prospect en rendez‑vous
# ────────────────────────────────────────────────────────────────────

RDV_CHECKLIST_THEMES = [
    {"key": "metiers_equipe",        "theme": "Métiers équipe",          "question": "Quels métiers dans l'équipe ?"},
    {"key": "outils",                "theme": "Outils",                  "question": "Quels outils (dev, gestion de projet, tests…) ?"},
    {"key": "taille_equipe",         "theme": "Taille équipe",           "question": "Nb pers dont internes / externes ?"},
    {"key": "projets_actuels",       "theme": "Projets actuels",         "question": "Projets en cours ?"},
    {"key": "projets_a_venir",       "theme": "Projets à venir",         "question": "Projets / roadmap à venir (3–12 mois) ?"},
    {"key": "societe",               "theme": "Société",                 "question": "Nb employés (site / groupe) ?"},
    {"key": "produits",              "theme": "Produits",                 "question": "Produits / systèmes principaux ?"},
    {"key": "autres_equipes",        "theme": "Autres équipes",          "question": "Autres équipes au même niveau ?"},
    {"key": "hierarchie",            "theme": "Hiérarchie",              "question": "Chefs / organisation (N+1, N+2…) ?"},
    {"key": "missions_externes",     "theme": "Missions externes",       "question": "Types de missions confiées aux consultants ?"},
    {"key": "duree_missions",        "theme": "Durée missions",          "question": "Durée moyenne des missions ?"},
    {"key": "vision_externalisation", "theme": "Vision externalisation", "question": "Vision sur l'externalisation (hausse, baisse, stable) ?"},
    {"key": "profils_recherches",    "theme": "Profils recherchés",      "question": "Profils types (ingé / tech, gestion de projet…) ?"},
    {"key": "xp_attendue",           "theme": "XP attendue",             "question": "Niveau d'XP (junior / confirmé / senior, exemples) ?"},
    {"key": "domaines_externalises", "theme": "Domaines externalisés",   "question": "Domaines / sujets le plus souvent externalisés ?"},
    {"key": "seniorite_consultants", "theme": "Séniorité consultants",   "question": "Séniorité moyenne / âge de l'équipe de consultants ?"},
    {"key": "formations_privilegiees","theme": "Formations privilégiées","question": "Écoles / formations préférées ?"},
    {"key": "xp_minimum",            "theme": "XP minimum",             "question": "Nb d'années d'XP minimum ?"},
    {"key": "origine_profils",       "theme": "Origine profils",         "question": "Origine habituelle des consultants (ESN, industrie…) ?"},
    {"key": "outils_indispensables", "theme": "Outils indispensables",   "question": "Outils / normes / environnements à maîtriser absolument ?"},
    {"key": "panel",                 "theme": "Panel",                   "question": "Panel ESN existant ? Partenaires principaux ?"},
    {"key": "process_achat",         "theme": "Process achat",           "question": "Comment se passe le process achat (demande, validation, délais) ?"},
    {"key": "validation_technique",  "theme": "Validation technique",    "question": "Comment est faite la validation technique des profils ?"},
    {"key": "appel_offre",           "theme": "Appel d'offre",           "question": "Appels d'offre ou consultations directes ?"},
    {"key": "criteres_partenaire",   "theme": "Critères partenaire",     "question": "Critères clés pour choisir un partenaire (réactivité, spécialisation, tarifs, etc.) ?"},
    {"key": "besoin_identifie",      "theme": "Besoin identifié",        "question": "Besoins ouverts / à venir ?"},
    {"key": "profils_a_proposer",    "theme": "Profils à proposer",      "question": "Typologie de profils à envoyer (compétences, techno, séniorité) ?"},
    {"key": "stakeholders",          "theme": "Stakeholders",            "question": "Décideurs / influenceurs impliqués ?"},
    {"key": "next_step",             "theme": "Next step",               "question": "Prochaine étape (envoi de profils, réunion technique…) + date / deadline ?"},
]


# ═══════════════════════════════════════════════════════════════════
# v26.3: Fonctions utilitaires pour "Avant réunion IA" — génération PDF
# ═══════════════════════════════════════════════════════════════════

def build_ollama_prompt_rdv(prospect: Dict[str, Any], company: Dict[str, Any] = None) -> str:
    """Construit le prompt Ollama pour analyser un profil LinkedIn et générer une fiche de préparation RDV.
    
    Args:
        prospect: Dict avec les champs du prospect (name, fonction, linkedin, etc.)
        company: Dict avec les infos de l'entreprise (groupe, site, etc.)
    
    Returns:
        String prompt structuré pour Ollama
    """
    nom_complet = prospect.get("name", "").strip()
    prenom = prospect.get("prenom", "").strip() or nom_complet.split()[0] if nom_complet else ""
    nom = prospect.get("nom", "").strip() or " ".join(nom_complet.split()[1:]) if len(nom_complet.split()) > 1 else nom_complet
    poste = prospect.get("fonction", "").strip()
    entreprise = ""
    ville = ""
    if company:
        entreprise = f"{company.get('groupe', '')} ({company.get('site', '')})".strip(" ()")
        ville = company.get("site", "").strip()
    linkedin = (prospect.get("linkedin") or "").strip()
    
    return f"""Tu es un expert en prospection B2B pour une ESN spécialisée en systèmes embarqués, robotique et ingénierie industrielle (société UpTechnologie, Lyon).

Tu dois analyser le profil LinkedIn suivant et générer une fiche de préparation RDV structurée en JSON.

--- PROFIL ---
Nom : {prenom} {nom}
Poste actuel : {poste}
Entreprise : {entreprise}
Ville : {ville}
URL LinkedIn : {linkedin}

--- FORMAT DE SORTIE ATTENDU (JSON strict) ---
{{
  "qui_est_il": {{
    "resume": "2-3 phrases de synthèse sur son profil, sa sensibilité, ses priorités",
    "titre_actuel": "...",
    "parcours": "résumé du parcours en 1-2 phrases",
    "stack_specialites": ["...", "..."],
    "activite_complementaire": "freelance / autre activité éventuelle"
  }},
  "contexte_entreprise": {{
    "description": "description de l'entreprise en 2-3 phrases",
    "taille": "...",
    "secteurs": ["...", "..."],
    "metiers_autour": ["...", "..."],
    "conclusion_matching": "pourquoi ces métiers matchent avec des candidats embarqué/robotique/IA"
  }},
  "besoins_probables": {{
    "data_referentiels": ["..."],
    "digital_bi2b": ["..."],
    "automatisation": ["..."],
    "ressources_contraintes": ["..."],
    "candidats_a_positionner": ["Ingé embarqué / industrie 4.0", "Dev back-end / data", "Ingé systèmes / intégration"]
  }},
  "interlocuteurs_potentiels": {{
    "marketing_digital": ["..."],
    "commerce_technique": ["..."],
    "technique_projet": ["..."],
    "conclusion": "..."
  }}
}}

Réponds UNIQUEMENT avec le JSON valide. Aucune source, aucun commentaire, aucune URL après le JSON. Commence directement par {{ et termine par }}.
"""


def build_fallback_prompt_rdv(prospect: Dict[str, Any], company: Dict[str, Any] = None) -> str:
    """Construit un prompt complet pour fallback (copier-coller dans une autre IA).
    
    Args:
        prospect: Dict avec les champs du prospect
        company: Dict avec les infos de l'entreprise
    
    Returns:
        String prompt complet
    """
    nom_complet = prospect.get("name", "").strip()
    prenom = prospect.get("prenom", "").strip() or nom_complet.split()[0] if nom_complet else ""
    nom = prospect.get("nom", "").strip() or " ".join(nom_complet.split()[1:]) if len(nom_complet.split()) > 1 else nom_complet
    poste = prospect.get("fonction", "").strip()
    entreprise = ""
    ville = ""
    if company:
        entreprise = f"{company.get('groupe', '')} ({company.get('site', '')})".strip(" ()")
        ville = company.get("site", "").strip()
    linkedin = (prospect.get("linkedin") or "").strip()
    
    return f"""Tu es un expert en prospection B2B pour une ESN spécialisée en systèmes embarqués, robotique et ingénierie industrielle (société UpTechnologie, Lyon).

Génère une fiche de préparation RDV complète au format JSON strict pour ce prospect :

Nom : {prenom} {nom}
Poste : {poste}
Entreprise : {entreprise}
Ville : {ville}
LinkedIn : {linkedin}

--- FORMAT DE SORTIE ATTENDU (JSON strict) ---
{{
  "qui_est_il": {{
    "resume": "2-3 phrases de synthèse sur son profil, sa sensibilité, ses priorités",
    "titre_actuel": "...",
    "parcours": "résumé du parcours en 1-2 phrases",
    "stack_specialites": ["...", "..."],
    "activite_complementaire": "freelance / autre activité éventuelle"
  }},
  "contexte_entreprise": {{
    "description": "description de l'entreprise en 2-3 phrases",
    "taille": "...",
    "secteurs": ["...", "..."],
    "metiers_autour": ["...", "..."],
    "conclusion_matching": "pourquoi ces métiers matchent avec des candidats embarqué/robotique/IA"
  }},
  "besoins_probables": {{
    "data_referentiels": ["..."],
    "digital_bi2b": ["..."],
    "automatisation": ["..."],
    "ressources_contraintes": ["..."],
    "candidats_a_positionner": ["Ingé embarqué / industrie 4.0", "Dev back-end / data", "Ingé systèmes / intégration"]
  }},
  "interlocuteurs_potentiels": {{
    "marketing_digital": ["..."],
    "commerce_technique": ["..."],
    "technique_projet": ["..."],
    "conclusion": "..."
  }}
}}

Réponds UNIQUEMENT avec le JSON, sans texte avant ni après.
"""


def build_fiche_rdv_pdf(prospect: Dict[str, Any], company: Dict[str, Any], ollama_data: Dict[str, Any]) -> BytesIO:
    """Génère un PDF A4 de fiche de préparation RDV avec ReportLab.
    
    Args:
        prospect: Dict avec les infos du prospect
        company: Dict avec les infos de l'entreprise
        ollama_data: Dict JSON parsé depuis la réponse Ollama
    
    Returns:
        BytesIO contenant le PDF généré
    """
    nom_complet = prospect.get("name", "").strip()
    # Extraire prénom et nom depuis name si prenom/nom ne sont pas définis
    if nom_complet:
        parts = nom_complet.split()
        prenom = prospect.get("prenom", "").strip() or (parts[0] if parts else "")
        nom = prospect.get("nom", "").strip() or (" ".join(parts[1:]) if len(parts) > 1 else "")
    else:
        prenom = prospect.get("prenom", "").strip() or ""
        nom = prospect.get("nom", "").strip() or ""
    
    # Fallback si toujours vide
    if not prenom and not nom:
        prenom = "Prospect"
        nom = ""
    
    poste = prospect.get("fonction", "").strip()
    entreprise_str = ""
    ville_str = ""
    if company:
        entreprise_str = f"{company.get('groupe', '')} ({company.get('site', '')})".strip(" ()")
        ville_str = company.get("site", "").strip()
    
    # Extraire les données Ollama
    qui_est_il = ollama_data.get("qui_est_il", {})
    contexte_entreprise = ollama_data.get("contexte_entreprise", {})
    besoins_probables = ollama_data.get("besoins_probables", {})
    interlocuteurs = ollama_data.get("interlocuteurs_potentiels", {})
    
    # Créer le buffer PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.8*cm,
        leftMargin=1.8*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
    )
    
    W, H = A4
    styles = getSampleStyleSheet()
    
    def S(name, parent='Normal', **kw):
        return ParagraphStyle(name, parent=styles[parent], **kw)
    
    # Couleurs
    GREY_DARK = colors.HexColor('#1A1A2E')
    GREY_MED = colors.HexColor('#2C3E50')
    BLUE_ACC = colors.HexColor('#2980B9')
    GREY_LINE = colors.HexColor('#BDC3C7')
    
    # Styles
    sMainTitle = S('MainTitle', fontName='Helvetica-Bold', fontSize=16, textColor=GREY_DARK,
                   spaceAfter=2, alignment=1, leading=20)
    sSubTitle = S('SubTitle', fontName='Helvetica', fontSize=9.5, textColor=GREY_MED,
                  spaceAfter=10, alignment=1, leading=14)
    sH1 = S('H1', fontName='Helvetica-Bold', fontSize=11.5, textColor=colors.white,
            spaceBefore=10, spaceAfter=4, leading=16)
    sH2 = S('H2', fontName='Helvetica-Bold', fontSize=10, textColor=GREY_DARK,
            spaceBefore=8, spaceAfter=2, leading=14)
    sH3 = S('H3', fontName='Helvetica-BoldOblique', fontSize=9, textColor=BLUE_ACC,
            spaceBefore=5, spaceAfter=2, leading=13)
    sBody = S('Body', fontName='Helvetica', fontSize=8.5, textColor=GREY_MED,
              spaceAfter=3, leading=13, alignment=4)
    sBullet = S('Bullet', fontName='Helvetica', fontSize=8.5, textColor=GREY_MED,
                spaceAfter=4, leading=14, leftIndent=10)
    sCheck = S('Check', fontName='Helvetica', fontSize=8.5, textColor=GREY_MED,
               spaceAfter=10, leading=18, leftIndent=12)
    sLink = S('Link', fontName='Helvetica-Bold', fontSize=8.5, textColor=GREY_DARK,
              spaceAfter=4, leading=14, leftIndent=10)
    
    def h1_block(text):
        tbl = Table([[Paragraph(text, sH1)]], colWidths=[W - 3.6*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), GREY_MED),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        return tbl
    
    def hr():
        return HRFlowable(width='100%', thickness=0.5, color=GREY_LINE, spaceAfter=3, spaceBefore=2)
    
    def b(text):
        return f'<b>{text}</b>'
    
    def bullet(text):
        return Paragraph('• ' + text, sBullet)
    
    def check(text):
        return Paragraph('[ ]  ' + text, sCheck)
    
    story = []
    
    # HEADER
    story.append(Paragraph('FICHE PRÉPARATION RDV PROSPECTION', sMainTitle))
    story.append(Paragraph(
        f'Prospect : {prenom} {nom} – {poste} – {entreprise_str} ({ville_str})',
        sSubTitle
    ))
    story.append(HRFlowable(width='100%', thickness=2, color=BLUE_ACC, spaceAfter=10))
    
    # SECTION 1
    story.append(h1_block('SECTION 1 – SYNTHÈSE PROSPECT'))
    story.append(Spacer(1, 6))
    
    # 1. Qui est [Prénom] et ce qu'il fait
    story.append(Paragraph(f'1. Qui est {prenom} et ce qu\'il fait', sH2))
    story.append(hr())
    resume = qui_est_il.get("resume", "")
    if resume:
        story.append(Paragraph(resume, sBody))
    if qui_est_il.get("titre_actuel"):
        story.append(bullet(b('Titre actuel :') + ' ' + qui_est_il.get("titre_actuel", "")))
    if qui_est_il.get("parcours"):
        story.append(bullet(b('Parcours :') + ' ' + qui_est_il.get("parcours", "")))
    if qui_est_il.get("stack_specialites"):
        specs = qui_est_il.get("stack_specialites", [])
        if isinstance(specs, list):
            story.append(bullet(b('Spécialités :') + ' ' + ', '.join(specs)))
    if qui_est_il.get("activite_complementaire"):
        story.append(bullet(b('Activité complémentaire :') + ' ' + qui_est_il.get("activite_complementaire", "")))
    if resume:
        story.append(Paragraph(
            f'<i>En clair : {resume}</i>',
            sBody
        ))
    story.append(Spacer(1, 5))
    
    # 2. Entreprise : environnement et métiers
    entreprise_nom = company.get('groupe', '') if company else entreprise_str
    story.append(Paragraph(f'2. {entreprise_nom} : environnement et métiers autour de lui', sH2))
    story.append(hr())
    if contexte_entreprise.get("description"):
        story.append(Paragraph(contexte_entreprise.get("description", ""), sBody))
    if contexte_entreprise.get("metiers_autour"):
        story.append(Paragraph(b('Métiers autour de lui :'), sBody))
        metiers = contexte_entreprise.get("metiers_autour", [])
        if isinstance(metiers, list):
            for m in metiers:
                story.append(bullet(m))
    if contexte_entreprise.get("conclusion_matching"):
        story.append(Paragraph(
            contexte_entreprise.get("conclusion_matching", ""),
            sLink
        ))
    story.append(Spacer(1, 5))
    
    # 3. Besoins probables
    story.append(Paragraph('3. Ses besoins probables (angle UpTechnologie)', sH2))
    story.append(hr())
    
    if besoins_probables.get("data_referentiels"):
        story.append(Paragraph('Data produits & référentiels', sH3))
        for item in besoins_probables.get("data_referentiels", []):
            if item:
                story.append(bullet(item))
    
    if besoins_probables.get("digital_bi2b"):
        story.append(Paragraph('E-commerce / Digital B2B', sH3))
        for item in besoins_probables.get("digital_bi2b", []):
            if item:
                story.append(bullet(item))
    
    if besoins_probables.get("automatisation"):
        story.append(Paragraph('Automatisation / outils internes', sH3))
        for item in besoins_probables.get("automatisation", []):
            if item:
                story.append(bullet(item))
    
    if besoins_probables.get("ressources_contraintes"):
        story.append(Paragraph('Ressources et contraintes', sH3))
        for item in besoins_probables.get("ressources_contraintes", []):
            if item:
                story.append(bullet(item))
    
    if besoins_probables.get("candidats_a_positionner"):
        story.append(Paragraph(b('C\'est là que je peux positionner mes candidats :'), sBody))
        for item in besoins_probables.get("candidats_a_positionner", []):
            if item:
                story.append(bullet(item))
    story.append(Spacer(1, 5))
    
    # 4. Métiers avec lesquels il travaille
    story.append(Paragraph('4. Métiers avec lesquels il travaille (interlocuteurs potentiels)', sH2))
    story.append(hr())
    if interlocuteurs.get("marketing_digital"):
        for item in interlocuteurs.get("marketing_digital", []):
            if item:
                story.append(bullet(item))
    if interlocuteurs.get("commerce_technique"):
        for item in interlocuteurs.get("commerce_technique", []):
            if item:
                story.append(bullet(item))
    if interlocuteurs.get("technique_projet"):
        for item in interlocuteurs.get("technique_projet", []):
            if item:
                story.append(bullet(item))
    if interlocuteurs.get("conclusion"):
        story.append(Paragraph(
            '<i>' + interlocuteurs.get("conclusion", "") + '</i>',
            sBody
        ))
    story.append(Spacer(1, 8))
    
    # SECTION 2
    story.append(h1_block('SECTION 2 – CHECKLIST RDV'))
    story.append(Spacer(1, 6))
    
    # Checklist fixe (8 sections)
    checklist_sections = [
        ('1. Contexte prospect', [
            'Vérifier son rôle exact : périmètre des projets et responsabilités.',
            'Confirmer s\'il gère aussi les outils internes (suivi projets, outils service, connecteurs SI).',
            'Identifier ses interlocuteurs principaux : commerce, technique, service, qualité, IT/IS.',
            'Comprendre les liens entre projets industriels, service client et activité business.',
        ]),
        ('2. Enjeux et priorités actuelles', [
            'Projets prioritaires 2025–2026 côté projets internationaux / modernisation / service.',
            'Objectifs business : satisfaction client, disponibilité des installations, marges projets, développement d\'offres.',
            'KPIs suivis : respect planning, coûts, pannes, temps d\'arrêt, taux de satisfaction.',
            'Contraintes majeures : budget, délais, ressources internes techniques / projet.',
        ]),
        ('3. Irritants et points de blocage', [
            'Manque de ressources techniques (ingénieurs automation / soft / data industrielle).',
            'Complexité / rigidité du SI projets / SAV (ERP, outils maison, PLM).',
            'Qualité, structuration, mise à jour de la donnée technique (installations, interventions, pannes).',
            'Difficultés à interfacer le digital (outils projet, service, IIoT) avec les systèmes terrain.',
            'Besoin d\'outils spécifiques pour les équipes internes (checklists, configurateurs, tableaux de bord).',
        ]),
        ('4. Organisation et recours aux ressources externes', [
            'Comment ils gèrent les besoins ponctuels : interne, freelances, intégrateurs, ESN.',
            'S\'ils ont déjà travaillé avec des sociétés de conseil / placement d\'ingénieurs.',
            'Leurs critères de choix d\'un partenaire technique (réactivité, expertise industrielle, proximité, mode d\'intervention).',
            'Le process de décision : qui décide, qui influence, qui utilise les solutions au quotidien.',
        ]),
        ('5. Positionnement UpTechnologie à présenter', [
            'Ton rôle : ingénieur d\'affaires spécialisé en systèmes embarqués, robotique, ingénierie industrielle.',
            'Ce que fait UpTechnologie : placement de consultants / ingénieurs pour renforcer les équipes sur des projets techniques.',
            'Capacité à intervenir à l\'interface terrain (automates, capteurs, lignes) / logiciel (SI, outils internes, supervision).',
            'Proximité géographique et connaissance du tissu industriel AURA.',
        ]),
        ('6. Types de besoins où tu peux aider', [
            'Solutions connectées : remontée de données des équipements vers le SI / outils projets / service.',
            'Automatisation de flux et fiabilisation de la donnée (scripts, ETL, API, connecteurs entre outils).',
            'Outils métiers pour les équipes internes (configurateurs, simulateurs, dashboards, portails clients).',
            'Projets industrie 4.0 nécessitant du logiciel embarqué / temps réel.',
        ]),
        ('7. Profils candidats à évoquer', [
            'Ingénieur systèmes embarqués / industrie 4.0 (automates, capteurs, équipements terrain).',
            'Profil logiciel / data back-end (scripts, API, intégration SI industriel).',
            'Profil passerelle terrain ↔ digital, à l\'aise en environnement industriel lourd.',
        ]),
        ('8. Next steps à sécuriser', [
            'Proposer l\'envoi d\'un court récap des échanges.',
            'Proposer 2–3 exemples de profils types alignés avec son environnement.',
            'Valider un point de suivi (après cadrage projet / avant pic d\'activité).',
            'Noter ses préférences de contact (mail, téléphone, LinkedIn) et disponibilités.',
        ]),
    ]
    
    for title, items in checklist_sections:
        story.append(Paragraph(title, sH2))
        story.append(hr())
        for item in items:
            story.append(check(item))
        story.append(Spacer(1, 3))
    
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width='100%', thickness=1.5, color=BLUE_ACC, spaceAfter=4))
    story.append(Paragraph(b('Notes libres / observations à chaud :'), sH2))
    for _ in range(4):
        story.append(HRFlowable(width='100%', thickness=0.4, color=GREY_LINE, spaceBefore=14, spaceAfter=0))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


# Routes /api/meetings/* + /api/rdv-checklist/* + /api/meeting-action-items/* — voir routes/meetings.py


@app.post("/api/prospect/<int:prospect_id>/summarize")
def api_prospect_summarize(prospect_id):
    """Génère un résumé IA de la fiche (cache en DB, force=1 pour régénérer)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _prospect_owned(prospect_id):
        return jsonify(ok=False, error="Accès refusé"), 403

    body = request.get_json(silent=True) or {}
    force = bool(body.get("force"))

    with _conn() as conn:
        if not force:
            cached = conn.execute(
                "SELECT summary, generatedAt FROM prospect_summaries WHERE prospect_id = ? AND owner_id = ?",
                (prospect_id, uid)
            ).fetchone()
            if cached and cached["summary"]:
                return jsonify(ok=True, summary=cached["summary"], generatedAt=cached["generatedAt"], cached=True)

        # Récolter les données nécessaires
        prow = conn.execute(
            "SELECT p.*, c.groupe AS company_groupe, c.site AS company_site "
            "FROM prospects p LEFT JOIN companies c ON c.id = p.company_id "
            "WHERE p.id = ? AND p.owner_id = ?",
            (prospect_id, uid)
        ).fetchone()
        if not prow:
            return jsonify(ok=False, error="Prospect introuvable"), 404
        prospect = dict(prow)

        # Events : on construit une liste similaire à api_prospect_timeline (light)
        events: list = []
        try:
            call_notes = json.loads(prospect.get("callNotes") or "[]")
            for n in call_notes:
                if isinstance(n, dict):
                    events.append({"type": "call_note", "date": n.get("date") or "",
                                   "title": "Note d'appel", "content": n.get("content") or ""})
        except Exception:
            pass
        try:
            for r in conn.execute(
                "SELECT date, type, title, content FROM prospect_events WHERE prospect_id = ? ORDER BY date DESC LIMIT 40",
                (prospect_id,)
            ).fetchall():
                events.append({"type": r["type"], "date": r["date"], "title": r["title"] or "", "content": r["content"] or ""})
        except Exception:
            pass
        try:
            for r in conn.execute(
                "SELECT m.date, m.title, m.summary, m.next_action, "
                "(SELECT COUNT(*) FROM meeting_action_items ai WHERE ai.meeting_id = m.id AND ai.status != 'done') AS pending "
                "FROM meetings m WHERE m.prospect_id = ? AND m.owner_id = ? ORDER BY m.date DESC LIMIT 20",
                (prospect_id, uid)
            ).fetchall():
                events.append({
                    "type": "cr", "date": r["date"], "title": r["title"] or "Compte-rendu",
                    "content": r["summary"] or "",
                    "meta": {"next_action": r["next_action"] or "", "action_pending": r["pending"] or 0}
                })
        except Exception:
            pass
        try:
            for r in conn.execute(
                "SELECT sentAt, channel, subject FROM push_logs WHERE prospect_id = ? ORDER BY id DESC LIMIT 10",
                (prospect_id,)
            ).fetchall():
                events.append({"type": "push", "date": r["sentAt"] or "",
                               "title": f"Push ({r['channel'] or 'email'})",
                               "content": r["subject"] or ""})
        except Exception:
            pass

        events = sorted(events, key=lambda e: str(e.get("date") or ""), reverse=True)

        attachments: list = []
        try:
            for r in conn.execute(
                "SELECT original_name, tags FROM prospect_attachments WHERE prospect_id = ? AND owner_id = ?",
                (prospect_id, uid)
            ).fetchall():
                tags = []
                try:
                    tags = json.loads(r["tags"] or "[]") or []
                except Exception:
                    pass
                attachments.append({"original_name": r["original_name"], "tags": tags})
        except Exception:
            pass

    prompt = _build_summary_prompt(prospect, events, attachments)

    try:
        text = _call_ai(prompt, timeout=120)
    except Exception as e:
        logger.warning("[summarize] IA call failed pid=%s: %s", prospect_id, e)
        return jsonify(ok=False, error=f"IA indisponible : {e}"), 503

    summary = (text or "").strip()
    now = datetime.datetime.now().isoformat(timespec="seconds")
    with _conn() as conn:
        conn.execute(
            """INSERT INTO prospect_summaries (prospect_id, owner_id, summary, generatedAt)
               VALUES (?, ?, ?, ?)
               ON CONFLICT(prospect_id) DO UPDATE SET summary = excluded.summary, generatedAt = excluded.generatedAt, owner_id = excluded.owner_id""",
            (prospect_id, uid, summary, now)
        )
    return jsonify(ok=True, summary=summary, generatedAt=now, cached=False)


@app.get("/api/prospect/upcoming-rdvs")
def api_prospect_upcoming_rdvs():
    """Liste les prospects dont le prochain RDV est dans les 48h.

    Utilisé par notifications.js côté client pour rappeler les RDV imminents.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    today = datetime.date.today().isoformat()
    plus2 = (datetime.date.today() + datetime.timedelta(days=2)).isoformat()

    with _conn() as conn:
        rows = conn.execute(
            """SELECT id, name, nextFollowUp, statut, fonction
               FROM prospects
               WHERE owner_id = ?
                 AND nextFollowUp IS NOT NULL
                 AND nextFollowUp != ''
                 AND date(substr(nextFollowUp,1,10)) BETWEEN date(?) AND date(?)
                 AND (deleted_at IS NULL OR deleted_at = '')
               ORDER BY nextFollowUp ASC
               LIMIT 30""",
            (uid, today, plus2)
        ).fetchall()
    items = [{
        "id": r["id"], "name": r["name"] or "", "nextFollowUp": r["nextFollowUp"],
        "statut": r["statut"] or "", "fonction": r["fonction"] or ""
    } for r in rows]
    return jsonify(ok=True, prospects=items)


# ═══════════════════════════════════════════════════════════════════
# v26.3: Routes API pour "Avant réunion IA" — streaming SSE et génération PDF
# ═══════════════════════════════════════════════════════════════════

@app.get("/api/prospect/<int:prospect_id>/infos-rdv-stream")
@login_required
def api_prospect_infos_rdv_stream(prospect_id: int):
    """Route SSE pour analyser un prospect via IA (Tavily+Ollama si configuré, sinon Ollama seul).

    Stream les tokens en temps réel, puis stocke la réponse complète dans
    _rdv_analysis_cache (et non en session, car la session n'est pas persistée
    dans les réponses SSE streaming) pour la génération PDF ultérieure.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    if not _prospect_owned(prospect_id):
        return jsonify(ok=False, error="Accès refusé"), 403

    with _conn() as conn:
        prospect_row = conn.execute(
            "SELECT * FROM prospects WHERE id=? AND owner_id=?;",
            (prospect_id, uid)
        ).fetchone()
        if not prospect_row:
            return jsonify(ok=False, error="Prospect introuvable"), 404

        prospect = dict(prospect_row)
        company = None
        if prospect.get("company_id"):
            company_row = conn.execute(
                "SELECT * FROM companies WHERE id=? AND owner_id=?;",
                (prospect["company_id"], uid)
            ).fetchone()
            if company_row:
                company = dict(company_row)

    prompt = build_ollama_prompt_rdv(prospect, company)
    fallback_prompt = build_fallback_prompt_rdv(prospect, company)
    cache_key = f"{uid}_{prospect_id}"

    def generate():
        full_response_parts: list[str] = []
        try:
            for sse_line in _stream_ai_web_sse(prompt, None, OLLAMA_TIMEOUT):
                if not sse_line.startswith("data: "):
                    yield sse_line
                    continue
                raw = sse_line[6:].strip()
                if not raw:
                    continue
                try:
                    evt = json.loads(raw)
                except json.JSONDecodeError:
                    yield sse_line
                    continue

                evt_type = evt.get("type")
                if evt_type == "token":
                    # Normalise la clé 'text' (Ollama/Tavily générique) → 'content' (frontend rdv)
                    token = evt.get("text") or evt.get("content") or ""
                    if token:
                        full_response_parts.append(token)
                    yield f"data: {json.dumps({'type': 'token', 'content': token}, ensure_ascii=False)}\n\n"
                elif evt_type == "end":
                    # Fin du stream : sauvegarder dans le cache et envoyer 'done'
                    full_text = "".join(full_response_parts)
                    if full_text:
                        _rdv_analysis_cache[cache_key] = full_text
                    yield f"data: {json.dumps({'type': 'done', 'pdf_url': f'/api/prospect/{prospect_id}/download-rdv-pdf'}, ensure_ascii=False)}\n\n"
                elif evt_type in ("start", "status"):
                    yield sse_line
                elif evt_type == "error":
                    yield f"data: {json.dumps({'type': 'error', 'fallback_prompt': fallback_prompt}, ensure_ascii=False)}\n\n"
                else:
                    yield sse_line
        except Exception:
            logger.exception("infos-rdv-stream failed")
            yield f"data: {json.dumps({'type': 'error', 'fallback_prompt': fallback_prompt}, ensure_ascii=False)}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.get("/api/prospect/<int:prospect_id>/download-rdv-pdf")
@login_required
def api_prospect_download_rdv_pdf(prospect_id: int):
    """Route pour télécharger le PDF de fiche de préparation RDV.
    
    Récupère la réponse Ollama stockée en session, parse le JSON, et génère le PDF.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    # Vérifier que le prospect appartient à l'utilisateur
    if not _prospect_owned(prospect_id):
        return jsonify(ok=False, error="Accès refusé"), 403
    
    # Récupérer la réponse IA depuis le cache en mémoire (session inutilisable en SSE)
    cache_key = f"{uid}_{prospect_id}"
    ollama_response = _rdv_analysis_cache.pop(cache_key, None)
    if not ollama_response:
        return jsonify(ok=False, error="Aucune analyse disponible. Relancez la génération."), 404
    
    # Récupérer le prospect et l'entreprise
    with _conn() as conn:
        prospect_row = conn.execute(
            "SELECT * FROM prospects WHERE id=? AND owner_id=?;",
            (prospect_id, uid)
        ).fetchone()
        if not prospect_row:
            return jsonify(ok=False, error="Prospect introuvable"), 404
        
        prospect = dict(prospect_row)
        company = None
        if prospect.get("company_id"):
            company_row = conn.execute(
                "SELECT * FROM companies WHERE id=? AND owner_id=?;",
                (prospect["company_id"], uid)
            ).fetchone()
            if company_row:
                company = dict(company_row)
    
    # Parser le JSON depuis la réponse Ollama
    # Extraction robuste : équilibrage des accolades pour ignorer le texte
    # autour du JSON (Sources, commentaires…) que Tavily/Ollama peut ajouter.
    def _extract_json_from_text(text):
        """Extrait le premier objet JSON complet depuis un texte potentiellement pollué."""
        import re as _re
        # 1. Tenter un bloc markdown ```json ... ```
        m = _re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, _re.DOTALL)
        if m:
            try:
                return json.loads(m.group(1))
            except Exception:
                pass
        # 2. Équilibrage des accolades : premier { ... } fermant correctement
        start = text.find('{')
        if start == -1:
            raise ValueError("Aucun JSON trouvé dans la réponse IA")
        depth = 0
        for i, c in enumerate(text[start:], start):
            if c == '{':
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0:
                    return json.loads(text[start:i + 1])
        raise ValueError("JSON incomplet dans la réponse IA")

    try:
        ollama_data = _extract_json_from_text(ollama_response)
    except (ValueError, json.JSONDecodeError) as e:
        logger.warning("Erreur parsing JSON Ollama RDV: %s — réponse brute: %s", e, ollama_response[:300])
        return jsonify(ok=False, error="Format de réponse IA invalide. Réessayez."), 400
    
    # Générer le PDF
    try:
        pdf_buffer = build_fiche_rdv_pdf(prospect, company, ollama_data)

        # Nom du fichier
        nom_complet = prospect.get("name", "").strip() or "prospect"
        nom_safe = "".join(c for c in nom_complet if c.isalnum() or c in (' ', '-', '_')).strip()[:50]
        filename = f"fiche_rdv_{nom_safe}.pdf"

        # Persiste le PDF + journalise l'événement IA "Avant RDV" pour
        # pouvoir le redonner plus tard (badge ✓ dans le picker IA).
        try:
            pdf_bytes = pdf_buffer.getvalue()
        except AttributeError:
            pdf_buffer.seek(0)
            pdf_bytes = pdf_buffer.read()
        try:
            ia_dir = DATA_DIR / "ia_pdfs" / str(uid) / str(prospect_id)
            ia_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            saved_name = f"fiche_rdv_{ts}.pdf"
            saved_path = ia_dir / saved_name
            saved_path.write_bytes(pdf_bytes)
            _log_ia_event(
                uid, prospect_id, "before",
                summary="Fiche prépa générée",
                meta={"pdf_path": str(saved_path), "filename": filename},
            )
        except Exception:
            logger.exception("Échec persistance PDF fiche RDV (non bloquant)")

        return send_file(
            BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.exception("Erreur génération PDF fiche RDV")
        return jsonify(ok=False, error=f"Erreur lors de la génération du PDF: {str(e)}"), 500


# ────────────────────────────────────────────────────────────────────
# IA — journal des analyses lancées sur un prospect
# ────────────────────────────────────────────────────────────────────

_IA_KIND_TITLE = {
    "scrap":  "Scraping IA",
    "before": "Fiche prépa IA",
    "after":  "Compte-rendu IA",
}


def _log_ia_event(uid: int, prospect_id: int, kind: str,
                  summary: str = "", meta: Dict[str, Any] | None = None) -> Dict[str, Any]:
    """Journalise une exécution d'IA dans prospect_events.

    type = "ia_<kind>" (ia_scrap, ia_before, ia_after).
    Retourne {ok, id, date, type, title}.
    """
    if kind not in _IA_KIND_TITLE:
        return {"ok": False, "error": "kind invalide"}
    title = _IA_KIND_TITLE[kind]
    etype = f"ia_{kind}"
    # Précision microseconde pour éviter les collisions sur la contrainte
    # UNIQUE (prospect_id, type, date) si plusieurs runs dans la seconde.
    date = datetime.datetime.now().isoformat(timespec="microseconds")
    meta_json = json.dumps(meta, ensure_ascii=False) if meta else None
    with _conn() as conn:
        cur = conn.execute(
            "INSERT OR IGNORE INTO prospect_events "
            "(prospect_id, date, type, title, content, meta, createdAt) "
            "VALUES (?, ?, ?, ?, ?, ?, ?);",
            (prospect_id, date, etype, title, summary or "", meta_json, date),
        )
        new_id = cur.lastrowid
    return {"ok": True, "id": new_id, "date": date, "type": etype, "title": title}


@app.post("/api/prospect/<int:prospect_id>/ia-log")
def api_prospect_ia_log(prospect_id: int):
    """Journalise une exécution d'IA pour le badge "✓ Fait" du picker."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _prospect_owned(prospect_id):
        return jsonify(ok=False, error="Accès refusé"), 403
    body = request.get_json(force=True, silent=True) or {}
    kind = (body.get("kind") or "").strip().lower()
    if kind not in _IA_KIND_TITLE:
        return jsonify(ok=False, error="kind doit être scrap|before|after"), 400
    summary = (body.get("summary") or "").strip()
    meta = body.get("meta") if isinstance(body.get("meta"), dict) else None
    res = _log_ia_event(uid, prospect_id, kind, summary, meta)
    if not res.get("ok"):
        return jsonify(res), 400
    return jsonify(res)


@app.get("/api/prospect/<int:prospect_id>/ia-pdf")
def api_prospect_ia_pdf(prospect_id: int):
    """Re-télécharge un PDF de fiche prépa déjà généré (via event_id)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _prospect_owned(prospect_id):
        return jsonify(ok=False, error="Accès refusé"), 403
    try:
        event_id = int(request.args.get("event_id", "0"))
    except (TypeError, ValueError):
        return jsonify(ok=False, error="event_id invalide"), 400
    if not event_id:
        return jsonify(ok=False, error="event_id requis"), 400
    with _conn() as conn:
        row = conn.execute(
            "SELECT meta FROM prospect_events "
            "WHERE id=? AND prospect_id=? AND type='ia_before';",
            (event_id, prospect_id),
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Fiche introuvable"), 404
    try:
        meta = json.loads(row["meta"] or "null") or {}
    except Exception:
        meta = {}
    pdf_path = meta.get("pdf_path") or ""
    if not pdf_path:
        return jsonify(ok=False, error="PDF non disponible"), 404
    p = Path(pdf_path)
    # Confine l'accès au dossier ia_pdfs de l'utilisateur courant.
    base = (DATA_DIR / "ia_pdfs" / str(uid)).resolve()
    try:
        if base not in p.resolve().parents:
            return jsonify(ok=False, error="Accès refusé"), 403
    except Exception:
        return jsonify(ok=False, error="Chemin PDF invalide"), 400
    if not p.exists():
        return jsonify(ok=False, error="Fichier PDF supprimé"), 404
    return send_file(
        str(p),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=meta.get("filename") or p.name,
    )


# ────────────────────────────────────────────────────────────────────
# EC1 Checklist – entretien de qualification candidat
# ────────────────────────────────────────────────────────────────────

EC1_CHECKLIST_ITEMS = [
    {"key": "mobilite_dispo_souhaits", "label": "Infos mobilité, disponibilité, souhaits"},
    {"key": "impression_generale", "label": "Impression générale du candidat"},
    {"key": "evaluation_technique", "label": "Évaluation technique"},
    {"key": "evaluation_personnalite", "label": "Évaluation personnalité"},
    {"key": "evaluation_communication", "label": "Évaluation communication"},
    {"key": "rappel_valeurs_up", "label": "Rappel des valeurs UpTechnologie"},
    {"key": "fourchette_salaire", "label": "Annonce fourchette salariale"},
    {"key": "reponse_questions_craintes", "label": "Réponse aux questions/craintes du candidat"},
    {"key": "process_prochaines_etapes", "label": "Détail du process et des prochaines étapes"},
]

def _blank_ec1_data() -> Dict[str, Any]:
    d = {t["key"]: {"checked": False, "note": ""} for t in EC1_CHECKLIST_ITEMS}
    d["__note"] = ""
    return d

def _ss(v: Any) -> str:
    return (str(v) if v is not None else "").strip()

@app.get("/api/ec1-checklist/themes")
def ec1_checklist_themes():
    return jsonify(ok=True, themes=EC1_CHECKLIST_ITEMS)

@app.get("/api/ec1-checklist")
def ec1_checklist_get():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    cid = request.args.get("candidate_id", type=int)
    if not cid:
        return jsonify(ok=False, error="candidate_id requis"), 400
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403
    with _conn() as conn:
        row = conn.execute(
            "SELECT interviewAt, data, updatedAt FROM candidate_ec1_checklists WHERE candidate_id=?;",
            (cid,),
        ).fetchone()
    if row:
        try:
            data = json.loads(row["data"]) if row["data"] else _blank_ec1_data()
        except Exception:
            data = _blank_ec1_data()
        return jsonify(ok=True, interviewAt=row["interviewAt"], data=data, updatedAt=row["updatedAt"])
    return jsonify(ok=True, interviewAt=None, data=_blank_ec1_data(), updatedAt=None)

@app.post("/api/ec1-checklist")
def ec1_checklist_save():
    """Upsert EC1 checklist for a candidate. Supports partial updates:
    - if 'data' is absent, keeps existing data
    - if 'interviewAt' is absent, keeps existing interviewAt
    """
    body = request.get_json(force=True, silent=True) or {}
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    cid = body.get("candidate_id") or body.get("id")
    if not cid:
        return jsonify(ok=False, error="candidate_id requis"), 400
    try:
        cid_i = int(cid)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="candidate_id invalide"), 400
    if not _candidate_owned(cid_i):
        return jsonify(ok=False, error="Accès refusé"), 403

    has_data = "data" in body
    has_interview = ("interviewAt" in body) or ("interview_at" in body)

    interviewAt = body.get("interviewAt", None)
    if interviewAt is None and "interview_at" in body:
        interviewAt = body.get("interview_at", None)
    if interviewAt is not None:
        interviewAt = _ss(interviewAt)
        if interviewAt == "":
            interviewAt = None

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with _conn() as conn:
        row = conn.execute(
            "SELECT interviewAt, data FROM candidate_ec1_checklists WHERE candidate_id=?;",
            (cid_i,),
        ).fetchone()

        current_interview = row["interviewAt"] if row else None
        current_data = None
        if row and row["data"]:
            try:
                current_data = json.loads(row["data"])
            except Exception:
                current_data = None
        if not isinstance(current_data, dict):
            current_data = _blank_ec1_data()

        new_interview = current_interview
        if has_interview:
            new_interview = interviewAt

        new_data = current_data
        if has_data:
            incoming = body.get("data")
            if isinstance(incoming, dict):
                # Keep only expected keys + __note
                blank = _blank_ec1_data()
                merged = {}
                for k in blank.keys():
                    if k == "__note":
                        merged[k] = _ss(incoming.get(k, blank[k]))
                    else:
                        v = incoming.get(k, blank[k])
                        if not isinstance(v, dict):
                            v = blank[k]
                        merged[k] = {
                            "checked": bool(v.get("checked", False)),
                            "note": _ss(v.get("note", "")),
                        }
                new_data = merged
            else:
                new_data = _blank_ec1_data()

        conn.execute(
            """INSERT INTO candidate_ec1_checklists (candidate_id, interviewAt, data, updatedAt)
               VALUES (?, ?, ?, ?)
               ON CONFLICT(candidate_id)
               DO UPDATE SET interviewAt=excluded.interviewAt, data=excluded.data, updatedAt=excluded.updatedAt""",
            (cid_i, new_interview, json.dumps(new_data, ensure_ascii=False), now),
        )

    return jsonify(ok=True, updatedAt=now)


# ────────────────────────────────────────────────────────────────────
# Candidate tabs (onglets fiche candidat: EC1 + note libre, v25)
# ────────────────────────────────────────────────────────────────────

@app.get("/api/candidate-tabs")
def api_candidate_tabs_list():
    """Liste des onglets d'un candidat (triés par sort_order)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    cid = request.args.get("candidate_id", type=int)
    if not cid:
        return jsonify(ok=False, error="candidate_id requis"), 400
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403
    with _conn() as conn:
        rows = conn.execute(
            "SELECT id, candidate_id, sort_order, type, title, payload, updated_at FROM candidate_tabs WHERE candidate_id=? ORDER BY sort_order ASC, id ASC;",
            (cid,),
        ).fetchall()
    tabs = []
    for r in rows:
        payload = None
        if r["payload"]:
            try:
                payload = json.loads(r["payload"])
            except Exception:
                payload = {}
        tabs.append({
            "id": r["id"],
            "candidate_id": r["candidate_id"],
            "sort_order": r["sort_order"],
            "type": r["type"],
            "title": r["title"],
            "payload": payload,
            "updated_at": r["updated_at"],
        })
    return jsonify(ok=True, tabs=tabs)


@app.post("/api/candidate-tabs")
def api_candidate_tabs_create():
    """Crée un nouvel onglet (ec1 ou note_libre)."""
    body = request.get_json(force=True, silent=True) or {}
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    cid = body.get("candidate_id")
    if not cid:
        return jsonify(ok=False, error="candidate_id requis"), 400
    try:
        cid_i = int(cid)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="candidate_id invalide"), 400
    if not _candidate_owned(cid_i):
        return jsonify(ok=False, error="Accès refusé"), 403
    tab_type = (body.get("type") or "").strip().lower()
    if tab_type not in ("ec1", "note_libre"):
        return jsonify(ok=False, error="type doit être 'ec1' ou 'note_libre'"), 400
    title = (body.get("title") or "").strip() or ("EC1" if tab_type == "ec1" else "Note")
    now = datetime.datetime.now().isoformat(timespec="seconds")
    if tab_type == "ec1":
        payload = {"interviewAt": None, "data": _blank_ec1_data()}
    else:
        payload = {"content": ""}
    payload_str = json.dumps(payload, ensure_ascii=False)
    with _conn() as conn:
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) + 1 AS next_order FROM candidate_tabs WHERE candidate_id=?;",
            (cid_i,),
        ).fetchone()["next_order"]
        cur = conn.execute(
            """INSERT INTO candidate_tabs (candidate_id, sort_order, type, title, payload, updated_at)
               VALUES (?, ?, ?, ?, ?, ?);""",
            (cid_i, max_order, tab_type, title, payload_str, now),
        )
        tab_id = cur.lastrowid
        conn.commit()
    return jsonify(ok=True, tab={"id": tab_id, "candidate_id": cid_i, "sort_order": max_order, "type": tab_type, "title": title, "payload": payload, "updated_at": now})


@app.put("/api/candidate-tabs/<int:tab_id>")
def api_candidate_tabs_update(tab_id: int):
    """Met à jour le titre et/ou le payload d'un onglet."""
    body = request.get_json(force=True, silent=True) or {}
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT id, candidate_id, sort_order, type, title, payload FROM candidate_tabs WHERE id=?;",
            (tab_id,),
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Onglet introuvable"), 404
        if not _candidate_owned(int(row["candidate_id"])):
            return jsonify(ok=False, error="Accès refusé"), 403
        now = datetime.datetime.now().isoformat(timespec="seconds")
        updates = []
        params = []
        if "title" in body:
            updates.append("title=?")
            params.append((body.get("title") or "").strip() or row["title"])
        if "payload" in body:
            pl = body["payload"]
            if isinstance(pl, dict):
                payload_str = json.dumps(pl, ensure_ascii=False)
            else:
                payload_str = str(pl) if pl is not None else row["payload"] or "{}"
            updates.append("payload=?")
            params.append(payload_str)
        if not updates:
            return jsonify(ok=True, updated_at=row.get("updated_at"))
        updates.append("updated_at=?")
        params.append(now)
        params.append(tab_id)
        conn.execute(
            "UPDATE candidate_tabs SET " + ", ".join(updates) + " WHERE id=?;",
            params,
        )
        conn.commit()
    return jsonify(ok=True, updated_at=now)

# ═══════════════════════════════════════════════════════
# App Settings API (v11)
# ═══════════════════════════════════════════════════════

# Routes /api/settings — déplacées dans routes/settings.py


# ═══════════════════════════════════════════════════════
# User Prefix / Teams Integration (v22.1)
# ═══════════════════════════════════════════════════════

def _compute_initials(display_name):
    """Compute default initials from display_name (e.g. 'Antoine Binet' → 'ABI')."""
    if not display_name:
        return "???"
    parts = display_name.strip().split()
    # Remove parenthetical like "Antoine (Admin)" → ["Antoine", "Binet"]
    parts = [p for p in parts if not p.startswith("(")]
    if len(parts) >= 2:
        # First letter of first name + first two letters of last name → ABI
        return (parts[0][0] + parts[-1][:2]).upper()
    elif len(parts) == 1:
        return parts[0][:3].upper()
    return "???"

def _build_adaptive_card(title: str, facts: list, actions: list = None, accent_color: str = "accent") -> dict:
    """Build an Adaptive Card v1.4 payload for Teams webhook.
    facts = [(label, value), ...], actions = [{title, url}, ...]"""
    body = [
        {"type": "TextBlock", "text": title, "weight": "Bolder", "size": "Medium", "color": accent_color},
        {"type": "FactSet", "facts": [{"title": f[0], "value": str(f[1]) if f[1] else "—"} for f in facts]}
    ]
    card = {
        "type": "AdaptiveCard",
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.4",
        "body": body,
    }
    if actions:
        card["actions"] = [{"type": "Action.OpenUrl", "title": a["title"], "url": a["url"]} for a in actions]
    return card


def _send_teams_webhook(card: dict, event_type: str = "notification"):
    """Teams webhook désactivé (section retirée des paramètres). No-op."""
    pass


# ═══════════════════════════════════════════════════════
# Manual KPI API (v16.5)
# ═══════════════════════════════════════════════════════

@app.post("/api/manual-kpi")
def api_manual_kpi_add():
    """Add a manual KPI entry (for actions done outside the app)."""
    payload = request.get_json(force=True, silent=False) or {}
    kpi_type = payload.get("type", "note")
    kpi_date = payload.get("date", datetime.datetime.now().strftime("%Y-%m-%d"))
    kpi_count = int(payload.get("count", 1))
    kpi_desc = payload.get("description", "")
    now = datetime.datetime.now().isoformat(timespec="seconds")

    user_id = None
    sess_user = session.get("user_id")
    if sess_user:
        user_id = sess_user

    with _conn() as conn:
        conn.execute(
            "INSERT INTO manual_kpi (user_id, type, date, count, description, createdAt) VALUES (?, ?, ?, ?, ?, ?);",
            (user_id, kpi_type, kpi_date, kpi_count, kpi_desc, now)
        )
    return jsonify(ok=True, message="KPI enregistré")


@app.get("/api/manual-kpi")
def api_manual_kpi_list():
    """List manual KPI entries (user's only), optionally filtered by date range."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    with _conn() as conn:
        query = "SELECT * FROM manual_kpi WHERE user_id=?"
        params = [uid]
        if date_from:
            query += " AND date >= ?"
            params.append(date_from)
        if date_to:
            query += " AND date <= ?"
            params.append(date_to)
        query += " ORDER BY date DESC, createdAt DESC LIMIT 200;"
        rows = conn.execute(query, params).fetchall()
    return jsonify(ok=True, entries=[dict(r) for r in rows])


# ═══════════════════════════════════════════════════════
# LinkedIn InMails (sourcing stats — dashboard objectifs)
# ═══════════════════════════════════════════════════════

@app.get("/api/linkedin-inmails")
def api_linkedin_inmails_list():
    """Liste les InMails LinkedIn de l'utilisateur, triés par date décroissante."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        rows = conn.execute(
            "SELECT * FROM linkedin_inmails WHERE owner_id=? ORDER BY sent_at DESC, created_at DESC LIMIT 500;",
            (uid,)
        ).fetchall()
    return jsonify(ok=True, entries=[dict(r) for r in rows])


@app.post("/api/linkedin-inmails")
def api_linkedin_inmails_add():
    """Enregistre un InMail LinkedIn envoyé (incrémente le compteur sourcing du jour)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    url = (payload.get("url") or "").strip()
    if not url:
        return jsonify(ok=False, error="URL manquante"), 400
    note = (payload.get("note") or "").strip()
    sent_at = (payload.get("sent_at") or datetime.datetime.now().strftime("%Y-%m-%d")).strip()
    now_ts = time.time()
    name = _parse_linkedin_name(url)
    with _conn() as conn:
        conn.execute(
            "INSERT INTO linkedin_inmails (url, note, name, sent_at, owner_id, created_at) VALUES (?, ?, ?, ?, ?, ?);",
            (url, note or None, name or None, sent_at, uid, now_ts)
        )
    return jsonify(ok=True, message="InMail enregistré")


@app.patch("/api/linkedin-inmails/<int:entry_id>")
def api_linkedin_inmails_update(entry_id: int):
    """Met à jour le nom affiché d'un InMail LinkedIn."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    name = (payload.get("name") or "").strip() or None
    with _conn() as conn:
        row = conn.execute(
            "SELECT id FROM linkedin_inmails WHERE id=? AND owner_id=?;", (entry_id, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Introuvable"), 404
        conn.execute("UPDATE linkedin_inmails SET name=? WHERE id=?;", (name, entry_id))
    return jsonify(ok=True)


@app.delete("/api/linkedin-inmails/<int:entry_id>")
def api_linkedin_inmails_delete(entry_id: int):
    """Supprime un InMail LinkedIn (seul le propriétaire peut le faire)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT id FROM linkedin_inmails WHERE id=? AND owner_id=?;", (entry_id, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Introuvable"), 404
        conn.execute("DELETE FROM linkedin_inmails WHERE id=?;", (entry_id,))
    return jsonify(ok=True)


# ═══════════════════════════════════════════════════════
# Candidate Folder API (v11)
# ═══════════════════════════════════════════════════════

def _get_setting(conn, key: str, default: str = "") -> str:
    row = conn.execute("SELECT value FROM app_settings WHERE key=?;", (key,)).fetchone()
    return row["value"] if row else default


def _build_candidate_folder_path(candidate_name: str, conn) -> Path | None:
    """Build the full path to a candidate's Windows folder."""
    base_path = _get_setting(conn, "candidate_folder_base", "")
    if not base_path:
        return None
    folder_format = _get_setting(conn, "candidate_folder_format", "{NOM} {Prenom}")

    # Parse name
    parts = candidate_name.strip().split()
    if len(parts) >= 2:
        prenom = parts[0]
        nom = " ".join(parts[1:])
    elif len(parts) == 1:
        prenom = parts[0]
        nom = parts[0]
    else:
        return None

    folder_name = folder_format.replace("{NOM}", nom.upper()).replace("{Prenom}", prenom.capitalize()).replace("{nom}", nom.lower()).replace("{prenom}", prenom.lower()).replace("{PRENOM}", prenom.upper()).replace("{Nom}", nom.capitalize())

    return Path(base_path) / folder_name


@app.get("/api/candidate-folder/<int:candidate_id>/files")
@login_required
@role_required('admin')
def api_candidate_folder_files(candidate_id: int):
    """List files in the candidate's Windows folder."""
    with _conn() as conn:
        row = conn.execute("SELECT name FROM candidates WHERE id=?;", (candidate_id,)).fetchone()
        if not row:
            return jsonify(ok=False, error="Candidate not found"), 404

        folder = _build_candidate_folder_path(row["name"], conn)

    if not folder:
        return jsonify(ok=False, error="Chemin de base non configuré. Allez dans Paramètres > Dossier candidats.", no_config=True)

    if not folder.exists():
        return jsonify(ok=True, folder=str(folder), files=[], exists=False)

    files = []
    try:
        for f in sorted(folder.iterdir()):
            if f.name.startswith(".") or f.name.startswith("~"):
                continue
            stat = f.stat()
            files.append({
                "name": f.name,
                "path": str(f),
                "is_dir": f.is_dir(),
                "size": stat.st_size,
                "ext": f.suffix.lower(),
                "modified": datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            })
    except PermissionError:
        return jsonify(ok=False, error="Accès refusé au dossier")

    return jsonify(ok=True, folder=str(folder), files=files, exists=True)


@app.get("/api/candidates/source-from-folder")
@login_required
def api_candidates_source_from_folder():
    """List subfolders of the candidate base path that do not yet have a candidate. For sourcing new names."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        base_path = _get_setting(conn, "candidate_folder_base", "").strip()
        if not base_path:
            return jsonify(ok=False, error="Configurez le chemin de base (Dossier candidats) sur la page Candidats.", no_config=True)
        existing = conn.execute("SELECT name FROM candidates WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at = '');", (uid,)).fetchall()
    existing_names = {(r["name"] or "").strip().lower() for r in existing}
    base = Path(base_path)
    if not base.exists() or not base.is_dir():
        return jsonify(ok=True, new=[], error="Dossier introuvable.")
    new_folders = []
    try:
        for name in sorted(base.iterdir()):
            if not name.is_dir() or name.name.startswith(".") or name.name.startswith("~"):
                continue
            fn = name.name.strip()
            if not fn:
                continue
            if fn.lower() in existing_names:
                continue
            files = []
            for f in sorted(name.iterdir()):
                if f.name.startswith(".") or f.name.startswith("~"):
                    continue
                try:
                    if f.is_file():
                        files.append({"name": f.name, "path": str(f)})
                except OSError:
                    pass
            new_folders.append({"folderName": fn, "path": str(name), "files": files[:50]})
    except OSError as e:
        return jsonify(ok=False, error=str(e)), 500
    return jsonify(ok=True, new=new_folders)


@app.post("/api/candidate-folder/<int:candidate_id>/open")
@login_required
@role_required('admin')
def api_candidate_folder_open(candidate_id: int):
    """Open the candidate's folder in Windows Explorer."""
    import subprocess
    chk = _require_same_origin()
    if chk:
        return chk
    with _conn() as conn:
        row = conn.execute("SELECT name FROM candidates WHERE id=?;", (candidate_id,)).fetchone()
        if not row:
            return jsonify(ok=False, error="Candidate not found"), 404
        folder = _build_candidate_folder_path(row["name"], conn)

    if not folder or not folder.exists():
        return jsonify(ok=False, error="Dossier introuvable")

    try:
        subprocess.Popen(["explorer", str(folder)])
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e))


@app.post("/api/candidate-folder/open-file")
@login_required
@role_required('admin')
def api_candidate_folder_open_file():
    """Open a specific file from a candidate's folder."""
    import subprocess
    chk = _require_same_origin()
    if chk:
        return chk
    payload = request.get_json(force=True, silent=False) or {}
    filepath = payload.get("path", "")
    candidate_id = payload.get("candidate_id")
    if not filepath:
        return jsonify(ok=False, error="path required"), 400

    p = Path(filepath)
    if not p.exists():
        return jsonify(ok=False, error="Fichier introuvable")

    # Safety: the path must be inside the configured candidate folder base (and ideally inside the candidate folder)
    try:
        with _conn() as conn:
            base_path = _get_setting(conn, "candidate_folder_base", "") or ""
            allowed_root = Path(base_path).resolve() if base_path else None

            cand_root = None
            if candidate_id:
                row = conn.execute("SELECT name FROM candidates WHERE id=?;", (int(candidate_id),)).fetchone()
                if row:
                    cand_root = _build_candidate_folder_path(row["name"], conn)
                    cand_root = cand_root.resolve() if cand_root else None

        rp = p.resolve()
        if cand_root and cand_root != rp and cand_root not in rp.parents:
            return jsonify(ok=False, error="Chemin non autorisé"), 403
        if allowed_root and allowed_root != rp and allowed_root not in rp.parents:
            return jsonify(ok=False, error="Chemin non autorisé"), 403
    except Exception:
        return jsonify(ok=False, error="Chemin non autorisé"), 403

    try:
        import os as _os
        _os.startfile(str(p))
        return jsonify(ok=True)
    except Exception as e:
        try:
            subprocess.Popen(["explorer", str(p)])
            return jsonify(ok=True)
        except Exception as e2:
            return jsonify(ok=False, error=str(e2))


# ═══════════════════════════════════════════════════════════════════
# Collaboration API (v25.5)
# ═══════════════════════════════════════════════════════════════════

# Routes /api/collab/* + /api/dashboard/adaptive + /api/dashboard/assistant* — voir routes/collab.py
# Routes /dc-generator/* + /api/dc/* + /candidates/<id>/dossier/download — voir routes/dc.py




# ── Update-check state (rempli par APScheduler toutes les 10 min) ─
_update_check_state: dict = {
    'update_available': False,
    'local_commit': '',
    'remote_commit': '',
    'checked_at': None,
    'error': None,
}
_update_check_lock = threading.Lock()


def _do_git_update_check():
    """Fetch origin/main et compare HEAD vs origin/main. Appelé par APScheduler."""
    global _update_check_state
    import subprocess as _sp
    try:
        _sp.run(
            ['git', 'fetch', 'origin', 'main'],
            cwd=str(APP_DIR), capture_output=True, timeout=20,
        )
        local = _sp.run(
            ['git', 'rev-parse', 'HEAD'],
            cwd=str(APP_DIR), capture_output=True, text=True, timeout=5,
        ).stdout.strip()
        remote = _sp.run(
            ['git', 'rev-parse', 'origin/main'],
            cwd=str(APP_DIR), capture_output=True, text=True, timeout=5,
        ).stdout.strip()
        available = bool(local and remote and local != remote)
        with _update_check_lock:
            _update_check_state = {
                'update_available': available,
                'local_commit': local[:8] if local else '',
                'remote_commit': remote[:8] if remote else '',
                'checked_at': datetime.datetime.now().isoformat(timespec='seconds'),
                'error': None,
            }
        if available:
            logger.info("Update check: nouvelle version disponible (local=%s remote=%s)",
                        local[:8], remote[:8])
    except Exception as exc:
        with _update_check_lock:
            _update_check_state['error'] = str(exc)
            _update_check_state['checked_at'] = datetime.datetime.now().isoformat(timespec='seconds')
        logger.warning("Update check échoué : %s", exc)


# ── Blueprints ────────────────────────────────────────────────────
# Importés en bas de fichier pour que tous les helpers soient déjà
# définis. Quand app.py est lancé comme script (__name__ == '__main__'),
# Python l'enregistre sous '__main__' et non 'app'. Les blueprints qui
# font `from app import ...` déclencheraient alors un import circulaire.
# Solution : on enregistre ce module sous le nom 'app' dans sys.modules
# avant les imports, ce qui évite un second chargement.
import sys as _sys  # noqa: E402
_sys.modules.setdefault('app', _sys.modules[__name__])

from routes.auth import auth_bp    # noqa: E402
# misc importé en premier car contient des helpers (deploy validation timer)
# utilisés par routes/deploy.py qui les ré-importe via `from app import ...`.
from routes.misc import (  # noqa: E402,F401
    _VALIDATION_TIMEOUT_SECONDS,
    _cancel_validation_timer,
    _schedule_restart,
    _start_validation_timer,
    _write_pending_validation,
    misc_bp,
)
from routes.deploy import deploy_bp  # noqa: E402
from routes.ai import ai_bp          # noqa: E402
from routes.transcription import transcription_bp, init_resume as _transcription_init_resume  # noqa: E402
from routes.besoins import besoins_bp  # noqa: E402
from routes.map import map_bp  # noqa: E402
from routes.companies import companies_bp  # noqa: E402
from routes.pages import pages_bp  # noqa: E402
from routes.settings import settings_bp  # noqa: E402
from routes.calendar import calendar_bp  # noqa: E402
from routes.attachments import attachments_bp  # noqa: E402
from routes.candidates import candidates_bp  # noqa: E402
from routes.push import push_bp  # noqa: E402
from routes.dashboard import dashboard_bp  # noqa: E402
from routes.prospects import prospects_bp  # noqa: E402
from routes.duplicates import duplicates_bp  # noqa: E402
from routes.push_logs import push_logs_bp  # noqa: E402
from routes.meetings import meetings_bp  # noqa: E402
from routes.bulk import bulk_bp  # noqa: E402
from routes.admin import admin_bp  # noqa: E402
from routes.dc import dc_bp  # noqa: E402
from routes.collab import collab_bp  # noqa: E402
app.register_blueprint(auth_bp)
app.register_blueprint(deploy_bp)
app.register_blueprint(ai_bp)
app.register_blueprint(transcription_bp)
app.register_blueprint(besoins_bp)
app.register_blueprint(map_bp)
app.register_blueprint(companies_bp)
app.register_blueprint(pages_bp)
app.register_blueprint(settings_bp)
app.register_blueprint(calendar_bp)
app.register_blueprint(attachments_bp)
app.register_blueprint(candidates_bp)
app.register_blueprint(push_bp)
app.register_blueprint(dashboard_bp)
app.register_blueprint(prospects_bp)
app.register_blueprint(duplicates_bp)
app.register_blueprint(push_logs_bp)
app.register_blueprint(meetings_bp)
app.register_blueprint(bulk_bp)
app.register_blueprint(admin_bp)
app.register_blueprint(misc_bp)
app.register_blueprint(dc_bp)
app.register_blueprint(collab_bp)


if __name__ == "__main__":
    DATA_DIR.mkdir(exist_ok=True)
    init_db()
    _migrate_users_schema()
    _migrate_candidate_statuses(DB_PATH)
    _migrate_all_user_dbs()
    _migrate_v30_all()
    load_initial_data_if_needed()
    # v32.1 — marque les transcriptions interrompues (crash/redémarrage) en erreur
    try:
        _transcription_init_resume()
    except Exception as _exc:
        logger.warning("init_resume transcription : %s", _exc)

    # Vérifier si une validation post-update est en attente (app redémarrée après un pull)
    if (APP_DIR / ".pending_validation").exists():
        _start_validation_timer()
        logger.info("Validation post-mise à jour en attente — rollback automatique dans %ds si non confirmée",
                    _VALIDATION_TIMEOUT_SECONDS)

    # Production mode with waitress (HTTPS via Cloudflare Tunnel)
    use_waitress = '--production' in sys.argv or '--prod' in sys.argv
    host = "0.0.0.0"  # Bind all interfaces for tunnel access
    port = int(os.environ.get("PORT", 8000))

    logger.info("Prosp'Up v%s starting (mode=%s, host=%s, port=%d)",
                APP_VERSION, "production" if use_waitress else "dev", host, port)

    # Scheduler backup journalier (3h00 chaque nuit) + purge soft-deleted (dimanche 4h00)
    # Ignoré dans le processus watcher de Werkzeug pour éviter le double démarrage
    if use_waitress or os.environ.get('WERKZEUG_RUN_MAIN'):
        try:
            from apscheduler.schedulers.background import BackgroundScheduler
            from backup import create_backup as _backup_create
            import atexit

            def _purge_old_soft_deletes():
                """v27.10: Supprime définitivement les enregistrements soft-deleted depuis plus de 30 jours."""
                cutoff = (datetime.datetime.now() - datetime.timedelta(days=30)).isoformat(timespec="seconds")
                try:
                    with _conn() as conn:
                        purged = {}
                        for tbl in ("prospects", "companies", "candidates"):
                            cur = conn.execute(f"DELETE FROM {tbl} WHERE deleted_at IS NOT NULL AND deleted_at < ?;", (cutoff,))
                            purged[tbl] = cur.rowcount
                    total = sum(purged.values())
                    if total:
                        logger.info("Purge soft-deleted: %s enregistrements supprimés (%s)", total, purged)
                    _audit_log("purge", "system", new_value=json.dumps(purged))
                except Exception as exc:
                    logger.error("Erreur purge soft-deleted: %s", exc)

            _scheduler = BackgroundScheduler()
            _scheduler.add_job(
                func=_backup_create,
                trigger='cron',
                hour=3, minute=0,
                id='daily_backup',
                replace_existing=True,
            )
            _scheduler.add_job(
                func=_purge_old_soft_deletes,
                trigger='cron',
                day_of_week='sun', hour=4, minute=0,
                id='weekly_purge_soft_deleted',
                replace_existing=True,
            )
            _scheduler.add_job(
                func=_do_git_update_check,
                trigger='interval',
                minutes=10,
                id='git_update_check',
                replace_existing=True,
            )
            _scheduler.start()
            atexit.register(lambda: _scheduler.shutdown())
            logger.info("Scheduler démarré — backup 3h00, purge soft-deleted dim. 4h00, update-check toutes les 10 min")
            # Premier check immédiat au démarrage (en thread pour ne pas bloquer)
            threading.Thread(target=_do_git_update_check, daemon=True, name='update_check_startup').start()
        except ImportError:
            logger.warning("apscheduler non installé — backup/purge automatique désactivés. Installer : pip install apscheduler")

    if use_waitress:
        try:
            from waitress import serve
            print(f"Prosp'Up v{APP_VERSION} en production (waitress) sur http://{host}:{port}")
            logger.info("Waitress server started with 4 threads")
            serve(app, host=host, port=port, threads=4)
        except ImportError:
            print("ATTENTION: waitress non installe, fallback sur Flask dev server")
            logger.warning("waitress not installed, falling back to Flask dev server")
            print(f"Prosp'Up demarre sur http://{host}:{port}")
            app.run(host=host, port=port, debug=False)
    else:
        print("ATTENTION: Mode developpement — NE PAS utiliser en production 24/7")
        print(f"    Lancer avec: python app.py --prod")
        print(f"Prosp'Up v{APP_VERSION} en dev sur http://127.0.0.1:{port}")
        logger.info("Dev server started (debug=True) — not for 24/7 use")
        app.run(host="127.0.0.1", port=port, debug=True)
