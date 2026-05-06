"""ProspUp — Blueprint Meetings (réunions, action items, opportunities, RDV checklist)."""
from __future__ import annotations

import datetime
import io
import json
import os
import re
from io import BytesIO
from pathlib import Path

from flask import Blueprint, Response, jsonify, request, send_file
from werkzeug.utils import secure_filename

from app import _audit_log, _create_auto_task, log_activity, logger
from config import APP_VERSION
from utils.auth import _prospect_owned, _uid, validate_payload
from utils.common import _now_iso
from utils.db import _conn
from utils.files import _validate_upload

meetings_bp = Blueprint("meetings", __name__)


@meetings_bp.get("/api/rdv-checklist/themes")
def rdv_checklist_themes():
    """Return the reference checklist themes (read-only list)."""
    return jsonify(ok=True, themes=RDV_CHECKLIST_THEMES)


@meetings_bp.get("/api/rdv-checklist")
def rdv_checklist_get():
    """Fetch saved checklist data for a prospect (owner only)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    pid = request.args.get("prospect_id", type=int)
    if not pid:
        return jsonify(ok=False, error="prospect_id requis"), 400
    if not _prospect_owned(pid):
        return jsonify(ok=False, error="Accès refusé"), 403
    with _conn() as conn:
        row = conn.execute(
            "SELECT data, updatedAt FROM rdv_checklists WHERE prospect_id=?", (pid,)
        ).fetchone()
    if row and row["data"]:
        return jsonify(ok=True, data=json.loads(row["data"]), updatedAt=row["updatedAt"])
    # Return blank structure
    blank = {t["key"]: {"reponse": "", "checked": False} for t in RDV_CHECKLIST_THEMES}
    return jsonify(ok=True, data=blank, updatedAt=None)


@meetings_bp.post("/api/rdv-checklist")
def rdv_checklist_save():
    """Save checklist data for a prospect (upsert, owner only)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    body = request.get_json(force=True)
    pid = body.get("prospect_id")
    data = body.get("data")
    if not pid or data is None:
        return jsonify(ok=False, error="prospect_id et data requis"), 400
    if not _prospect_owned(int(pid)):
        return jsonify(ok=False, error="Accès refusé"), 403
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with _conn() as conn:
        conn.execute(
            """INSERT INTO rdv_checklists (prospect_id, data, updatedAt)
               VALUES (?, ?, ?)
               ON CONFLICT(prospect_id)
               DO UPDATE SET data=excluded.data, updatedAt=excluded.updatedAt""",
            (pid, json.dumps(data, ensure_ascii=False), now),
        )
    return jsonify(ok=True, updatedAt=now)


@meetings_bp.post("/api/rdv-checklist/parse-file")
def rdv_checklist_parse_file():
    """Parse uploaded file (PDF, Word, Excel) and extract text content."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    if 'file' not in request.files:
        return jsonify(ok=False, error="Aucun fichier fourni"), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify(ok=False, error="Fichier vide"), 400

    ok_upload, err_upload = _validate_upload(file, "document_or_excel")
    if not ok_upload:
        return jsonify(ok=False, error=err_upload[0]), err_upload[1]

    filename = file.filename.lower()
    text = None

    try:
        if filename.endswith('.pdf'):
            from pypdf import PdfReader
            raw = file.read()
            reader = PdfReader(BytesIO(raw))
            text_parts = []
            for page in reader.pages:
                text_parts.append(page.extract_text() or '')
            text = '\n'.join(text_parts)
        
        elif filename.endswith(('.doc', '.docx')):
            from docx import Document
            doc = Document(BytesIO(file.read()))
            text_parts = []
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)
            text = '\n'.join(text_parts)
        
        elif filename.endswith(('.xls', '.xlsx')):
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file.read()), read_only=True)
            text_parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            text = '\n'.join(text_parts)
        
        elif filename.endswith('.txt'):
            text = file.read().decode('utf-8', errors='ignore')
        
        else:
            return jsonify(ok=False, error=f"Format de fichier non supporté: {filename}"), 400
        
        if not text or not text.strip():
            return jsonify(ok=False, error="Aucun texte extrait du fichier"), 400
        
        return jsonify(ok=True, text=text.strip())
    
    except Exception as e:
        import traceback
        _log_handler.handle(logging.LogRecord(
            name='prospup', level=logging.ERROR, pathname=__file__, lineno=0,
            msg=f"Erreur parsing fichier: {str(e)}\n{traceback.format_exc()}", args=(), exc_info=None
        ))
        return jsonify(ok=False, error=f"Erreur lors de l'extraction: {str(e)}"), 500

# ────────────────────────────────────────────────────────────────────
# Meetings – historique des réunions avec grille de qualification
# ────────────────────────────────────────────────────────────────────

@meetings_bp.post("/api/meetings")
def meetings_create():
    """Créer une nouvelle réunion (CR de RDV) avec snapshot grille + IA fields + tâches inline."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    body, err = validate_payload({'prospect_id': (int, str), 'title': str})
    if err:
        return err
    prospect_id = body.get("prospect_id")
    title = body.get("title", "").strip()
    checklist_data = body.get("checklist_data")
    notes = body.get("notes", "").strip()
    raw_transcript = (body.get("raw_transcript") or "").strip()
    summary = (body.get("summary") or "").strip()
    next_action = (body.get("next_action") or "").strip()
    tags = body.get("tags") or []
    documents = (body.get("documents") or "").strip()
    date_override = (body.get("date") or "").strip()
    action_items = body.get("action_items") or []

    if not prospect_id:
        return jsonify(ok=False, error="prospect_id requis"), 400
    if not title:
        return jsonify(ok=False, error="Titre requis"), 400
    if not _prospect_owned(int(prospect_id)):
        return jsonify(ok=False, error="Accès refusé"), 403

    if isinstance(tags, str):
        tags = [t.strip() for t in tags.split(",") if t.strip()]
    elif isinstance(tags, list):
        tags = [str(t).strip() for t in tags if str(t).strip()]
    else:
        tags = []

    now = datetime.datetime.now().isoformat(timespec="seconds")
    today = date_override if date_override else datetime.datetime.now().strftime("%Y-%m-%d")

    with _conn() as conn:
        cursor = conn.execute(
            """INSERT INTO meetings (prospect_id, owner_id, date, title, checklist_data, notes,
                                    summary, raw_transcript, next_action, tags, documents, createdAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                prospect_id, uid, today, title,
                json.dumps(checklist_data, ensure_ascii=False) if checklist_data else None,
                notes, summary, raw_transcript, next_action,
                json.dumps(tags, ensure_ascii=False) if tags else None,
                documents, now,
            )
        )
        meeting_id = cursor.lastrowid

        # Action items inline (créés en même temps que le CR)
        for ai in action_items:
            if not isinstance(ai, dict):
                continue
            task_txt = (ai.get("task") or "").strip()
            if not task_txt:
                continue
            conn.execute(
                """INSERT INTO meeting_action_items (meeting_id, prospect_id, task, assignee, due_date, priority, status, owner_id, createdAt)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    meeting_id, prospect_id, task_txt,
                    (ai.get("assignee") or None),
                    (ai.get("due_date") or None),
                    (ai.get("priority") or None),
                    (ai.get("status") or "pending"),
                    uid, now,
                )
            )

        # Hook: réunion créée (meeting_done)
        try:
            p_row = conn.execute(
                "SELECT id, name, email, telephone, linkedin, statut, pertinence, nextFollowUp, company_id FROM prospects WHERE id=? AND owner_id=?;",
                (prospect_id, uid)
            ).fetchone()
            if p_row:
                context = {
                    "prospect_id": p_row["id"],
                    "name": p_row["name"] or "",
                    "email": p_row["email"],
                    "telephone": p_row["telephone"],
                    "linkedin": p_row["linkedin"],
                    "statut": p_row["statut"],
                    "pertinence": p_row["pertinence"],
                    "nextFollowUp": p_row["nextFollowUp"],
                    "company_id": p_row["company_id"],
                    "meeting_title": title,
                    "meeting_notes": notes,
                }
                if context.get("company_id"):
                    c_row = conn.execute(
                        "SELECT groupe FROM companies WHERE id=? AND owner_id=?;",
                        (context["company_id"], uid)
                    ).fetchone()
                    if c_row:
                        context["company_groupe"] = c_row["groupe"] or ""
                _create_auto_task("meeting_done", context)
        except Exception as e:
            logger.warning("Erreur hook tâche auto pour réunion: %s", e)

        # v32.0 : rattacher des pièces jointes existantes au CR
        attachment_ids = body.get("attachment_ids") or []
        if isinstance(attachment_ids, list):
            for aid in attachment_ids:
                try:
                    aid_i = int(aid)
                    conn.execute(
                        "UPDATE prospect_attachments SET meeting_id = ? WHERE id = ? AND owner_id = ? AND prospect_id = ?",
                        (meeting_id, aid_i, uid, prospect_id)
                    )
                except (TypeError, ValueError):
                    continue

    return jsonify(ok=True, id=meeting_id, date=today)


def _meeting_row_to_dict(row, with_checklist=True):
    checklist = None
    if with_checklist and row["checklist_data"]:
        try:
            checklist = json.loads(row["checklist_data"])
        except Exception:
            pass
    tags = []
    try:
        if "tags" in row.keys() and row["tags"]:
            tags = json.loads(row["tags"]) or []
    except Exception:
        tags = []
    out = {
        "id": row["id"],
        "date": row["date"],
        "title": row["title"],
        "notes": row["notes"] or "",
        "summary": (row["summary"] if "summary" in row.keys() else "") or "",
        "raw_transcript": (row["raw_transcript"] if "raw_transcript" in row.keys() else "") or "",
        "next_action": (row["next_action"] if "next_action" in row.keys() else "") or "",
        "tags": tags,
        "documents": (row["documents"] if "documents" in row.keys() else "") or "",
        "createdAt": row["createdAt"],
    }
    if with_checklist:
        out["checklist_data"] = checklist
    return out


@meetings_bp.get("/api/meetings")
def meetings_list():
    """Lister les réunions d'un prospect (owner only) — légère, sans grille détaillée."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    prospect_id = request.args.get("prospect_id", type=int)
    if not prospect_id:
        return jsonify(ok=False, error="prospect_id requis"), 400
    if not _prospect_owned(prospect_id):
        return jsonify(ok=False, error="Accès refusé"), 403

    with _conn() as conn:
        rows = conn.execute(
            """SELECT m.id, m.date, m.title, m.checklist_data, m.notes,
                      m.summary, m.raw_transcript, m.next_action, m.tags, m.documents, m.createdAt,
                      (SELECT COUNT(*) FROM meeting_action_items ai WHERE ai.meeting_id = m.id) AS action_count,
                      (SELECT COUNT(*) FROM meeting_action_items ai WHERE ai.meeting_id = m.id AND ai.status != 'done') AS action_pending
               FROM meetings m
               WHERE m.prospect_id = ? AND m.owner_id = ?
               ORDER BY m.date DESC, m.createdAt DESC""",
            (prospect_id, uid)
        ).fetchall()

    meetings = []
    for row in rows:
        m = _meeting_row_to_dict(row, with_checklist=False)
        m["action_count"] = row["action_count"] or 0
        m["action_pending"] = row["action_pending"] or 0
        meetings.append(m)

    return jsonify(ok=True, meetings=meetings)


@meetings_bp.get("/api/meetings/<int:meeting_id>")
def meetings_get(meeting_id):
    """Détail d'une réunion : CR + grille snapshot + action items."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    with _conn() as conn:
        row = conn.execute(
            """SELECT id, prospect_id, date, title, checklist_data, notes,
                      summary, raw_transcript, next_action, tags, documents, createdAt
               FROM meetings WHERE id = ? AND owner_id = ?""",
            (meeting_id, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Réunion introuvable"), 404
        meeting = _meeting_row_to_dict(row)
        meeting["prospect_id"] = row["prospect_id"]

        ai_rows = conn.execute(
            """SELECT id, task, assignee, due_date, priority, status, createdAt
               FROM meeting_action_items
               WHERE meeting_id = ? AND owner_id = ?
               ORDER BY status ASC, due_date ASC, createdAt ASC""",
            (meeting_id, uid)
        ).fetchall()
        meeting["action_items"] = [{
            "id": r["id"], "task": r["task"], "assignee": r["assignee"],
            "due_date": r["due_date"], "priority": r["priority"],
            "status": r["status"], "createdAt": r["createdAt"],
        } for r in ai_rows]

        # v32.0 : pièces jointes liées au CR
        att_rows = conn.execute(
            """SELECT id, original_name, size, mime_type, thumbnail
               FROM prospect_attachments
               WHERE meeting_id = ? AND owner_id = ?
               ORDER BY createdAt DESC""",
            (meeting_id, uid)
        ).fetchall()
        meeting["attachments"] = [{
            "id": r["id"], "original_name": r["original_name"],
            "size": r["size"] or 0, "mime_type": r["mime_type"] or "",
            "has_thumbnail": bool(r["thumbnail"]),
        } for r in att_rows]

    return jsonify(ok=True, meeting=meeting)


@meetings_bp.put("/api/meetings/<int:meeting_id>")
def meetings_update(meeting_id):
    """Mettre à jour un CR existant. Le payload remplace les action_items si fourni."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    body = request.get_json(force=True) or {}

    with _conn() as conn:
        existing = conn.execute(
            "SELECT id, prospect_id FROM meetings WHERE id = ? AND owner_id = ?",
            (meeting_id, uid)
        ).fetchone()
        if not existing:
            return jsonify(ok=False, error="Réunion introuvable"), 404

        prospect_id = existing["prospect_id"]
        sets = []
        vals = []

        for key in ("title", "date", "notes", "summary", "raw_transcript", "next_action", "documents"):
            if key in body:
                sets.append(f"{key} = ?")
                vals.append((body.get(key) or "").strip() if isinstance(body.get(key), str) else body.get(key))

        if "checklist_data" in body:
            cd = body.get("checklist_data")
            sets.append("checklist_data = ?")
            vals.append(json.dumps(cd, ensure_ascii=False) if cd else None)

        if "tags" in body:
            tags = body.get("tags") or []
            if isinstance(tags, str):
                tags = [t.strip() for t in tags.split(",") if t.strip()]
            elif isinstance(tags, list):
                tags = [str(t).strip() for t in tags if str(t).strip()]
            else:
                tags = []
            sets.append("tags = ?")
            vals.append(json.dumps(tags, ensure_ascii=False) if tags else None)

        if not sets and "action_items" not in body:
            return jsonify(ok=False, error="Aucun champ à mettre à jour"), 400

        if sets:
            vals.append(meeting_id)
            conn.execute(f"UPDATE meetings SET {', '.join(sets)} WHERE id = ?", vals)

        if "action_items" in body:
            now = datetime.datetime.now().isoformat(timespec="seconds")
            conn.execute("DELETE FROM meeting_action_items WHERE meeting_id = ? AND owner_id = ?", (meeting_id, uid))
            for ai in (body.get("action_items") or []):
                if not isinstance(ai, dict):
                    continue
                task_txt = (ai.get("task") or "").strip()
                if not task_txt:
                    continue
                conn.execute(
                    """INSERT INTO meeting_action_items (meeting_id, prospect_id, task, assignee, due_date, priority, status, owner_id, createdAt)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        meeting_id, prospect_id, task_txt,
                        (ai.get("assignee") or None),
                        (ai.get("due_date") or None),
                        (ai.get("priority") or None),
                        (ai.get("status") or "pending"),
                        uid, now,
                    )
                )

        # v32.0 : remplacer la liste de pièces jointes liées au CR
        if "attachment_ids" in body:
            attachment_ids = body.get("attachment_ids") or []
            # Détacher tout ce qui était lié
            conn.execute(
                "UPDATE prospect_attachments SET meeting_id = NULL WHERE meeting_id = ? AND owner_id = ?",
                (meeting_id, uid)
            )
            if isinstance(attachment_ids, list):
                for aid in attachment_ids:
                    try:
                        aid_i = int(aid)
                        conn.execute(
                            "UPDATE prospect_attachments SET meeting_id = ? WHERE id = ? AND owner_id = ? AND prospect_id = ?",
                            (meeting_id, aid_i, uid, prospect_id)
                        )
                    except (TypeError, ValueError):
                        continue

    return jsonify(ok=True, id=meeting_id)


@meetings_bp.delete("/api/meetings/<int:meeting_id>")
def meetings_delete(meeting_id):
    """Supprimer un CR (cascade sur action_items et opportunities via FK)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    with _conn() as conn:
        row = conn.execute(
            "SELECT id FROM meetings WHERE id = ? AND owner_id = ?",
            (meeting_id, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Réunion introuvable"), 404
        conn.execute("DELETE FROM meeting_action_items WHERE meeting_id = ? AND owner_id = ?", (meeting_id, uid))
        conn.execute("DELETE FROM meeting_opportunities WHERE meeting_id = ? AND owner_id = ?", (meeting_id, uid))
        conn.execute("DELETE FROM meetings WHERE id = ? AND owner_id = ?", (meeting_id, uid))

    return jsonify(ok=True)


@meetings_bp.put("/api/meeting-action-items/<int:item_id>")
def meeting_action_item_update(item_id):
    """Mettre à jour un action item (cocher fait, modifier libellé/date/priorité)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    body = request.get_json(force=True) or {}

    with _conn() as conn:
        row = conn.execute(
            "SELECT id FROM meeting_action_items WHERE id = ? AND owner_id = ?",
            (item_id, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Tâche introuvable"), 404

        sets = []
        vals = []
        for key in ("task", "assignee", "due_date", "priority", "status"):
            if key in body:
                sets.append(f"{key} = ?")
                v = body.get(key)
                vals.append(v.strip() if isinstance(v, str) else v)
        if not sets:
            return jsonify(ok=False, error="Aucun champ à mettre à jour"), 400
        vals.append(item_id)
        conn.execute(f"UPDATE meeting_action_items SET {', '.join(sets)} WHERE id = ?", vals)

    return jsonify(ok=True)


@meetings_bp.delete("/api/meeting-action-items/<int:item_id>")
def meeting_action_item_delete(item_id):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    with _conn() as conn:
        row = conn.execute(
            "SELECT id FROM meeting_action_items WHERE id = ? AND owner_id = ?",
            (item_id, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Tâche introuvable"), 404
        conn.execute("DELETE FROM meeting_action_items WHERE id = ? AND owner_id = ?", (item_id, uid))

    return jsonify(ok=True)


@meetings_bp.get("/api/meetings/<int:meeting_id>/action-items")
def meetings_action_items_list(meeting_id):
    """Lister les action items d'une réunion."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    with _conn() as conn:
        # Vérifier que la réunion existe et appartient à l'utilisateur
        meeting = conn.execute(
            "SELECT id, prospect_id FROM meetings WHERE id = ? AND owner_id = ?",
            (meeting_id, uid)
        ).fetchone()
        if not meeting:
            return jsonify(ok=False, error="Réunion introuvable"), 404
        
        rows = conn.execute(
            """SELECT id, task, assignee, due_date, priority, status, createdAt
               FROM meeting_action_items
               WHERE meeting_id = ? AND owner_id = ?
               ORDER BY due_date ASC, priority DESC, createdAt ASC""",
            (meeting_id, uid)
        ).fetchall()
    
    action_items = []
    for row in rows:
        action_items.append({
            "id": row["id"],
            "task": row["task"],
            "assignee": row["assignee"],
            "due_date": row["due_date"],
            "priority": row["priority"],
            "status": row["status"],
            "createdAt": row["createdAt"]
        })
    
    return jsonify(ok=True, action_items=action_items)


@meetings_bp.post("/api/meetings/<int:meeting_id>/action-items")
def meetings_action_items_create(meeting_id):
    """Créer un action item pour une réunion."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    body = request.get_json(force=True)
    task = body.get("task", "").strip()
    assignee = body.get("assignee")
    due_date = body.get("due_date")
    priority = body.get("priority")
    
    if not task:
        return jsonify(ok=False, error="task requis"), 400
    
    with _conn() as conn:
        # Vérifier que la réunion existe et appartient à l'utilisateur
        meeting = conn.execute(
            "SELECT id, prospect_id FROM meetings WHERE id = ? AND owner_id = ?",
            (meeting_id, uid)
        ).fetchone()
        if not meeting:
            return jsonify(ok=False, error="Réunion introuvable"), 404
        
        prospect_id = meeting["prospect_id"]
        now = datetime.datetime.now().isoformat(timespec="seconds")
        
        cursor = conn.execute(
            """INSERT INTO meeting_action_items (meeting_id, prospect_id, task, assignee, due_date, priority, status, owner_id, createdAt)
               VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?)""",
            (meeting_id, prospect_id, task, assignee, due_date, priority, uid, now)
        )
        action_item_id = cursor.lastrowid
    
    return jsonify(ok=True, id=action_item_id)


@meetings_bp.get("/api/meetings/<int:meeting_id>/opportunities")
def meetings_opportunities_list(meeting_id):
    """Lister les opportunités d'une réunion."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    with _conn() as conn:
        # Vérifier que la réunion existe et appartient à l'utilisateur
        meeting = conn.execute(
            "SELECT id, prospect_id FROM meetings WHERE id = ? AND owner_id = ?",
            (meeting_id, uid)
        ).fetchone()
        if not meeting:
            return jsonify(ok=False, error="Réunion introuvable"), 404
        
        rows = conn.execute(
            """SELECT id, type, estimated_value, probability, description, createdAt
               FROM meeting_opportunities
               WHERE meeting_id = ? AND owner_id = ?
               ORDER BY estimated_value DESC, probability DESC, createdAt ASC""",
            (meeting_id, uid)
        ).fetchall()
    
    opportunities = []
    for row in rows:
        opportunities.append({
            "id": row["id"],
            "type": row["type"],
            "estimated_value": row["estimated_value"],
            "probability": row["probability"],
            "description": row["description"],
            "createdAt": row["createdAt"]
        })
    
    return jsonify(ok=True, opportunities=opportunities)


@meetings_bp.post("/api/meetings/<int:meeting_id>/opportunities")
def meetings_opportunities_create(meeting_id):
    """Créer une opportunité pour une réunion."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    body = request.get_json(force=True)
    type_opp = body.get("type", "").strip()
    estimated_value = body.get("estimated_value")
    probability = body.get("probability")
    description = body.get("description")
    
    if not type_opp:
        return jsonify(ok=False, error="type requis"), 400
    
    with _conn() as conn:
        # Vérifier que la réunion existe et appartient à l'utilisateur
        meeting = conn.execute(
            "SELECT id, prospect_id FROM meetings WHERE id = ? AND owner_id = ?",
            (meeting_id, uid)
        ).fetchone()
        if not meeting:
            return jsonify(ok=False, error="Réunion introuvable"), 404
        
        prospect_id = meeting["prospect_id"]
        now = datetime.datetime.now().isoformat(timespec="seconds")
        
        cursor = conn.execute(
            """INSERT INTO meeting_opportunities (meeting_id, prospect_id, type, estimated_value, probability, description, owner_id, createdAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (meeting_id, prospect_id, type_opp, estimated_value, probability, description, uid, now)
        )
        opportunity_id = cursor.lastrowid
    
    return jsonify(ok=True, id=opportunity_id)


@meetings_bp.get("/api/meetings/<int:meeting_id>/pdf")
def meetings_export_pdf(meeting_id):
    """Exporter une réunion en PDF."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    with _conn() as conn:
        row = conn.execute(
            """SELECT m.id, m.prospect_id, m.date, m.title, m.checklist_data, m.notes, m.createdAt,
                      p.name as prospect_name, p.fonction, c.groupe as company_name, c.site
               FROM meetings m
               JOIN prospects p ON m.prospect_id = p.id
               LEFT JOIN companies c ON p.company_id = c.id
               WHERE m.id = ? AND m.owner_id = ?""",
            (meeting_id, uid)
        ).fetchone()
    
    if not row:
        return jsonify(ok=False, error="Réunion introuvable"), 404
    
    # Parse checklist data
    checklist = None
    if row["checklist_data"]:
        try:
            checklist = json.loads(row["checklist_data"])
        except Exception:
            pass
    
    # Load themes for display
    themes_dict = {t["key"]: t for t in RDV_CHECKLIST_THEMES}
    
    # Generate HTML for PDF
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; padding: 40px; line-height: 1.6; color: #333; }}
            h1 {{ color: #f59e0b; border-bottom: 3px solid #f59e0b; padding-bottom: 10px; }}
            h2 {{ color: #6366f1; margin-top: 30px; }}
            .header {{ margin-bottom: 30px; }}
            .info-row {{ margin: 8px 0; }}
            .info-label {{ font-weight: bold; display: inline-block; width: 150px; }}
            .section {{ margin: 20px 0; padding: 15px; background: #f9fafb; border-left: 4px solid #6366f1; }}
            .theme-title {{ font-weight: bold; color: #6366f1; margin-top: 15px; }}
            .theme-question {{ color: #666; font-style: italic; margin: 5px 0; }}
            .theme-answer {{ margin: 10px 0 20px 20px; white-space: pre-wrap; }}
            .notes {{ margin-top: 30px; padding: 15px; background: #fff3cd; border-left: 4px solid #f59e0b; }}
        </style>
    </head>
    <body>
        <h1>Compte-rendu de réunion</h1>
        <div class="header">
            <div class="info-row"><span class="info-label">Date :</span> {escape_html(str(row["date"] or ""))}</div>
            <div class="info-row"><span class="info-label">Titre :</span> {escape_html(str(row["title"] or ""))}</div>
            <div class="info-row"><span class="info-label">Prospect :</span> {escape_html(str(row["prospect_name"] or ""))}</div>
            <div class="info-row"><span class="info-label">Fonction :</span> {escape_html(str(row["fonction"] or ""))}</div>
            <div class="info-row"><span class="info-label">Entreprise :</span> {escape_html(str((row["company_name"] or "") + (" (" + (row["site"] or "") + ")" if row["site"] else "")))}</div>
        </div>
    """
    
    if checklist:
        html_content += "<h2>Grille de qualification</h2>"
        for key, data in checklist.items():
            if not data or not isinstance(data, dict):
                continue
            theme = themes_dict.get(key)
            if not theme:
                continue
            reponse = data.get("reponse", "").strip()
            if not reponse:
                continue
            html_content += f"""
            <div class="section">
                <div class="theme-title">{escape_html(str(theme["theme"]))}</div>
                <div class="theme-question">{escape_html(str(theme["question"]))}</div>
                <div class="theme-answer">{escape_html(str(reponse))}</div>
            </div>
            """
    
    if row["notes"]:
        html_content += f'<div class="notes"><strong>Notes complémentaires :</strong><br>{escape_html(str(row["notes"]))}</div>'
    
    html_content += """
    </body>
    </html>
    """
    
    # Convert HTML to PDF using weasyprint (fallback to HTML if not available)
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return send_file(
            BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"reunion_{row['date']}_{meeting_id}.pdf"
        )
    except ImportError:
        # Fallback: return HTML that can be printed to PDF by browser (Ctrl+P > Enregistrer en PDF)
        return Response(
            html_content,
            mimetype="text/html",
            headers={"Content-Disposition": f'inline; filename="reunion_{row["date"]}_{meeting_id}.html"'}
        )


# ═══════════════════════════════════════════════════════════════════
# v31.9: Pièces jointes prospect — upload, liste, téléchargement, suppression
# ═══════════════════════════════════════════════════════════════════

# Routes /api/prospect/attachments* — voir routes/attachments.py


# ─── Lot 4 : Résumé IA d'une fiche prospect ────────────────────────

def _build_summary_prompt(prospect: dict, events: list, attachments: list) -> str:
    """Construit le prompt pour résumer un prospect à partir de son historique."""
    lines = []
    lines.append("Tu es un assistant CRM qui synthétise des fiches prospect B2B.")
    lines.append("Génère un résumé en 5 lignes maximum, factuel, actionnable, en français.")
    lines.append("Structure : 1) Qui (rôle, entreprise) 2) Statut commercial 3) Derniers échanges clés 4) Points d'accroche 5) Prochaine action recommandée.")
    lines.append("")
    lines.append("=== PROSPECT ===")
    lines.append(f"Nom : {prospect.get('name') or '—'}")
    lines.append(f"Fonction : {prospect.get('fonction') or '—'}")
    lines.append(f"Entreprise : {prospect.get('company_groupe') or '—'}{(' · ' + prospect.get('company_site')) if prospect.get('company_site') else ''}")
    lines.append(f"Statut : {prospect.get('statut') or '—'}")
    lines.append(f"Pertinence : {prospect.get('pertinence') or 0}/5")
    if prospect.get("notes"):
        lines.append(f"Notes : {(prospect.get('notes') or '')[:500]}")
    lines.append("")
    lines.append("=== HISTORIQUE (du plus récent au plus ancien) ===")
    for e in events[:30]:
        date = (e.get("date") or "")[:10]
        title = e.get("title") or e.get("type") or ""
        content = (e.get("content") or "")[:300]
        meta = e.get("meta") or {}
        extras = []
        if meta.get("next_action"):
            extras.append(f"prochaine action: {meta.get('next_action')[:200]}")
        if meta.get("action_pending"):
            extras.append(f"{meta['action_pending']} tâches en attente")
        ex_str = (" | " + " · ".join(extras)) if extras else ""
        lines.append(f"- [{date}] {title}{(': ' + content) if content else ''}{ex_str}")
    if attachments:
        lines.append("")
        lines.append("=== DOCUMENTS ASSOCIÉS ===")
        for a in attachments[:15]:
            tags = a.get("tags") or []
            tag_str = (" [" + ", ".join(tags) + "]") if tags else ""
            lines.append(f"- {a.get('original_name')}{tag_str}")
    lines.append("")
    lines.append("=== RÉSUMÉ (5 lignes max) ===")
    return "\n".join(lines)
