"""
ProspUp — Blueprint Traitement Besoin

Pages :
  GET  /v30/besoins                          — Liste des besoins
  GET  /v30/besoins/<id>                     — Détail / édition d'un besoin

API :
  POST   /api/besoins                        — Créer un besoin
  GET    /api/besoins                        — Lister mes besoins (filtre statut)
  GET    /api/besoins/<id>                   — Détail
  PUT    /api/besoins/<id>                   — Mettre à jour
  DELETE /api/besoins/<id>                   — Soft delete
  GET    /api/besoins/<id>/export.xlsx       — Export Excel (format=recto|verso)
"""
from __future__ import annotations

import datetime
import json
import re
from io import BytesIO

from flask import Blueprint, Response, jsonify, render_template, request, send_file

from app import (
    APP_VERSION,
    _conn,
    _get_current_user,
    _uid,
    logger,
)
from utils.db import _sidebar_counts

besoins_bp = Blueprint("besoins", __name__)


STATUTS = ("ouvert", "en_cours", "pourvu", "abandonne")


# ─── Helpers ──────────────────────────────────────────────────────────


def _enrich_candidats(uid: int, candidats: list) -> list:
    """Enrichit chaque entrée candidate avec les infos de la fiche (vsa_url,
    role, location, linkedin, seniority) si `cand_id` est défini.
    Le résultat est exposé sous la clé `_ref` côté front (lecture seule)."""
    if not candidats or not uid:
        return candidats
    cand_ids: list[int] = []
    for c in candidats:
        cid = c.get("cand_id") if isinstance(c, dict) else None
        if cid:
            try:
                cand_ids.append(int(cid))
            except Exception:
                pass
    if not cand_ids:
        return candidats
    placeholders = ",".join("?" * len(cand_ids))
    by_id: dict = {}
    try:
        with _conn() as conn:
            rows = conn.execute(
                f"SELECT id, name, role, location, vsa_url, linkedin, tech, "
                f"seniority, email, phone "
                f"FROM candidates WHERE owner_id=? AND id IN ({placeholders}) "
                f"AND deleted_at IS NULL;",
                (uid, *cand_ids),
            ).fetchall()
            for r in rows:
                by_id[int(r["id"])] = {
                    "id": r["id"],
                    "name": r["name"],
                    "role": r["role"],
                    "location": r["location"],
                    "vsa_url": r["vsa_url"],
                    "linkedin": r["linkedin"],
                    "tech": r["tech"],
                    "seniority": r["seniority"],
                    "email": r["email"],
                    "phone": r["phone"],
                }
    except Exception as e:
        logger.warning("besoins: enrichment failed (%s)", e)
        return candidats
    for c in candidats:
        if not isinstance(c, dict):
            continue
        cid = c.get("cand_id")
        if not cid:
            continue
        try:
            ref = by_id.get(int(cid))
        except Exception:
            ref = None
        if ref:
            c["_ref"] = ref
    return candidats


def _row_to_dict(row, uid: int | None = None) -> dict:
    if row is None:
        return {}
    d = dict(row)
    # Décoder candidats_json
    raw = d.pop("candidats_json", None)
    candidats: list = []
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, list):
                candidats = data
        except Exception:
            candidats = []
    if uid is not None:
        candidats = _enrich_candidats(uid, candidats)
    d["candidats"] = candidats
    return d


def _payload_clean(payload: dict) -> dict:
    """Filtre + sanitise un payload de création / update."""
    allowed = {
        "client", "localisation", "contact", "date_appel",
        "intitule", "date_besoin", "duree_mission",
        "descriptif", "competences", "connaissances",
        "experience", "profil_type", "commentaires",
        "preparation_rt",
        "statut", "priority",
        "prospect_id", "company_id",
    }
    out: dict = {}
    for k, v in (payload or {}).items():
        if k in allowed:
            if isinstance(v, str):
                v = v.strip()
            out[k] = v
    # Statut : normalisation
    s = out.get("statut")
    if s and s not in STATUTS:
        out["statut"] = "ouvert"
    # IDs : entiers ou None
    for k in ("prospect_id", "company_id", "priority"):
        if k in out:
            v = out[k]
            if v in ("", None):
                out[k] = None
            else:
                try:
                    out[k] = int(v)
                except Exception:
                    out[k] = None
    return out


def _resolve_prospect_context(uid: int, prospect_id: int) -> dict:
    """Retourne un dict {client, contact, company_id, localisation} pré-rempli depuis un prospect."""
    out = {}
    try:
        with _conn() as conn:
            row = conn.execute(
                """
                SELECT p.id, p.name, p.company_id,
                       c.groupe AS company_name,
                       c.city AS company_city,
                       c.country AS company_country
                FROM prospects p
                LEFT JOIN companies c ON c.id = p.company_id
                WHERE p.id=? AND p.owner_id=?
                  AND (p.deleted_at IS NULL OR p.deleted_at='')
                LIMIT 1;
                """,
                (prospect_id, uid),
            ).fetchone()
            if row:
                out["contact"] = row["name"] or ""
                out["client"] = row["company_name"] or ""
                out["company_id"] = row["company_id"]
                loc = " ".join(p for p in [row["company_city"], row["company_country"]] if p).strip()
                if loc:
                    out["localisation"] = loc
    except Exception:
        pass
    return out


# ─── Pages ────────────────────────────────────────────────────────────


@besoins_bp.get("/v30/besoins")
def page_besoins_list():
    uid = _uid()
    if not uid:
        return Response(status=302, headers={"Location": "/login"})
    u = _get_current_user() or {}
    dn = (u.get("display_name") or u.get("username") or "").strip()
    parts = [p for p in dn.split() if p]
    user_initials = "".join(p[0].upper() for p in parts[:2]) or "AB"
    return render_template(
        "v30/besoins.html",
        active="besoins",
        crumbs=["Prosp'Up", "Traitement Besoin"],
        counts=_sidebar_counts(),
        pinned=[],
        user_initials=user_initials,
        app_version=APP_VERSION,
    )


@besoins_bp.get("/v30/besoins/<int:bid>")
def page_besoins_detail(bid: int):
    uid = _uid()
    if not uid:
        return Response(status=302, headers={"Location": "/login"})
    with _conn() as conn:
        row = conn.execute(
            "SELECT id, intitule, client FROM besoins "
            "WHERE id=? AND owner_id=? AND (deleted_at IS NULL OR deleted_at='') LIMIT 1;",
            (bid, uid),
        ).fetchone()
    if not row:
        return Response(status=302, headers={"Location": "/v30/besoins"})
    u = _get_current_user() or {}
    dn = (u.get("display_name") or u.get("username") or "").strip()
    parts = [p for p in dn.split() if p]
    user_initials = "".join(p[0].upper() for p in parts[:2]) or "AB"
    title = (row["intitule"] or "").strip() or (row["client"] or "").strip() or f"Besoin #{bid}"
    return render_template(
        "v30/besoin_detail.html",
        active="besoins",
        crumbs=[
            {"label": "Prosp'Up", "href": "/v30/dashboard"},
            {"label": "Besoins", "href": "/v30/besoins"},
            title,
        ],
        counts=_sidebar_counts(),
        pinned=[],
        user_initials=user_initials,
        besoin_id=bid,
        besoin_title=title,
        app_version=APP_VERSION,
    )


# ─── API ──────────────────────────────────────────────────────────────


@besoins_bp.post("/api/besoins")
def api_create_besoin():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(silent=True) or {}
    data = _payload_clean(payload)

    # Préfill depuis prospect_id si fourni et pas de client/contact explicite
    pid = data.get("prospect_id")
    if pid:
        ctx = _resolve_prospect_context(uid, pid)
        for k, v in ctx.items():
            if not data.get(k):
                data[k] = v

    now = datetime.datetime.now().isoformat(timespec="seconds")
    statut = data.get("statut") or "ouvert"
    if statut not in STATUTS:
        statut = "ouvert"

    candidats = payload.get("candidats")
    candidats_json = None
    if isinstance(candidats, list):
        candidats_json = json.dumps(candidats, ensure_ascii=False)

    cols = (
        "client", "localisation", "contact", "date_appel",
        "intitule", "date_besoin", "duree_mission",
        "descriptif", "competences", "connaissances",
        "experience", "profil_type", "commentaires",
        "preparation_rt",
        "statut", "priority",
        "candidats_json",
        "prospect_id", "company_id",
        "owner_id", "created_at", "updated_at",
    )
    values = (
        data.get("client") or "",
        data.get("localisation") or "",
        data.get("contact") or "",
        data.get("date_appel") or "",
        data.get("intitule") or "",
        data.get("date_besoin") or "",
        data.get("duree_mission") or "",
        data.get("descriptif") or "",
        data.get("competences") or "",
        data.get("connaissances") or "",
        data.get("experience") or "",
        data.get("profil_type") or "",
        data.get("commentaires") or "",
        data.get("preparation_rt") or "",
        statut,
        data.get("priority"),
        candidats_json,
        data.get("prospect_id"),
        data.get("company_id"),
        uid,
        now,
        now,
    )
    placeholders = ",".join("?" * len(cols))
    sql = f"INSERT INTO besoins ({','.join(cols)}) VALUES ({placeholders});"
    with _conn() as conn:
        cur = conn.execute(sql, values)
        bid = cur.lastrowid
        row = conn.execute("SELECT * FROM besoins WHERE id=?", (bid,)).fetchone()
    return jsonify(ok=True, besoin=_row_to_dict(row, uid=uid))


@besoins_bp.get("/api/besoins")
def api_list_besoins():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    statut = (request.args.get("statut") or "").strip()
    prospect_id = (request.args.get("prospect_id") or "").strip()

    where = ["b.owner_id=?", "(b.deleted_at IS NULL OR b.deleted_at='')"]
    params: list = [uid]
    if statut and statut in STATUTS:
        where.append("b.statut=?")
        params.append(statut)
    if prospect_id:
        try:
            where.append("b.prospect_id=?")
            params.append(int(prospect_id))
        except Exception:
            pass

    sql = (
        "SELECT b.*, p.name AS prospect_name, c.groupe AS company_name "
        "FROM besoins b "
        "LEFT JOIN prospects p ON p.id = b.prospect_id "
        "LEFT JOIN companies c ON c.id = b.company_id "
        f"WHERE {' AND '.join(where)} "
        "ORDER BY b.created_at DESC, b.id DESC LIMIT 500;"
    )
    with _conn() as conn:
        rows = conn.execute(sql, params).fetchall()
    items = []
    for r in rows:
        d = _row_to_dict(r, uid=uid)
        # Stats candidats
        cands = d.get("candidats") or []
        d["candidats_count"] = len(cands)
        items.append(d)
    return jsonify(ok=True, items=items)


@besoins_bp.get("/api/besoins/<int:bid>")
def api_get_besoin(bid: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT b.*, p.name AS prospect_name, c.groupe AS company_name "
            "FROM besoins b "
            "LEFT JOIN prospects p ON p.id = b.prospect_id "
            "LEFT JOIN companies c ON c.id = b.company_id "
            "WHERE b.id=? AND b.owner_id=? AND (b.deleted_at IS NULL OR b.deleted_at='') LIMIT 1;",
            (bid, uid),
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Introuvable"), 404
    return jsonify(ok=True, besoin=_row_to_dict(row, uid=uid))


@besoins_bp.put("/api/besoins/<int:bid>")
def api_update_besoin(bid: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(silent=True) or {}
    data = _payload_clean(payload)

    with _conn() as conn:
        existing = conn.execute(
            "SELECT id, client, contact, localisation, company_id "
            "FROM besoins WHERE id=? AND owner_id=? "
            "AND (deleted_at IS NULL OR deleted_at='') LIMIT 1;",
            (bid, uid),
        ).fetchone()
        if not existing:
            return jsonify(ok=False, error="Introuvable"), 404

        # Si on lie un nouveau prospect, pré-remplir les champs vides
        if data.get("prospect_id"):
            ctx = _resolve_prospect_context(uid, data["prospect_id"])
            existing_cols = set(existing.keys())
            for k, v in ctx.items():
                if k in data:
                    continue
                cur_val = existing[k] if k in existing_cols else None
                if isinstance(cur_val, str):
                    is_empty = not cur_val.strip()
                else:
                    is_empty = cur_val is None
                if is_empty:
                    data[k] = v

        sets = []
        values: list = []
        for k, v in data.items():
            sets.append(f"{k}=?")
            values.append(v)

        if "candidats" in payload:
            cands = payload.get("candidats")
            # Nettoie les clés enrichies côté front (lecture seule) avant
            # persistance — `_ref` est rebuilé à chaque GET via JOIN.
            if isinstance(cands, list):
                cleaned = []
                for c in cands:
                    if isinstance(c, dict):
                        cleaned.append({k: v for k, v in c.items() if not k.startswith("_")})
                    else:
                        cleaned.append(c)
                cands = cleaned
            sets.append("candidats_json=?")
            values.append(json.dumps(cands, ensure_ascii=False) if isinstance(cands, list) else None)

        sets.append("updated_at=?")
        values.append(datetime.datetime.now().isoformat(timespec="seconds"))
        values.extend([bid, uid])

        if sets:
            sql = f"UPDATE besoins SET {','.join(sets)} WHERE id=? AND owner_id=?;"
            conn.execute(sql, values)
        row = conn.execute(
            "SELECT b.*, p.name AS prospect_name, c.groupe AS company_name "
            "FROM besoins b "
            "LEFT JOIN prospects p ON p.id = b.prospect_id "
            "LEFT JOIN companies c ON c.id = b.company_id "
            "WHERE b.id=?;",
            (bid,),
        ).fetchone()
    return jsonify(ok=True, besoin=_row_to_dict(row, uid=uid))


@besoins_bp.delete("/api/besoins/<int:bid>")
def api_delete_besoin(bid: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    now = datetime.datetime.now().isoformat(timespec="seconds")
    with _conn() as conn:
        cur = conn.execute(
            "UPDATE besoins SET deleted_at=? WHERE id=? AND owner_id=? "
            "AND (deleted_at IS NULL OR deleted_at='');",
            (now, bid, uid),
        )
        if cur.rowcount == 0:
            return jsonify(ok=False, error="Introuvable"), 404
    return jsonify(ok=True)


# ─── Export Excel ─────────────────────────────────────────────────────


def _safe_filename(s: str) -> str:
    s = (s or "").strip() or "besoin"
    s = re.sub(r"[^A-Za-z0-9._ -]+", "", s)
    s = re.sub(r"\s+", "_", s)
    return s[:80]


def _build_sheet(ws, b: dict, cands: list):
    """Construit la feuille unique 'recto verso' selon le template Scintil (13 colonnes A-M)."""
    from openpyxl.styles import Alignment, Border, Font, Side

    THIN = Side(border_style="thin", color="000000")
    MED = Side(border_style="medium", color="000000")
    BOLD = Font(name="Calibri", size=11, bold=True)
    REG = Font(name="Calibri", size=11, bold=False)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    WRAP = Alignment(vertical="center", wrap_text=True)

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.236
    ws.page_margins.right = 0.236
    ws.page_margins.top = 0.354
    ws.page_margins.bottom = 0.354

    widths = {
        "A": 33.1, "B": 33.2, "C": 8.8, "D": 8.4, "E": 8.4,
        "F": 12.2, "G": 16.3, "H": 8.6, "I": 8.8, "J": 27.8,
        "K": 39.9, "L": 31.7, "M": 28.6, "N": 11.4,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 28.5
    ws.row_dimensions[4].height = 105.6
    ws.row_dimensions[5].height = 63.0
    ws.row_dimensions[6].height = 43.5
    ws.row_dimensions[7].height = 76.8
    ws.row_dimensions[8].height = 43.2
    ws.row_dimensions[9].height = 15.0
    for r in range(10, 63):
        ws.row_dimensions[r].height = 31.5

    # Fusions header
    ws.merge_cells("B1:E1")
    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:J1")
    ws.merge_cells("K1:L1")
    ws.merge_cells("B2:G2")
    ws.merge_cells("I2:J2")
    ws.merge_cells("L2:M2")
    ws.merge_cells("B3:M3")
    ws.merge_cells("B4:M4")
    ws.merge_cells("B5:M5")
    ws.merge_cells("B6:M6")
    ws.merge_cells("B7:F7")
    ws.merge_cells("G7:H7")
    ws.merge_cells("I7:M7")
    ws.merge_cells("B8:M8")

    # Cellules header
    ws["A1"] = "Client";          ws["B1"] = b.get("client") or ""
    ws["F1"] = "Localisation";    ws["H1"] = b.get("localisation") or ""
    ws["K1"] = "Durée mission";   ws["M1"] = b.get("duree_mission") or ""
    ws["A2"] = "Contact";         ws["B2"] = b.get("contact") or ""
    ws["H2"] = "Date appel";      ws["I2"] = b.get("date_appel") or ""
    ws["K2"] = "Date besoin";     ws["L2"] = b.get("date_besoin") or ""
    ws["A3"] = "Besoin";          ws["B3"] = b.get("intitule") or ""
    ws["A4"] = "Descriptif";              ws["B4"] = b.get("descriptif") or ""
    ws["A5"] = "Compétences requises";    ws["B5"] = b.get("competences") or ""
    ws["A6"] = "Connaissances attendues"; ws["B6"] = b.get("connaissances") or ""
    ws["A7"] = "Expérience";      ws["B7"] = b.get("experience") or ""
    ws["G7"] = "Ingénieur et / ou Technicien ?"; ws["I7"] = b.get("profil_type") or ""
    ws["A8"] = "Commentaires";    ws["B8"] = b.get("commentaires") or ""

    # Header candidats (row 9, colonnes A-M)
    headers = ["Candidat", "Commentaires", "Dispo", "Appel", "DT",
               "RDV1", "RDV2", "RT", "Envoi DT", "Propal", "RT",
               "Lieux Habitation", "Diplome"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=i, value=h)
        cell.font = BOLD
        cell.alignment = CENTER

    # Données candidats (rows 10-62, 53 lignes, 13 colonnes)
    keys = ["candidat", "commentaires", "dispo", "appel", "dt",
            "rdv1", "rdv2", "rt", "envoi_dt", "propal", "rt_client",
            "lieu_habitation", "diplome"]
    for ri in range(53):
        r = 10 + ri
        c = cands[ri] if ri < len(cands) else {}
        for ci, k in enumerate(keys, start=1):
            ws.cell(row=r, column=ci, value=c.get(k) or "")

    # Bordures (A1:M62)
    for r in range(1, 63):
        for ci in range(1, 14):
            cell = ws.cell(row=r, column=ci)
            top = MED if r == 1 or r == 9 else THIN
            bottom = MED if r == 8 or r == 9 else THIN
            left = MED if ci == 1 else THIN
            right = MED if ci == 13 else THIN
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            if r <= 8 and ci == 1:
                cell.font = BOLD
                cell.alignment = WRAP
            elif r == 9:
                cell.font = BOLD
                cell.alignment = CENTER
            elif r <= 8 and ci > 1:
                cell.font = BOLD
                cell.alignment = LEFT
            else:
                cell.font = REG
                cell.alignment = LEFT

    ws.print_area = "A1:M62"


# ─── Export PDF ───────────────────────────────────────────────────────


# Couleurs reprises du design system v30 (équivalents hex des oklch tokens).
_PDF_COLORS = {
    "ink_950":   "#1A1F2E",
    "ink_800":   "#3A4055",
    "ink_700":   "#52596B",
    "ink_600":   "#6D7282",
    "ink_500":   "#8B8F9C",
    "ink_400":   "#B3B7C0",
    "ink_300":   "#D6D8DE",
    "ink_200":   "#E5E7EB",
    "ink_100":   "#F2F3F5",
    "ink_50":    "#FAFAFB",
    "accent":    "#3B5BDB",
    "accent_2":  "#5C7CFA",
    "accent_soft": "#E7ECFB",
    "success":   "#22C55E",
    "success_soft": "#E6F7EE",
    "info":      "#3B82F6",
    "info_soft": "#E5F0FE",
    "warn":      "#F59E0B",
    "warn_soft": "#FEF3DC",
    "danger":    "#EF4444",
    "danger_soft": "#FCE8E8",
}


_STATUT_PILL = {
    "ouvert":    {"label": "Ouvert",    "bg": "#E5F0FE", "fg": "#3B82F6"},
    "en_cours":  {"label": "En cours",  "bg": "#FEF3DC", "fg": "#B45309"},
    "pourvu":    {"label": "Pourvu",    "bg": "#E6F7EE", "fg": "#15803D"},
    "abandonne": {"label": "Abandonné", "bg": "#F2F3F5", "fg": "#6D7282"},
}


_CAND_STATUS_PILL = {
    "dispo": {"label": "Disponible",    "bg": "#E6F7EE", "fg": "#15803D"},
    "msg":   {"label": "Messagerie",    "bg": "#E5F0FE", "fg": "#1E40AF"},
    "":      {"label": "Pas contacté",  "bg": "#F2F3F5", "fg": "#52596B"},
    "nope":  {"label": "Non disponible", "bg": "#FCE8E8", "fg": "#B91C1C"},
}


def _pdf_esc(value) -> str:
    """Échappe les caractères réservés XML/RML pour ReportLab Paragraph."""
    if value is None:
        return ""
    s = str(value)
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
    )


def _fmt_fr_date(value: str) -> str:
    """Convertit '2026-05-11' en '11/05/2026' si possible, sinon renvoie tel quel."""
    if not value:
        return ""
    s = str(value).strip()
    try:
        d = datetime.date.fromisoformat(s[:10])
        return d.strftime("%d/%m/%Y")
    except Exception:
        return s


def _build_besoin_pdf(b: dict, cands: list) -> BytesIO:
    """Génère un PDF A4 portrait de la fiche besoin (entête + mission + candidats)."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        HRFlowable,
        KeepTogether,
        PageTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    C = {k: colors.HexColor(v) for k, v in _PDF_COLORS.items()}

    bio = BytesIO()
    W, H = A4
    margin_x = 1.6 * cm
    margin_top = 2.3 * cm
    margin_bottom = 1.7 * cm

    intitule = (b.get("intitule") or "").strip() or "Besoin sans intitulé"
    client_str = (b.get("client") or b.get("company_name") or "").strip()
    contact_str = (b.get("contact") or b.get("prospect_name") or "").strip()

    styles = getSampleStyleSheet()

    def S(name, parent="Normal", **kw):
        return ParagraphStyle(name, parent=styles[parent], **kw)

    # Styles inspirés du design v30
    s_title = S("title", fontName="Helvetica-Bold", fontSize=20, leading=24,
                textColor=C["ink_950"], spaceAfter=2)
    s_subtitle = S("subtitle", fontName="Helvetica", fontSize=10, leading=13,
                   textColor=C["ink_600"], spaceAfter=12)
    s_section = S("section", fontName="Helvetica-Bold", fontSize=10.5,
                  leading=14, textColor=C["accent"], spaceBefore=10,
                  spaceAfter=4, letterSpacing=0.5)
    s_label = S("label", fontName="Helvetica-Bold", fontSize=7.5, leading=10,
                textColor=C["ink_500"], spaceAfter=1, letterSpacing=0.4)
    s_value = S("value", fontName="Helvetica", fontSize=9.5, leading=13,
                textColor=C["ink_800"], spaceAfter=4)
    s_value_strong = S("valueStrong", fontName="Helvetica-Bold", fontSize=10,
                       leading=14, textColor=C["ink_950"], spaceAfter=4)
    s_body = S("body", fontName="Helvetica", fontSize=9.5, leading=14,
               textColor=C["ink_800"], spaceAfter=4, alignment=TA_LEFT)
    s_muted = S("muted", fontName="Helvetica-Oblique", fontSize=9, leading=12,
                textColor=C["ink_500"], spaceAfter=4)
    s_cand_name = S("candName", fontName="Helvetica-Bold", fontSize=11.5,
                    leading=14, textColor=C["ink_950"])
    s_cand_meta = S("candMeta", fontName="Helvetica", fontSize=8.5, leading=11,
                    textColor=C["ink_600"])
    s_track_label = S("trackLabel", fontName="Helvetica-Bold", fontSize=7,
                      leading=9, textColor=C["ink_500"], letterSpacing=0.5)
    s_track_value = S("trackValue", fontName="Helvetica", fontSize=9,
                      leading=11.5, textColor=C["ink_800"])
    s_pill = S("pill", fontName="Helvetica-Bold", fontSize=8, leading=10,
               textColor=colors.white, alignment=TA_LEFT)

    story: list = []

    # ── HEADER (titre + chip statut + ligne accent) ────────────────────
    statut_pill = _STATUT_PILL.get(
        (b.get("statut") or "").strip(),
        _STATUT_PILL["ouvert"],
    )
    pill_cell = Paragraph(
        f'<font color="{statut_pill["fg"]}"><b>&nbsp;&nbsp;{_pdf_esc(statut_pill["label"]).upper()}&nbsp;&nbsp;</b></font>',
        S("pillInner", fontName="Helvetica-Bold", fontSize=8, leading=11),
    )
    pill_tbl = Table([[pill_cell]], colWidths=[3.3 * cm])
    pill_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(statut_pill["bg"])),
        ("BOX", (0, 0), (-1, -1), 0, colors.HexColor(statut_pill["bg"])),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
    ]))

    header_left = [
        Paragraph(f'<font color="{_PDF_COLORS["ink_500"]}" size="8.5"><b>FICHE BESOIN — PROSP\'UP</b></font>',
                  S("eyebrow", fontName="Helvetica-Bold", fontSize=8.5,
                    leading=11, textColor=C["ink_500"], letterSpacing=1)),
        Spacer(1, 2),
        Paragraph(_pdf_esc(intitule), s_title),
    ]
    sub_parts = []
    if client_str:
        sub_parts.append(f'<b>{_pdf_esc(client_str)}</b>')
    if contact_str:
        sub_parts.append(_pdf_esc(contact_str))
    if b.get("localisation"):
        sub_parts.append(_pdf_esc(b.get("localisation")))
    if sub_parts:
        header_left.append(Paragraph(" · ".join(sub_parts), s_subtitle))

    header_tbl = Table(
        [[header_left, pill_tbl]],
        colWidths=[W - 2 * margin_x - 3.4 * cm, 3.4 * cm],
    )
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (0, 0), "TOP"),
        ("VALIGN", (1, 0), (1, 0), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_tbl)
    story.append(HRFlowable(width="100%", thickness=2, color=C["accent"],
                            spaceBefore=2, spaceAfter=10))

    # ── BLOC INFOS GÉNÉRALES (key/value grid 2 colonnes) ──────────────
    def kv(label: str, value) -> list:
        return [
            Paragraph(label.upper(), s_label),
            Paragraph(_pdf_esc(value) if value else
                      f'<font color="{_PDF_COLORS["ink_400"]}">—</font>',
                      s_value),
        ]

    infos_rows = []
    line1 = kv("Client", client_str) + kv("Contact", contact_str)
    line2 = kv("Localisation", b.get("localisation")) + \
            kv("Profil recherché", b.get("profil_type"))
    line3 = kv("Date appel", _fmt_fr_date(b.get("date_appel"))) + \
            kv("Date besoin", _fmt_fr_date(b.get("date_besoin")))
    line4 = kv("Durée mission", b.get("duree_mission")) + \
            kv("Lié au prospect", b.get("prospect_name"))

    info_data = [
        [line1[0], line1[1], line1[2], line1[3]],
        [line2[0], line2[1], line2[2], line2[3]],
        [line3[0], line3[1], line3[2], line3[3]],
        [line4[0], line4[1], line4[2], line4[3]],
    ]
    # Tableau 4 colonnes : label | value | label | value
    col_w = (W - 2 * margin_x) / 4
    infos_tbl = Table(info_data,
                      colWidths=[col_w * 0.7, col_w * 1.3, col_w * 0.7, col_w * 1.3])
    infos_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, C["ink_200"]),
    ]))
    story.append(infos_tbl)
    story.append(Spacer(1, 6))

    # ── BLOC MISSION (multi-paragraphes) ──────────────────────────────
    mission_fields = [
        ("Descriptif",            b.get("descriptif")),
        ("Compétences requises",  b.get("competences")),
        ("Connaissances attendues", b.get("connaissances")),
        ("Expérience",            b.get("experience")),
        ("Commentaires",          b.get("commentaires")),
    ]
    has_mission = any((v or "").strip() for _, v in mission_fields)
    if has_mission:
        story.append(Paragraph("MISSION", s_section))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C["ink_200"],
                                spaceBefore=0, spaceAfter=6))
        for label, val in mission_fields:
            if not (val or "").strip():
                continue
            story.append(Paragraph(label.upper(), s_label))
            text = _pdf_esc(val).replace("\n", "<br/>")
            story.append(Paragraph(text, s_body))
            story.append(Spacer(1, 2))
        story.append(Spacer(1, 4))

    # ── BLOC CANDIDATS POSITIONNÉS ────────────────────────────────────
    cand_list = [c for c in (cands or []) if isinstance(c, dict)]
    story.append(Paragraph(
        f"CANDIDATS POSITIONNÉS ({len(cand_list)})", s_section))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C["ink_200"],
                            spaceBefore=0, spaceAfter=6))

    if not cand_list:
        story.append(Paragraph(
            "Aucun candidat positionné sur ce besoin.",
            s_muted))
    else:
        for idx, c in enumerate(cand_list, start=1):
            status_key = (c.get("cand_status") or "").strip().lower()
            if status_key not in _CAND_STATUS_PILL:
                status_key = ""
            spill = _CAND_STATUS_PILL[status_key]

            name = (c.get("candidat") or "").strip() or f"Candidat #{idx}"
            ref = c.get("_ref") or {}

            # Ligne 1 — numéro + nom + chip statut
            num_para = Paragraph(
                f'<font color="{_PDF_COLORS["ink_400"]}" size="9">'
                f'<b>#{idx:02d}</b></font>',
                S("num", fontName="Helvetica-Bold", fontSize=9, leading=12,
                  textColor=C["ink_400"]),
            )
            name_para = Paragraph(_pdf_esc(name), s_cand_name)
            chip_para = Paragraph(
                f'<font color="{spill["fg"]}"><b>'
                f'&nbsp;&nbsp;{_pdf_esc(spill["label"]).upper()}&nbsp;&nbsp;</b></font>',
                S("chipInner", fontName="Helvetica-Bold", fontSize=7.5,
                  leading=10),
            )
            chip_tbl = Table([[chip_para]], colWidths=[3.0 * cm])
            chip_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(spill["bg"])),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            # Sous-titre meta (rôle, localisation, séniorité, tech)
            meta_parts = []
            if ref.get("role"):       meta_parts.append(_pdf_esc(ref.get("role")))
            if ref.get("seniority"):  meta_parts.append(_pdf_esc(ref.get("seniority")))
            if ref.get("location"):   meta_parts.append(_pdf_esc(ref.get("location")))
            if c.get("lieu_habitation") and not ref.get("location"):
                meta_parts.append(_pdf_esc(c.get("lieu_habitation")))
            if c.get("diplome"):      meta_parts.append(_pdf_esc(c.get("diplome")))
            meta_str = " · ".join(meta_parts)

            # Contact (email/phone) — depuis _ref ou champs libres
            contact_parts = []
            phone = (ref.get("phone") or c.get("phone") or "").strip()
            email = (ref.get("email") or "").strip()
            if phone:
                contact_parts.append(
                    f'<font color="{_PDF_COLORS["ink_500"]}"><b>Tél.</b></font> '
                    f'{_pdf_esc(phone)}')
            if email:
                contact_parts.append(
                    f'<font color="{_PDF_COLORS["ink_500"]}"><b>Email</b></font> '
                    f'{_pdf_esc(email)}')
            profile_url = (ref.get("vsa_url") or c.get("profile_url") or "").strip()
            if profile_url:
                display = profile_url if len(profile_url) <= 60 else profile_url[:57] + "…"
                contact_parts.append(
                    f'<font color="{_PDF_COLORS["ink_500"]}"><b>Profil</b></font> '
                    f'<font color="{_PDF_COLORS["accent"]}"><u>{_pdf_esc(display)}</u></font>')
            contact_str = "  ·  ".join(contact_parts)

            top_inner = [name_para]
            if meta_str:
                top_inner.append(Paragraph(meta_str, s_cand_meta))
            if contact_str:
                top_inner.append(Paragraph(contact_str, s_cand_meta))

            top_row = Table(
                [[num_para, top_inner, chip_tbl]],
                colWidths=[1.0 * cm, W - 2 * margin_x - 1.0 * cm - 3.1 * cm,
                           3.1 * cm],
            )
            top_row.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))

            # Tableau tracking (3 colonnes × 4 lignes max)
            TRACK = [
                ("Dispo",      c.get("dispo")),
                ("Appel",      c.get("appel")),
                ("DT",         c.get("dt")),
                ("RDV 1",      c.get("rdv1")),
                ("RDV 2",      c.get("rdv2")),
                ("RT",         c.get("rt")),
                ("Envoi DT",   c.get("envoi_dt")),
                ("Propal",     c.get("propal")),
                ("RT client",  c.get("rt_client")),
            ]
            track_cells = []
            for label, val in TRACK:
                lbl = Paragraph(label.upper(), s_track_label)
                v = (val or "").strip()
                vpara = Paragraph(
                    _pdf_esc(v) if v else
                    f'<font color="{_PDF_COLORS["ink_400"]}">—</font>',
                    s_track_value,
                )
                track_cells.append([lbl, vpara])

            # Mise en grille 3 colonnes
            grid_rows = []
            for i in range(0, len(track_cells), 3):
                chunk = track_cells[i:i + 3]
                while len(chunk) < 3:
                    chunk.append([Paragraph("", s_track_label),
                                  Paragraph("", s_track_value)])
                # chaque "cellule" est un mini-Table label+value
                cells = []
                for lbl_p, val_p in chunk:
                    cell_tbl = Table([[lbl_p], [val_p]],
                                     colWidths=[(W - 2 * margin_x - 1.0 * cm) / 3 - 8])
                    cell_tbl.setStyle(TableStyle([
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 1),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                    ]))
                    cells.append(cell_tbl)
                grid_rows.append(cells)

            grid_w = (W - 2 * margin_x - 1.0 * cm) / 3
            tracking_tbl = Table(grid_rows,
                                 colWidths=[grid_w, grid_w, grid_w])
            tracking_tbl.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("BACKGROUND", (0, 0), (-1, -1), C["ink_50"]),
                ("BOX", (0, 0), (-1, -1), 0.5, C["ink_200"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, C["ink_200"]),
            ]))

            # Bloc commentaires
            commentaire = (c.get("commentaires") or "").strip()
            comm_block = None
            if commentaire:
                comm_block = Table(
                    [[Paragraph("COMMENTAIRES", s_label)],
                     [Paragraph(_pdf_esc(commentaire).replace("\n", "<br/>"),
                                s_value)]],
                    colWidths=[W - 2 * margin_x - 1.0 * cm],
                )
                comm_block.setStyle(TableStyle([
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (0, 0), 6),
                    ("BOTTOMPADDING", (0, 0), (0, 0), 1),
                    ("TOPPADDING", (0, 1), (0, 1), 0),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                    ("BACKGROUND", (0, 0), (-1, -1), C["accent_soft"]),
                    ("BOX", (0, 0), (-1, -1), 0.5,
                     colors.HexColor("#C9D4F2")),
                ]))

            # Compose la carte candidat (avec bande gauche colorée selon statut)
            card_inner = [top_row, Spacer(1, 2), tracking_tbl]
            if comm_block is not None:
                card_inner.append(Spacer(1, 4))
                card_inner.append(comm_block)

            card_tbl = Table(
                [[Paragraph("", s_value), card_inner]],
                colWidths=[0.18 * cm, W - 2 * margin_x - 0.18 * cm],
            )
            card_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0),
                 colors.HexColor(spill["fg"])),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 10),
                ("RIGHTPADDING", (1, 0), (1, 0), 10),
                ("TOPPADDING", (1, 0), (1, 0), 8),
                ("BOTTOMPADDING", (1, 0), (1, 0), 10),
                ("BACKGROUND", (1, 0), (1, 0), colors.white),
                ("LINEABOVE", (1, 0), (1, 0), 0.5, C["ink_200"]),
                ("LINEBELOW", (1, 0), (1, 0), 0.5, C["ink_200"]),
                ("LINEAFTER", (1, 0), (1, 0), 0.5, C["ink_200"]),
            ]))
            story.append(KeepTogether(card_tbl))
            story.append(Spacer(1, 8))

    # ── PRÉPARATION RT ────────────────────────────────────────────────
    prep = (b.get("preparation_rt") or "").strip()
    if prep:
        story.append(Spacer(1, 6))
        story.append(Paragraph("PRÉPARATION AVANT LA RT", s_section))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C["ink_200"],
                                spaceBefore=0, spaceAfter=6))
        story.append(Paragraph(_pdf_esc(prep).replace("\n", "<br/>"), s_body))

    # ── PAGE TEMPLATE + FOOTER ────────────────────────────────────────
    def on_page(canv, _doc):
        canv.saveState()
        # Bandeau haut (accent fin)
        canv.setFillColor(C["accent"])
        canv.rect(0, H - 0.45 * cm, W, 0.45 * cm, stroke=0, fill=1)
        # Eyebrow ProspUp
        canv.setFillColor(C["ink_500"])
        canv.setFont("Helvetica-Bold", 7.5)
        canv.drawString(margin_x, H - 0.95 * cm,
                        "PROSP'UP · TRAITEMENT BESOIN")
        # Footer : date + page
        now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        canv.setFillColor(C["ink_500"])
        canv.setFont("Helvetica", 8)
        canv.drawString(margin_x, 1.0 * cm,
                        f"Export généré le {now_str}")
        canv.drawRightString(W - margin_x, 1.0 * cm,
                             f"Page {canv.getPageNumber()}")
        canv.setStrokeColor(C["ink_200"])
        canv.setLineWidth(0.4)
        canv.line(margin_x, 1.35 * cm, W - margin_x, 1.35 * cm)
        canv.restoreState()

    doc = BaseDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=margin_x,
        rightMargin=margin_x,
        topMargin=margin_top,
        bottomMargin=margin_bottom,
        title=f"Besoin — {intitule}",
        author="Prosp'Up",
    )
    frame = Frame(margin_x, margin_bottom, W - 2 * margin_x,
                  H - margin_top - margin_bottom, id="main",
                  leftPadding=0, rightPadding=0, topPadding=0,
                  bottomPadding=0)
    doc.addPageTemplates([
        PageTemplate(id="default", frames=[frame], onPage=on_page),
    ])
    doc.build(story)
    bio.seek(0)
    return bio


@besoins_bp.get("/api/besoins/<int:bid>/export.pdf")
def api_export_besoin_pdf(bid: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT b.*, p.name AS prospect_name, c.groupe AS company_name "
            "FROM besoins b "
            "LEFT JOIN prospects p ON p.id = b.prospect_id "
            "LEFT JOIN companies c ON c.id = b.company_id "
            "WHERE b.id=? AND b.owner_id=? AND (b.deleted_at IS NULL OR b.deleted_at='') LIMIT 1;",
            (bid, uid),
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Introuvable"), 404

    b = _row_to_dict(row, uid=uid)
    cands = b.get("candidats") or []

    if not b.get("client") and b.get("company_name"):
        b["client"] = b["company_name"]
    if not b.get("contact") and b.get("prospect_name"):
        b["contact"] = b["prospect_name"]

    try:
        bio = _build_besoin_pdf(b, cands)
    except Exception as e:
        logger.exception("besoins: PDF export failed (%s)", e)
        return jsonify(ok=False, error=f"Erreur génération PDF : {e}"), 500

    intitule = _safe_filename(b.get("intitule") or b.get("client") or f"besoin_{bid}")
    filename = f"fiche_besoin_{intitule}.pdf"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@besoins_bp.get("/api/besoins/<int:bid>/export.xlsx")
def api_export_besoin_xlsx(bid: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT b.*, p.name AS prospect_name, c.groupe AS company_name "
            "FROM besoins b "
            "LEFT JOIN prospects p ON p.id = b.prospect_id "
            "LEFT JOIN companies c ON c.id = b.company_id "
            "WHERE b.id=? AND b.owner_id=? AND (b.deleted_at IS NULL OR b.deleted_at='') LIMIT 1;",
            (bid, uid),
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Introuvable"), 404

    b = _row_to_dict(row)
    cands = b.get("candidats") or []

    if not b.get("client") and b.get("company_name"):
        b["client"] = b["company_name"]
    if not b.get("contact") and b.get("prospect_name"):
        b["contact"] = b["prospect_name"]

    from openpyxl import Workbook
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws = wb.create_sheet(title="recto verso")
    _build_sheet(ws, b, cands)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    intitule = _safe_filename(b.get("intitule") or b.get("client") or f"besoin_{bid}")
    filename = f"03_traitement_besoin_{intitule}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
