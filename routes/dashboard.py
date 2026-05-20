"""ProspUp — Blueprint Dashboard / Stats / Photos.

Le plus gros bloc de routes : stats KPI (filtre temporel, charts,
export XLSX, predictions IA), insights IA, photos prospect, pipeline
dashboard. Reste à découper plus finement dans une session ultérieure
(stats vs dashboard vs photos)."""
from __future__ import annotations

import csv
import datetime
import io
import json
import os
import re
import shutil
import unicodedata
from io import BytesIO
from pathlib import Path

from flask import Blueprint, Response, jsonify, request, send_file, stream_with_context
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename

from app import _audit_log, log_activity, logger
from config import APP_DIR, DATA_DIR
from services.dashboard_goals import build_goals_payload as _build_goals_payload, get_goals_config as _get_goals_config
from utils.ai_helpers import _call_ai, _load_ai_config
from utils.auth import _get_current_user, _prospect_owned, _uid, login_required, role_required
from utils.common import _now_iso, _today_iso
from utils.db import _conn
from utils.files import _validate_upload

dashboard_bp = Blueprint("dashboard", __name__)


@dashboard_bp.get("/api/stats")
def api_stats():
    # Range modes:
    # - /api/stats?days=30
    # - /api/stats?range=all
    # - /api/stats?start=YYYY-MM-DD&end=YYYY-MM-DD  (inclusive)
    today = datetime.date.today()

    def _parse_iso_date(s: str):
        try:
            return datetime.date.fromisoformat((s or "").strip())
        except Exception:
            return None

    mode = "days"
    start_d = None
    end_d = None

    if (request.args.get("range") or "").strip().lower() == "all":
        mode = "all"
    else:
        start_q = request.args.get("start")
        end_q = request.args.get("end")
        if start_q and end_q:
            s = _parse_iso_date(start_q)
            e = _parse_iso_date(end_q)
            if s and e:
                mode = "custom"
                start_d, end_d = (s, e) if s <= e else (e, s)
        if start_d is None or end_d is None:
            days = request.args.get("days") or "30"
            try:
                days_i = max(1, min(365, int(days)))
            except Exception:
                days_i = 30
            mode = "days"
            end_d = today
            start_d = today - datetime.timedelta(days=days_i - 1)

    start_iso = start_d.isoformat() if start_d else ""
    end_iso = end_d.isoformat() if end_d else ""
    today_iso = _today_iso()
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    with _conn() as conn:
        # BUG 27 : total = actifs (non supprimés, non archivés) pour cohérence avec /v30/prospects
        total_prospects = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? "
            "AND (deleted_at IS NULL OR deleted_at='') "
            "AND (is_archived IS NULL OR is_archived=0);",
            (uid,),
        ).fetchone()["n"]
        total_companies = conn.execute(
            "SELECT COUNT(*) AS n FROM companies WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='');",
            (uid,),
        ).fetchone()["n"]

        # status counts (all time) — prospects de l'utilisateur uniquement
        rdv_total = conn.execute("SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND statut='Rendez-vous';", (uid,)).fetchone()["n"]
        recall_total = conn.execute("SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND statut='À rappeler';", (uid,)).fetchone()["n"]

        # RDV obtenus DANS LA PÉRIODE sélectionnée (≠ rdv_total, qui est le
        # pipeline tout-temps). Source : events rdv_taken datés dans la plage
        # + fallback lastContact pour les RDV pris avant l'instrumentation des
        # events. Même logique que le graphe « RDV obtenus » (rdvPerMonth).
        if mode == "all":
            rdv_obtained = conn.execute(
                """SELECT COUNT(DISTINCT pid) AS n FROM (
                     SELECT e.prospect_id AS pid FROM prospect_events e
                     JOIN prospects p ON p.id=e.prospect_id
                     WHERE p.owner_id=? AND e.type='rdv_taken'
                       AND (p.deleted_at IS NULL OR p.deleted_at='')
                     UNION
                     SELECT p.id AS pid FROM prospects p
                     WHERE p.owner_id=? AND p.statut='Rendez-vous'
                       AND (p.deleted_at IS NULL OR p.deleted_at='')
                       AND NOT EXISTS (SELECT 1 FROM prospect_events e2
                                       WHERE e2.prospect_id=p.id AND e2.type='rdv_taken')
                   );""",
                (uid, uid),
            ).fetchone()["n"]
        else:
            rdv_obtained = conn.execute(
                """SELECT COUNT(DISTINCT pid) AS n FROM (
                     SELECT e.prospect_id AS pid FROM prospect_events e
                     JOIN prospects p ON p.id=e.prospect_id
                     WHERE p.owner_id=? AND e.type='rdv_taken'
                       AND substr(e.date,1,10)>=? AND substr(e.date,1,10)<=?
                       AND (p.deleted_at IS NULL OR p.deleted_at='')
                     UNION
                     SELECT p.id AS pid FROM prospects p
                     WHERE p.owner_id=? AND p.statut='Rendez-vous'
                       AND (p.deleted_at IS NULL OR p.deleted_at='')
                       AND substr(p.lastContact,1,10)>=? AND substr(p.lastContact,1,10)<=?
                       AND NOT EXISTS (SELECT 1 FROM prospect_events e2
                                       WHERE e2.prospect_id=p.id AND e2.type='rdv_taken')
                   );""",
                (uid, start_iso, end_iso, uid, start_iso, end_iso),
            ).fetchone()["n"]

        # followups (always relative to today)
        late = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND nextFollowUp IS NOT NULL AND nextFollowUp != '' AND nextFollowUp < ?;",
            (uid, today_iso),
        ).fetchone()["n"]
        due_today = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND nextFollowUp = ?;",
            (uid, today_iso),
        ).fetchone()["n"]

        # activity (in selected range) — push_logs des prospects de l'utilisateur uniquement
        if mode == "all":
            pushes = conn.execute(
                "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=?;",
                (uid,),
            ).fetchone()["n"]
        else:
            pushes = conn.execute(
                "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=? WHERE substr(l.sentAt,1,10) >= ? AND substr(l.sentAt,1,10) <= ?;",
                (uid, start_iso, end_iso),
            ).fetchone()["n"]

        call_rows = conn.execute(
            "SELECT callNotes FROM prospects WHERE owner_id=? AND callNotes IS NOT NULL AND callNotes != '' AND (deleted_at IS NULL OR deleted_at = '');",
            (uid,),
        ).fetchall()
        call_notes = 0
        for r in call_rows:
            try:
                notes = json.loads(r["callNotes"] or "[]")
                if isinstance(notes, list):
                    for n in notes:
                        d = (n.get("date") if isinstance(n, dict) else "") or ""
                        d = d[:10]
                        if not d:
                            continue
                        if mode == "all":
                            call_notes += 1
                        else:
                            if start_iso <= d <= end_iso:
                                call_notes += 1
            except Exception:
                continue

        # Notes stockées dans prospect_events (mpAddNote, prospect_detail "+ Note", etc.)
        try:
            if mode == "all":
                call_notes += conn.execute(
                    """SELECT COUNT(*) AS n FROM prospect_events e
                       JOIN prospects p ON p.id=e.prospect_id
                       WHERE p.owner_id=? AND e.type IN ('note','note_libre','call_note')
                         AND (p.deleted_at IS NULL OR p.deleted_at='');""",
                    (uid,),
                ).fetchone()["n"]
            else:
                call_notes += conn.execute(
                    """SELECT COUNT(*) AS n FROM prospect_events e
                       JOIN prospects p ON p.id=e.prospect_id
                       WHERE p.owner_id=? AND e.type IN ('note','note_libre','call_note')
                         AND substr(e.date,1,10) >= ? AND substr(e.date,1,10) <= ?
                         AND (p.deleted_at IS NULL OR p.deleted_at='');""",
                    (uid, start_iso, end_iso),
                ).fetchone()["n"]
        except Exception:
            pass

        # Appels tracés (call_logs — clics bouton Appeler)
        try:
            if mode == "all":
                calls_count = conn.execute(
                    "SELECT COUNT(*) AS n FROM call_logs WHERE owner_id=?;", (uid,)
                ).fetchone()["n"]
            else:
                calls_count = conn.execute(
                    "SELECT COUNT(*) AS n FROM call_logs WHERE owner_id=? AND date>=? AND date<=?;",
                    (uid, start_iso, end_iso),
                ).fetchone()["n"]
        except Exception:
            calls_count = 0

        # Hot companies scoring (range for pushes, but late followups are always "today")
        hot = []
        if mode == "all":
            push_range_cond = "1=1"
            push_params = ()
        else:
            # Same robustness for hot companies scoring
            push_range_cond = "substr(l.sentAt,1,10) >= ? AND substr(l.sentAt,1,10) <= ?"
            push_params = (start_iso, end_iso)

        rows = conn.execute(
            f'''
            SELECT c.id, c.groupe, c.site,
                   COUNT(p.id) AS prospect_count,
                   SUM(CASE WHEN p.statut='Rendez-vous' THEN 1 ELSE 0 END) AS rdv_count,
                   SUM(CASE WHEN p.nextFollowUp IS NOT NULL AND p.nextFollowUp != '' AND p.nextFollowUp < ? THEN 1 ELSE 0 END) AS overdue_count,
                   (
                     SELECT COUNT(*)
                     FROM push_logs l
                     JOIN prospects p2 ON p2.id=l.prospect_id AND p2.owner_id=?
                     WHERE p2.company_id=c.id AND {push_range_cond}
                   ) AS pushes_recent
            FROM companies c
            LEFT JOIN prospects p ON p.company_id=c.id AND p.owner_id=? AND (p.deleted_at IS NULL OR p.deleted_at='')
            WHERE c.owner_id=? AND (c.deleted_at IS NULL OR c.deleted_at='')
            GROUP BY c.id
            ORDER BY (rdv_count*5 + overdue_count*3 + pushes_recent*2) DESC
            LIMIT 10;
            ''',
            (today_iso, uid, *push_params, uid, uid),
        ).fetchall()
        for r in rows:
            score = int((r["rdv_count"] or 0) * 5 + (r["overdue_count"] or 0) * 3 + (r["pushes_recent"] or 0) * 2)
            hot.append(
                {
                    "company_id": r["id"],
                    "groupe": r["groupe"],
                    "site": r["site"],
                    "score": score,
                    "prospectCount": r["prospect_count"] or 0,
                    "rdvCount": r["rdv_count"] or 0,
                    "lateFollowups": r["overdue_count"] or 0,
                }
            )

    payload = {
        "ok": True,
        "range": {"mode": mode, "from": start_iso if mode != "all" else "", "to": end_iso if mode != "all" else ""},
        "totals": {"prospects": total_prospects, "companies": total_companies},
        "activity": {"pushes": pushes, "callNotes": call_notes, "calls": calls_count},
        "followups": {"late": late, "dueToday": due_today},
        "statusCounts": {"Rendezvous": rdv_total, "A_rappeler": recall_total},
        "hotCompanies": hot,
        # legacy fields (compat)
        "total_prospects": total_prospects,
        # RDV obtenus dans la période sélectionnée (et non le pipeline
        # tout-temps) — la page Stats affiche cette valeur sous « RDV obtenus ».
        "rdv": rdv_obtained,
        "pushes": pushes,
        "calls": call_notes,
        "overdue": late,
        "hot_companies": hot,
    }
    return jsonify(payload)


@dashboard_bp.post("/api/stats/insights")
def api_stats_insights():
    """Génère des insights IA à partir des statistiques actuelles et historiques."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    # Récupérer les paramètres de période (même logique que /api/stats)
    today = datetime.date.today()
    
    def _parse_iso_date(s: str):
        try:
            return datetime.date.fromisoformat((s or "").strip())
        except Exception:
            return None
    
    req_data = request.json if request.is_json else request.form
    mode = req_data.get("mode", "days")
    start_d = None
    end_d = None
    
    if mode == "all":
        start_d = None
        end_d = today
        prev_start_d = None
        prev_end_d = None
    elif mode == "custom":
        start_q = req_data.get("start")
        end_q = req_data.get("end")
        if start_q and end_q:
            s = _parse_iso_date(start_q)
            e = _parse_iso_date(end_q)
            if s and e:
                start_d, end_d = (s, e) if s <= e else (e, s)
        if start_d is None or end_d is None:
            # Fallback sur 30 jours
            end_d = today
            start_d = today - datetime.timedelta(days=29)
    else:
        # mode == "days"
        days = req_data.get("days", 30)
        try:
            days_i = max(1, min(365, int(days)))
        except Exception:
            days_i = 30
        end_d = today
        start_d = today - datetime.timedelta(days=days_i - 1)
    
    # Période précédente pour comparaison (si pas "all")
    if mode != "all" and start_d and end_d:
        period_days = (end_d - start_d).days + 1
        prev_end_d = start_d - datetime.timedelta(days=1)
        prev_start_d = prev_end_d - datetime.timedelta(days=period_days - 1)
    else:
        prev_start_d = None
        prev_end_d = None
    
    start_iso = start_d.isoformat() if start_d else ""
    end_iso = end_d.isoformat() if end_d else ""
    prev_start_iso = prev_start_d.isoformat() if prev_start_d else ""
    prev_end_iso = prev_end_d.isoformat() if prev_end_d else ""
    today_iso = _today_iso()

    with _conn() as conn:
        # Stats actuelles
        current_stats = {
            "totals": {
                "prospects": conn.execute("SELECT COUNT(*) AS n FROM prospects WHERE owner_id=?;", (uid,)).fetchone()["n"],
                "companies": conn.execute("SELECT COUNT(*) AS n FROM companies WHERE owner_id=?;", (uid,)).fetchone()["n"],
            },
            "activity": {},
            "followups": {},
            "statusCounts": {},
            "hotCompanies": [],
        }
        
        # Activity (période actuelle)
        if mode == "all":
            current_stats["activity"]["pushes"] = conn.execute(
                "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=?;",
                (uid,),
            ).fetchone()["n"]
        else:
            current_stats["activity"]["pushes"] = conn.execute(
                "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=? WHERE substr(l.sentAt,1,10) >= ? AND substr(l.sentAt,1,10) <= ?;",
                (uid, start_iso, end_iso),
            ).fetchone()["n"]
        
        # Call notes (période actuelle) — callNotes JSON + prospect_events de type note
        call_rows = conn.execute(
            "SELECT callNotes FROM prospects WHERE owner_id=? AND callNotes IS NOT NULL AND callNotes != '' AND (deleted_at IS NULL OR deleted_at = '');",
            (uid,),
        ).fetchall()
        call_notes = 0
        for r in call_rows:
            try:
                notes = json.loads(r["callNotes"] or "[]")
                if isinstance(notes, list):
                    for n in notes:
                        d = (n.get("date") if isinstance(n, dict) else "") or ""
                        d = d[:10]
                        if not d:
                            continue
                        if mode == "all":
                            call_notes += 1
                        else:
                            if start_iso <= d <= end_iso:
                                call_notes += 1
            except Exception:
                continue
        try:
            event_note_rows = conn.execute(
                """SELECT substr(e.date,1,10) AS d FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id
                   WHERE p.owner_id=? AND e.type IN ('note','note_libre','call_note')
                     AND (p.deleted_at IS NULL OR p.deleted_at='');""",
                (uid,),
            ).fetchall()
        except Exception:
            event_note_rows = []
        for r in event_note_rows:
            d = r["d"] or ""
            if not d:
                continue
            if mode == "all" or (start_iso <= d <= end_iso):
                call_notes += 1
        current_stats["activity"]["callNotes"] = call_notes
        
        # Followups
        current_stats["followups"]["late"] = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND nextFollowUp IS NOT NULL AND nextFollowUp != '' AND nextFollowUp < ?;",
            (uid, today_iso),
        ).fetchone()["n"]
        current_stats["followups"]["dueToday"] = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND nextFollowUp = ?;",
            (uid, today_iso),
        ).fetchone()["n"]
        
        # Status counts
        current_stats["statusCounts"]["Rendezvous"] = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND statut='Rendez-vous';", (uid,)
        ).fetchone()["n"]
        current_stats["statusCounts"]["A_rappeler"] = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND statut='À rappeler';", (uid,)
        ).fetchone()["n"]
        
        # Hot companies (top 5)
        hot_rows = conn.execute(
            '''
            SELECT c.id, c.groupe, c.site,
                   COUNT(p.id) AS prospect_count,
                   SUM(CASE WHEN p.statut='Rendez-vous' THEN 1 ELSE 0 END) AS rdv_count,
                   SUM(CASE WHEN p.nextFollowUp IS NOT NULL AND p.nextFollowUp != '' AND p.nextFollowUp < ? THEN 1 ELSE 0 END) AS overdue_count
            FROM companies c
            LEFT JOIN prospects p ON p.company_id=c.id AND p.owner_id=?
            WHERE c.owner_id=?
            GROUP BY c.id
            ORDER BY (rdv_count*5 + overdue_count*3) DESC
            LIMIT 5;
            ''',
            (today_iso, uid, uid),
        ).fetchall()
        current_stats["hotCompanies"] = [
            {
                "groupe": r["groupe"],
                "site": r["site"],
                "prospectCount": r["prospect_count"] or 0,
                "rdvCount": r["rdv_count"] or 0,
                "lateFollowups": r["overdue_count"] or 0,
            }
            for r in hot_rows
        ]
        
        # Stats période précédente (pour comparaison)
        prev_stats = {}
        if prev_start_d and prev_end_d:
            prev_stats["activity"] = {}
            if mode != "all":
                prev_stats["activity"]["pushes"] = conn.execute(
                    "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=? WHERE substr(l.sentAt,1,10) >= ? AND substr(l.sentAt,1,10) <= ?;",
                    (uid, prev_start_iso, prev_end_iso),
                ).fetchone()["n"]
                
                prev_call_notes = 0
                for r in call_rows:
                    try:
                        notes = json.loads(r["callNotes"] or "[]")
                        if isinstance(notes, list):
                            for n in notes:
                                d = (n.get("date") if isinstance(n, dict) else "") or ""
                                d = d[:10]
                                if prev_start_iso <= d <= prev_end_iso:
                                    prev_call_notes += 1
                    except Exception:
                        continue
                for r in event_note_rows:
                    d = r["d"] or ""
                    if d and prev_start_iso <= d <= prev_end_iso:
                        prev_call_notes += 1
                prev_stats["activity"]["callNotes"] = prev_call_notes
                
                prev_stats["statusCounts"] = {}
                prev_stats["statusCounts"]["Rendezvous"] = conn.execute(
                    "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND statut='Rendez-vous' AND lastContact >= ? AND lastContact <= ?;",
                    (uid, prev_start_iso, prev_end_iso),
                ).fetchone()["n"]
        
        # Calcul du taux de conversion
        conversion_rate = 0
        if current_stats["totals"]["prospects"] > 0:
            conversion_rate = round((current_stats["statusCounts"]["Rendezvous"] / current_stats["totals"]["prospects"]) * 100, 1)
        
        # Construction du prompt pour Ollama
        prompt = f"""Tu es un analyste expert en prospection B2B. Analyse les statistiques suivantes et génère des insights structurés en JSON.

STATISTIQUES ACTUELLES (période: {start_iso} → {end_iso if end_iso else 'all time'}):
- Total prospects: {current_stats["totals"]["prospects"]}
- Total entreprises: {current_stats["totals"]["companies"]}
- Push envoyés: {current_stats["activity"]["pushes"]}
- Notes d'appel: {current_stats["activity"]["callNotes"]}
- Relances en retard: {current_stats["followups"]["late"]}
- Relances aujourd'hui: {current_stats["followups"]["dueToday"]}
- Prospects en RDV: {current_stats["statusCounts"]["Rendezvous"]}
- Prospects à rappeler: {current_stats["statusCounts"]["A_rappeler"]}
- Taux de conversion (RDV/Total): {conversion_rate}%
- Top entreprises chaudes: {json.dumps(current_stats["hotCompanies"], ensure_ascii=False)}"""

        if prev_stats:
            prompt += f"""

STATISTIQUES PÉRIODE PRÉCÉDENTE (période: {prev_start_iso} → {prev_end_iso}):
- Push envoyés: {prev_stats.get("activity", {}).get("pushes", 0)}
- Notes d'appel: {prev_stats.get("activity", {}).get("callNotes", 0)}
- Prospects en RDV: {prev_stats.get("statusCounts", {}).get("Rendezvous", 0)}"""

        prompt += """

ANALYSE À EFFECTUER:
1. Résumé automatique: Décris l'évolution du pipeline en 2-3 phrases (augmentation/diminution, points forts).
2. Points d'attention: Liste 2-4 alertes concrètes (ex: "3 prospects n'ont pas été contactés depuis 30 jours", "Relances en retard à traiter").
3. Suggestions stratégiques: Propose 2-3 recommandations actionnables basées sur les données (ex: "Les prospects du secteur X convertissent mieux", "Augmenter la fréquence de relance").
4. Benchmarking: Compare avec la période précédente si disponible, sinon avec les meilleures pratiques.

RÉPONSE ATTENDUE (JSON strict, pas de markdown):
{
  "summary": "Résumé en 2-3 phrases",
  "alerts": ["Alerte 1", "Alerte 2"],
  "recommendations": ["Recommandation 1", "Recommandation 2"],
  "benchmarks": {"current": X, "best": Y, "period": "description"}
}

IMPORTANT: Réponds UNIQUEMENT avec le JSON, sans texte avant ou après."""

        try:
            # Appel à l'IA
            ai_response = _call_ai(prompt, timeout=120)
            
            # Nettoyage de la réponse (enlever markdown si présent)
            ai_response = ai_response.strip()
            if ai_response.startswith("```json"):
                ai_response = ai_response[7:]
            if ai_response.startswith("```"):
                ai_response = ai_response[3:]
            if ai_response.endswith("```"):
                ai_response = ai_response[:-3]
            ai_response = ai_response.strip()
            
            # Parse JSON
            insights = json.loads(ai_response)
            
            # Validation de la structure
            if not isinstance(insights, dict):
                raise ValueError("Réponse IA n'est pas un objet JSON")
            
            # Structure par défaut si champs manquants
            result = {
                "summary": insights.get("summary", "Analyse en cours..."),
                "alerts": insights.get("alerts", []),
                "recommendations": insights.get("recommendations", []),
                "benchmarks": insights.get("benchmarks", {}),
            }
            
            return jsonify({"ok": True, "insights": result})
            
        except json.JSONDecodeError as e:
            logger.error("Erreur parsing JSON insights IA: %s", e)
            logger.error("Réponse brute: %s", ai_response[:500])
            return jsonify({
                "ok": False,
                "error": "Erreur parsing réponse IA",
                "raw": ai_response[:500] if 'ai_response' in locals() else "",
            }), 500
        except Exception as e:
            logger.error("Erreur génération insights: %s", e)
            return jsonify({"ok": False, "error": str(e)}), 500


@dashboard_bp.get("/api/stats/predictions")
def api_stats_predictions():
    """Génère des prédictions IA basées sur les statistiques historiques (tendances futures, conversions prévues)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    today = datetime.date.today()
    today_iso = _today_iso()
    
    # Récupérer les statistiques historiques (12 dernières semaines)
    weeks_data = []
    with _conn() as conn:
        for i in range(12, 0, -1):
            week_end = today - datetime.timedelta(days=(i - 1) * 7)
            week_start = week_end - datetime.timedelta(days=6)
            week_start_iso = week_start.isoformat()
            week_end_iso = week_end.isoformat()
            
            # Compter les actions de cette semaine
            pushes = conn.execute(
                "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id = l.prospect_id AND p.owner_id=? WHERE substr(l.sentAt,1,10) >= ? AND substr(l.sentAt,1,10) <= ?;",
                (uid, week_start_iso, week_end_iso),
            ).fetchone()["n"]
            
            # Compter les notes d'appel
            call_notes_count = 0
            prospects_rows = conn.execute(
                "SELECT callNotes FROM prospects WHERE owner_id=? AND callNotes IS NOT NULL AND callNotes != '';",
                (uid,),
            ).fetchall()
            for r in prospects_rows:
                try:
                    notes = json.loads(r["callNotes"] or "[]")
                    if isinstance(notes, list):
                        for n in notes:
                            note_date = (n.get("date") or "")[:10]
                            if note_date >= week_start_iso and note_date <= week_end_iso:
                                call_notes_count += 1
                except Exception:
                    pass
            
            # Compter les RDV
            rdv_count = conn.execute(
                "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND rdvDate IS NOT NULL AND rdvDate != '' AND substr(rdvDate,1,10) >= ? AND substr(rdvDate,1,10) <= ?;",
                (uid, week_start_iso, week_end_iso),
            ).fetchone()["n"]
            
            weeks_data.append({
                "week": f"S{week_end.isocalendar()[1]}",
                "pushes": pushes,
                "call_notes": call_notes_count,
                "rdv": rdv_count,
            })
    
    # Récupérer les totaux actuels (BUG 27 : exclure aussi les archivés)
    with _conn() as conn:
        total_prospects = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND deleted_at IS NULL "
            "AND (is_archived IS NULL OR is_archived=0);",
            (uid,),
        ).fetchone()["n"]
        total_companies = conn.execute("SELECT COUNT(*) AS n FROM companies WHERE owner_id=? AND deleted_at IS NULL;", (uid,)).fetchone()["n"]
        rdv_prospects = conn.execute("SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND statut='Rendez-vous' AND deleted_at IS NULL;", (uid,)).fetchone()["n"]
        overdue_count = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND nextFollowUp IS NOT NULL AND nextFollowUp != '' AND nextFollowUp < ? AND deleted_at IS NULL;",
            (uid, today_iso),
        ).fetchone()["n"]

    # Early return : sans données, l'IA ne peut rien prédire d'utile et l'appel
    # prend 60-120 s pour rien. On répond directement avec un état "no_data".
    has_activity = total_prospects > 0 or any(
        w.get("pushes", 0) + w.get("call_notes", 0) + w.get("rdv", 0) > 0 for w in weeks_data
    )
    if not has_activity:
        return jsonify({
            "ok": True,
            "predictions": {
                "trends": {"pushes": "stabilité", "call_notes": "stabilité", "rdv": "stabilité"},
                "conversion_rate": {
                    "current": 0,
                    "predicted": 0,
                    "explanation": "Pas encore d'historique — ajoutez des prospects et des actions pour générer des prédictions.",
                },
                "recommendations": [
                    "Importez ou ajoutez vos premiers prospects.",
                    "Loggez vos premiers appels et envois push pour alimenter le modèle.",
                ],
                "forecast": {f"week_{i}": {"pushes": 0, "call_notes": 0, "rdv": 0} for i in range(1, 5)},
            },
            "no_data": True,
        })

    # Construire le prompt pour les prédictions
    prompt = f"""Tu es un assistant pour un CRM de prospection B2B. Analyse les données historiques et génère des prédictions pour les 4 prochaines semaines.

DONNÉES HISTORIQUES (12 dernières semaines):
{json.dumps(weeks_data, indent=2, ensure_ascii=False)}

SITUATION ACTUELLE:
- Total prospects: {total_prospects}
- Total entreprises: {total_companies}
- Prospects en RDV: {rdv_prospects}
- Relances en retard: {overdue_count}

PRÉDICTIONS À GÉNÉRER:
1. "trends": Tendances prévues pour les 4 prochaines semaines (pushes, notes, RDV)
2. "conversion_rate": Taux de conversion prévu (prospects → RDV)
3. "recommendations": 2-3 recommandations pour optimiser les résultats
4. "forecast": Prévisions chiffrées pour les 4 prochaines semaines

RÉPONSE ATTENDUE (JSON strict, pas de markdown):
{{
  "trends": {{
    "pushes": "tendance (augmentation/diminution/stabilité)",
    "call_notes": "tendance",
    "rdv": "tendance"
  }},
  "conversion_rate": {{
    "current": X,
    "predicted": Y,
    "explanation": "explication courte"
  }},
  "recommendations": ["Recommandation 1", "Recommandation 2"],
  "forecast": {{
    "week_1": {{"pushes": X, "call_notes": Y, "rdv": Z}},
    "week_2": {{"pushes": X, "call_notes": Y, "rdv": Z}},
    "week_3": {{"pushes": X, "call_notes": Y, "rdv": Z}},
    "week_4": {{"pushes": X, "call_notes": Y, "rdv": Z}}
  }}
}}

IMPORTANT: Réponds UNIQUEMENT avec le JSON, sans texte avant ou après."""
    
    try:
        # Appel à l'IA
        ai_response = _call_ai(prompt, timeout=120)
        
        # Nettoyage de la réponse (enlever markdown si présent)
        ai_response = ai_response.strip()
        if ai_response.startswith("```json"):
            ai_response = ai_response[7:]
        if ai_response.startswith("```"):
            ai_response = ai_response[3:]
        if ai_response.endswith("```"):
            ai_response = ai_response[:-3]
        ai_response = ai_response.strip()
        
        # Parse JSON
        predictions = json.loads(ai_response)
        
        # Validation de la structure
        if not isinstance(predictions, dict):
            raise ValueError("Réponse IA n'est pas un objet JSON")
        
        # Structure par défaut si champs manquants
        result = {
            "trends": predictions.get("trends", {}),
            "conversion_rate": predictions.get("conversion_rate", {}),
            "recommendations": predictions.get("recommendations", []),
            "forecast": predictions.get("forecast", {}),
        }
        
        return jsonify({"ok": True, "predictions": result})
        
    except json.JSONDecodeError as e:
        logger.error("Erreur parsing JSON predictions IA: %s", e)
        logger.error("Réponse brute: %s", ai_response[:500] if 'ai_response' in locals() else "")
        return jsonify({
            "ok": False,
            "error": "Erreur parsing réponse IA",
            "raw": ai_response[:500] if 'ai_response' in locals() else "",
        }), 500
    except Exception as e:
        logger.error("Erreur génération predictions: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# ====== Prospect Photo Upload ======
import uuid as _uuid

from config import AVATARS_DIR, PHOTOS_DIR

os.makedirs(PHOTOS_DIR, exist_ok=True)
os.makedirs(AVATARS_DIR, exist_ok=True)

# Migration: déplacer les photos existantes de static/photos/ vers data/photos/
_old_photos_dir = APP_DIR / "static" / "photos"
if _old_photos_dir.exists():
    for _f in _old_photos_dir.iterdir():
        if _f.is_file():
            _dest = PHOTOS_DIR / _f.name
            if not _dest.exists():
                _f.rename(_dest)
    try:
        _old_photos_dir.rmdir()
    except OSError:
        pass

@dashboard_bp.post("/api/prospect/photo")
def api_prospect_photo():
    """Upload a photo for a prospect. Saves to static/photos/ and updates DB."""
    pid = request.form.get("prospect_id")
    if not pid:
        return jsonify({"ok": False, "error": "prospect_id required"}), 400
    pid = int(pid)
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _prospect_owned(pid):
        return jsonify({"ok": False, "error": "Accès refusé"}), 403

    f = request.files.get("photo")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400

    ok_upload, err_upload = _validate_upload(f, "image")
    if not ok_upload:
        return jsonify(ok=False, error=err_upload[0]), err_upload[1]
    ext = os.path.splitext(f.filename)[1].lower()

    # Save with unique name
    fname = f"prospect_{pid}{ext}"
    fpath = os.path.join(PHOTOS_DIR, fname)
    try:
        f.save(fpath)
    except OSError as e:
        logger.error("Photo save failed for prospect %s: %s", pid, e)
        return jsonify({"ok": False, "error": "Erreur sauvegarde fichier"}), 500

    photo_url = f"/api/photos/prospect/{pid}"

    with _conn() as conn:
        conn.execute("UPDATE prospects SET photo_url = ? WHERE id = ? AND owner_id=?;", (photo_url, pid, uid))

    return jsonify({"ok": True, "photo_url": photo_url})

@dashboard_bp.get("/api/photos/prospect/<int:prospect_id>")
def api_prospect_photo_serve(prospect_id):
    """Serve a prospect photo with ownership check (authenticated route)."""
    if not _prospect_owned(prospect_id):
        return jsonify({"error": "Accès non autorisé"}), 403
    _mimetypes = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif"}
    for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        fpath = os.path.join(PHOTOS_DIR, f"prospect_{prospect_id}{ext}")
        if os.path.isfile(fpath):
            return send_file(fpath, mimetype=_mimetypes[ext])
    return jsonify({"error": "Photo non trouvée"}), 404

@dashboard_bp.delete("/api/prospect/photo")
def api_prospect_photo_delete():
    """Remove a prospect's photo."""
    pid = request.args.get("prospect_id") or request.form.get("prospect_id")
    if not pid:
        return jsonify({"ok": False, "error": "prospect_id required"}), 400
    pid = int(pid)
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _prospect_owned(pid):
        return jsonify({"ok": False, "error": "Accès refusé"}), 403

    for _ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        _old_path = PHOTOS_DIR / f"prospect_{pid}{_ext}"
        if _old_path.is_file():
            _old_path.unlink()
            break
    with _conn() as conn:
        conn.execute("UPDATE prospects SET photo_url = NULL WHERE id = ? AND owner_id=?;", (pid, uid))

    return jsonify({"ok": True})


# ====== Stats Charts API ======
@dashboard_bp.get("/api/stats/charts")
def api_stats_charts():
    """Provide aggregated data for Chart.js graphs on the stats page."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    today = datetime.date.today()
    today_iso = _today_iso()

    with _conn() as conn:
        # 1) Status distribution — prospects de l'utilisateur uniquement (hors supprimés)
        status_rows = conn.execute(
            "SELECT statut, COUNT(*) AS n FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') GROUP BY statut ORDER BY n DESC;",
            (uid,),
        ).fetchall()
        status_dist = {r["statut"]: r["n"] for r in status_rows}

        # 2) Push + calls + callNotes per week (last 12 weeks)
        # Pre-load note dates (callNotes JSON + prospect_events type note) pour bucketing rapide
        _cn_dates = []
        for r in conn.execute(
            "SELECT callNotes FROM prospects WHERE owner_id=? AND callNotes IS NOT NULL AND callNotes!='' AND (deleted_at IS NULL OR deleted_at='');",
            (uid,),
        ).fetchall():
            try:
                for n in (json.loads(r["callNotes"] or "[]") or []):
                    ds = (n.get("date") or "")[:10]
                    if ds:
                        _cn_dates.append(ds)
            except Exception:
                pass
        try:
            for r in conn.execute(
                """SELECT e.date FROM prospect_events e
                   JOIN prospects p ON p.id=e.prospect_id
                   WHERE p.owner_id=? AND e.type IN ('note','note_libre','call_note')
                     AND (p.deleted_at IS NULL OR p.deleted_at='');""",
                (uid,),
            ).fetchall():
                ds = (r["date"] or "")[:10]
                if ds:
                    _cn_dates.append(ds)
        except Exception:
            pass

        weeks = []
        activity_weeks = []
        for i in range(11, -1, -1):
            d = today - datetime.timedelta(weeks=i)
            mon = d - datetime.timedelta(days=d.weekday())
            sun = mon + datetime.timedelta(days=6)
            mon_iso, sun_iso = mon.isoformat(), sun.isoformat()
            label = f"S{mon.isocalendar()[1]}"
            push_n = conn.execute(
                "SELECT COUNT(*) AS n FROM push_logs l JOIN prospects p ON p.id=l.prospect_id AND p.owner_id=? WHERE substr(l.sentAt,1,10)>=? AND substr(l.sentAt,1,10)<=?;",
                (uid, mon_iso, sun_iso),
            ).fetchone()["n"]
            try:
                calls_n = conn.execute(
                    "SELECT COUNT(*) AS n FROM call_logs WHERE owner_id=? AND date>=? AND date<=?;",
                    (uid, mon_iso, sun_iso),
                ).fetchone()["n"]
            except Exception:
                calls_n = 0
            notes_n = sum(1 for ds in _cn_dates if mon_iso <= ds <= sun_iso)
            weeks.append({"label": label, "count": push_n})
            activity_weeks.append({"label": label, "calls": calls_n, "callNotes": notes_n, "push": push_n})

        # 3) RDV pris par mois (6 derniers mois) — source primaire : prospect_events rdv_taken
        #    fallback : lastContact des prospects RDV sans événement (rétro-compatibilité)
        months_rdv = []
        for i in range(5, -1, -1):
            first = (today.replace(day=1) - datetime.timedelta(days=i * 28)).replace(day=1)
            if first.month == 12:
                last = first.replace(year=first.year + 1, month=1, day=1) - datetime.timedelta(days=1)
            else:
                last = first.replace(month=first.month + 1, day=1) - datetime.timedelta(days=1)
            count = conn.execute(
                """SELECT COUNT(DISTINCT pid) AS n FROM (
                     SELECT e.prospect_id AS pid
                     FROM prospect_events e
                     JOIN prospects p ON p.id=e.prospect_id
                     WHERE p.owner_id=? AND e.type='rdv_taken'
                       AND substr(e.date,1,10)>=? AND substr(e.date,1,10)<=?
                       AND (p.deleted_at IS NULL OR p.deleted_at='')
                     UNION
                     SELECT p.id AS pid
                     FROM prospects p
                     WHERE p.owner_id=? AND p.statut='Rendez-vous'
                       AND (p.deleted_at IS NULL OR p.deleted_at='')
                       AND p.lastContact>=? AND p.lastContact<=?
                       AND NOT EXISTS (
                         SELECT 1 FROM prospect_events e2
                         WHERE e2.prospect_id=p.id AND e2.type='rdv_taken'
                       )
                   )""",
                (uid, first.isoformat(), last.isoformat(),
                 uid, first.isoformat(), last.isoformat()),
            ).fetchone()["n"]
            months_rdv.append({"label": first.strftime("%b %Y"), "count": count})

        # 4) Top 8 companies by prospect count (prospects de l'utilisateur)
        top_companies = conn.execute(
            """SELECT c.groupe || CASE WHEN c.site IS NOT NULL AND c.site != '' THEN ' (' || c.site || ')' ELSE '' END AS name,
                      COUNT(p.id) AS n
               FROM companies c JOIN prospects p ON p.company_id = c.id AND p.owner_id=?
               GROUP BY c.id ORDER BY n DESC LIMIT 8;""",
            (uid,),
        ).fetchall()
        top_comp = [{"name": r["name"], "count": r["n"]} for r in top_companies]

        # 5) Pertinence distribution
        pert_rows = conn.execute(
            "SELECT pertinence, COUNT(*) AS n FROM prospects WHERE owner_id=? GROUP BY pertinence ORDER BY pertinence DESC;",
            (uid,),
        ).fetchall()
        pert_dist = {str(r["pertinence"]): r["n"] for r in pert_rows}

        # 6) Top consultants pushés (tout l'historique, top 6) — agrège candidate_id1 + candidate_id2
        top_pushed_rows = conn.execute(
            """SELECT ca.id AS cid, ca.name AS cname, COUNT(*) AS n FROM (
                   SELECT l.candidate_id1 AS cid FROM push_logs l
                     JOIN prospects p ON p.id=l.prospect_id AND p.owner_id=?
                     WHERE l.candidate_id1 IS NOT NULL
                   UNION ALL
                   SELECT l.candidate_id2 AS cid FROM push_logs l
                     JOIN prospects p ON p.id=l.prospect_id AND p.owner_id=?
                     WHERE l.candidate_id2 IS NOT NULL
               ) pc
               JOIN candidates ca ON ca.id = pc.cid AND ca.owner_id=?
               GROUP BY ca.id, ca.name
               ORDER BY n DESC LIMIT 6;""",
            (uid, uid, uid),
        ).fetchall()
        top_pushed = [{"name": r["cname"] or f"Candidat {r['cid']}", "count": r["n"]} for r in top_pushed_rows]

        # 7) Urgence des prospects (répartition pour Priorités IA)
        urgent_overdue = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') AND nextAction IS NOT NULL AND nextAction!='' AND nextAction<?;",
            (uid, today_iso),
        ).fetchone()["n"]
        urgent_today = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') AND nextAction=?;",
            (uid, today_iso),
        ).fetchone()["n"]
        week_end = (today + datetime.timedelta(days=7)).isoformat()
        urgent_week = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') AND nextAction>? AND nextAction<=?;",
            (uid, today_iso, week_end),
        ).fetchone()["n"]
        urgent_later = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE owner_id=? AND (deleted_at IS NULL OR deleted_at='') AND nextAction>?;",
            (uid, week_end),
        ).fetchone()["n"]
        urgency_dist = [
            {"label": "En retard", "count": urgent_overdue},
            {"label": "Aujourd'hui", "count": urgent_today},
            {"label": "Cette semaine", "count": urgent_week},
            {"label": "Plus tard", "count": urgent_later},
        ]

        # 8) Top tags / compétences (12 plus fréquents)
        tag_counts: dict[str, int] = {}
        for r in conn.execute(
            "SELECT tags FROM prospects WHERE owner_id=? AND tags IS NOT NULL AND tags!='' AND (deleted_at IS NULL OR deleted_at='');",
            (uid,),
        ).fetchall():
            raw = r["tags"] or ""
            parsed: list[str] = []
            try:
                j = json.loads(raw)
                if isinstance(j, list):
                    parsed = [str(x).strip() for x in j if str(x).strip()]
            except Exception:
                parsed = [t.strip() for t in raw.split(",") if t.strip()]
            for t in parsed:
                tag_counts[t] = tag_counts.get(t, 0) + 1
        top_tags = [
            {"name": k, "count": v}
            for k, v in sorted(tag_counts.items(), key=lambda kv: kv[1], reverse=True)[:12]
        ]

        # 9) Daily activity (56 derniers jours = 8 semaines × 7) pour heatmap
        # On agrège push_logs + call_logs + callNotes par jour.
        start_56 = today - datetime.timedelta(days=55)
        start_56_iso = start_56.isoformat()
        daily_counts: dict[str, int] = {}
        # init avec 0 sur la fenêtre
        for i in range(56):
            d = (start_56 + datetime.timedelta(days=i)).isoformat()
            daily_counts[d] = 0
        # push_logs
        for r in conn.execute(
            """SELECT substr(l.sentAt,1,10) AS d, COUNT(*) AS n
               FROM push_logs l JOIN prospects p ON p.id=l.prospect_id AND p.owner_id=?
               WHERE substr(l.sentAt,1,10)>=?
               GROUP BY substr(l.sentAt,1,10);""",
            (uid, start_56_iso),
        ).fetchall():
            d = r["d"] or ""
            if d in daily_counts:
                daily_counts[d] += r["n"] or 0
        # call_logs
        try:
            for r in conn.execute(
                "SELECT substr(date,1,10) AS d, COUNT(*) AS n FROM call_logs WHERE owner_id=? AND date>=? GROUP BY substr(date,1,10);",
                (uid, start_56_iso),
            ).fetchall():
                d = r["d"] or ""
                if d in daily_counts:
                    daily_counts[d] += r["n"] or 0
        except Exception:
            pass
        # callNotes (réutilise _cn_dates déjà collecté)
        for ds in _cn_dates:
            if ds in daily_counts:
                daily_counts[ds] += 1
        daily_activity = [{"date": d, "count": daily_counts[d]} for d in sorted(daily_counts.keys())]

        # 10) Portfolio per week — taille du portefeuille en fin de chaque semaine (12 sem.)
        # Fallback : prospects créés à `lastContact` (champ disponible). Cumul croissant.
        portfolio_per_week: list[dict] = []
        for i in range(11, -1, -1):
            d = today - datetime.timedelta(weeks=i)
            sun = d - datetime.timedelta(days=d.weekday()) + datetime.timedelta(days=6)
            sun_iso = sun.isoformat()
            label = f"S{(sun - datetime.timedelta(days=6)).isocalendar()[1]}"
            n = conn.execute(
                """SELECT COUNT(*) AS n FROM prospects p
                   WHERE p.owner_id=? AND (p.deleted_at IS NULL OR p.deleted_at='')
                     AND (
                       (p.lastContact IS NOT NULL AND p.lastContact!='' AND p.lastContact<=?)
                       OR p.lastContact IS NULL OR p.lastContact='' )""",
                (uid, sun_iso),
            ).fetchone()["n"]
            portfolio_per_week.append({"label": label, "count": n})

    return jsonify({
        "ok": True,
        "statusDistribution": status_dist,
        "pushPerWeek": weeks,
        "activityPerWeek": activity_weeks,
        "rdvPerMonth": months_rdv,
        "topCompanies": top_comp,
        "pertinenceDistribution": pert_dist,
        "topPushedConsultants": top_pushed,
        "urgencyDistribution": urgency_dist,
        "topTags": top_tags,
        "dailyActivity": daily_activity,
        "portfolioPerWeek": portfolio_per_week,
    })


# ────────────────────────────────────────────────────────────────────
# Export Excel hebdomadaire – suivi activité (v22.1)
# ────────────────────────────────────────────────────────────────────

@dashboard_bp.get("/api/stats/export_weekly_xlsx")
def api_stats_export_weekly_xlsx():
    """Generate an XLSX file following the exact 'Suivi activité' template for a given ISO week.
    Query params:
      - week: ISO week like 2026-W10  (defaults to current week)
      - ollama: 1 to enable Ollama enrichment (normalize métiers, extract besoins, generate codes notes)
    Format: 15 columns (A-O), zipped rows (candidate + prospection + push on same row),
    thick border on column G, goals on first data row (M-N-O).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    # ── Parse week param ──
    week_param = request.args.get("week", "").strip()
    use_ollama = request.args.get("ollama", "").strip() == "1"
    today = datetime.date.today()

    if week_param:
        try:
            year, w = week_param.split("-W")
            year, w = int(year), int(w)
            jan4 = datetime.date(year, 1, 4)
            start_of_w1 = jan4 - datetime.timedelta(days=jan4.isoweekday() - 1)
            monday = start_of_w1 + datetime.timedelta(weeks=w - 1)
        except Exception:
            monday = today - datetime.timedelta(days=today.weekday())
    else:
        monday = today - datetime.timedelta(days=today.weekday())

    sunday = monday + datetime.timedelta(days=6)
    start = monday.isoformat()
    end = sunday.isoformat()
    week_num = monday.isocalendar()[1]
    week_label = f"S{week_num}"

    # ── Helper: Call AI if enabled ──
    def _call_ollama(prompt: str) -> str:
        if not use_ollama:
            return ""
        try:
            return _call_ai(prompt, timeout=OLLAMA_TIMEOUT)
        except Exception:
            return ""

    with _conn() as conn:
        # ── 1) Candidats EC2 (passage à EC2 dans la semaine) ──
        ec2_rows = conn.execute(
            """SELECT DISTINCT ca.id, ca.name, ca.role, ca.sector, ca.seniority, ca.years_experience, ca.status, ca.notes,
                      COALESCE(e.date, substr(ca.updatedAt, 1, 10)) AS ec2_date
               FROM candidates ca
               LEFT JOIN candidate_events e ON e.candidate_id = ca.id AND e.type = 'ec2' AND e.date >= ? AND e.date <= ?
               WHERE ca.owner_id = ? AND ca.status = 'ec2'
               AND (e.date IS NOT NULL OR (substr(ca.updatedAt, 1, 10) >= ? AND substr(ca.updatedAt, 1, 10) <= ?))
               ORDER BY COALESCE(e.date, ca.updatedAt);""",
            (start, end, uid, start, end),
        ).fetchall()
        ec2_list = [dict(r) for r in ec2_rows]

        # ── 2) Candidats EC1 (entretiens de la semaine) ──
        ec1_rows = conn.execute(
            """SELECT ca.id, ca.name, ca.role, ca.sector, ca.seniority, ca.years_experience, ca.status,
                      json_extract(t.payload, '$.interviewAt') AS interviewAt,
                      json_extract(t.payload, '$.data') AS ec1_data,
                      json_extract(t.payload, '$.availability') AS availability,
                      json_extract(t.payload, '$.notes') AS tab_notes
               FROM candidate_tabs t
               JOIN candidates ca ON ca.id = t.candidate_id AND ca.owner_id = ?
               WHERE t.type = 'ec1'
                 AND json_extract(t.payload, '$.interviewAt') IS NOT NULL
                 AND substr(json_extract(t.payload, '$.interviewAt'), 1, 10) >= ?
                 AND substr(json_extract(t.payload, '$.interviewAt'), 1, 10) <= ?
               ORDER BY json_extract(t.payload, '$.interviewAt');""",
            (uid, start, end),
        ).fetchall()
        ec1_list = [dict(r) for r in ec1_rows]

        # ── 3) Candidats Sourcing (ajoutés cette semaine, hors EC1/EC2) ──
        ec1_ec2_ids = {r["id"] for r in ec2_list} | {r["id"] for r in ec1_rows}
        sourcing_rows = conn.execute(
            """SELECT id, name, role, sector, seniority, years_experience, status, notes, createdAt
               FROM candidates
               WHERE owner_id = ?
                 AND substr(createdAt, 1, 10) >= ? AND substr(createdAt, 1, 10) <= ?
               ORDER BY createdAt;""",
            (uid, start, end),
        ).fetchall()
        sourcing_list = [dict(r) for r in sourcing_rows if r["id"] not in ec1_ec2_ids]

        # ── 4) Prospections (RDV pris) ──
        prosp_rdv_rows = conn.execute(
            """SELECT DISTINCT p.id, p.name AS prospect_name, COALESCE(c.groupe, '') AS company_name,
                      COALESCE(e.date, substr(p.lastContact, 1, 10)) AS rdv_date
               FROM prospects p
               LEFT JOIN companies c ON c.id = p.company_id
               LEFT JOIN prospect_events e ON e.prospect_id = p.id AND e.type = 'rdv_taken' AND e.date >= ? AND e.date <= ?
               WHERE p.owner_id = ? AND p.statut = 'Rendez-vous' AND (
                   (e.date IS NOT NULL) OR
                   (p.lastContact >= ? AND p.lastContact <= ?)
               )
               ORDER BY COALESCE(e.date, p.lastContact);""",
            (start, end, uid, start, end),
        ).fetchall()
        prosp_rdv_list = [dict(r) for r in prosp_rdv_rows]

        # ── 5) Clients vus (RDV effectué) ──
        clients_vus_rows = conn.execute(
            """SELECT DISTINCT p.id, p.name AS prospect_name, COALESCE(c.groupe, '') AS company_name,
                      p.notes, p.callNotes, p.lastContact,
                      COALESCE(e.date, substr(p.lastContact, 1, 10)) AS meeting_date
               FROM prospects p
               LEFT JOIN companies c ON c.id = p.company_id
               LEFT JOIN prospect_events e ON e.prospect_id = p.id
                   AND e.type IN ('meeting', 'reunion', 'rdv_done')
                   AND e.date >= ? AND e.date <= ?
               WHERE p.owner_id = ? AND p.statut = 'Rendez-vous' AND (
                   (e.date IS NOT NULL) OR
                   (p.lastContact >= ? AND p.lastContact <= ?)
               )
               ORDER BY COALESCE(e.date, p.lastContact);""",
            (start, end, uid, start, end),
        ).fetchall()
        clients_vus_list = [dict(r) for r in clients_vus_rows]

        # ── 6) Pushs (groupés par candidat, triés par nb desc) ──
        push_rows = conn.execute(
            """SELECT l.candidate_id1, l.candidate_id2, ca1.name AS candidate1_name, ca2.name AS candidate2_name,
                      l.sentAt
               FROM push_logs l
               JOIN prospects p ON p.id = l.prospect_id AND p.owner_id = ?
               LEFT JOIN candidates ca1 ON ca1.id = l.candidate_id1 AND ca1.owner_id = ?
               LEFT JOIN candidates ca2 ON ca2.id = l.candidate_id2 AND ca2.owner_id = ?
               WHERE substr(l.sentAt, 1, 10) >= ? AND substr(l.sentAt, 1, 10) <= ?
               ORDER BY l.sentAt;""",
            (uid, uid, uid, start, end),
        ).fetchall()
        push_list = [dict(r) for r in push_rows]
        push_by_candidate: dict = {}
        for pl in push_list:
            for cid_key, cname_key in [("candidate_id1", "candidate1_name"), ("candidate_id2", "candidate2_name")]:
                cid = pl.get(cid_key)
                if cid:
                    cname = pl.get(cname_key) or f"Candidat {cid}"
                    entry = push_by_candidate.setdefault(cid, {"name": cname, "count": 0})
                    entry["count"] += 1
        push_consultants = sorted(push_by_candidate.values(), key=lambda x: -x["count"])

        # ── 7) Objectifs ──
        goals_cfg = _get_goals_config(conn)
        weekly_goals = goals_cfg.get("weekly", {})
        attendus_prosp = weekly_goals.get("rdv", {}).get("target", 5)
        attendus_entretiens = weekly_goals.get("sourcing_solid", {}).get("target", 3)
        attendus_pushs = weekly_goals.get("push", {}).get("target", 15)

    # ── Enrichissement Ollama (optionnel) ──
    if use_ollama:
        all_cands = [(ec, "ec2") for ec in ec2_list] + [(ec, "ec1") for ec in ec1_list] + [(ec, "sourcing") for ec in sourcing_list]
        for item, _type in all_cands:
            metier = item.get("role") or item.get("sector") or ""
            if not metier or len(metier) < 3:
                p = f"Normalise ce métier en un nom court et standard: '{metier}'. Réponds uniquement avec le métier normalisé."
                normalized = _call_ollama(p)
                item["_normalized_metier"] = (normalized or metier)[:50]
            else:
                item["_normalized_metier"] = metier

        for client in clients_vus_list:
            notes = (client.get("notes") or "") + " " + (client.get("callNotes") or "")
            if notes.strip():
                p = f"Extrais les besoins exprimés par ce client (une ligne par besoin, format court):\n{notes[:500]}"
                besoins = _call_ollama(p)
                client["_besoins"] = (besoins or "")[:200]
            else:
                client["_besoins"] = ""

        for ec1 in ec1_list:
            ec1_data_str = ec1.get("ec1_data") or "{}"
            try:
                ec1_data = json.loads(ec1_data_str) if ec1_data_str else {}
            except Exception:
                ec1_data = {}
            parts = []
            if ec1.get("role"):
                parts.append(f"Métier: {ec1['role']}")
            if ec1.get("years_experience"):
                parts.append(f"Expérience: {ec1['years_experience']} ans")
            if ec1_data:
                parts.append(f"Données EC1: {json.dumps(ec1_data, ensure_ascii=False)[:200]}")
            if parts:
                p = "Génère un code note court (ex: 'B OKS') pour ce candidat:\n" + "\n".join(parts) + "\nRéponds uniquement avec le code."
                code = _call_ollama(p)
                ec1["_code_note"] = (code or "")[:20]
            else:
                ec1["_code_note"] = ""

    # ── Helper: extraire exp numérique ──
    def _parse_exp(ec):
        exp = ec.get("years_experience") or ec.get("seniority") or ""
        try:
            if isinstance(exp, str) and exp.strip():
                m = re.search(r'\d+', exp)
                return int(m.group()) if m else exp
        except Exception:
            pass
        return exp

    # ── Construire la liste ordonnée des candidats (EC2 d'abord, puis EC1, puis Sourcing) ──
    all_candidates = (
        [("EC2", ec) for ec in ec2_list] +
        [("EC1", ec) for ec in ec1_list] +
        [("Sourcing", ec) for ec in sourcing_list]
    )

    # ══════════════════════════════════════════════════════
    # Build the XLSX workbook — 15 colonnes A-O, layout zip
    # ══════════════════════════════════════════════════════
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste"

    # Styles
    header_fill = PatternFill(start_color="2B3A4E", end_color="2B3A4E", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="thick", color="000000"), right=Side(style="thick", color="000000"),
        top=Side(style="thick", color="000000"), bottom=Side(style="thick", color="000000"),
    )

    # ── En-têtes (15 colonnes A-O) ──
    headers = [
        "Semaine",             # A
        "Entretiens",          # B  (EC1 / EC2 / Sourcing)
        "Métier",              # C  (Nom - Rôle)
        "Exp",                 # D
        "Dispo",               # E
        "Notes",               # F
        "Commenta",            # G  (séparateur — bordure épaisse)
        "Prospections RDV pris",  # H
        "Clients vus",         # I
        "Besoins",             # J
        "Pushs consultant",    # K
        "Nb pushs",            # L
        "Attendus Prosp",      # M
        "Attendus Entretiens", # N
        "Attendus Pushs",      # O
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Largeurs de colonnes
    col_widths = [10, 12, 35, 6, 12, 18, 18, 28, 22, 30, 28, 9, 14, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Layout zip : chaque ligne i combine le i-ème candidat + i-ème prospection + i-ème push ──
    total_rows = max(1, len(all_candidates), len(prosp_rdv_list), len(clients_vus_list), len(push_consultants))
    week_start_row = 2

    for i in range(total_rows):
        row = week_start_row + i

        # A: Semaine (sera fusionné verticalement à la fin)
        ws.cell(row=row, column=1, value=week_label)

        # B-G : candidat
        if i < len(all_candidates):
            ctype, ec = all_candidates[i]
            ws.cell(row=row, column=2, value=ctype)  # B: EC1 / EC2 / Sourcing
            # C: "Nom - Rôle"
            name = ec.get("name") or ""
            role = ec.get("_normalized_metier") if use_ollama else (ec.get("role") or ec.get("sector") or "")
            metier_str = f"{name} - {role}" if role else name
            ws.cell(row=row, column=3, value=metier_str)
            ws.cell(row=row, column=4, value=_parse_exp(ec))  # D: Exp
            # E: Dispo (depuis tab EC1 si disponible, sinon 'asap')
            dispo = ec.get("availability") or "asap"
            ws.cell(row=row, column=5, value=dispo)
            # F: Notes courtes
            if ctype == "EC1":
                ws.cell(row=row, column=6, value=ec.get("_code_note") if use_ollama else (ec.get("tab_notes") or ""))
            else:
                ws.cell(row=row, column=6, value=(ec.get("notes") or "")[:120])
            # G: Commenta (bordure épaisse — séparateur visuel)
            if ctype == "EC1":
                ec1_data_str = ec.get("ec1_data") or "{}"
                try:
                    ec1_data = json.loads(ec1_data_str) if ec1_data_str else {}
                    commenta = json.dumps(ec1_data, ensure_ascii=False)[:300] if ec1_data else ""
                except Exception:
                    commenta = ""
            else:
                commenta = (ec.get("notes") or "")[:300]
            ws.cell(row=row, column=7, value=commenta)

        # H: Prospections RDV pris
        if i < len(prosp_rdv_list):
            prosp = prosp_rdv_list[i]
            company = prosp.get("company_name", "")
            prosp_text = f"{prosp.get('prospect_name', '')} - {company}" if company else prosp.get("prospect_name", "")
            ws.cell(row=row, column=8, value=prosp_text)

        # I-J: Clients vus + Besoins
        if i < len(clients_vus_list):
            client = clients_vus_list[i]
            company = client.get("company_name", "")
            client_text = f"{client.get('prospect_name', '')} - {company}" if company else client.get("prospect_name", "")
            ws.cell(row=row, column=9, value=client_text)
            besoins = client.get("_besoins") if use_ollama else ""
            if not besoins:
                notes_raw = (client.get("notes") or "") + " " + (client.get("callNotes") or "")
                besoins = notes_raw.strip()[:200]
            ws.cell(row=row, column=10, value=besoins)

        # K-L: Pushs consultant + Nb pushs
        if i < len(push_consultants):
            pc = push_consultants[i]
            cnt = pc.get("count", 0)
            ws.cell(row=row, column=11, value=f"{pc.get('name', '')} ({cnt}x)")
            ws.cell(row=row, column=12, value=cnt)

        # M-N-O: Objectifs (première ligne de la semaine uniquement)
        if i == 0:
            ws.cell(row=row, column=13, value=attendus_prosp)
            ws.cell(row=row, column=14, value=attendus_entretiens)
            ws.cell(row=row, column=15, value=attendus_pushs)

        # Bordures
        for col in range(1, 16):
            ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=7).border = thick_border  # Séparateur G

    # ── Fusionner la colonne A (Semaine) sur toutes les lignes de la semaine ──
    week_end_row = week_start_row + total_rows - 1
    if week_end_row > week_start_row:
        ws.merge_cells(f'A{week_start_row}:A{week_end_row}')
        merged_cell = ws.cell(row=week_start_row, column=1)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Wrap text / alignement vertical ──
    for r in range(week_start_row, week_end_row + 1):
        for col in [3, 7, 10, 11]:
            ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Stream le fichier ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Suivi_activite_{week_label}_{monday.isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ====== Duplicates API ======
def _norm_phone(s: str) -> str:
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits[-10:] if len(digits) > 10 else digits


def _normalize(s: str) -> str:
    """Lowercase + strip + remove accents + collapse whitespace.

    Used for duplicate detection keys; must be deterministic across OS/timezone.
    """
    s = (s or "").strip().lower()
    # Remove accents/diacritics
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    # Keep it simple (avoid funky separators)
    s = re.sub(r"\s+", " ", s)
    return s


def _name_key_for_duplicate(name: str) -> str:
    """Normalise un nom pour comparaison doublons: INITIALES NOM (ex. PY CAMPION).

    Prénom(s) → initiales (chaque sous-mot - . - espace donne une lettre).
    Dernière partie = nom de famille.
    """
    s = _normalize(name or "")
    if not s:
        return ""
    parts = re.split(r"[\s,;]+", s)
    parts = [x for x in parts if x]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0].upper()
    lastname = parts[-1]
    first_parts = " ".join(parts[:-1])
    initials = []
    for sub in re.split(r"[\s.\-]+", first_parts):
        if sub:
            initials.append(sub[0])
    initials_str = "".join(initials).upper()
    return f"{initials_str} {lastname.upper()}".strip()


def _split_name_for_dup(name: str) -> tuple[str, str]:
    """Retourne (lastname_norm, firstname_norm) pour comparaison doublons stricte.

    Sépare explicitement le nom de famille (dernière partie) du prénom (reste).
    Exemple : "Jean-Pierre DUPONT" → ("dupont", "jean pierre")
    Utilisé pour éviter les faux positifs : exige même nom ET même initiale prénom.
    """
    s = _normalize(name or "")
    parts = [x for x in re.split(r"[\s,;]+", s) if x]
    if not parts:
        return ("", "")
    if len(parts) == 1:
        return (parts[0], "")
    lastname = parts[-1]
    firstname = " ".join(parts[:-1])
    # Normaliser tirets/points dans le prénom → "jean-pierre" devient "jean pierre"
    firstname = re.sub(r"[\.\-]+", " ", firstname).strip()
    firstname = re.sub(r"\s+", " ", firstname)
    return (lastname, firstname)


# ────────────────────────────────────────────────────────────────────
# Stats v30 — données pour charts interactifs (période mensuelle + filtres)
# ────────────────────────────────────────────────────────────────────

@dashboard_bp.get("/api/stats/data")
def api_stats_data():
    """Agrégats pour les 4 charts v30 : RDV/mois, Appels/mois, Funnel, Top entreprises.
    Query params:
      - period : YYYY-MM (month-based sliding window, défaut = mois courant)
      - start / end : YYYY-MM-DD (custom range — prioritaire sur period)
      - tags : CSV de tags à filtrer
      - statuts : CSV de statuts à filtrer
      - user_id : int (admin only — filtrer par utilisateur)
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    today = datetime.date.today()

    # ── Résolution de la période ──
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    period = (request.args.get("period") or "").strip()  # YYYY-MM

    if start_s and end_s:
        try:
            start_d = datetime.date.fromisoformat(start_s)
            end_d = datetime.date.fromisoformat(end_s)
            if start_d > end_d:
                start_d, end_d = end_d, start_d
        except Exception:
            start_d = today.replace(day=1)
            end_d = today
    elif period:
        try:
            y, m = int(period[:4]), int(period[5:7])
            start_d = datetime.date(y, m, 1)
            if m == 12:
                end_d = datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_d = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
        except Exception:
            start_d = today.replace(day=1)
            end_d = today
    else:
        start_d = today.replace(day=1)
        end_d = today

    # ── Filtres optionnels ──
    tags_filter = [t.strip() for t in (request.args.get("tags") or "").split(",") if t.strip()]
    statuts_filter = [s.strip() for s in (request.args.get("statuts") or "").split(",") if s.strip()]

    # Admin peut filtrer par utilisateur
    target_uid = uid
    user_id_param = request.args.get("user_id", "").strip()
    if user_id_param:
        u = _get_current_user()
        if u and u.get("role") == "admin":
            try:
                target_uid = int(user_id_param)
            except Exception:
                pass

    # Construire des clauses SQL dynamiques
    base_cond = ("p.owner_id=? AND (p.deleted_at IS NULL OR p.deleted_at='') "
                 "AND (p.is_archived IS NULL OR p.is_archived=0)")
    base_params: list = [target_uid]

    if statuts_filter:
        ph = ",".join("?" * len(statuts_filter))
        base_cond += f" AND p.statut IN ({ph})"
        base_params.extend(statuts_filter)

    # RDV par mois (6 derniers mois se terminant par end_d)
    months_rdv = []
    months_calls = []
    for i in range(5, -1, -1):
        ref = end_d.replace(day=1)
        # reculer i mois
        y_off = ref.year + (ref.month - 1 - i) // 12
        m_off = (ref.month - 1 - i) % 12 + 1
        first = datetime.date(y_off, m_off, 1)
        if m_off == 12:
            last = datetime.date(y_off + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            last = datetime.date(y_off, m_off + 1, 1) - datetime.timedelta(days=1)
        label = first.strftime("%b %Y")
        fi, li = first.isoformat(), last.isoformat()

        with _conn() as conn:
            rdv_n = conn.execute(
                f"""SELECT COUNT(DISTINCT e.prospect_id) AS n
                    FROM prospect_events e
                    JOIN prospects p ON p.id=e.prospect_id
                    WHERE {base_cond} AND e.type='rdv_taken'
                      AND substr(e.date,1,10)>=? AND substr(e.date,1,10)<=?""",
                base_params + [fi, li],
            ).fetchone()["n"]

            try:
                calls_n = conn.execute(
                    "SELECT COUNT(*) AS n FROM call_logs WHERE owner_id=? AND date>=? AND date<=?",
                    (target_uid, fi, li),
                ).fetchone()["n"]
            except Exception:
                calls_n = 0

        months_rdv.append({"label": label, "count": rdv_n})
        months_calls.append({"label": label, "count": calls_n})

    # Funnel
    with _conn() as conn:
        total_p = conn.execute(
            f"SELECT COUNT(*) AS n FROM prospects p WHERE {base_cond}", base_params
        ).fetchone()["n"]
        rdv_p = conn.execute(
            f"SELECT COUNT(*) AS n FROM prospects p WHERE {base_cond} AND p.statut='Rendez-vous'",
            base_params,
        ).fetchone()["n"]
        conv_rate = round(rdv_p / total_p, 4) if total_p > 0 else 0.0

        # Top entreprises
        top_rows = conn.execute(
            f"""SELECT c.groupe AS name, COUNT(p.id) AS n
                FROM companies c
                JOIN prospects p ON p.company_id=c.id
                WHERE {base_cond}
                GROUP BY c.id ORDER BY n DESC LIMIT 10""",
            base_params,
        ).fetchall()
        top_companies = [{"name": r["name"] or "—", "count": r["n"]} for r in top_rows]

    return jsonify(
        ok=True,
        period={"start": start_d.isoformat(), "end": end_d.isoformat()},
        rdv_by_month=[m["count"] for m in months_rdv],
        rdv_labels=[m["label"] for m in months_rdv],
        calls_by_month=[m["count"] for m in months_calls],
        calls_labels=[m["label"] for m in months_calls],
        funnel={"prospects": total_p, "rdv": rdv_p, "conversion_rate": conv_rate},
        top_companies=top_companies,
    )


@dashboard_bp.get("/api/stats/export")
def api_stats_export():
    """Export des données stats en JSON ou CSV.
    Query params:
      - period : YYYY-MM
      - start / end : YYYY-MM-DD
      - format : json | csv  (défaut json)
      - tags / statuts / user_id : mêmes filtres que /api/stats/data
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    fmt = (request.args.get("format") or "json").lower().strip()

    today = datetime.date.today()
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()
    period = (request.args.get("period") or "").strip()
    tags_filter = [t.strip() for t in (request.args.get("tags") or "").split(",") if t.strip()]
    statuts_filter = [s.strip() for s in (request.args.get("statuts") or "").split(",") if s.strip()]

    if start_s and end_s:
        try:
            start_d = datetime.date.fromisoformat(start_s)
            end_d = datetime.date.fromisoformat(end_s)
        except Exception:
            start_d = today.replace(day=1); end_d = today
    elif period:
        try:
            y, m = int(period[:4]), int(period[5:7])
            start_d = datetime.date(y, m, 1)
            end_d = (datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)) if m < 12 else datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
        except Exception:
            start_d = today.replace(day=1); end_d = today
    else:
        start_d = today.replace(day=1); end_d = today

    target_uid = uid
    user_id_param = request.args.get("user_id", "").strip()
    if user_id_param:
        u = _get_current_user()
        if u and u.get("role") == "admin":
            try:
                target_uid = int(user_id_param)
            except Exception:
                pass

    base_cond = ("p.owner_id=? AND (p.deleted_at IS NULL OR p.deleted_at='') "
                 "AND (p.is_archived IS NULL OR p.is_archived=0)")
    base_params: list = [target_uid]
    if statuts_filter:
        ph = ",".join("?" * len(statuts_filter))
        base_cond += f" AND p.statut IN ({ph})"
        base_params.extend(statuts_filter)

    # Données brutes prospects pour l'export
    with _conn() as conn:
        rows = conn.execute(
            f"""SELECT p.id, p.name, p.statut, p.lastContact, p.nextFollowUp,
                       c.groupe AS company
                FROM prospects p
                LEFT JOIN companies c ON c.id=p.company_id AND c.owner_id=?
                WHERE {base_cond}
                  AND (p.lastContact>=? OR p.nextFollowUp>=?)
                ORDER BY p.name""",
            [target_uid] + base_params + [start_d.isoformat(), start_d.isoformat()],
        ).fetchall()
        data_rows = [dict(r) for r in rows]

    filename_base = f"stats_{start_d}_{end_d}"

    if fmt == "csv":
        import io as _io
        import csv as _csv
        out = _io.StringIO()
        writer = _csv.DictWriter(out, fieldnames=["id", "name", "company", "statut", "lastContact", "nextFollowUp"])
        writer.writeheader()
        writer.writerows(data_rows)
        csv_bytes = out.getvalue().encode("utf-8-sig")
        return Response(
            csv_bytes,
            mimetype="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )
    else:
        payload = json.dumps({"period": {"start": start_d.isoformat(), "end": end_d.isoformat()}, "prospects": data_rows}, ensure_ascii=False, indent=2)
        return Response(
            payload.encode("utf-8"),
            mimetype="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.json"'},
        )


