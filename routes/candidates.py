"""ProspUp — Blueprint Candidates (CRUD + sous-ressources : experiences,
educations, certifications, skills, availability, dossier de compétence,
DC upload/extract/rename/delete)."""
from __future__ import annotations

import datetime
import json
import os
import re
from pathlib import Path

from flask import Blueprint, jsonify, request, send_file
from werkzeug.utils import secure_filename

from app import _audit_log, _dump_json_list, _maybe_log_candidate_events, _parse_json_int_list, _parse_json_str_list, log_activity, logger
from config import APP_DIR, DATA_DIR, APP_VERSION
from utils.ai_helpers import _call_ai, _call_ollama_direct, _load_ai_config
from utils.auth import _candidate_owned, _require_same_origin, _uid, login_required, role_required, validate_payload
from utils.common import _now_iso, _today_iso
from utils.db import _conn
from utils.files import _candidate_attachment_dir, _extract_pdf_text, _validate_upload

candidates_bp = Blueprint("candidates", __name__)


@candidates_bp.get("/api/candidates")
def api_candidates_list():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    # v23.5: optional pagination via ?page=&limit=
    page_param = request.args.get("page")
    if page_param is not None:
        try:
            page = max(1, int(page_param))
            limit = min(500, max(1, int(request.args.get("limit") or 100)))
        except (TypeError, ValueError):
            page, limit = 1, 100
        offset = (page - 1) * limit
        with _conn() as conn:
            total = int(conn.execute("SELECT COUNT(*) FROM candidates WHERE owner_id=? AND deleted_at IS NULL;", (uid,)).fetchone()[0])
            rows = conn.execute(
                "SELECT * FROM candidates WHERE owner_id=? AND deleted_at IS NULL ORDER BY COALESCE(updatedAt, createdAt) DESC, id DESC LIMIT ? OFFSET ?;",
                (uid, limit, offset),
            ).fetchall()
            gen_ids = {
                int(r[0]) for r in conn.execute(
                    "SELECT DISTINCT candidate_id FROM dc_generations WHERE owner_id=? AND deleted_at IS NULL AND candidate_id IS NOT NULL;",
                    (uid,),
                ).fetchall()
            }
        out = []
        for r in rows:
            d = dict(r)
            d["skills"] = _parse_json_str_list(d.get("skills"))
            d["company_ids"] = _parse_json_int_list(d.get("company_ids"))
            d["has_dc"] = _candidate_has_dc(uid, d, gen_ids)
            out.append(d)
        from math import ceil
        return jsonify(ok=True, candidates=out, pagination={"page": page, "limit": limit, "total": total, "pages": ceil(total / limit) if limit else 1})
    # Non-paginated (backward compatible)
    with _conn() as conn:
        rows = conn.execute(
            "SELECT * FROM candidates WHERE owner_id=? AND deleted_at IS NULL ORDER BY COALESCE(updatedAt, createdAt) DESC, id DESC;",
            (uid,),
        ).fetchall()
        gen_ids = {
            int(r[0]) for r in conn.execute(
                "SELECT DISTINCT candidate_id FROM dc_generations WHERE owner_id=? AND deleted_at IS NULL AND candidate_id IS NOT NULL;",
                (uid,),
            ).fetchall()
        }
    out: List[Dict[str, Any]] = []
    for r in rows:
        d = dict(r)
        d["skills"] = _parse_json_str_list(d.get("skills"))
        d["company_ids"] = _parse_json_int_list(d.get("company_ids"))
        d["has_dc"] = _candidate_has_dc(uid, d, gen_ids)
        out.append(d)
    return jsonify(out)


def _candidate_has_dc(uid: int, cand: dict, gen_ids: set) -> bool:
    """Vrai si le candidat a un DC: PDF uploadé (champ ou fichier sur disque)
    OU un DC généré via le générateur (table dc_generations)."""
    if cand.get("dossier_competence_pdf"):
        return True
    cid = cand.get("id")
    if cid in gen_ids:
        return True
    dc_dir = DATA_DIR / "dossiers_candidats" / str(uid) / str(cid)
    return dc_dir.is_dir() and any(dc_dir.glob("*.pdf"))


@candidates_bp.get("/api/candidates/<int:candidate_id>")
def api_candidate_get(candidate_id: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute("SELECT * FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "not_found"}), 404
        cand = dict(row)

        cand["skills"] = _parse_json_str_list(cand.get("skills"))
        cand["company_ids"] = _parse_json_int_list(cand.get("company_ids"))

        companies: List[Dict[str, Any]] = []
        if cand["company_ids"]:
            placeholders = ",".join(["?"] * len(cand["company_ids"]))
            rows2 = conn.execute(
                f"SELECT * FROM companies WHERE owner_id=? AND id IN ({placeholders}) ORDER BY groupe, site;",
                (uid, *cand["company_ids"]),
            ).fetchall()
            companies = [dict(r) for r in rows2]

    return jsonify({"ok": True, "candidate": cand, "companies": companies})


@candidates_bp.put("/api/candidates/<int:candidate_id>")
def api_candidate_put(candidate_id: int):
    """Partial-update a candidate (inline edit from v30 fiche candidat)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _candidate_owned(candidate_id):
        return jsonify(ok=False, error="Accès refusé"), 403
    body = request.get_json(force=True, silent=True) or {}
    if not body:
        return jsonify(ok=False, error="Body vide"), 400

    ALLOWED = {
        "name", "role", "location", "seniority", "tech", "linkedin", "source", "status", "notes",
        "phone", "email", "sector", "prenom", "titre", "years_experience", "annees_experience",
        "domaine_principal", "description_push", "disponibilite", "mobilite", "permis_travail",
        "fonctions_recherchees", "motif_recherche", "avancement_recherches",
        "remuneration_actuelle", "pretentions_salariales", "propal_a", "langues",
        "eval_technique", "eval_personnalite", "eval_communication",
        "references_candidat", "avis_perso",
        "entretien_date", "entretien_lieu", "entretien_notes",
        "vsa_url", "onenote_url",
    }
    updates = {k: v for k, v in body.items() if k in ALLOWED}
    if not updates:
        return jsonify(ok=True)

    now = datetime.datetime.now().isoformat(timespec="seconds")
    set_clause = ", ".join(k + "=?" for k in updates) + ", updatedAt=?"
    params = list(updates.values()) + [now, candidate_id, uid]

    with _conn() as conn:
        cur = conn.execute(
            "UPDATE candidates SET " + set_clause + " WHERE id=? AND owner_id=?;",
            params,
        )
        if cur.rowcount == 0:
            return jsonify(ok=False, error="Candidat introuvable"), 404
        conn.commit()

    return jsonify(ok=True)


@candidates_bp.get("/api/candidates/<int:candidate_id>/experiences")
def api_candidate_experiences_get(candidate_id: int):
    """Get all experiences for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        # Verify candidate exists and belongs to user
        cand = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not cand:
            return jsonify({"ok": False, "error": "not_found"}), 404
        rows = conn.execute(
            "SELECT * FROM candidate_experiences WHERE candidate_id=? AND owner_id=? ORDER BY start_date DESC, id DESC;",
            (candidate_id, uid)
        ).fetchall()
        experiences = []
        for row in rows:
            exp = dict(row)
            # Parse technologies JSON if present
            if exp.get("technologies"):
                try:
                    exp["technologies"] = json.loads(exp["technologies"])
                except:
                    exp["technologies"] = []
            else:
                exp["technologies"] = []
            experiences.append(exp)
    return jsonify({"ok": True, "experiences": experiences})


@candidates_bp.post("/api/candidates/<int:candidate_id>/experiences")
def api_candidate_experiences_post(candidate_id: int):
    """Create a new experience for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=False) or {}
    company_name = (payload.get("company_name") or "").strip()
    if not company_name:
        return jsonify({"ok": False, "error": "company_name is required"}), 400
    with _conn() as conn:
        # Verify candidate exists and belongs to user
        cand = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not cand:
            return jsonify({"ok": False, "error": "not_found"}), 404
        now = datetime.datetime.now().isoformat()
        technologies = payload.get("technologies")
        if isinstance(technologies, (list, dict)):
            technologies = json.dumps(technologies, ensure_ascii=False)
        elif not isinstance(technologies, str):
            technologies = ""
        cur = conn.execute(
            """INSERT INTO candidate_experiences
               (candidate_id, company_name, role, start_date, end_date, description, technologies, owner_id, createdAt, updatedAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);""",
            (
                candidate_id,
                company_name,
                (payload.get("role") or "").strip() or None,
                (payload.get("start_date") or "").strip() or None,
                (payload.get("end_date") or "").strip() or None,
                (payload.get("description") or "").strip() or None,
                technologies,
                uid,
                now,
                now,
            ),
        )
        exp_id = cur.lastrowid
    return jsonify({"ok": True, "id": exp_id})


@candidates_bp.get("/api/candidates/<int:candidate_id>/educations")
def api_candidate_educations_get(candidate_id: int):
    """Get all educations for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        # Verify candidate exists and belongs to user
        cand = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not cand:
            return jsonify({"ok": False, "error": "not_found"}), 404
        rows = conn.execute(
            "SELECT * FROM candidate_educations WHERE candidate_id=? AND owner_id=? ORDER BY year DESC, id DESC;",
            (candidate_id, uid)
        ).fetchall()
        educations = [dict(row) for row in rows]
    return jsonify({"ok": True, "educations": educations})


@candidates_bp.post("/api/candidates/<int:candidate_id>/educations")
def api_candidate_educations_post(candidate_id: int):
    """Create a new education for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=False) or {}
    school = (payload.get("school") or "").strip()
    if not school:
        return jsonify({"ok": False, "error": "school is required"}), 400
    with _conn() as conn:
        # Verify candidate exists and belongs to user
        cand = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not cand:
            return jsonify({"ok": False, "error": "not_found"}), 404
        now = datetime.datetime.now().isoformat()
        cur = conn.execute(
            """INSERT INTO candidate_educations
               (candidate_id, degree, school, year, specialization, owner_id, createdAt, updatedAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?);""",
            (
                candidate_id,
                (payload.get("degree") or "").strip() or None,
                school,
                (payload.get("year") or "").strip() or None,
                (payload.get("specialization") or "").strip() or None,
                uid,
                now,
                now,
            ),
        )
        edu_id = cur.lastrowid
    return jsonify({"ok": True, "id": edu_id})


@candidates_bp.get("/api/candidates/<int:candidate_id>/certifications")
def api_candidate_certifications_get(candidate_id: int):
    """Get all certifications for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        # Verify candidate exists and belongs to user
        cand = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not cand:
            return jsonify({"ok": False, "error": "not_found"}), 404
        rows = conn.execute(
            "SELECT * FROM candidate_certifications WHERE candidate_id=? AND owner_id=? ORDER BY obtained_date DESC, id DESC;",
            (candidate_id, uid)
        ).fetchall()
        certifications = [dict(row) for row in rows]
    return jsonify({"ok": True, "certifications": certifications})


@candidates_bp.post("/api/candidates/<int:candidate_id>/certifications")
def api_candidate_certifications_post(candidate_id: int):
    """Create a new certification for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=False) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400
    with _conn() as conn:
        # Verify candidate exists and belongs to user
        cand = conn.execute("SELECT id FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not cand:
            return jsonify({"ok": False, "error": "not_found"}), 404
        now = datetime.datetime.now().isoformat()
        cur = conn.execute(
            """INSERT INTO candidate_certifications
               (candidate_id, name, issuer, obtained_date, expiry_date, owner_id, createdAt, updatedAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?);""",
            (
                candidate_id,
                name,
                (payload.get("issuer") or "").strip() or None,
                (payload.get("obtained_date") or "").strip() or None,
                (payload.get("expiry_date") or "").strip() or None,
                uid,
                now,
                now,
            ),
        )
        cert_id = cur.lastrowid
    return jsonify({"ok": True, "id": cert_id})


# ════════════════════════════════════════════════════════════
# v30 — Candidate skills + availability (granularité fine)
# ════════════════════════════════════════════════════════════

def _cand_owned_row(conn, candidate_id: int, uid: int):
    return conn.execute(
        "SELECT id, tech FROM candidates WHERE id=? AND owner_id=?;",
        (candidate_id, uid),
    ).fetchone()


@candidates_bp.get("/api/candidates/<int:candidate_id>/skills")
def api_cand_skills_get(candidate_id: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        cand = _cand_owned_row(conn, candidate_id, uid)
        if not cand:
            return jsonify(ok=False, error="not_found"), 404
        rows = conn.execute(
            "SELECT id, name, category, level FROM candidate_skills "
            "WHERE candidate_id=? ORDER BY category, name;",
            (candidate_id,),
        ).fetchall()
        skills = [dict(r) for r in rows]
        # Backfill depuis candidates.tech si aucune skill enregistrée
        cand_tech = cand["tech"] if "tech" in cand.keys() else None
        if not skills and cand_tech:
            tech_list = [t.strip() for t in str(cand_tech).split(",") if t.strip()]
            for name in tech_list[:40]:
                try:
                    conn.execute(
                        "INSERT OR IGNORE INTO candidate_skills (candidate_id, name, category, level) "
                        "VALUES (?,?,?,?);",
                        (candidate_id, name, "Compétences", 3),
                    )
                except Exception:
                    pass
            skills = [dict(r) for r in conn.execute(
                "SELECT id, name, category, level FROM candidate_skills "
                "WHERE candidate_id=? ORDER BY category, name;", (candidate_id,)
            ).fetchall()]
    return jsonify(ok=True, skills=skills)


@candidates_bp.post("/api/candidates/<int:candidate_id>/skills")
def api_cand_skills_post(candidate_id: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    err = _require_same_origin()
    if err:
        return err
    payload = request.get_json(force=True, silent=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify(ok=False, error="name is required"), 400
    category = (payload.get("category") or "").strip() or None
    level = max(1, min(5, int(payload.get("level") or 3)))
    with _conn() as conn:
        cand = _cand_owned_row(conn, candidate_id, uid)
        if not cand:
            return jsonify(ok=False, error="not_found"), 404
        try:
            conn.execute(
                "INSERT INTO candidate_skills (candidate_id, name, category, level) "
                "VALUES (?,?,?,?) "
                "ON CONFLICT(candidate_id, name) DO UPDATE SET "
                "  category=excluded.category, level=excluded.level;",
                (candidate_id, name, category, level),
            )
        except Exception as e:
            return jsonify(ok=False, error=str(e)), 400
        row = conn.execute(
            "SELECT id, name, category, level FROM candidate_skills "
            "WHERE candidate_id=? AND name=?;",
            (candidate_id, name),
        ).fetchone()
    return jsonify(ok=True, skill=dict(row) if row else None)


@candidates_bp.delete("/api/candidates/<int:candidate_id>/skills/<int:skill_id>")
def api_cand_skills_delete(candidate_id: int, skill_id: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    err = _require_same_origin()
    if err:
        return err
    with _conn() as conn:
        cand = _cand_owned_row(conn, candidate_id, uid)
        if not cand:
            return jsonify(ok=False, error="not_found"), 404
        conn.execute(
            "DELETE FROM candidate_skills WHERE id=? AND candidate_id=?;",
            (skill_id, candidate_id),
        )
    return jsonify(ok=True)


@candidates_bp.get("/api/candidates/<int:candidate_id>/availability")
def api_cand_avail_get(candidate_id: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        cand = _cand_owned_row(conn, candidate_id, uid)
        if not cand:
            return jsonify(ok=False, error="not_found"), 404
        rows = conn.execute(
            "SELECT week_iso, status FROM candidate_availability "
            "WHERE candidate_id=? ORDER BY week_iso;",
            (candidate_id,),
        ).fetchall()
    return jsonify(ok=True, availability=[dict(r) for r in rows])


@candidates_bp.post("/api/candidates/<int:candidate_id>/availability")
def api_cand_avail_post(candidate_id: int):
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    err = _require_same_origin()
    if err:
        return err
    payload = request.get_json(force=True, silent=True) or {}
    week_iso = (payload.get("week_iso") or "").strip()
    status = (payload.get("status") or "").strip().lower()
    if not week_iso or status not in ("libre", "busy", "placed"):
        return jsonify(ok=False, error="week_iso + status (libre|busy|placed) requis"), 400
    with _conn() as conn:
        cand = _cand_owned_row(conn, candidate_id, uid)
        if not cand:
            return jsonify(ok=False, error="not_found"), 404
        conn.execute(
            "INSERT INTO candidate_availability (candidate_id, week_iso, status) "
            "VALUES (?,?,?) "
            "ON CONFLICT(candidate_id, week_iso) DO UPDATE SET status=excluded.status;",
            (candidate_id, week_iso, status),
        )
    return jsonify(ok=True)


@candidates_bp.get("/api/candidates/<int:candidate_id>/dossier-competence")
def api_candidate_dossier_competence(candidate_id: int):
    """Serve the competence dossier PDF for a candidate."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    
    with _conn() as conn:
        row = conn.execute("SELECT dossier_competence_pdf FROM candidates WHERE id=? AND owner_id=?;", (candidate_id, uid)).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "not_found"}), 404
        
        pdf_path = row["dossier_competence_pdf"]
        if not pdf_path or not pdf_path.strip():
            return jsonify({"ok": False, "error": "Aucun dossier de compétence renseigné"}), 404
        
        # Chemin du PDF (peut être relatif ou absolu)
        pdf_file = Path(pdf_path)
        if not pdf_file.is_absolute():
            # Si relatif, chercher dans le dossier dossiers_competence à la racine
            pdf_file = APP_DIR / "dossiers_competence" / pdf_file
        
        if not pdf_file.exists() or not pdf_file.is_file():
            return jsonify({"ok": False, "error": "Fichier PDF introuvable"}), 404
        
        # Vérifier que c'est bien un PDF
        if pdf_file.suffix.lower() != ".pdf":
            return jsonify({"ok": False, "error": "Le fichier n'est pas un PDF"}), 400
        
        try:
            return send_file(str(pdf_file), mimetype="application/pdf", as_attachment=True, download_name=pdf_file.name)
        except Exception as e:
            logger.error(f"Error serving PDF: {e}")
            return jsonify({"ok": False, "error": f"Erreur lors du chargement du PDF: {str(e)}"}), 500


@candidates_bp.post("/api/candidates/save")
def api_candidates_save():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload, err = validate_payload({'name': str})
    if err:
        return err
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400

    # Normalize fields
    def _t(k): 
        v = payload.get(k)
        return (str(v).strip() if v is not None else None) or None

    cid = payload.get("id")
    _candidate_action = 'update' if cid else 'create'
    template_id = payload.get("template_id")
    template_name = (payload.get("template_name") or "").strip() or None
    try:
        template_id = int(template_id) if template_id not in (None, "", "null") else None
    except Exception:
        template_id = None

    now = datetime.datetime.now().isoformat(timespec="seconds")
    event_date = now[:10]

    skills_json = _dump_json_list(payload.get("skills"), as_int=False)
    company_ids_json = _dump_json_list(payload.get("company_ids"), as_int=True)

    is_archived = 0
    try:
        is_archived = int(payload.get("is_archived") or 0)
    except Exception:
        is_archived = 0

    years_experience = None
    try:
        ye = payload.get("years_experience")
        years_experience = int(ye) if ye not in (None, "", "null") else None
    except Exception:
        years_experience = None

    with _conn() as conn:
        cur = conn.cursor()
        old_status = None
        existing_dc_path = None
        if cid:
            try:
                r0 = conn.execute("SELECT status, dossier_competence_pdf FROM candidates WHERE id=? AND owner_id=?;", (int(cid), uid)).fetchone()
                old_status = r0["status"] if r0 else None
                existing_dc_path = r0["dossier_competence_pdf"] if r0 else None
            except Exception:
                old_status = None
        # v27.x: champs prenom, titre, annees_experience, domaine_principal
        annees_exp_v = None
        try:
            ae = payload.get("annees_experience")
            annees_exp_v = int(ae) if ae not in (None, "", "null") else None
        except Exception:
            annees_exp_v = None

        # v28.1: champs fiche entretien
        permis_conduire_v = None
        try:
            pc = payload.get("permis_conduire")
            permis_conduire_v = int(pc) if pc not in (None, "", "null") else None
        except Exception:
            permis_conduire_v = None
        vehicule_v = None
        try:
            vh = payload.get("vehicule")
            vehicule_v = int(vh) if vh not in (None, "", "null") else None
        except Exception:
            vehicule_v = None

        if cid:
            cur.execute(
                '''
                UPDATE candidates
                SET name=?, role=?, location=?, seniority=?, tech=?, linkedin=?, source=?, status=?, notes=?,
                    onenote_url=?, vsa_url=?, skills=?, company_ids=?, is_archived=?,
                    years_experience=?, sector=?, phone=?, email=?, dossier_competence_pdf=?,
                    prenom=?, titre=?, annees_experience=?, domaine_principal=?,
                    description_push=?,
                    disponibilite=?, mobilite=?, permis_conduire=?, vehicule=?, permis_travail=?,
                    fonctions_recherchees=?, motif_recherche=?, avancement_recherches=?,
                    remuneration_actuelle=?, pretentions_salariales=?, propal_a=?,
                    eval_technique=?, eval_personnalite=?, eval_communication=?,
                    langues=?, references_candidat=?, avis_perso=?,
                    entretien_date=?, entretien_lieu=?, entretien_notes=?,
                    updatedAt=?
                WHERE id=? AND owner_id=?;
                ''',
                (
                    name,
                    _t("role"),
                    _t("location"),
                    _t("seniority"),
                    _t("tech"),
                    _t("linkedin"),
                    _t("source"),
                    _t("status"),
                    _t("notes"),
                    _t("onenote_url"),
                    _t("vsa_url"),
                    skills_json,
                    company_ids_json,
                    is_archived,
                    years_experience,
                    _t("sector"),
                    _t("phone"),
                    _t("email"),
                    _t("dossier_competence_pdf") or existing_dc_path,
                    _t("prenom"),
                    _t("titre"),
                    annees_exp_v,
                    _t("domaine_principal"),
                    _t("description_push"),
                    _t("disponibilite"),
                    _t("mobilite"),
                    permis_conduire_v,
                    vehicule_v,
                    _t("permis_travail"),
                    _t("fonctions_recherchees"),
                    _t("motif_recherche"),
                    _t("avancement_recherches"),
                    _t("remuneration_actuelle"),
                    _t("pretentions_salariales"),
                    _t("propal_a"),
                    _t("eval_technique"),
                    _t("eval_personnalite"),
                    _t("eval_communication"),
                    _t("langues"),
                    _t("references_candidat"),
                    _t("avis_perso"),
                    _t("entretien_date"),
                    _t("entretien_lieu"),
                    _t("entretien_notes"),
                    now,
                    int(cid),
                    uid,
                ),
            )
            if cur.rowcount == 0:
                cid = None  # fallback to insert
        if not cid:
            cur.execute(
                '''
                INSERT INTO candidates (
                    name, role, location, seniority, tech, linkedin, source, status, notes,
                    onenote_url, vsa_url, skills, company_ids, is_archived,
                    years_experience, sector, phone, email, dossier_competence_pdf,
                    prenom, titre, annees_experience, domaine_principal,
                    description_push,
                    disponibilite, mobilite, permis_conduire, vehicule, permis_travail,
                    fonctions_recherchees, motif_recherche, avancement_recherches,
                    remuneration_actuelle, pretentions_salariales, propal_a,
                    eval_technique, eval_personnalite, eval_communication,
                    langues, references_candidat, avis_perso,
                    entretien_date, entretien_lieu, entretien_notes,
                    createdAt, updatedAt, owner_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                ''',
                (
                    name,
                    _t("role"),
                    _t("location"),
                    _t("seniority"),
                    _t("tech"),
                    _t("linkedin"),
                    _t("source"),
                    _t("status"),
                    _t("notes"),
                    _t("onenote_url"),
                    _t("vsa_url"),
                    skills_json,
                    company_ids_json,
                    is_archived,
                    years_experience,
                    _t("sector"),
                    _t("phone"),
                    _t("email"),
                    _t("dossier_competence_pdf"),
                    _t("prenom"),
                    _t("titre"),
                    annees_exp_v,
                    _t("domaine_principal"),
                    _t("description_push"),
                    _t("disponibilite"),
                    _t("mobilite"),
                    permis_conduire_v,
                    vehicule_v,
                    _t("permis_travail"),
                    _t("fonctions_recherchees"),
                    _t("motif_recherche"),
                    _t("avancement_recherches"),
                    _t("remuneration_actuelle"),
                    _t("pretentions_salariales"),
                    _t("propal_a"),
                    _t("eval_technique"),
                    _t("eval_personnalite"),
                    _t("eval_communication"),
                    _t("langues"),
                    _t("references_candidat"),
                    _t("avis_perso"),
                    _t("entretien_date"),
                    _t("entretien_lieu"),
                    _t("entretien_notes"),
                    now,
                    now,
                    uid,
                ),
            )
            cid = cur.lastrowid

        # Candidate KPI events (contacted / solid) based on status transition
        try:
            _maybe_log_candidate_events(conn, int(cid), old_status, _t("status"), event_date)
        except Exception:
            pass

    log_activity(_candidate_action, 'candidat', cid, name)
    return jsonify({"ok": True, "id": cid})


@candidates_bp.get("/api/candidates/fiche-entretien-template")
def api_candidates_fiche_entretien_template():
    """Télécharge le modèle de fiche entretien Excel.

    Si un fichier modèle existe dans docs/sample/, on le sert. Sinon on
    génère un modèle minimal à la volée pour que l'endpoint réponde toujours."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    # Cherche un fichier modèle existant (peut être nommé différemment)
    candidates_paths = [
        APP_DIR / "docs" / "Fiche entretien NEW Prenom NOM - EC1 XXX  JJMMAAAA.xlsx",
        APP_DIR / "sample" / "fiche_entretien.xlsx",
    ]
    for p in candidates_paths:
        if p.exists():
            return send_file(str(p), as_attachment=True, download_name="fiche_entretien_Up.xlsx")

    # Fallback : génération à la volée d'un template minimal
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiche entretien"

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="EFF2F8", end_color="EFF2F8", fill_type="solid")
    section_font = Font(bold=True, size=11)

    ws["A1"] = "Fiche entretien candidat — Up Technologies"
    ws["A1"].font = Font(bold=True, size=14, color="2563EB")
    ws.merge_cells("A1:B1")

    rows = [
        ("", ""),
        ("IDENTITÉ", ""),
        ("Prénom", ""),
        ("Nom", ""),
        ("Email", ""),
        ("Téléphone", ""),
        ("LinkedIn", ""),
        ("Localisation", ""),
        ("", ""),
        ("PROFIL PROFESSIONNEL", ""),
        ("Métier / Poste recherché", ""),
        ("Expérience (années)", ""),
        ("Compétences techniques principales", ""),
        ("Langues", ""),
        ("", ""),
        ("DISPONIBILITÉ & MOBILITÉ", ""),
        ("Disponibilité", ""),
        ("TJM / Salaire visé", ""),
        ("Mobilité géographique", ""),
        ("Type de contrat souhaité", ""),
        ("", ""),
        ("ENTRETIEN EC1", ""),
        ("Date d'entretien", ""),
        ("Recruteur", ""),
        ("Notes EC1", ""),
        ("Décision", ""),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if label and not value and label.isupper():
            ws.cell(row=i, column=1).fill = section_fill
            ws.cell(row=i, column=1).font = section_font

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="fiche_entretien_Up.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@candidates_bp.post("/api/candidates/parse-fiche-entretien")
def api_candidates_parse_fiche_entretien():
    """Parse une fiche entretien Excel (format Up Technologies) via Ollama.
    Retourne un JSON avec les champs extraits pour pré-remplir la fiche candidat.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify(ok=False, error="Fichier manquant"), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify(ok=False, error="Format non supporté (xlsx/xls requis)"), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                non_none = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if non_none:
                    lines.append(" | ".join(non_none))
        excel_text = "\n".join(lines[:150])  # limiter à 150 lignes
    except Exception as e:
        return jsonify(ok=False, error=f"Erreur lecture Excel : {e}"), 400

    prompt = f"""Tu es un assistant RH expert. Voici le contenu brut d'une fiche d'entretien candidat extraite d'un fichier Excel :

---
{excel_text}
---

Extrait les informations suivantes et retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après, sans balises markdown :
{{
  "disponibilite": "date ou délai de disponibilité du candidat (ex: Immédiate, 1 mois de préavis, 3 mois)",
  "mobilite": "zones géographiques acceptées (ex: Lyon, Paris, Grenoble, Nationale, Internationale)",
  "permis_conduire": 1 ou 0,
  "vehicule": 1 ou 0,
  "permis_travail": "type de permis de travail si mentionné (ex: Salarié, CDI, indépendant)",
  "fonctions_recherchees": "postes et secteurs recherchés",
  "motif_recherche": "motif de départ / motivations",
  "avancement_recherches": "avancement des autres pistes (ex: ED, EP, Std By, actif discret)",
  "remuneration_actuelle": "rémunération actuelle (fixe + variable + avantages)",
  "pretentions_salariales": "prétentions salariales souhaitées",
  "propal_a": "salaire proposé par l'entreprise si mentionné",
  "eval_technique": "évaluation technique (note ou commentaire)",
  "eval_personnalite": "évaluation personnalité",
  "eval_communication": "évaluation communication",
  "langues": "langues et niveaux (ex: Anglais B2 testé, Espagnol A1)",
  "references_candidat": "références transmises (nom, fonction, société, contact)",
  "avis_perso": "avis personnel du consultant sur le candidat — texte libre de la case Détails / commentaires de la section Évaluation"
}}

Si une information est absente, mets null pour les champs numériques et \"\" pour les champs texte."""

    try:
        config = _load_ai_config()
        timeout = int(config.get("ollama_timeout", 120))
        raw = _call_ollama_direct(prompt, config, timeout)
        # Extraire le JSON de la réponse
        json_match = re.search(r'\{[\s\S]*\}', raw)
        if not json_match:
            return jsonify(ok=False, error="L'IA n'a pas retourné de JSON valide", raw=raw[:500]), 422
        parsed = json.loads(json_match.group(0))
        return jsonify(ok=True, fields=parsed)
    except json.JSONDecodeError as e:
        return jsonify(ok=False, error=f"JSON invalide retourné par l'IA : {e}", raw=raw[:500] if 'raw' in dir() else ""), 422
    except Exception as e:
        return jsonify(ok=False, error=f"Erreur IA : {e}"), 500


@candidates_bp.post("/api/candidates/delete")
def api_candidates_delete():
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=False) or {}
    cid = payload.get("id")
    if not cid:
        return jsonify({"ok": False, "error": "id is required"}), 400
    _cand_name = None
    with _conn() as conn:
        _cand_row = conn.execute("SELECT name FROM candidates WHERE id=? AND owner_id=?;", (int(cid), uid)).fetchone()
        _cand_name = _cand_row["name"] if _cand_row else None
        # v23.5: soft delete instead of hard delete
        conn.execute("UPDATE candidates SET deleted_at=? WHERE id=? AND owner_id=?;", (_now_iso(), int(cid), uid))
    _audit_log("soft_delete", "candidate", int(cid))
    log_activity('delete', 'candidat', int(cid), _cand_name)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════
# v27.4: Helper extraction texte PDF — voir utils/files.py (doublon supprimé en phase A3)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# v32.84 — Extraction texte multi-format + enrichissement IA de fiche
# ═══════════════════════════════════════════════════════════════════

# Extensions dont on sait extraire le texte pour l'enrichissement IA.
_ENRICHABLE_EXTS = {".pdf", ".docx", ".txt", ".xlsx", ".csv"}

# Champs candidat que l'IA peut renseigner depuis un document. Tous présents
# dans la liste ALLOWED de api_candidate_put (sinon le PUT les ignorerait).
_DOC_ENRICH_KEYS = (
    "name", "prenom", "role", "location", "sector", "domaine_principal",
    "tech", "years_experience", "phone", "email", "linkedin", "langues",
    "disponibilite", "mobilite", "permis_travail", "pretentions_salariales",
    "remuneration_actuelle", "motif_recherche", "fonctions_recherchees",
    "eval_technique", "eval_personnalite", "eval_communication",
)


def _extract_text_from_file(file_path: Path, ext: str, max_chars: int = 8000) -> str:
    """Extrait le texte d'un document (PDF, DOCX, TXT, XLSX, CSV) pour l'IA.

    Retourne une chaîne vide si le format est non supporté ou en cas d'échec.
    """
    ext = (ext or "").lower()
    try:
        if ext == ".pdf":
            return _extract_pdf_text(file_path, max_chars=max_chars)
        if ext == ".docx":
            from docx import Document as _Docx
            doc = _Docx(str(file_path))
            parts = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
            for table in doc.tables:
                for trow in table.rows:
                    cells = [c.text.strip() for c in trow.cells if c.text and c.text.strip()]
                    if cells:
                        parts.append(" | ".join(cells))
            return "\n".join(parts)[:max_chars].strip()
        if ext in (".txt", ".csv"):
            raw = file_path.read_bytes()
            for enc in ("utf-8", "cp1252", "latin-1"):
                try:
                    return raw.decode(enc)[:max_chars].strip()
                except UnicodeDecodeError:
                    continue
            return raw.decode("utf-8", errors="ignore")[:max_chars].strip()
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            lines, total = [], 0
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v).strip() for v in row if v not in (None, "")]
                    if not vals:
                        continue
                    line = " | ".join(vals)
                    lines.append(line)
                    total += len(line) + 1
                    if total > max_chars:
                        break
                if total > max_chars:
                    break
            wb.close()
            return "\n".join(lines)[:max_chars].strip()
    except Exception as e:
        logger.warning("_extract_text_from_file(%s) error: %s", file_path, e)
    return ""


def _ai_extract_candidate_fields(doc_text: str, source_label: str = "document",
                                 timeout: int = 120) -> dict:
    """Analyse le texte d'un document via l'IA locale et renvoie les champs candidat.

    Lève json.JSONDecodeError / ValueError si la réponse IA est inexploitable,
    ou propage l'exception réseau si l'IA est indisponible.
    """
    doc_text = (doc_text or "").strip()
    if not doc_text:
        return {}
    prompt = f"""Tu es un assistant RH expert. Analyse le {source_label} ci-dessous (il peut s'agir d'un CV, d'un dossier de compétences, d'une fiche d'entretien ou d'un profil LinkedIn) et extrais les informations concernant le candidat.

--- DÉBUT DU DOCUMENT ---
{doc_text}
--- FIN DU DOCUMENT ---

Retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après, sans balise markdown. Mets null pour toute information réellement absente du document — n'invente jamais une valeur :
{{
  "name": "Prénom NOM complet du candidat",
  "prenom": "Prénom seul",
  "role": "Poste ou titre principal (ex: Ingénieur Systèmes Embarqués)",
  "location": "Ville ou région de résidence",
  "sector": "Secteur d'activité principal (ex: Industrie, Défense, IT)",
  "domaine_principal": "Domaine d'expertise détaillé",
  "tech": "Compétences techniques principales, séparées par des virgules",
  "years_experience": <nombre entier d'années d'expérience professionnelle, ou null>,
  "phone": "Numéro de téléphone",
  "email": "Adresse email",
  "linkedin": "URL complète du profil LinkedIn",
  "langues": "Langues parlées et niveaux (ex: Français natif, Anglais courant)",
  "disponibilite": "Disponibilité ou date de prise de poste",
  "mobilite": "Zones de mobilité géographique",
  "permis_travail": "Permis de travail / nationalité si le document le précise",
  "pretentions_salariales": "Prétentions salariales",
  "remuneration_actuelle": "Rémunération actuelle",
  "motif_recherche": "Motivation de la recherche ou du changement de poste",
  "fonctions_recherchees": "Postes ou fonctions recherchés",
  "eval_technique": "Synthèse de l'évaluation technique (seulement si fiche d'entretien)",
  "eval_personnalite": "Synthèse personnalité / savoir-être (seulement si fiche d'entretien)",
  "eval_communication": "Synthèse de la communication (seulement si fiche d'entretien)"
}}"""
    result_text = _call_ai(prompt, timeout=timeout)
    clean = (result_text or "").strip()
    if clean.startswith("```"):
        clean = re.sub(r'^```[^\n]*\n?', '', clean)
        clean = re.sub(r'\n?```$', '', clean)
    json_match = re.search(r'\{[\s\S]*\}', clean)
    if not json_match:
        raise ValueError("La réponse de l'IA ne contient pas de JSON exploitable")
    raw = json.loads(json_match.group(0))
    if not isinstance(raw, dict):
        raise ValueError("La réponse de l'IA n'est pas un objet JSON")
    fields: dict = {}
    for key in _DOC_ENRICH_KEYS:
        if key not in raw or raw[key] is None:
            continue
        val = str(raw[key]).strip()
        if val and val.lower() not in ("null", "none", "n/a", "na", "-", "—", "non précisé"):
            fields[key] = val
    return fields


# ═══════════════════════════════════════════════════════════════════
# v27.x PARTIE 3: Extraction DC PDF + upload DC
# ═══════════════════════════════════════════════════════════════════

@candidates_bp.post("/api/candidates/extract-dc")
def api_candidates_extract_dc():
    """Extrait les champs d'un DC PDF via Ollama (local, données confidentielles).
    Retourne: { ok, fields: { name, prenom, titre, annees_experience, domaine_principal, tags, role } }
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    if 'dc' not in request.files:
        return jsonify(ok=False, error="Champ 'dc' manquant"), 400

    pdf_file = request.files['dc']
    if not pdf_file.filename:
        return jsonify(ok=False, error="Nom de fichier vide"), 400

    # Lire le PDF et extraire le texte
    pdf_text = ""
    try:
        import io
        pdf_bytes = pdf_file.read()

        # Tenter avec pdfminer (souvent disponible)
        try:
            from pdfminer.high_level import extract_text as _extract_pdf  # type: ignore
            pdf_text = _extract_pdf(io.BytesIO(pdf_bytes), maxpages=5) or ""
        except ImportError:
            pass

        # Fallback: pypdf
        if not pdf_text:
            try:
                import pypdf  # type: ignore
                reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
                for page in reader.pages[:5]:
                    pdf_text += page.extract_text() or ""
            except ImportError:
                pass

        if not pdf_text.strip():
            return jsonify(ok=False, error="Impossible d'extraire le texte du PDF (bibliothèque PDF manquante)"), 422
    except Exception as e:
        return jsonify(ok=False, error=f"Erreur lecture PDF: {e}"), 500

    # Limiter le texte pour Ollama
    pdf_text_short = pdf_text[:4000]

    prompt = f"""Tu es un assistant qui extrait des informations d'un dossier de compétences (CV) d'un ingénieur.
Dossier de compétences (extrait) :
{pdf_text_short}

Extrait les informations suivantes au format JSON strict (sans commentaire, sans markdown) :
{{
  "name": "Prénom NOM",
  "prenom": "Prénom",
  "titre": "Titre ou poste principal",
  "annees_experience": <nombre entier d'années d'expérience ou null>,
  "domaine_principal": "domaine principal (ex: Systèmes embarqués, Automotive, Défense...)",
  "tags": ["tag1", "tag2", ...],
  "role": "rôle court pour la liste (ex: Embedded Software Engineer)"
}}
Réponds uniquement avec le JSON, sans aucun texte autour."""

    try:
        result_text = _call_ai(prompt, timeout=60)
        # Parser le JSON
        # Nettoyer les éventuels marqueurs markdown
        clean = result_text.strip()
        if clean.startswith("```"):
            clean = re.sub(r'^```[^\n]*\n?', '', clean)
            clean = re.sub(r'\n?```$', '', clean)
        fields = json.loads(clean)
        return jsonify(ok=True, fields=fields)
    except (json.JSONDecodeError, ValueError) as e:
        logger.warning("DC extraction JSON parse error: %s — response: %s", e, result_text[:200] if 'result_text' in dir() else '')
        return jsonify(ok=False, error="L'IA n'a pas retourné un JSON valide"), 422
    except Exception as e:
        logger.warning("DC extraction error: %s", e)
        return jsonify(ok=False, error=f"IA indisponible: {e}"), 503


@candidates_bp.post("/api/candidates/<int:cid>/dc-enrich")
def api_candidate_dc_enrich(cid):
    """Lit le DC PDF existant sur disque et l'analyse via IA pour enrichir la fiche.
    Retourne: { ok, fields: {...} }. La validation se fait côté client.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT name FROM candidates WHERE id=? AND owner_id=? AND deleted_at IS NULL;",
            (cid, uid)
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Candidat introuvable"), 404

    dc_dir = DATA_DIR / "dossiers_candidats" / str(uid) / str(cid)
    pdf_files = sorted(dc_dir.glob("*.pdf")) if dc_dir.is_dir() else []
    if not pdf_files:
        return jsonify(ok=False, error="Aucun DC trouvé pour ce candidat. Chargez un PDF d'abord."), 404

    pdf_text = _extract_text_from_file(pdf_files[0], ".pdf", max_chars=8000)
    if not pdf_text or len(pdf_text) < 20:
        return jsonify(ok=False, error="Impossible d'extraire le texte du PDF (PDF scanné, protégé ou bibliothèque PDF manquante)."), 422

    try:
        fields = _ai_extract_candidate_fields(pdf_text, source_label="dossier de compétences")
    except (json.JSONDecodeError, ValueError) as e:
        logger.warning("DC enrich JSON parse error cid=%s: %s", cid, e)
        return jsonify(ok=False, error="L'IA n'a pas retourné un résultat exploitable."), 422
    except Exception as e:
        logger.warning("DC enrich error cid=%s: %s", cid, e)
        return jsonify(ok=False, error=f"IA indisponible : {e}"), 503

    return jsonify(ok=True, fields=fields)


@candidates_bp.post("/api/candidates/upload-dc")
def api_candidates_upload_dc():
    """Sauvegarde le DC PDF d'un candidat dans data/dossiers_candidats/{uid}/{cid}/.
    Attend multipart/form-data: 'dc' (PDF) + 'candidate_id' (int).
    Sécurisé : un user ne peut uploader que pour ses propres candidats.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401

    if 'dc' not in request.files:
        return jsonify(ok=False, error="Champ 'dc' manquant"), 400

    candidate_id = request.form.get("candidate_id")
    if not candidate_id:
        return jsonify(ok=False, error="candidate_id manquant"), 400

    try:
        cid_int = int(candidate_id)
    except ValueError:
        return jsonify(ok=False, error="candidate_id invalide"), 400

    # Vérifier que le candidat appartient à l'utilisateur
    with _conn() as conn:
        cand_row = conn.execute(
            "SELECT id, name FROM candidates WHERE id=? AND owner_id=? AND deleted_at IS NULL;",
            (cid_int, uid)
        ).fetchone()
    if not cand_row:
        return jsonify(ok=False, error="Candidat introuvable"), 404

    pdf_file = request.files['dc']
    if not pdf_file.filename:
        return jsonify(ok=False, error="Nom de fichier vide"), 400

    filename = "".join(c for c in pdf_file.filename if c.isalnum() or c in "._- ")
    if not filename.lower().endswith('.pdf'):
        return jsonify(ok=False, error="Seuls les fichiers PDF sont acceptés"), 400

    # Dossier sécurisé par user + candidat
    dc_dir = DATA_DIR / "dossiers_candidats" / str(uid) / str(cid_int)
    dc_dir.mkdir(parents=True, exist_ok=True)
    target = dc_dir / filename

    # Vérification path traversal
    try:
        if not str(target.resolve()).startswith(str((DATA_DIR / "dossiers_candidats" / str(uid)).resolve())):
            return jsonify(ok=False, error="Chemin invalide"), 403
    except Exception:
        return jsonify(ok=False, error="Chemin invalide"), 403

    try:
        pdf_file.save(str(target))
        # Mettre à jour le champ dossier_competence_pdf du candidat
        with _conn() as conn:
            conn.execute(
                "UPDATE candidates SET dossier_competence_pdf=?, updatedAt=? WHERE id=? AND owner_id=?;",
                (str(target), _now_iso(), cid_int, uid)
            )
        logger.info("DC uploadé: user=%s cand=%s file=%s", uid, cid_int, filename)
        return jsonify(ok=True, filename=filename, path=str(target))
    except Exception as e:
        logger.error("Erreur upload DC: %s", e)
        return jsonify(ok=False, error=str(e)), 500


@candidates_bp.get("/api/candidates/<int:cid>/dc-status")
def api_candidate_dc_status(cid):
    """Vérifie si un DC est présent pour le candidat.
    Retourne: { ok, has_dc, files: [filename], generated: [{id, filename, generated_at, download_url}] }
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT dossier_competence_pdf FROM candidates WHERE id=? AND owner_id=? AND deleted_at IS NULL;",
            (cid, uid)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Candidat introuvable"), 404
        # DC générés via le générateur (table dc_generations)
        gen_rows = conn.execute(
            "SELECT id, filename, generated_at, used_ollama FROM dc_generations "
            "WHERE candidate_id=? AND owner_id=? AND deleted_at IS NULL "
            "ORDER BY generated_at DESC LIMIT 5;",
            (cid, uid)
        ).fetchall()

    dc_dir = DATA_DIR / "dossiers_candidats" / str(uid) / str(cid)
    files = sorted([f.name for f in dc_dir.glob("*.pdf")]) if dc_dir.is_dir() else []
    # Fallback: ancien champ texte
    if not files and row["dossier_competence_pdf"]:
        files = [Path(row["dossier_competence_pdf"]).name]

    generated = []
    for g in (gen_rows or []):
        try:
            gen_dt = datetime.datetime.fromisoformat(g["generated_at"])
            gen_label = gen_dt.strftime('%d/%m/%Y à %H:%M')
        except Exception:
            gen_label = g["generated_at"] or ''
        generated.append({
            'id':           g["id"],
            'filename':     g["filename"] or '',
            'generated_at': gen_label,
            'used_ollama':  bool(g["used_ollama"]),
            'download_url': f'/api/dc/{g["id"]}/download',
        })

    has_dc = bool(files) or bool(generated)
    return jsonify(ok=True, has_dc=has_dc, files=files, generated=generated)


@candidates_bp.post("/api/candidates/<int:cid>/dc-rename")
def api_candidate_dc_rename(cid):
    """Renomme le fichier DC d'un candidat sur le disque et met à jour la DB."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    payload = request.get_json(force=True, silent=True) or {}
    new_name = (payload.get("new_name") or "").strip()
    if not new_name:
        return jsonify(ok=False, error="Nouveau nom manquant"), 400
    # Sécurité : nom de fichier simple, sans chemin
    new_name = "".join(c for c in new_name if c.isalnum() or c in "._- ")
    if not new_name.lower().endswith(".pdf"):
        new_name += ".pdf"

    with _conn() as conn:
        row = conn.execute(
            "SELECT dossier_competence_pdf FROM candidates WHERE id=? AND owner_id=? AND deleted_at IS NULL;",
            (cid, uid)
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Candidat introuvable"), 404

    dc_dir = DATA_DIR / "dossiers_candidats" / str(uid) / str(cid)
    # Trouver le fichier existant
    existing_files = sorted(dc_dir.glob("*.pdf")) if dc_dir.is_dir() else []
    if not existing_files:
        return jsonify(ok=False, error="Aucun fichier DC trouvé"), 404

    old_file = existing_files[0]
    new_file = dc_dir / new_name

    # Vérification path traversal
    try:
        base = str((DATA_DIR / "dossiers_candidats" / str(uid)).resolve())
        if not str(new_file.resolve()).startswith(base):
            return jsonify(ok=False, error="Chemin invalide"), 403
    except Exception:
        return jsonify(ok=False, error="Chemin invalide"), 403

    try:
        old_file.rename(new_file)
        with _conn() as conn:
            conn.execute(
                "UPDATE candidates SET dossier_competence_pdf=?, updatedAt=? WHERE id=? AND owner_id=?;",
                (str(new_file), _now_iso(), cid, uid)
            )
        logger.info("DC renommé: user=%s cand=%s %s→%s", uid, cid, old_file.name, new_name)
        return jsonify(ok=True, filename=new_name)
    except Exception as e:
        logger.error("Erreur renommage DC: %s", e)
        return jsonify(ok=False, error=str(e)), 500


@candidates_bp.post("/api/candidates/<int:cid>/dc-delete")
def api_candidate_dc_delete(cid):
    """Supprime le(s) fichier(s) DC d'un candidat du disque et efface le champ en DB."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT id FROM candidates WHERE id=? AND owner_id=? AND deleted_at IS NULL;",
            (cid, uid)
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Candidat introuvable"), 404

    dc_dir = DATA_DIR / "dossiers_candidats" / str(uid) / str(cid)
    deleted = []
    if dc_dir.is_dir():
        for f in dc_dir.glob("*.pdf"):
            try:
                f.unlink()
                deleted.append(f.name)
            except Exception as e:
                logger.error("Erreur suppression DC %s: %s", f, e)

    with _conn() as conn:
        conn.execute(
            "UPDATE candidates SET dossier_competence_pdf=NULL, updatedAt=? WHERE id=? AND owner_id=?;",
            (_now_iso(), cid, uid)
        )
    logger.info("DC supprimé: user=%s cand=%s files=%s", uid, cid, deleted)
    return jsonify(ok=True, deleted=deleted)


# ═══════════════════════════════════════════════════════════════════
# v32.75 — Pièces jointes candidat (CV, fiche entretien, suivi…)
# ═══════════════════════════════════════════════════════════════════

_ALLOWED_ATT_KINDS = {"cv", "ec1", "suivi", "autre"}


@candidates_bp.post("/api/candidates/<int:cid>/attachments")
def api_candidate_attachment_upload(cid):
    """Upload d'une pièce jointe pour un candidat (autre que le DC).

    multipart/form-data : file (obligatoire), title, kind (cv|ec1|suivi|autre), description.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify(ok=False, error="Fichier manquant"), 400

    ok_upload, err_upload = _validate_upload(file, "prospect_attachment")
    if not ok_upload:
        msg, code = err_upload
        return jsonify(ok=False, error=msg), code

    title = (request.form.get("title") or "").strip()[:120] or None
    description = (request.form.get("description") or "").strip()[:500] or None
    kind_raw = (request.form.get("kind") or "autre").strip().lower()
    kind = kind_raw if kind_raw in _ALLOWED_ATT_KINDS else "autre"

    import uuid as _uuid
    ext = os.path.splitext(file.filename or "")[1].lower()
    safe_orig = os.path.basename(file.filename or "").strip() or "fichier"
    stored_name = f"{_uuid.uuid4().hex}{ext}"

    attach_dir = _candidate_attachment_dir(uid, cid)
    target = attach_dir / stored_name
    # Path traversal guard
    try:
        base = str((Path("data") / f"user_{uid}" / "candidate_attachments").resolve())
        if not str(target.resolve()).startswith(base):
            return jsonify(ok=False, error="Chemin invalide"), 403
    except Exception:
        return jsonify(ok=False, error="Chemin invalide"), 403

    data = file.read()
    target.write_bytes(data)

    now = _now_iso()
    mime = file.mimetype or ""

    with _conn() as conn:
        cur = conn.execute(
            """INSERT INTO candidate_attachments
               (candidate_id, owner_id, filename, original_name, size, mime_type,
                description, title, kind, createdAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (cid, uid, stored_name, safe_orig, len(data), mime, description, title, kind, now),
        )
        att_id = cur.lastrowid

    logger.info("Pièce jointe candidat: user=%s cand=%s kind=%s file=%s", uid, cid, kind, safe_orig)
    return jsonify(ok=True, id=att_id, original_name=safe_orig, size=len(data),
                   createdAt=now, kind=kind, title=title)


@candidates_bp.get("/api/candidates/<int:cid>/attachments")
def api_candidate_attachment_list(cid):
    """Liste des pièces jointes (hors DC) d'un candidat."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403
    with _conn() as conn:
        rows = conn.execute(
            "SELECT id, filename, original_name, size, mime_type, description, title, kind, createdAt "
            "FROM candidate_attachments WHERE candidate_id=? AND owner_id=? ORDER BY createdAt DESC;",
            (cid, uid),
        ).fetchall()
    out = [dict(r) for r in rows]
    return jsonify(ok=True, attachments=out)


@candidates_bp.get("/api/candidate-attachments/<int:att_id>/file")
def api_candidate_attachment_file(att_id):
    """Sert le fichier d'une pièce jointe candidat (download)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT * FROM candidate_attachments WHERE id=? AND owner_id=?;",
            (att_id, uid),
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Pièce jointe introuvable"), 404

    attach_dir = _candidate_attachment_dir(uid, row["candidate_id"])
    file_path = attach_dir / row["filename"]
    if not file_path.exists():
        return jsonify(ok=False, error="Fichier introuvable"), 404

    inline_mimes = {"application/pdf", "image/jpeg", "image/png", "image/webp", "text/plain"}
    disposition = "inline" if row["mime_type"] in inline_mimes else "attachment"
    return send_file(
        str(file_path),
        mimetype=row["mime_type"] or "application/octet-stream",
        as_attachment=(disposition == "attachment"),
        download_name=row["original_name"],
    )


@candidates_bp.patch("/api/candidate-attachments/<int:att_id>")
def api_candidate_attachment_update(att_id):
    """Met à jour les métadonnées d'une pièce jointe candidat (title, description, kind)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    body = request.get_json(force=True, silent=True) or {}
    sets, vals = [], []
    if "title" in body:
        v = (body.get("title") or "").strip()[:120] or None
        sets.append("title=?"); vals.append(v)
    if "description" in body:
        v = (body.get("description") or "").strip()[:500] or None
        sets.append("description=?"); vals.append(v)
    if "kind" in body:
        v = (body.get("kind") or "autre").strip().lower()
        if v not in _ALLOWED_ATT_KINDS:
            v = "autre"
        sets.append("kind=?"); vals.append(v)
    if not sets:
        return jsonify(ok=False, error="Aucun champ à modifier"), 400
    vals.extend([att_id, uid])
    with _conn() as conn:
        cur = conn.execute(
            f"UPDATE candidate_attachments SET {', '.join(sets)} WHERE id=? AND owner_id=?;",
            vals,
        )
        if cur.rowcount == 0:
            return jsonify(ok=False, error="Pièce jointe introuvable"), 404
    return jsonify(ok=True)


@candidates_bp.delete("/api/candidate-attachments/<int:att_id>")
def api_candidate_attachment_delete(att_id):
    """Supprime une pièce jointe candidat (fichier + DB)."""
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT candidate_id, filename FROM candidate_attachments WHERE id=? AND owner_id=?;",
            (att_id, uid),
        ).fetchone()
        if not row:
            return jsonify(ok=False, error="Pièce jointe introuvable"), 404
        attach_dir = _candidate_attachment_dir(uid, row["candidate_id"])
        file_path = attach_dir / row["filename"]
        try:
            if file_path.exists():
                file_path.unlink()
        except Exception as e:
            logger.warning("[candidate-attachment] delete file error att=%s: %s", att_id, e)
        conn.execute("DELETE FROM candidate_attachments WHERE id=? AND owner_id=?;", (att_id, uid))
    return jsonify(ok=True)


@candidates_bp.post("/api/candidate-attachments/<int:att_id>/enrich")
def api_candidate_attachment_enrich(att_id):
    """Analyse une pièce jointe candidat (CV, fiche d'entretien, DC, profil
    LinkedIn exporté en PDF…) via l'IA locale et renvoie les champs détectés.

    Retourne: { ok, fields: {...} } — aucun champ n'est appliqué ici, la
    validation manuelle se fait côté client dans une modale de comparaison.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    with _conn() as conn:
        row = conn.execute(
            "SELECT id, candidate_id, filename, original_name FROM candidate_attachments "
            "WHERE id=? AND owner_id=?;",
            (att_id, uid),
        ).fetchone()
    if not row:
        return jsonify(ok=False, error="Pièce jointe introuvable"), 404

    ext = os.path.splitext(row["original_name"] or row["filename"] or "")[1].lower()
    if ext not in _ENRICHABLE_EXTS:
        return jsonify(ok=False, error="Format non analysable par l'IA. Formats acceptés : "
                       "PDF, DOCX, XLSX, TXT. Pour un profil LinkedIn, exportez-le en PDF."), 415

    file_path = _candidate_attachment_dir(uid, row["candidate_id"]) / row["filename"]
    if not file_path.is_file():
        return jsonify(ok=False, error="Fichier introuvable sur le disque"), 404

    doc_text = _extract_text_from_file(file_path, ext, max_chars=8000)
    if not doc_text or len(doc_text) < 20:
        return jsonify(ok=False, error="Impossible d'extraire du texte de ce document "
                       "(fichier vide, scanné en image ou protégé)."), 422

    try:
        fields = _ai_extract_candidate_fields(doc_text, source_label="document")
    except (json.JSONDecodeError, ValueError) as e:
        logger.warning("Attachment enrich JSON parse error att=%s: %s", att_id, e)
        return jsonify(ok=False, error="L'IA n'a pas retourné un résultat exploitable."), 422
    except Exception as e:
        logger.warning("Attachment enrich error att=%s: %s", att_id, e)
        return jsonify(ok=False, error=f"IA indisponible : {e}"), 503

    return jsonify(ok=True, fields=fields)


# ═══════════════════════════════════════════════════════════════════
# v32.83 — Fiche entretien EC1 — formulaire éditable + export Excel
# Génération par manipulation directe du template .xlsx (zip) : préserve
# les cases à cocher, images, mise en forme et formules des 4 pages.
# ═══════════════════════════════════════════════════════════════════

def _ec1_template_path() -> Path | None:
    """Localise le template Excel de fiche entretien EC1."""
    for p in (
        APP_DIR / "exemples" / "Fiche entretien NEW Prenom NOM - EC1 XXX  JJMMAAAA.xlsx",
        APP_DIR / "sample" / "fiche_entretien.xlsx",
        APP_DIR / "docs" / "Fiche entretien NEW Prenom NOM - EC1 XXX  JJMMAAAA.xlsx",
    ):
        if p.exists():
            return p
    return None


def _slugify_for_filename(s: str) -> str:
    """Convertit une chaîne en nom de fichier Excel safe."""
    s = re.sub(r"[^A-Za-z0-9_\-\. ]+", "", s or "").strip()
    return re.sub(r"\s+", "_", s) or "candidat"


# ── Modèle de données de la fiche EC1 ──────────────────────────────
# Clé du formulaire → colonne candidat (uniquement les champs persistés).
_EC1_DB_FIELDS = {
    "name": "name",
    "phone": "phone",
    "email": "email",
    "source": "source",
    "ec1_date": "entretien_date",
    "permis_conduire": "permis_conduire",
    "vehicule": "vehicule",
    "permis_travail": "permis_travail",
    "disponibilite": "disponibilite",
    "mobilite": "mobilite",
    "fonctions_recherchees": "fonctions_recherchees",
    "motif_recherche": "motif_recherche",
    "remuneration_actuelle": "remuneration_actuelle",
    "pretentions_salariales": "pretentions_salariales",
    "propal_a": "propal_a",
    "avancement_recherches": "avancement_recherches",
    "eval_technique": "eval_technique",
    "eval_personnalite": "eval_personnalite",
    "eval_communication": "eval_communication",
    "langues": "langues",
    "references_candidat": "references_candidat",
    "avis_perso": "avis_perso",
    "entretien_notes": "entretien_notes",
}
# Champs stockés en entier 0/1 (Oui/Non).
_EC1_INT_FIELDS = {"permis_conduire", "vehicule"}
# Cases de la checklist EC1.
_EC1_CHECKLIST_KEYS = (
    "mobilite_dispo_souhaits", "impression_generale", "evaluation_technique",
    "evaluation_personnalite", "evaluation_communication", "rappel_valeurs_up",
    "fourchette_salaire", "reponse_questions_craintes", "process_prochaines_etapes",
)
# Zones de mobilité — calquées sur le template Excel (cases à cocher).
_EC1_MOBILITY_ZONES = (
    "Banlieue parisienne", "Lyon", "Aix", "Sophia", "Paris", "Grenoble",
    "Toulon", "Province", "Nationale", "Valence", "Montpellier", "Rennes",
    "Internationale",
)
# Zones sans case à cocher dans le template → on préfixe la cellule d'un ✓.
_EC1_MOBILITY_TEXT_CELLS = {
    "Valence": "F14", "Montpellier": "G14", "Rennes": "H14", "Internationale": "J14",
}


def _ec1_yesno_to_int(v):
    """Convertit une valeur Oui/Non du formulaire en 0/1 (None si indéterminé)."""
    s = str(v if v is not None else "").strip().lower()
    if s in ("oui", "yes", "1", "true", "o", "on"):
        return 1
    if s in ("non", "no", "0", "false", "n", "off"):
        return 0
    return None


def _ec1_yesno_label(v) -> str:
    """Convertit 0/1/Oui/Non en libellé Oui/Non (chaîne vide si indéterminé)."""
    iv = _ec1_yesno_to_int(v)
    return "Oui" if iv == 1 else ("Non" if iv == 0 else "")


def _ec1_fmt_date(s) -> str:
    """Formatte une date ISO en JJ/MM/AAAA pour l'Excel."""
    s = str(s or "").strip()
    if not s:
        return ""
    try:
        return datetime.datetime.fromisoformat(s).strftime("%d/%m/%Y")
    except Exception:
        return s[:10]


def _ec1_recruteur_trigramme(user) -> str:
    """Trigramme recruteur déduit du username."""
    try:
        uname = (user["username"] if user else "") or ""
    except Exception:
        uname = ""
    return uname[:3].upper() if uname else "XXX"


def _ec1_diplomes_default(c: dict) -> str:
    """Ligne « Diplômes et expérience » déduite de la fiche candidat."""
    dip = c.get("titre") or c.get("role") or ""
    annees = c.get("annees_experience") or c.get("years_experience") or c.get("seniority") or ""
    s = str(dip or "").strip()
    if annees:
        s = (s + f" — {annees} ans").strip(" —")
    return s


def _ec1_zone_set(mobilite) -> set:
    """Parse la chaîne `mobilite` (zones séparées par des virgules) en set de
    libellés normalisés sur la liste de référence."""
    out = set()
    if not mobilite:
        return out
    parts = [p.strip() for p in str(mobilite).replace(";", ",").split(",")]
    lower_ref = {z.lower(): z for z in _EC1_MOBILITY_ZONES}
    for p in parts:
        if not p:
            continue
        z = lower_ref.get(p.lower())
        if z:
            out.add(z)
    return out


def _ec1_load_context(cid: int, uid: int):
    """Charge le candidat, la checklist EC1 et l'utilisateur courant.

    Retour : (candidat_dict | None, ec1_data, interview_at, user)
    """
    with _conn() as conn:
        cand = conn.execute(
            "SELECT * FROM candidates WHERE id=? AND owner_id=? AND deleted_at IS NULL;",
            (cid, uid),
        ).fetchone()
        if not cand:
            return None, {}, None, None
        ec1_row = conn.execute(
            "SELECT interviewAt, data FROM candidate_ec1_checklists WHERE candidate_id=?;",
            (cid,),
        ).fetchone()
    ec1_data = {}
    if ec1_row and ec1_row["data"]:
        try:
            ec1_data = json.loads(ec1_row["data"]) or {}
        except Exception:
            ec1_data = {}
    interview_at = ec1_row["interviewAt"] if ec1_row else None
    user = None
    try:
        from utils.db import _auth_conn
        with _auth_conn() as aconn:
            user = aconn.execute(
                "SELECT username, display_name FROM users WHERE id=?;", (uid,),
            ).fetchone()
    except Exception:
        user = None
    return dict(cand), ec1_data, interview_at, user


def _ec1_data_from_db(c: dict, ec1_data: dict, user, interview_at=None) -> dict:
    """Construit le dictionnaire plat de la fiche EC1 depuis la base."""
    ec1_data = ec1_data or {}
    ec1_date = c.get("entretien_date") or interview_at or ""
    return {
        "name": c.get("name") or "",
        "phone": c.get("phone") or "",
        "email": c.get("email") or "",
        "diplomes_experience": _ec1_diplomes_default(c),
        "date_lieu_naissance": "",
        "etat_civil": "",
        "source": c.get("source") or "",
        "recruteur_trigramme": _ec1_recruteur_trigramme(user),
        "ec1_date": str(ec1_date)[:10],
        "permis_conduire": _ec1_yesno_label(c.get("permis_conduire")),
        "vehicule": _ec1_yesno_label(c.get("vehicule")),
        "permis_travail": c.get("permis_travail") or "",
        "demarches_administratives": "",
        "disponibilite": c.get("disponibilite") or "",
        "domicile": "",
        "mobilite": c.get("mobilite") or "",
        "fonctions_recherchees": c.get("fonctions_recherchees") or "",
        "motif_recherche": c.get("motif_recherche") or "",
        "remuneration_actuelle": c.get("remuneration_actuelle") or "",
        "pretentions_salariales": c.get("pretentions_salariales") or "",
        "propal_a": c.get("propal_a") or "",
        "mail_recap": "",
        "montant_recap": "",
        "avancement_recherches": c.get("avancement_recherches") or "",
        "eval_technique": c.get("eval_technique") or "",
        "eval_personnalite": c.get("eval_personnalite") or "",
        "eval_communication": c.get("eval_communication") or "",
        "ec1_statut": c.get("status") or "",
        "avis_perso": c.get("avis_perso") or "",
        "langues": c.get("langues") or "",
        "references_candidat": c.get("references_candidat") or "",
        "entretien_notes": c.get("entretien_notes") or "",
        "__checklist": ec1_data,
        "__free_note": str(ec1_data.get("__note") or "") if isinstance(ec1_data, dict) else "",
    }


# ── Génération Excel par manipulation directe du zip .xlsx ──────────

def _ec1_xml_escape(s: str) -> str:
    """Échappe une valeur texte pour insertion dans du XML."""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _ec1_set_cell(sheet_xml: str, ref: str, value: str) -> str:
    """Écrit `value` (chaîne) dans la cellule `ref` du XML de la feuille,
    sous forme d'inline string. Insère la cellule si elle est absente."""
    value = _ec1_xml_escape(value)
    pat = r'<c r="' + ref + r'"( s="\d+")?[^>]*?(/>|>.*?</c>)'
    m = re.search(pat, sheet_xml, re.S)
    if m:
        style = m.group(1) or ""
        new = ('<c r="' + ref + '"' + style + ' t="inlineStr"><is>'
               '<t xml:space="preserve">' + value + '</t></is></c>')
        return sheet_xml[:m.start()] + new + sheet_xml[m.end():]
    # Cellule absente → on l'insère en tête de sa ligne (colonne A).
    row_num = re.match(r'[A-Z]+(\d+)', ref)
    if row_num:
        rm = re.search(r'<row r="' + row_num.group(1) + r'"[^>]*>', sheet_xml)
        if rm:
            cell = ('<c r="' + ref + '" t="inlineStr"><is>'
                    '<t xml:space="preserve">' + value + '</t></is></c>')
            return sheet_xml[:rm.end()] + cell + sheet_xml[rm.end():]
    return sheet_xml


def _ec1_shared_strings(ss_xml: str) -> list:
    """Liste des chaînes partagées (texte brut) du classeur."""
    out = []
    for si in re.findall(r'<si>(.*?)</si>', ss_xml, re.S):
        out.append(re.sub(r'<[^>]+>', '', si))
    return out


def _ec1_cell_text(sheet_xml: str, ref: str, strings: list) -> str:
    """Texte affiché d'une cellule (résout les chaînes partagées)."""
    m = re.search(r'<c r="' + ref + r'"[^>]*?(?:/>|>(.*?)</c>)', sheet_xml, re.S)
    if not m:
        return ""
    seg, inner = m.group(0), m.group(1)
    if inner is None:
        return ""
    v = re.search(r'<v>(.*?)</v>', inner, re.S)
    if v and 't="s"' in seg:
        try:
            return strings[int(v.group(1))]
        except Exception:
            return ""
    if v:
        return v.group(1)
    isr = re.search(r'<t[^>]*>(.*?)</t>', inner, re.S)
    return isr.group(1) if isr else ""


def _ec1_resolve_checkboxes(sheet_xml: str, rels_xml: str, strings: list) -> dict:
    """Cartographie les cases à cocher du template.

    Chaque case est repérée par (chemin ctrlProp, shapeId) — le shapeId
    correspond à l'attribut `o:spid` du dessin VML.

    Retour : { 'permis': [(ctrlProp, shapeId)], 'vehicule': [...],
               'zones': {zone: [(ctrlProp, shapeId)]},
               'chk': {'A48': [...], 'A49': [...]} }
    """
    from openpyxl.utils import get_column_letter
    rel_map = dict(re.findall(
        r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
    res = {"permis": [], "vehicule": [], "zones": {}, "chk": {"A48": [], "A49": []}}
    zone_lower = {z.lower(): z for z in _EC1_MOBILITY_ZONES}
    pattern = (r'<control shapeId="(\d+)" r:id="(rId\d+)" name="[^"]*".*?'
               r'<from><xdr:col>(\d+)</xdr:col><xdr:colOff>\d+</xdr:colOff>'
               r'<xdr:row>(\d+)</xdr:row>')
    for shape_id, rid, col, row in re.findall(pattern, sheet_xml, re.S):
        col, row = int(col), int(row)
        target = rel_map.get(rid, "")
        if not target:
            continue
        cp = "xl/" + target.replace("../", "")
        box = (cp, shape_id)
        if col == 1 and row == 8:
            res["permis"].append(box)
        elif col == 3 and row == 8:
            res["vehicule"].append(box)
        elif col == 0 and row == 47:
            res["chk"]["A48"].append(box)
        elif col == 0 and row == 48:
            res["chk"]["A49"].append(box)
        else:
            city_ref = get_column_letter(col + 2) + str(row + 1)
            city = _ec1_cell_text(sheet_xml, city_ref, strings).strip()
            zone = zone_lower.get(city.lower())
            if zone:
                res["zones"].setdefault(zone, []).append(box)
    return res


def _ec1_build_xlsx(d: dict) -> bytes:
    """Génère la fiche EC1 .xlsx en modifiant directement le template (zip).

    Préserve toutes les cases à cocher, images et mises en forme ; coche
    permis / véhicule / zones de mobilité / checklist selon `d`.
    """
    import zipfile
    from io import BytesIO

    template = _ec1_template_path()
    if not template:
        raise RuntimeError("Template Excel introuvable")

    with zipfile.ZipFile(str(template), "r") as zin:
        infos = zin.infolist()
        parts = {n: zin.read(n) for n in zin.namelist()}

    sheet = parts["xl/worksheets/sheet1.xml"].decode("utf-8")
    rels = parts.get("xl/worksheets/_rels/sheet1.xml.rels", b"").decode("utf-8")
    strings = _ec1_shared_strings(parts.get("xl/sharedStrings.xml", b"").decode("utf-8"))

    def g(k):
        v = d.get(k)
        return "" if v is None else str(v).strip()

    # ── Valeurs texte des cellules ──
    ec1_date = g("ec1_date")
    cells = [
        ("A3", "Prénom Nom : " + g("name")),
        ("C4", g("phone")),
        ("C5", g("email")),
        ("C6", g("diplomes_experience")),
        ("C7", g("date_lieu_naissance")),
        ("C8", g("etat_civil")),
        ("K4", g("source")),
        ("G6", g("recruteur_trigramme") or "XXX"),
        ("H6", "OK" if ec1_date else ""),
        ("J6", _ec1_fmt_date(ec1_date)),
        ("B10", g("permis_travail")),
        ("A12", g("demarches_administratives")),
        ("G10", g("disponibilite")),
        ("I11", g("domicile")),
        ("A15", g("fonctions_recherchees")),
        ("F18", g("motif_recherche")),
        ("B20", g("remuneration_actuelle")),
        ("B21", g("pretentions_salariales")),
        ("B22", g("propal_a")),
        ("B23", g("mail_recap")),
        ("E23", g("montant_recap")),
        ("F21", g("avancement_recherches")),
        ("B26", g("eval_technique")),
        ("B27", g("eval_personnalite")),
        ("B28", g("eval_communication")),
        ("B29", g("ec1_statut")),
        ("D25", g("avis_perso")),
        ("B36", g("langues")),
        ("A42", g("references_candidat")),
    ]
    free_note = g("entretien_notes") or g("__free_note")
    if free_note:
        cells.append(("A52", "Notes : " + free_note[:400]))

    zones = _ec1_zone_set(d.get("mobilite"))
    for zone, ref in _EC1_MOBILITY_TEXT_CELLS.items():
        if zone in zones:
            cells.append((ref, "✓ " + zone))

    for ref, val in cells:
        if val == "":
            continue
        sheet = _ec1_set_cell(sheet, ref, val)

    # ── Cases à cocher ──
    cb = _ec1_resolve_checkboxes(sheet, rels, strings)
    to_check = []  # (ctrlProp_path, vml_id)
    if _ec1_yesno_to_int(d.get("permis_conduire")) == 1:
        to_check += cb["permis"]
    if _ec1_yesno_to_int(d.get("vehicule")) == 1:
        to_check += cb["vehicule"]
    for zone in zones:
        to_check += cb["zones"].get(zone, [])
    chk = d.get("__checklist") if isinstance(d.get("__checklist"), dict) else {}

    def _checked(key):
        return isinstance(chk.get(key), dict) and bool(chk[key].get("checked"))

    if _checked("mobilite_dispo_souhaits"):
        to_check += cb["chk"]["A48"]
    if _checked("fourchette_salaire"):
        to_check += cb["chk"]["A49"]

    parts["xl/worksheets/sheet1.xml"] = sheet.encode("utf-8")

    # ctrlProps : ajoute checked="Checked"
    shape_ids = set()
    for cp_path, shape_id in to_check:
        shape_ids.add(shape_id)
        if cp_path in parts:
            x = parts[cp_path].decode("utf-8")
            if "checked=" not in x:
                x = x.replace('objectType="CheckBox"',
                              'objectType="CheckBox" checked="Checked"', 1)
            parts[cp_path] = x.encode("utf-8")

    # VML : ajoute <x:Checked>1</x:Checked> dans le ClientData de la case.
    # Les formes VML sont repérées par leur o:spid (= shapeId du contrôle).
    vml_key = "xl/drawings/vmlDrawing1.vml"
    if shape_ids and vml_key in parts:
        vml = parts[vml_key].decode("utf-8")
        for sid in shape_ids:
            vml = re.sub(
                r'(_x0000_s' + re.escape(sid) + r'"[\s\S]*?</x:Anchor>)',
                lambda m: m.group(1) + "<x:Checked>1</x:Checked>",
                vml, count=1, flags=re.S,
            )
        parts[vml_key] = vml.encode("utf-8")

    # Force le recalcul des formules (pages 2-4) à l'ouverture.
    wb_key = "xl/workbook.xml"
    if wb_key in parts:
        wb = parts[wb_key].decode("utf-8")
        if "<calcPr" in wb and "fullCalcOnLoad" not in wb:
            wb = wb.replace("<calcPr ", '<calcPr fullCalcOnLoad="1" ', 1)
        parts[wb_key] = wb.encode("utf-8")

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            zout.writestr(info, parts[info.filename])
    return bio.getvalue()


def _ec1_persist(cid: int, uid: int, fields: dict, checklist) -> None:
    """Applique les champs EC1 éditables au candidat + enregistre la checklist."""
    fields = fields or {}
    sets, vals = [], []
    for fkey, col in _EC1_DB_FIELDS.items():
        if fkey not in fields:
            continue
        raw = fields.get(fkey)
        if fkey in _EC1_INT_FIELDS:
            iv = _ec1_yesno_to_int(raw)
            if iv is None:
                continue
            sets.append(f"{col}=?")
            vals.append(iv)
        else:
            v = "" if raw is None else str(raw).strip()
            if fkey == "name" and not v:
                continue  # ne jamais écraser le nom par du vide
            if fkey == "ec1_date":
                v = v[:10]
            sets.append(f"{col}=?")
            vals.append(v)
    with _conn() as conn:
        if sets:
            conn.execute(
                f"UPDATE candidates SET {', '.join(sets)}, updatedAt=? WHERE id=? AND owner_id=?;",
                vals + [_now_iso(), cid, uid],
            )
        if isinstance(checklist, dict) and checklist:
            row = conn.execute(
                "SELECT interviewAt, data FROM candidate_ec1_checklists WHERE candidate_id=?;",
                (cid,),
            ).fetchone()
            existing = {}
            if row and row["data"]:
                try:
                    existing = json.loads(row["data"]) or {}
                except Exception:
                    existing = {}
            merged = {k: {"checked": False, "note": ""} for k in _EC1_CHECKLIST_KEYS}
            merged["__note"] = str(existing.get("__note") or "").strip()
            for k, v in checklist.items():
                if k in _EC1_CHECKLIST_KEYS and isinstance(v, dict):
                    merged[k] = {
                        "checked": bool(v.get("checked", False)),
                        "note": str(v.get("note") or "").strip(),
                    }
            interview_at = (row["interviewAt"] if row else None) or _today_iso()
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """INSERT INTO candidate_ec1_checklists (candidate_id, interviewAt, data, updatedAt)
                   VALUES (?, ?, ?, ?)
                   ON CONFLICT(candidate_id)
                   DO UPDATE SET data=excluded.data, updatedAt=excluded.updatedAt""",
                (cid, interview_at, json.dumps(merged, ensure_ascii=False), now),
            )


@candidates_bp.route("/api/candidates/<int:cid>/ec1-export.xlsx", methods=["GET", "POST"])
def api_candidate_ec1_export(cid):
    """Génère et télécharge la fiche entretien EC1 Excel pré-remplie.

    - GET  : génère depuis la fiche candidat en base.
    - POST : génère depuis le formulaire édité ({fields, checklist}) et
             enregistre au passage les champs persistables.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403

    c, ec1_data, interview_at, user = _ec1_load_context(cid, uid)
    if c is None:
        return jsonify(ok=False, error="Candidat introuvable"), 404

    if request.method == "POST":
        body = request.get_json(force=True, silent=True) or {}
        fields = body.get("fields") if isinstance(body.get("fields"), dict) else {}
        checklist = body.get("checklist") if isinstance(body.get("checklist"), dict) else {}
        try:
            _ec1_persist(cid, uid, fields, checklist)
        except Exception as e:
            logger.warning("[ec1] persist on export failed: %s", e)
        d = _ec1_data_from_db(c, ec1_data, user, interview_at)
        for k, v in fields.items():
            d[k] = v
        d["__checklist"] = checklist or ec1_data
        d["__free_note"] = str(fields.get("entretien_notes") or d.get("__free_note") or "")
    else:
        d = _ec1_data_from_db(c, ec1_data, user, interview_at)

    try:
        xlsx_bytes = _ec1_build_xlsx(d)
    except Exception as e:
        logger.error("EC1 export failed: %s", e)
        return jsonify(ok=False, error=f"Erreur génération Excel : {e}"), 500

    from io import BytesIO
    safe_name = _slugify_for_filename(c.get("name") or "candidat")
    trig = _slugify_for_filename(d.get("recruteur_trigramme") or "XXX")
    today = datetime.datetime.now().strftime("%d%m%Y")
    download_name = f"Fiche_entretien_{safe_name}_EC1_{trig}_{today}.xlsx"
    return send_file(
        BytesIO(xlsx_bytes),
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@candidates_bp.post("/api/candidates/<int:cid>/ec1-apply")
def api_candidate_ec1_apply(cid):
    """Enregistre les champs édités de la fiche EC1 sur le candidat.

    Body JSON : { fields: {...}, checklist: {...} }
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403

    body = request.get_json(force=True, silent=True) or {}
    fields = body.get("fields")
    checklist = body.get("checklist")
    if not isinstance(fields, dict):
        return jsonify(ok=False, error="Payload invalide"), 400
    try:
        _ec1_persist(cid, uid, fields, checklist if isinstance(checklist, dict) else {})
    except Exception as e:
        logger.error("EC1 apply failed: %s", e)
        return jsonify(ok=False, error=f"Erreur : {e}"), 500
    return jsonify(ok=True, applied=True)


@candidates_bp.post("/api/candidates/<int:cid>/ec1-from-transcript")
def api_candidate_ec1_from_transcript(cid):
    """Analyse une transcription d'entretien EC1 via IA.

    Body JSON : { transcript: "..." }
    Retour : { ok, fields: {...}, checklist: {...}, candidate: {...}, meta: {...} }
    Aucune écriture en base : l'enregistrement passe par /ec1-apply ou par
    /ec1-export.xlsx (POST), après vérification dans le formulaire éditable.
    """
    uid = _uid()
    if not uid:
        return jsonify(ok=False, error="Non authentifié"), 401
    if not _candidate_owned(cid):
        return jsonify(ok=False, error="Accès refusé"), 403

    body = request.get_json(force=True, silent=True) or {}
    transcript = (body.get("transcript") or "").strip()
    if not transcript:
        return jsonify(ok=False, error="Transcription vide"), 400
    if len(transcript) > 40000:
        transcript = transcript[:40000]

    zones_list = ", ".join(_EC1_MOBILITY_ZONES)
    prompt = f"""Tu es un assistant RH expert. Voici la transcription d'un entretien EC1 (entretien de qualification candidat) chez Up Technologies.

---
{transcript}
---

Extrait les informations suivantes et retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après, sans balises markdown :
{{
  "fields": {{
    "diplomes_experience": "diplômes et années d'expérience (ex: Ingénieur ESTP, 5 ans)",
    "permis_conduire": "Oui ou Non (chaîne vide si non abordé)",
    "vehicule": "Oui ou Non (chaîne vide si non abordé)",
    "permis_travail": "permis / autorisation de travail (ex: Carte de séjour, UE)",
    "demarches_administratives": "détails des démarches administratives en cours",
    "disponibilite": "date ou délai de disponibilité (ex: Immédiate, 3 mois)",
    "domicile": "ville de domicile du candidat",
    "mobilite": "zones de mobilité, séparées par des virgules, UNIQUEMENT parmi : {zones_list}",
    "fonctions_recherchees": "postes et secteurs visés",
    "motif_recherche": "motif de départ / motivations principales",
    "remuneration_actuelle": "rémunération actuelle (fixe + variable + avantages)",
    "pretentions_salariales": "prétentions salariales",
    "propal_a": "propositions / offres déjà reçues ailleurs",
    "avancement_recherches": "avancement des autres pistes (ED/EP/Std By/discret)",
    "eval_technique": "évaluation technique (1-2 phrases)",
    "eval_personnalite": "évaluation personnalité (1-2 phrases)",
    "eval_communication": "évaluation communication (1-2 phrases)",
    "langues": "langues et niveaux (ex: Anglais B2 testé)",
    "references_candidat": "références transmises (nom, fonction, société, contact)",
    "avis_perso": "synthèse personnelle / avis du recruteur (3-4 phrases)",
    "entretien_notes": "compte-rendu structuré (notes EC1 : points clés, prochaines étapes)"
  }},
  "checklist": {{
    "mobilite_dispo_souhaits": {{"checked": true|false, "note": ""}},
    "impression_generale": {{"checked": true|false, "note": ""}},
    "evaluation_technique": {{"checked": true|false, "note": ""}},
    "evaluation_personnalite": {{"checked": true|false, "note": ""}},
    "evaluation_communication": {{"checked": true|false, "note": ""}},
    "rappel_valeurs_up": {{"checked": true|false, "note": ""}},
    "fourchette_salaire": {{"checked": true|false, "note": ""}},
    "reponse_questions_craintes": {{"checked": true|false, "note": ""}},
    "process_prochaines_etapes": {{"checked": true|false, "note": ""}}
  }}
}}

Mets les champs absents à la chaîne vide. Pour permis_conduire et vehicule, réponds exactement « Oui » ou « Non » (ou chaîne vide). Pour mobilite, n'utilise que les libellés de la liste fournie. Pour la checklist, mets `checked` à true si le sujet a été abordé pendant l'entretien."""

    try:
        config = _load_ai_config()
        timeout = int(config.get("ollama_timeout", 120))
        raw = _call_ollama_direct(prompt, config, timeout)
        clean = raw.strip()
        if clean.startswith("```"):
            clean = re.sub(r'^```[^\n]*\n?', '', clean)
            clean = re.sub(r'\n?```$', '', clean)
        json_match = re.search(r'\{[\s\S]*\}', clean)
        if not json_match:
            return jsonify(ok=False, error="L'IA n'a pas retourné de JSON valide", raw=raw[:500]), 422
        parsed = json.loads(json_match.group(0))
        fields = parsed.get("fields") or {}
        checklist = parsed.get("checklist") or {}
    except json.JSONDecodeError as e:
        return jsonify(ok=False, error=f"JSON invalide : {e}"), 422
    except Exception as e:
        logger.warning("EC1 from-transcript IA error: %s", e)
        return jsonify(ok=False, error=f"IA indisponible : {e}"), 503

    c, ec1_data, interview_at, user = _ec1_load_context(cid, uid)
    if c is None:
        return jsonify(ok=False, error="Candidat introuvable"), 404
    candidate_snapshot = _ec1_data_from_db(c, ec1_data, user, interview_at)
    meta = {
        "recruteur_trigramme": _ec1_recruteur_trigramme(user),
        "ec1_date": candidate_snapshot.get("ec1_date") or _today_iso()[:10],
    }
    candidate_snapshot.pop("__checklist", None)
    candidate_snapshot.pop("__free_note", None)
    return jsonify(
        ok=True, fields=fields, checklist=checklist,
        candidate=candidate_snapshot, meta=meta,
    )
